"""Parse the emailed financial workbooks.

Three layouts are auto-detected per file and merged into one result, so the
user can drop all of them in the same upload folder:

  - **FINANCIAL SUMMARY** — the standard: many offices stacked on one sheet,
    office header in column B '(OWNER-STATE)', metric label in column C, the
    4 week-ending columns D-G.
  - **MONTHLY REPORT** — one office, a sheet per ~6-week block; weeks run
    across columns, metric labels down column A (e.g. German Lopez's file).
  - **CASH FLOW (per-week)** — one office, one sheet per week; each sheet a
    cash-flow + P&L statement (e.g. Coel Reif's file).

Every parser emits the same shape — {normalized owner: office} where each
office is {"office", "owner", "metrics": {METRIC: {week_date: value}}}. The
metric values are date-keyed so the focus-report fill can slot each into the
right week column regardless of how the source file laid its weeks out.
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# Focus-sheet row label  ->  the canonical metric key.
# Aptel / Arcadia Consulting are only written where the tab has that row.
# The 'PNL -Profit %' row is intentionally NOT mapped — no source yet (skip).
FINANCIAL_METRICS: Dict[str, str] = {
    "Total Funds Available": "TOTAL FUNDS AVAILABLE",
    "Owners Payroll":        "OWNERS PAYROLL",
    "Total Expenses":        "TOTAL EXPENSES",
    "Indeed":                "INDEED",
    "Aptel":                 "APTEL",
    "Arcadia Consulting":    "ARCADIA CONSULTING",
    "Owners Withdrawal":     "OWNERS WITHDRAWAL",
    "Profit/Loss":           "PROFIT/LOSS",
    "Operating %":           "OPERATING PERCENTAGE %",
}

# Bespoke-format row label (normalized via _lbl) -> canonical metric key.
# Only the metrics Megan confirmed are mapped; anything left out (Operating %,
# and for Coel also Aptel/Arcadia/Owners Withdrawal) is simply never
# populated, so the fill doesn't write it. Coel's Indeed spend is NOT a labeled
# row — it's summed from the P&L detail transactions (see _coel_indeed_total).
_GERMAN_METRICS: Dict[str, str] = {
    "current bank balance total": "TOTAL FUNDS AVAILABLE",
    "owner payroll":              "OWNERS PAYROLL",
    "total costs for week":       "TOTAL EXPENSES",
    "ksw/aptel":                  "APTEL",
    "recruiter - arcadia/nlr":    "ARCADIA CONSULTING",
    "owners draw/atm":            "OWNERS WITHDRAWAL",
    "bottom line profit/loss":    "PROFIT/LOSS",
}
_COEL_METRICS: Dict[str, str] = {
    "local balance":  "TOTAL FUNDS AVAILABLE",
    "owner payroll":  "OWNERS PAYROLL",
    "total expense":  "TOTAL EXPENSES",
    "bl profit/loss": "PROFIT/LOSS",
}

_NAME_SUFFIXES = {"jr", "sr", "ii", "iii", "iv"}
# Coel-style sheet name: a per-week tab like '5.9.26'.
_DATE_SHEET = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})\s*$")


def norm_name(s) -> str:
    """Normalize a person name for matching — lowercase, drop punctuation,
    hyphens to spaces, and a trailing generational suffix (Jr / III / ...)."""
    s = str(s or "").lower().replace("'", "").replace(".", "").replace("-", " ")
    return " ".join(w for w in s.split() if w not in _NAME_SUFFIXES)


def _lbl(s) -> str:
    """Normalize a row label for matching — lowercase, trim, collapse
    whitespace, drop apostrophes (so 'OWNER'S DRAW' == 'owners draw')."""
    s = str(s or "").strip().lower().replace("'", "").replace("’", "")
    return re.sub(r"\s+", " ", s)


def _owner_from_office(header: str) -> str:
    """'ABYL ACQUISITION GROUP, INC. (ABEL DRAPER-CT)' -> 'ABEL DRAPER'.
    Takes the trailing parenthetical and strips a 2-letter state code."""
    m = re.search(r"\(([^)]+)\)\s*$", header or "")
    inside = (m.group(1) if m else "").strip()
    return re.sub(r"\s*-\s*[A-Za-z]{2}$", "", inside).strip()


def _state_from_office(header: str) -> str:
    """The 2-letter state code on the office header, or '' if none.
    '...(RYAN MCSPADDEN-CO)' -> 'CO'. Some owners run more than one
    location (Ryan: CO + TX) — the state distinguishes their office
    blocks so each lands in its own row-set on the tab."""
    m = re.search(r"\(([^)]+)\)\s*$", header or "")
    inside = (m.group(1) if m else "")
    sm = re.search(r"-\s*([A-Za-z]{2})\s*$", inside)
    return sm.group(1).upper() if sm else ""


def _detect_format(wb) -> str:
    """'german' | 'coel' | 'summary' — by a signature cell in column A."""
    for name in wb.sheetnames:
        a4 = _lbl(wb[name]["A4"].value)
        if "monthly report" in a4:
            return "german"
        if "statement of cash flows" in a4:
            return "coel"
    return "summary"


def _num(v):
    """Normalize a summary value cell so both layouts store the same type:
    '90,036.50' -> 90036.50, '$1,200' -> 1200, '49.33%' -> 0.4933 (fraction,
    matching the old numeric % like 0.6125). Numbers pass through; unparseable
    text is returned as-is."""
    if v is None or isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1].replace(",", "").strip()) / 100.0
        except ValueError:
            return v
    try:
        return float(s.replace(",", "").replace("$", "").strip())
    except ValueError:
        return v


def _date_cell(v) -> Optional[dt.date]:
    """A date-header cell -> date, from a datetime OR an 'MM/DD/YYYY' string."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return dt.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _date_header(cells) -> Tuple[Optional[List[dt.date]], Optional[int]]:
    """Longest contiguous run (>=2) of date cells in a row -> (dates, start col
    index). Handles the old layout (datetimes in D-F) and the new one
    (MM/DD/YYYY strings in B-E)."""
    parsed = [_date_cell(c) for c in cells]
    best_start, best_len, i = None, 0, 0
    while i < len(parsed):
        if parsed[i] is not None:
            j = i
            while j < len(parsed) and parsed[j] is not None:
                j += 1
            if j - i > best_len:
                best_start, best_len = i, j - i
            i = j
        else:
            i += 1
    if best_len >= 2:
        return parsed[best_start:best_start + best_len], best_start
    return None, None


def _parse_summary(wb) -> Tuple[List[dict], List[dt.date]]:
    """FINANCIAL SUMMARY layout — offices stacked on the first sheet. Handles
    BOTH hubtruth layouts by locating the value columns dynamically off the
    date-header row: the old one (office header col B, metric col C, datetime
    headers/values in D+) and the new 'ORG FINANCIAL SUMMARY' one (office +
    metric both col A, 'MM/DD/YYYY' string headers/values shifted to B+)."""
    ws = wb[wb.sheetnames[0]]
    weeks: List[dt.date] = []
    vcol: Optional[int] = None       # 0-based col where the week values start
    offices: List[dict] = []
    cur: Optional[dict] = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cells = [c.value for c in row]
        if vcol is None:
            dates, start = _date_header(cells)
            if dates:
                weeks, vcol = dates, start
            continue
        label_region = cells[:vcol]
        vals = [(_num(cells[vcol + i]) if vcol + i < len(cells) else None)
                for i in range(len(weeks))]
        # office header: a '(OWNER-STATE)' cell in the label region
        office = next((str(v).strip() for v in label_region
                       if v and "(" in str(v) and ")" in str(v)), None)
        if office:
            cur = {"office": office,
                   "owner": _owner_from_office(office),
                   "state": _state_from_office(office),
                   "metrics": {}}
            offices.append(cur)
            continue
        # metric row: the last non-empty label cell before the values
        label = next((str(v).strip() for v in reversed(label_region)
                      if v and str(v).strip()), None)
        if label and cur is not None:
            cur["metrics"][label.upper()] = {
                weeks[i]: vals[i] for i in range(len(weeks)) if vals[i] is not None}
    _resign_profit_loss(offices)
    return offices, weeks


def _resign_profit_loss(offices: List[dict]) -> None:
    """Restore the +/- on PROFIT/LOSS from the movement in TOTAL FUNDS
    AVAILABLE. In this (hubtruth summary) format, "PROFIT/LOSS" equals the
    week-over-week change in TOTAL FUNDS AVAILABLE — verified to the cent on
    ~80% of owner-weeks (the rest are restatements). Amber's ORG template
    reports only the MAGNITUDE (always positive), so a LOSING week reads as a
    gain. Re-sign each week by whether the owner's funds rose (profit, +) or
    fell (loss, -), keeping the file's reported magnitude. The oldest in-file
    week is left as-is (no prior week here to diff) — it gets corrected on the
    next weekly file. Scoped to summary format only: the German/Coel templates
    have their own parsers and are already correctly signed, so they never
    reach here."""
    for off in offices:
        tfa = off["metrics"].get("TOTAL FUNDS AVAILABLE")
        pl = off["metrics"].get("PROFIT/LOSS")
        if not tfa or not pl:
            continue
        common = sorted(w for w in pl if w in tfa)
        for i in range(1, len(common)):
            w, wp = common[i], common[i - 1]
            if not isinstance(pl[w], (int, float)):
                continue
            dtfa = tfa[w] - tfa[wp]
            pl[w] = -abs(pl[w]) if dtfa < 0 else abs(pl[w])
        # The OLDEST in-file week can't be signed — its prior week (needed for
        # the funds delta) isn't in this file. And because a given week sits in
        # ~4 successive weekly files and the fill re-writes its column each
        # time, the LAST write to that column is the run where the week is
        # oldest — which would stamp the unsigned (magnitude-only) value and
        # silently undo the correction. So DON'T write P/L for the oldest week:
        # drop it, leaving the correctly-signed value from the run where this
        # week was newer. (Only P/L is dropped; the week's other metrics still
        # fill. Needs >=2 weeks so a signed week remains.)
        if len(common) >= 2:
            pl.pop(common[0], None)


def _parse_german(wb) -> Tuple[List[dict], List[dt.date]]:
    """MONTHLY REPORT layout — one office, weeks across columns, metric
    labels down column A. Parses every sheet (one per ~6-week block) and
    merges; the latest sheet wins a boundary week."""
    owner = office = None
    metrics: Dict[str, Dict[dt.date, object]] = {}
    weeks_seen: set = set()
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(min_row=1, max_row=wb[name].max_row))
        for r in rows[:4]:
            a = _lbl(r[0].value)
            if a.startswith("company name") and len(r) > 1:
                owner = owner or str(r[1].value or "").strip()
            elif a.startswith("owner") and "name" in a and len(r) > 1:
                office = office or str(r[1].value or "").strip()
        # The DATE row — first row with 4+ datetimes in columns B-G.
        date_cols: Dict[int, dt.date] = {}
        for r in rows:
            cand = {i: r[i].value for i in range(1, 8)
                    if i < len(r) and isinstance(r[i].value, dt.datetime)}
            if len(cand) >= 4:
                date_cols = {i: v.date() for i, v in cand.items()}
                break
        if not date_cols:
            continue
        weeks_seen.update(date_cols.values())
        for r in rows:
            canon = _GERMAN_METRICS.get(_lbl(r[0].value)) if r else None
            if not canon:
                continue
            m = metrics.setdefault(canon, {})
            for ci, d in date_cols.items():
                v = r[ci].value if ci < len(r) else None
                if isinstance(v, (int, float)):
                    m.setdefault(d, v)            # latest sheet wins
    if not owner:
        return [], []
    return ([{"office": office or owner, "owner": owner, "metrics": metrics}],
            sorted(weeks_seen))


def _coel_indeed_total(ws) -> Optional[float]:
    """Coel's cash-flow file has no 'Indeed' line — his Indeed spend lives as
    one or more checks inside the P&L DETAIL block (category '905 Advertising'),
    each a transaction row whose Name column reads 'Indeed'. Find the detail
    header row (the one carrying BOTH a 'Name' and an 'Amount' label — no
    hardcoded columns, the template can move them), then sum every transaction
    amount whose Name is Indeed. Returns None when there's no detail table, so
    a week without the section never writes a spurious 0. Jobs2Me / ARS
    Recruiting / Aptel checks sit in the same block but are deliberately left
    out — they belong to other metrics, not Indeed."""
    name_col = amt_col = None
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        labels = {_lbl(c.value): c.column - 1 for c in r if c.value is not None}
        if "name" in labels and "amount" in labels:
            name_col, amt_col = labels["name"], labels["amount"]
            break
    if name_col is None:
        return None
    total: Optional[float] = None
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if name_col >= len(r) or amt_col >= len(r):
            continue
        nm, amt = r[name_col].value, r[amt_col].value
        if nm and "indeed" in _lbl(nm) and isinstance(amt, (int, float)):
            total = (total or 0) + amt
    return total


def _parse_coel(wb) -> Tuple[List[dict], List[dt.date]]:
    """CASH FLOW layout — one office, one sheet per week. Each metric label
    is found anywhere on the sheet; its value is the first number to the
    right on the same row. Indeed is the exception — summed from the P&L
    detail transactions (_coel_indeed_total), since it has no labeled row."""
    owner = office = None
    metrics: Dict[str, Dict[dt.date, object]] = {}
    weeks_seen: set = set()
    for name in wb.sheetnames:
        m = _DATE_SHEET.match(name)
        if not m:
            continue
        mon, day, yr = (int(x) for x in m.groups())
        try:
            week = dt.date(yr + 2000 if yr < 100 else yr, mon, day)
        except ValueError:
            continue
        weeks_seen.add(week)
        ws = wb[name]
        if owner is None:
            owner = str(ws["A2"].value or "").strip()
            office = str(ws["A1"].value or "").strip()
        for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
            for ci, cell in enumerate(r):
                canon = _COEL_METRICS.get(_lbl(cell.value))
                if not canon:
                    continue
                val = next((r[k].value for k in range(ci + 1, len(r))
                            if isinstance(r[k].value, (int, float))), None)
                if val is not None:
                    metrics.setdefault(canon, {}).setdefault(week, val)
        indeed = _coel_indeed_total(ws)
        if indeed is not None:
            metrics.setdefault("INDEED", {}).setdefault(week, indeed)
    if not owner:
        return [], []
    return ([{"office": office or owner, "owner": owner, "metrics": metrics}],
            sorted(weeks_seen))


def parse_one_file(path) -> Tuple[List[dict], List[dt.date], str]:
    """Parse one workbook — detects its layout and dispatches. Returns
    (offices, weeks, format-name)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    fmt = _detect_format(wb)
    offices, weeks = {"german": _parse_german,
                      "coel": _parse_coel}.get(fmt, _parse_summary)(wb)
    return offices, weeks, fmt


def parse_financial_files(paths, logfn=print) -> Tuple[Dict[str, List[dict]], List[dt.date], List[Tuple[str, str]]]:
    """Parse + merge any number of financial workbooks — any mix of the three
    layouts — into one {normalized owner: [offices]} map. A duplicate
    (owner, state) lets the later file win; a multi-location owner keeps one
    office per state (Ryan: CO + TX). Returns (by_owner, weeks, problems).

    `weeks` is the reporting window the fill writes into: the FINANCIAL
    SUMMARY week-endings when any summary file is present (Eve's checked
    weeks), otherwise the 4 most recent dates seen.

    `problems` is a list of (filename, reason) tuples for files that didn't
    yield usable data — surfaced to the Hub so an unknown / corrupted
    template doesn't silently slip through. Cases:
      - 'can't open' — workbook is corrupted, locked, or wrong file type.
      - '0 offices parsed' — file opened but our parser found no office rows
        (often signals a brand-new template layout the auto-detector
        misclassified as 'summary'; needs a new layout added)."""
    by_owner: Dict[str, dict] = {}
    summary_weeks: List[dt.date] = []
    all_weeks: set = set()
    problems: List[Tuple[str, str]] = []
    for p in paths:
        name = Path(p).name
        if name.startswith("~$"):
            continue                       # Excel lock/temp file
        try:
            offices, wk, fmt = parse_one_file(p)
        except Exception as e:
            msg = f"can't open ({type(e).__name__}: {str(e)[:80]})"
            logfn(f"financial: ❌ SKIPPED {name} — {msg}")
            problems.append((name, msg))
            continue
        if not offices:
            msg = (f"0 offices parsed (detected as {fmt!r} — likely a "
                   f"new template layout)")
            logfn(f"financial: ⚠️  {name} — {msg}")
            problems.append((name, msg))
            continue
        logfn(f"financial: {name} — {fmt} format, {len(offices)} office(s)")
        all_weeks.update(wk)
        if fmt == "summary" and wk:
            summary_weeks = wk
        for o in offices:
            key = norm_name(o["owner"])
            if not key:
                continue
            # Dedup by (owner, state) so a re-uploaded file's later copy wins,
            # but a multi-location owner (Ryan: CO + TX) keeps BOTH offices.
            st = o.get("state", "")
            bucket = by_owner.setdefault(key, [])
            existing = next((i for i, x in enumerate(bucket)
                             if x.get("state", "") == st), None)
            if existing is None:
                bucket.append(o)
            else:
                bucket[existing] = o            # later file wins
    weeks = summary_weeks or sorted(all_weeks)[-4:]
    return by_owner, weeks, problems
