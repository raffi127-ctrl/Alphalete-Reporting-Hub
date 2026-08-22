"""Board helper for Shikamaru's Slack commission review (Carlos, 2026-07-23).

Shikamaru posts the week's commission summary + workbook to Carlos on Slack;
Carlos replies with changes in plain English; Shikamaru parses them, gets a
button-confirm, then shells THIS helper to apply. All sheet logic lives here,
under the recruiting-report auth (same split as promo_checkin.helper).

  .venv/bin/python -m automations.payroll_review.helper summary
      -> {"week", "locked", "reps": [{"name","brought","paid"}...], "totals"}
  .venv/bin/python -m automations.payroll_review.helper xlsx /path/out.xlsx
      -> {"path", "reps", "week"}   (Summary + Payout Detail sheets)
  .venv/bin/python -m automations.payroll_review.helper png /path/out.png
      -> {"path", "week", "locked"}  screenshot of the Commission tab's
         left-side boxes (payout/P&L table, ACTIVE REPS - NO REVENUE,
         REVENUE - NOT ACTIVE, first/second-week paycheck counts) rendered
         via the Sheets PDF export -> pypdfium2 at high DPI. Carlos prefers
         this picture over a text summary in Slack (2026-07-23).
  .venv/bin/python -m automations.payroll_review.helper apply '<json>'
      -> applies [{"rep","amount","label","type"}] ATOMICALLY:
         * every rep must resolve EXACTLY (Commission roster + Name Aliases,
           prefix rule); any ambiguity/miss -> nothing applied, candidates
           returned (Carlos's if-unsure-ask-me rule)
         * type "Bonus" (default; amount may be negative) or "NOPAY"
           (-> REMOVEALL row)
         * labels written RAW (a numeric-looking label like "0" silently
           breaks REMOVELINE parsing — learned 2026-07-23)
         * refuses if Commission F1 says the current week is LOCKED
         * triggers the web-app refresh, returns before/after payouts
"""
from __future__ import annotations

import json
import sys
import time
import unicodedata

SHEET_ID = "1Hltk25zTudsaoYJFKvKqWlpT_4MF5_ZZq734XKVCJKY"


def _nrm(s) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def _sheet():
    from automations.recruiting_report.fill import open_by_key
    return open_by_key(SHEET_ID)


def _read_summary(sh):
    cm = sh.worksheet("Commission")
    week = str(cm.acell("B1").value or "").strip()
    locked_note = str(cm.acell("F1").value or "").strip()
    locked = week and week in locked_note and "LOCK" in locked_note.upper()
    reps, total = [], None
    for r in cm.get("A4:C120"):
        name = str(r[0]).strip() if r else ""
        if not name:
            continue
        if name == "TOTAL":
            total = {"brought": str(r[1]).strip() if len(r) > 1 else "",
                     "paid": str(r[2]).strip() if len(r) > 2 else ""}
            break
        reps.append({"name": name,
                     "brought": str(r[1]).strip() if len(r) > 1 else "",
                     "paid": str(r[2]).strip() if len(r) > 2 else ""})
    return cm, week, bool(locked), reps, total


def summary() -> dict:
    sh = _sheet()
    _cm, week, locked, reps, total = _read_summary(sh)
    return {"week": week, "locked": locked, "reps": reps, "totals": total}


def xlsx(path: str) -> dict:
    import openpyxl
    from openpyxl.styles import Font
    sh = _sheet()
    cm, week, _locked, reps, total = _read_summary(sh)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"Commission — week ending {week}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Rep", "Brought In", "Commission"])
    for c in ws[2]:
        c.font = Font(bold=True)
    for r in reps:
        ws.append([r["name"], r["brought"], r["paid"]])
    if total:
        ws.append(["TOTAL", total["brought"], total["paid"]])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    det = wb.create_sheet("Payout Detail")
    for row in cm.get("G3:N600"):
        vals = [str(c).strip() for c in (row + [""] * 8)[:8]]
        if any(vals):
            det.append(vals)
    det.column_dimensions["A"].width = 28
    det.column_dimensions["D"].width = 30
    wb.save(path)
    return {"path": path, "reps": len(reps), "week": week}


def png(path: str) -> dict:
    """Screenshot of the Commission tab's left-side boxes (cols A:E — the
    per-rep payout/P&L table with TOTAL, the no-revenue / not-active boxes,
    and the first/second-week paycheck counts). The rep payout sheets live
    in G:N and are deliberately excluded. Renders the sheet's own PDF export
    so it looks exactly like the tab (fills, borders, fonts)."""
    import io

    import pypdfium2 as pdfium
    sh = _sheet()
    cm, week, locked, _reps, _total = _read_summary(sh)
    # last non-empty row in A:E, padded a little, so the capture tracks
    # roster growth instead of hardcoding row 80
    last = 44
    for i, row in enumerate(cm.get("A1:E200"), 1):
        if any(str(c).strip() for c in row):
            last = max(last, i)
    url = (f"https://docs.google.com/spreadsheets/d/{sh.id}/export"
           f"?format=pdf&gid={cm.id}&range=A1:E{last + 2}"
           "&portrait=true&fitw=true&gridlines=false&size=letter"
           "&fzr=false"  # don't repeat the frozen row on page 2+
           "&top_margin=0.25&bottom_margin=0.25"
           "&left_margin=0.25&right_margin=0.25")
    from automations.recruiting_report.fill import _client
    r = _client().http_client.request("get", url)
    r.raise_for_status()
    if "pdf" not in str(r.headers.get("content-type", "")):
        return {"error": "export did not return a PDF (auth scope?)"}
    pdf = pdfium.PdfDocument(io.BytesIO(r.content))
    from PIL import Image, ImageChops

    def _trim(im):  # cut white margins top/bottom (keep a little air)
        bbox = ImageChops.invert(im.convert("L")).getbbox()
        if not bbox:
            return None  # fully blank page
        pad = 20
        return im.crop((0, max(0, bbox[1] - pad),
                        im.width, min(im.height, bbox[3] + pad)))

    pages = [p for p in (_trim(pg.render(scale=200 / 72).to_pil())
                         for pg in pdf) if p is not None]
    if len(pages) == 1:
        img = pages[0]
    else:  # stitch trimmed pages vertically (roster outgrew one page)
        img = Image.new("RGB", (max(p.width for p in pages),
                                sum(p.height for p in pages)), "white")
        y = 0
        for p in pages:
            img.paste(p, (0, y))
            y += p.height
    img.save(path, "PNG")
    return {"path": path, "week": week, "locked": locked,
            "rows": last, "pages": len(pdf),
            "size": [img.width, img.height]}


def _webapp_refresh() -> str:
    import pathlib

    import requests
    cfg = (pathlib.Path(__file__).resolve().parents[2]
           / "vantura-payroll-webapp.json")
    try:
        url = json.loads(cfg.read_text()).get("webapp_url", "")
    except Exception:  # noqa: BLE001
        url = ""
    if not url:
        return "refresh skipped (no web-app config)"
    r = requests.get(url, params={"action": "refresh"}, timeout=600)
    r.raise_for_status()
    return f"refresh: {r.text[:80]}"


def apply(payload: str) -> dict:
    items = json.loads(payload)
    if not isinstance(items, list) or not items:
        return {"error": "apply needs a non-empty JSON list"}
    sh = _sheet()
    _cm, week, locked, reps, _total = _read_summary(sh)
    if locked:
        return {"error": f"week {week} is LOCKED — no changes possible"}

    roster = {_nrm(r["name"]): r["name"] for r in reps}
    # Name Aliases: alias -> canonical
    try:
        for row in sh.worksheet("Name Aliases").get_all_values():
            if len(row) > 1 and str(row[0]).strip() and str(row[1]).strip():
                canon = _nrm(row[1])
                if canon in roster:
                    roster.setdefault(_nrm(row[0]), roster[canon])
    except Exception:  # noqa: BLE001
        pass

    def resolve(name):
        n = _nrm(name)
        if n in roster:
            return roster[n]
        hits = {v for k, v in roster.items()
                if k.startswith(n + " ") or n.startswith(k + " ") or n in k}
        return hits.pop() if len(hits) == 1 else sorted(hits)

    resolved, problems = [], []
    for it in items:
        who = resolve(it.get("rep", ""))
        if isinstance(who, str):
            resolved.append((who, it))
        else:
            problems.append({"rep": it.get("rep"), "candidates": who})
    if problems:
        return {"error": "ambiguous/unknown rep(s) — NOTHING applied",
                "problems": problems}

    before = {r["name"]: r["paid"] for r in reps}
    adj = sh.worksheet("Adjustments")
    rows = []
    for who, it in resolved:
        typ = str(it.get("type", "Bonus")).strip() or "Bonus"
        if typ.upper() == "NOPAY":
            rows.append([week, who, "REMOVEALL", "",
                         str(it.get("label", "no pay (Slack review)"))])
        else:
            rows.append([week, who, "Bonus", float(it.get("amount", 0)),
                         str(it.get("label", "Bonus (Slack review)"))])
    adj.append_rows(rows, value_input_option="RAW")
    note = _webapp_refresh()
    time.sleep(6)
    _cm2, _w2, _l2, reps2, total2 = _read_summary(sh)
    after = {r["name"]: r["paid"] for r in reps2}
    touched = sorted({who for who, _ in resolved})
    return {"applied": len(rows), "refresh": note, "week": week,
            "changes": [{"rep": w, "before": before.get(w, "(not listed)"),
                         "after": after.get(w, "(removed — no pay)")}
                        for w in touched],
            "new_total": total2}


def main() -> int:
    mode = sys.argv[1] if len(sys.argv) > 1 else "summary"
    if mode == "summary":
        print(json.dumps(summary()))
    elif mode == "xlsx":
        print(json.dumps(xlsx(sys.argv[2])))
    elif mode == "png":
        print(json.dumps(png(sys.argv[2])))
    elif mode == "apply":
        print(json.dumps(apply(sys.argv[2])))
    else:
        print(json.dumps({"error": f"unknown mode {mode}"}))
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
