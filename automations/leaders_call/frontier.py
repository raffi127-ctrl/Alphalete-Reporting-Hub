"""Frontier section for the Leader's Call — from the emailed
"Sales Verification - Frontier …" workbook (NOT Tableau).

Flow (preflight upload, #16): before the run, Maud uploads the Frontier
"Sales Verification" xlsx (or csv) via the Hub card. It lands in UPLOAD_DIR and
this module sums each agent's sales for the just-completed Mon–Sun week and
returns the Leader's Call Frontier rows: [(rep, owner, apps)] for reps with 8+,
owner = the Frontier ICD (Abel Draper).

Week realignment: the "Sales by Week" sheet has a Current-week block and a
Prior-week block, each with DATED daily columns, and Frontier's Sun/Mon column
labels are offset a day from the calendar. So we ignore the labels and sum each
agent's daily values by ACTUAL DATE for the 7 calendar dates of the target
week (which straddles both blocks) — the robust version of the loom's manual
"add prior-week Monday onto current-week Sunday".
"""
from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path

UPLOAD_DIR = (Path(__file__).resolve().parent.parent
              / "uploaded" / "leaders_call_frontier")
FRONTIER_OWNER = "Abel Draper"
THRESHOLD = 8
SALES_SHEET = "Sales by Week"


def _newest_upload() -> Path:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    files = [p for p in UPLOAD_DIR.iterdir()
             if p.suffix.lower() in (".xlsx", ".xls", ".csv")]
    if not files:
        raise FileNotFoundError(
            f"No Frontier upload in {UPLOAD_DIR}. Upload the Frontier 'Sales "
            "Verification' workbook via the Hub card's preflight step first.")
    return max(files, key=lambda p: p.stat().st_mtime)


def _target_week_dates(today: dt.date | None = None) -> set:
    """The 7 calendar dates Mon..Sun of the just-completed week (the WE-Sunday
    on/before today — same target every campaign uses)."""
    from automations.alphalete_org_report.opt_nds import _current_target_week_end
    sun = _current_target_week_end(today)
    return {sun - dt.timedelta(days=i) for i in range(7)}


def parse_file(path: Path, today: dt.date | None = None) -> list[tuple]:
    """Sum each agent's sales over the target week's dates from the
    'Sales by Week' sheet. Returns [(rep, owner, apps)] for apps >= 8 desc."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheet = SALES_SHEET if SALES_SHEET in wb.sheetnames else wb.sheetnames[0]
    grid = list(wb[sheet].iter_rows(values_only=True))

    agent_by_date: dict[str, dict] = {}
    col_dates: dict[int, dt.date] = {}
    SKIP = ("sales agent name", "grand total", "week ending",
            "current sales week", "prior sales week", "daily sales")
    for row in grid:
        cells = list(row)
        dts = {i: c.date() for i, c in enumerate(cells)
               if isinstance(c, dt.datetime)}
        if len(dts) >= 3:                      # a dated daily-column header row
            col_dates = dts
            continue
        name = cells[0].strip() if cells and isinstance(cells[0], str) else ""
        if not name or not col_dates or name.lower() in SKIP \
                or any(name.lower().startswith(s) for s in SKIP):
            continue
        for i, d in col_dates.items():
            v = cells[i] if i < len(cells) else None
            if isinstance(v, (int, float)):
                agent_by_date.setdefault(name, {})
                agent_by_date[name][d] = agent_by_date[name].get(d, 0) + v

    week = _target_week_dates(today)
    out = []
    for name, dd in agent_by_date.items():
        apps = int(sum(v for d, v in dd.items() if d in week))
        if apps >= THRESHOLD:
            out.append((name, FRONTIER_OWNER, apps))
    out.sort(key=lambda t: t[2], reverse=True)
    return out


def _parse_csv(path: Path, today: dt.date | None = None) -> list[tuple]:
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        rows = list(csv.reader(f))
    # Fallback for a flat rep|sales csv export.
    if not rows:
        return []
    hdr = [(c or "").strip().lower() for c in rows[0]]
    rep_i = next((i for i, h in enumerate(hdr) if "agent" in h or "rep" in h
                  or "name" in h), 0)
    val_i = next((i for i, h in enumerate(hdr) if "sale" in h or "total" in h
                  or "app" in h), len(hdr) - 1)
    out = []
    for r in rows[1:]:
        rep = (r[rep_i] if rep_i < len(r) else "").strip()
        try:
            apps = int(float(re.sub(r"[,\s]", "", r[val_i])))
        except (ValueError, IndexError):
            continue
        if rep and apps >= THRESHOLD:
            out.append((rep, FRONTIER_OWNER, apps))
    out.sort(key=lambda t: t[2], reverse=True)
    return out


def parse_uploaded(today: dt.date | None = None) -> list[tuple]:
    """Parse the newest Frontier upload -> [(rep, owner, apps)] (apps >= 8).
    Raises FileNotFoundError if nothing uploaded (run_all treats that as
    'Frontier left as-is')."""
    path = _newest_upload()
    if path.suffix.lower() == ".csv":
        return _parse_csv(path, today)
    return parse_file(path, today)
