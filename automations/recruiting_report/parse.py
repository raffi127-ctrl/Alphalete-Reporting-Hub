"""Extract recruiting funnel counts from an ApplicantStream .xlsx export.

Final field positions are confirmed against a real export during Phase 1.
For now this matches by row label, which is robust to column drift but
requires the export to use these exact label strings. If the labels differ,
update LABEL_ALIASES.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

# Funnel metrics we extract from each office's export. Keys are the canonical
# names used everywhere downstream; values are the label strings (lowercased,
# stripped) that may appear in the export's label column.
LABEL_ALIASES: Dict[str, List[str]] = {
    "pull":                 ["apps / pull", "apps/pull", "pull", "applies"],
    "first_booked":         ["1st booked", "1st rds booked", "first round booked"],
    "first_showed":         ["1st showed", "1st rds showed", "first round showed"],
    "second_booked":        ["2nd booked", "2nd rds booked", "second round booked"],
    "second_showed":        ["2nd showed", "2nd rds showed", "second round showed"],
    "job_offered":          ["job offered", "offer extended"],
    "bob":                  ["bob", "back of bus", "back-of-bus"],
    "new_starts_scheduled": ["new starts scheduled", "new start scheduled"],
    "new_starts_showed":    ["new starts showed", "new start showed"],
}

# Reverse lookup: alias -> canonical
_ALIAS_TO_CANON = {alias: canon for canon, aliases in LABEL_ALIASES.items() for alias in aliases}


def parse_xlsx(path: Path) -> Dict[str, Optional[int]]:
    """Return {canonical_metric: count}. None for any metric not found."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    counts: Dict[str, Optional[int]] = {k: None for k in LABEL_ALIASES}

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label = (str(row[0]) if row[0] is not None else "").strip().lower()
        if label in _ALIAS_TO_CANON:
            value = next((c for c in row[1:] if isinstance(c, (int, float))), None)
            counts[_ALIAS_TO_CANON[label]] = int(value) if value is not None else None

    wb.close()
    return counts


def main() -> int:
    ap = argparse.ArgumentParser(description="Parse one ApplicantStream .xlsx export.")
    ap.add_argument("path", type=Path, help="Path to the .xlsx file.")
    args = ap.parse_args()

    if not args.path.exists():
        print(f"File not found: {args.path}", file=sys.stderr)
        return 1

    counts = parse_xlsx(args.path)
    print(json.dumps(counts, indent=2))

    missing = [k for k, v in counts.items() if v is None]
    if missing:
        print(f"\nWARNING: missing metrics: {missing}", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
