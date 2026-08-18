"""The daily BOX Order Log workbook: one summary tab + one tab per rep.

The Fiber counterpart of this is `order_log._append_rep_breakdown_tabs`, and
this follows its conventions deliberately — Georgia 12pt, a dark header band,
colored week banners, one tab per rep (not per rep-week) so a rep opens their
single tab and sees everything.

This is the DAILY artifact and is separate from the rolling six-week sheet
(see sheet.py). It covers the full pull, not just the six-week window.
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import clean

# Megan's house look for report xlsx — matches the Fiber order log exactly.
FONT_NAME = "Georgia"
FONT_SIZE = 12
WIDTH_FACTOR = 1.3
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

HEADER_BG = "434343"
WEEK_BG = "2563EB"
REP_BG = "EDEDED"         # the per-rep bands inside the Pending Orders tab

COLUMNS = (
    "Sale Date", "Business Name", "Contract ID", "Status", "Contr. Sub-status",
    "Secondary Status", "Accepted Date", "BF Tier", "Term", "Complete Sales",
    "Sales (All) kWH+Therms",
)
# The summary tab names the rep; the per-rep tabs don't need to repeat it.
SUMMARY_COLUMNS = ("Rep Name",) + COLUMNS

# Excel forbids : \ / ? * [ ] in sheet titles, and caps them at 31 chars.
_BAD_TITLE = re.compile(r"[:\\/?*\[\]]")


def _font(color: str = "000000", *, bold: bool = False,
          italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, italic=italic,
                color=color)


def _border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_title(name: str, used: set) -> str:
    title = _BAD_TITLE.sub("-", name).strip() or "Rep"
    title = title[:31]
    base, n = title, 2
    while title.lower() in used:
        suffix = " ({})".format(n)
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _fmt(value) -> str:
    return "" if value is None else str(value).strip()


def _cell_value(s, column: str):
    """One cell for a sale, by column label."""
    if column == "Rep Name":
        return _fmt(s.fields.get("Rep Name"))
    if column == "Sale Date":
        return s.sale_date or ""
    if column == "Accepted Date":
        return clean._parse_date(_fmt(s.fields.get("Accepted Date"))) or ""
    if column == "Status":
        return s.status
    if column == "Contr. Sub-status":
        return s.sub_status
    if column == "Secondary Status":
        return s.secondary
    if column == "Sales (All) kWH+Therms":
        raw = _fmt(s.fields.get(column)).replace(",", "")
        try:
            return int(raw)
        except ValueError:
            return raw
    return _fmt(s.fields.get(column))


def _write_header(sh, row: int, columns: Sequence[str]) -> None:
    border = _border()
    for c, label in enumerate(columns, start=1):
        cell = sh.cell(row=row, column=c, value=label)
        cell.font = _font("FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment, cell.border = CENTER, border


def _write_row(sh, row: int, s, columns: Sequence[str]) -> None:
    border = _border()
    # History-aware: a Verification sale reads as "waiting" if it was already
    # submitted and "ours to chase" if it wasn't. Same rule as the sheet.
    fill_hex = clean.color_for(s.status, s.history)
    fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None
    for c, column in enumerate(columns, start=1):
        cell = sh.cell(row=row, column=c, value=_cell_value(s, column))
        cell.font = _font()
        cell.alignment = LEFT if column in ("Business Name", "Rep Name") else CENTER
        cell.border = border
        if fill is not None:
            cell.fill = fill
        if column in ("Sale Date", "Accepted Date"):
            cell.number_format = "mm/dd/yyyy"
        elif column == "Sales (All) kWH+Therms":
            cell.number_format = "#,##0"


def _banner(sh, row: int, text: str, ncol: int) -> None:
    cell = sh.cell(row=row, column=1, value=text)
    cell.font = _font("FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor=WEEK_BG)
    cell.alignment, cell.border = CENTER, _border()
    sh.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)


def _autosize(sh, columns: Sequence[str], rows: Sequence) -> None:
    for c, column in enumerate(columns, start=1):
        widest = len(column)
        for s in rows:
            widest = max(widest, len(str(_cell_value(s, column))))
        sh.column_dimensions[get_column_letter(c)].width = min(
            46, max(10, widest * WIDTH_FACTOR))


def _legend(sh, row: int) -> int:
    """A short color key, so a rep opening their tab knows what red means."""
    sh.cell(row=row, column=1, value="What the colors mean").font = _font(bold=True)
    row += 1
    for status in clean.STATUS_PRIORITY:
        meaning = clean.STATUS_MEANING.get(status)
        if not meaning:
            continue
        swatch = sh.cell(row=row, column=1, value=status)
        hex_v = clean.STATUS_COLORS.get(status)
        if hex_v:
            swatch.fill = PatternFill("solid", fgColor=hex_v)
        swatch.font, swatch.alignment, swatch.border = _font(), CENTER, _border()
        note = sh.cell(row=row, column=2, value=meaning)
        note.font, note.alignment = _font(italic=True), LEFT
        row += 1
    return row


# --- pay-structure estimate (Carlos's office; the 'box' campaign) -----------
# BOX has no pricing of its own (see payout.py), so if Carlos has filled in his
# Pay Structure link we price each accepted deal by its BF Tier x Term rate.
_PAY_CAMPAIGN = "box"
_PAY_SHEET_ID = "1eJ3-BeOvbGaWV5XZ8BNgJT9QrgbaToAf9W2PdMABTAw"   # AUTOMATION MASTER
_MONEY_FMT = '"$"#,##0'
_BOX_PAYS = ("Accepted by Supplier",)    # the status that pays (box_order_log.payout)


def _load_pay_grid():
    """Carlos's box pay structure, or None. Never let it break the log."""
    try:
        import os
        os.environ.setdefault("PAY_STRUCTURE_SHEET_ID", _PAY_SHEET_ID)
        from automations.pay_structure import store as _ps
        if not _ps.estimate_live():          # gated OFF until go-live (PAY_ESTIMATE_LIVE=1)
            return None
        g = _ps.load("carlos")
        if g and _PAY_CAMPAIGN in g.rates:
            return g
    except Exception:
        pass
    return None


def _sale_key(s) -> str:
    """'BF 1 — 36mo' — the exact sale-type key the editor prices against."""
    bf = _fmt(s.fields.get("BF Tier"))
    term = _fmt(s.fields.get("Term"))
    return "{} — {}mo".format(bf, term) if bf and term else ""


def _line_rates(s, grid):
    key = _sale_key(s)
    if not key:
        return [0.0] * len(grid.levels), True
    vals = [grid.rate(_PAY_CAMPAIGN, key, lvl) for lvl in grid.levels]
    return vals, (not any(vals))


def _write_pay_cells(sh, r, s, grid, start_col):
    """Per-level pay for one deal (NYP in red when the office hasn't priced it)."""
    border = _border()
    fill_hex = clean.color_for(s.status, s.history)
    fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None
    vals, unpriced = _line_rates(s, grid)
    for i, _lvl in enumerate(grid.levels):
        cell = sh.cell(row=r, column=start_col + i)
        cell.alignment, cell.border = CENTER, border
        if fill is not None:
            cell.fill = fill
        if unpriced:
            cell.value, cell.font = "NYP", _font("CC0000")
        else:
            cell.value, cell.font = vals[i], _font()
            cell.number_format = _MONEY_FMT


def _write_week_total(sh, r, group, grid, base_ncol, ncol):
    """TOTAL row: estimated pay from the ACCEPTED deals in this week, per level."""
    totals = [0.0] * len(grid.levels)
    for s in group:
        if s.status in _BOX_PAYS:
            vals, unpriced = _line_rates(s, grid)
            if not unpriced:
                for i in range(len(grid.levels)):
                    totals[i] += vals[i]
    accent = (grid.accent or "#1F9D57").lstrip("#") or "1F9D57"
    border = _border()
    label = sh.cell(row=r, column=1,
                    value="\U0001F4B5 TOTAL — est. pay (accepted deals)")
    label.font = _font("333333", bold=True)
    label.fill = PatternFill("solid", fgColor="F2F2F2")
    label.alignment = Alignment(horizontal="right", vertical="center")
    label.border = border
    sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=base_ncol)
    for i in range(len(grid.levels)):
        cell = sh.cell(row=r, column=base_ncol + 1 + i, value=totals[i])
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = _font("FFFFFF", bold=True)
        cell.alignment, cell.border = CENTER, border
        cell.number_format = _MONEY_FMT
    return r + 1


def build(sales: Sequence, out_path: Path, *,
          today: Optional[dt.date] = None) -> Path:
    """Write the workbook: 'All Reps' summary, then one tab per rep."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    today = today or dt.date.today()
    pay_grid = _load_pay_grid()          # None unless Carlos filled in his link

    wb = Workbook()
    used: set = set()

    # ---- summary tab ----------------------------------------------------
    sh = wb.active
    sh.title = _safe_title("All Reps", used)
    ordered = sorted(
        sales,
        key=lambda s: (-(s.week_ending.toordinal() if s.week_ending else 0),
                       _fmt(s.fields.get("Rep Name")),
                       clean._priority(s.level)))
    row = 1
    for week, group in clean.by_week(ordered).items():
        label = week.strftime("%m/%d/%Y") if week else "No sale date"
        _banner(sh, row, "Week Ending {}  •  {} sale{}".format(
            label, len(group), "" if len(group) == 1 else "s"),
            len(SUMMARY_COLUMNS))
        row += 1
        _write_header(sh, row, SUMMARY_COLUMNS)
        row += 1
        for s in group:
            _write_row(sh, row, s, SUMMARY_COLUMNS)
            row += 1
        row += 2
    _autosize(sh, SUMMARY_COLUMNS, ordered)
    sh.freeze_panes = "A3"

    # ---- payout by week --------------------------------------------------
    # Reps down, weeks across, PAID sales in the cells. "Paid" is Accepted by
    # Supplier counted in the week it was accepted, because that is the week
    # it pays (Carlos, 2026-07-18).
    from . import payout as _payout
    reps_ranked, weeks_desc, posted, pending = _payout.by_week_matrix(ordered)
    if reps_ranked:
        psh = wb.create_sheet(_safe_title("Payout by Week", used))
        headers = (["Rep"] + [w.strftime("%m/%d") for w in weeks_desc]
                   + ["Accepted Total", "Still Open"])
        psh.cell(row=1, column=1,
                 value="Sales accepted by the supplier, by week ending"
                 ).font = _font(bold=True)
        psh.cell(row=2, column=1,
                 value="Each column is the week the supplier ACCEPTED the "
                       "sale — not the week it was sold, so these won't match "
                       "the log's week totals. Acceptance is what next week's "
                       "pay is based on. STILL OPEN is not a week figure — "
                       "it's every deal still waiting on acceptance, whenever "
                       "it was sold."
                 ).font = _font(italic=True)
        _write_header(psh, 4, headers)
        r = 5
        for rep in reps_ranked:
            row_vals = [rep]
            paid_total = 0
            for w in weeks_desc:
                n = posted.get((rep, w), 0)
                paid_total += n
                row_vals.append(n)
            row_vals += [paid_total, pending.get(rep, 0)]
            for c, v in enumerate(row_vals, start=1):
                cell = psh.cell(row=r, column=c, value=v)
                cell.font = _font(bold=(c >= len(headers) - 1))
                cell.alignment = LEFT if c == 1 else CENTER
                cell.border = _border()
            r += 1
        # TOTAL strip — COMPUTED VALUES, not formulas. openpyxl writes a
        # formula with no cached result, so anything that doesn't recalculate
        # (Slack's file preview, Quick Look, a Google Sheets import) renders
        # the row blank. Carlos saw exactly that: "the totals aren't showing".
        col_totals = [0] * (len(headers) - 1)
        for rep in reps_ranked:
            vals = [posted.get((rep, w), 0) for w in weeks_desc]
            vals.append(sum(vals))                 # Paid Total
            vals.append(pending.get(rep, 0))       # Pending
            for i, v in enumerate(vals):
                col_totals[i] += v
        for c in range(1, len(headers) + 1):
            v = "TOTAL" if c == 1 else col_totals[c - 2]
            cell = psh.cell(row=r, column=c, value=v)
            cell.font = _font(bold=True)
            cell.alignment = LEFT if c == 1 else CENTER
            cell.border = _border()
            cell.fill = PatternFill("solid", fgColor="EDEDED")
        psh.column_dimensions["A"].width = 30
        for c in range(2, len(headers) + 1):
            psh.column_dimensions[get_column_letter(c)].width = 12
        psh.freeze_panes = "B5"

    # ---- pending orders --------------------------------------------------
    # Every deal still in flight — not yet accepted, not dead — in one tab, so
    # Carlos has a single worklist to chase (his ask, 2026-07-22: "a tab with
    # all of the pending orders like ATT has"). "Pending" is exactly the
    # payout image's "Still Open": status is neither Accepted by Supplier nor a
    # terminal cancel/reject/drop.
    from . import payout as _payout
    pend = [s for s in ordered
            if s.status not in _payout.POSTED_STATUSES
            and s.status not in _payout.CANCEL_STATUSES]

    def _next_step(s) -> str:
        """Plain, action-focused — what has to happen next, no color-speak."""
        submitted = any(h.startswith(clean.SUBMITTED) for h in s.history)
        if s.status == "Ready For Booking":
            return "Send the bill copy / ETF document"
        if s.status == "Incomplete":
            return "Fix the missing contract data"
        if s.status == clean.SUBMITTED:
            return "Waiting on the supplier — nothing for us to do"
        if s.status == "Verification":
            return ("Waiting on the supplier" if submitted
                    else "Submit a bill or ETF document")
        return clean.STATUS_MEANING.get(s.status, "")

    P_COLS = ("Rep Name", "Sale Date", "Days Waiting", "Business Name",
              "Contract ID", "Status", "Next step")

    # Two sections, each grouped by rep (Carlos, 2026-08-18: "all the yellow
    # ones in their own section... the first section can be all the orders that
    # aren't in yellow, organized by sales rep").
    #
    # "Yellow" is whatever color_for() actually paints yellow, NOT a status
    # name: that's Submitted to Supplier, plus a Verification sale whose
    # HISTORY shows it was already submitted (an un-submitted Verification is
    # orange). Splitting on the paint is also the who-has-the-ball split —
    # section 1 is every deal with something still to do on our end, section 2
    # is the ones sitting with the supplier.
    ours, waiting = [], []
    for s in pend:
        if clean.color_for(s.status, s.history) == clean.YELLOW:
            waiting.append(s)
        else:
            ours.append(s)

    def _by_rep(rows):
        """[(rep, sales)] — reps A-Z, oldest deal first inside each rep.

        Oldest first because this is a chase list: within a rep's block the
        stalest deal is the one that most needs a call.
        """
        groups: Dict[str, List] = {}
        for s in rows:
            rep = _fmt(s.fields.get("Rep Name")) or "(no rep)"
            groups.setdefault(rep, []).append(s)
        for rep in groups:
            groups[rep].sort(key=lambda s: (s.sale_date or dt.date.max,
                                            _fmt(s.fields.get("Business Name"))))
        return [(rep, groups[rep]) for rep in sorted(groups)]

    def _plural(n):
        return "" if n == 1 else "s"

    def _rep_band(sh, row, text):
        cell = sh.cell(row=row, column=1, value=text)
        cell.font = _font(bold=True)
        cell.fill = PatternFill("solid", fgColor=REP_BG)
        cell.alignment, cell.border = LEFT, _border()
        sh.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=len(P_COLS))

    def _pending_row(sh, row, s):
        waited = (today - s.sale_date).days if s.sale_date else ""
        vals = [_fmt(s.fields.get("Rep Name")), s.sale_date or "", waited,
                _fmt(s.fields.get("Business Name")),
                _fmt(s.fields.get("Contract ID")), s.status, _next_step(s)]
        fill_hex = clean.color_for(s.status, s.history)
        fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None
        for c, v in enumerate(vals, start=1):
            cell = sh.cell(row=row, column=c, value=v)
            cell.font = _font()
            cell.alignment = LEFT if c in (1, 4, 7) else CENTER
            cell.border = _border()
            if fill is not None:
                cell.fill = fill
            if c == 2:
                cell.number_format = "mm/dd/yyyy"

    def _pending_section(sh, row, title, rows, empty_note):
        _banner(sh, row, title, len(P_COLS))
        row += 1
        _write_header(sh, row, P_COLS)
        row += 1
        if not rows:
            sh.cell(row=row, column=1, value=empty_note).font = _font(italic=True)
            return row + 2
        for rep, rep_rows in _by_rep(rows):
            _rep_band(sh, row, "{}  •  {} order{}".format(
                rep, len(rep_rows), _plural(len(rep_rows))))
            row += 1
            for s in rep_rows:
                _pending_row(sh, row, s)
                row += 1
            row += 1                      # a blank line between reps
        return row + 1

    psh = wb.create_sheet(_safe_title("Pending Orders", used))
    psh.cell(row=1, column=1, value="Pending orders — not yet accepted").font = \
        _font(bold=True)
    psh.cell(row=2, column=1,
             value="Every deal still in flight (cancelled and rejected are "
                   "excluded). {} as of {}. Two sections, each by sales rep: "
                   "what we still have to work first, then the yellow ones "
                   "sitting with the supplier.".format(
                       "{} pending".format(len(pend)) if pend else "none pending",
                       today.strftime("%m/%d/%Y"))).font = _font(italic=True)
    if not pend:
        psh.cell(row=4, column=1, value="Nothing pending — every deal is "
                 "accepted or closed.").font = _font(italic=True)
    else:
        r = _pending_section(
            psh, 4,
            "OURS TO WORK  •  {} order{}  —  something still to do on our "
            "end".format(len(ours), _plural(len(ours))),
            ours, "Nothing here — every open deal is with the supplier.")
        _pending_section(
            psh, r,
            "WAITING ON THE SUPPLIER  •  {} order{}  —  the yellow ones, "
            "nothing for us to do".format(len(waiting), _plural(len(waiting))),
            waiting, "Nothing here — no deal is sitting with the supplier.")
    widths = [22, 12, 13, 30, 12, 22, 40]
    for c, w in enumerate(widths, start=1):
        psh.column_dimensions[get_column_letter(c)].width = w
    psh.freeze_panes = "A6"

    # ---- one tab per rep -------------------------------------------------
    by_rep: Dict[str, List] = {}
    for s in ordered:
        by_rep.setdefault(_fmt(s.fields.get("Rep Name")) or "(no rep)", []).append(s)

    cols = COLUMNS + (tuple(pay_grid.levels) if pay_grid else ())
    ncol = len(cols)
    for rep in sorted(by_rep):
        rep_sales = by_rep[rep]
        rsh = wb.create_sheet(_safe_title(rep, used))
        r = 1
        rsh.cell(row=r, column=1, value=rep).font = _font(bold=True)
        r += 1
        rsh.cell(row=r, column=1,
                 value="{} sale{} • BOX Order Log • {}".format(
                     len(rep_sales), "" if len(rep_sales) == 1 else "s",
                     today.strftime("%m/%d/%Y"))).font = _font(italic=True)
        r += 2
        for week, group in clean.by_week(rep_sales).items():
            label = week.strftime("%m/%d/%Y") if week else "No sale date"
            _banner(rsh, r, "Week Ending {}  •  {} sale{}".format(
                label, len(group), "" if len(group) == 1 else "s"), ncol)
            r += 1
            _write_header(rsh, r, cols)
            r += 1
            for s in group:
                _write_row(rsh, r, s, COLUMNS)
                if pay_grid:
                    _write_pay_cells(rsh, r, s, pay_grid, len(COLUMNS) + 1)
                r += 1
            if pay_grid:
                r = _write_week_total(rsh, r, group, pay_grid, len(COLUMNS), ncol)
            r += 2
        _legend(rsh, r)
        _autosize(rsh, COLUMNS, rep_sales)
        for i in range(len(pay_grid.levels) if pay_grid else 0):
            rsh.column_dimensions[get_column_letter(len(COLUMNS) + 1 + i)].width = 11

    wb.save(out_path)
    return out_path
