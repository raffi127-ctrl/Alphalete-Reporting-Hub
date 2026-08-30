"""Residential Rep Count TRACKER — weekly append + screenshot email (Carlos).

Feeds Carlos's standalone "Residential Rep Count Tracker" workbook from
Archey's weekly "Residential Rep Counts" email (the same alphaletereporting@
IMAP source Megan's residential_rep_count report already reads):

  1. parse the xlsx attachment's 'Org Snapshot (by Campaign)' tab — one row per
     Org Leader with Unique Headcount (+ ICD count when the tab carries it,
     else derived by counting that leader's rows on 'ICD Headcount (by
     Campaign)');
  2. append a new week-ending column to the tracker's three trend tabs
     (Unique Headcount / ICDs / Reps Per ICD), matching leaders BY NAME (each
     tab may be user-sorted independently), auto-adding new leaders;
  3. move the amber "latest week" highlight, re-issue the sort filters and the
     Week View picker range, and point the Week View picker at the new week;
  4. screenshot the Week View tab + the Unique Headcount tab (last 4 weeks)
     via the Sheets PDF-export engine and Slack-DM both from Lucy (the bot
     token — same delivery as team_tree; Carlos asked for Slack, NOT email).

Runs on LUCY 2 via com.alphalete.rcnc-tracker-friday (Fri 8am/11am/2pm + Sat
9am CT — Archey usually sends Thu night, occasionally Fri morning). Every
firing is idempotent: once this week's column is written AND the email sent, a
flag file short-circuits the later firings.

    python -m automations.rcnc_tracker.run --probe      # env + email + xlsx recon, NO writes
    python -m automations.rcnc_tracker.run --dry-run    # parse + plan, NO writes/sends
    python -m automations.rcnc_tracker.run --no-send    # write sheet, skip the Slack DM
    python -m automations.rcnc_tracker.run              # full run (sends to TEST list)
    python -m automations.rcnc_tracker.run --week-ending 2026-08-29 --force
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
import tempfile
import traceback
from pathlib import Path
from typing import Dict, List, Optional, Tuple

REPO_ROOT = Path(__file__).resolve().parents[2]
FLAG_DIR = REPO_ROOT / "output" / "rcnc_tracker"

TRACKER_ID = "1hOP-zt2r2vI5PloTRn2W41oj02dIwJ7GRAt2wAKTwmA"
TRACKER_URL = f"https://docs.google.com/spreadsheets/d/{TRACKER_ID}/edit"
TREND_TABS = ["Unique Headcount", "ICDs", "Reps Per ICD"]
WEEK_VIEW_TAB = "Week View"
TOTAL_NAMES = ["Total (Legacy D2D)", "Non Legacy Org Leaders", "Grand Total"]
HDR_ROW = 3                      # 1-based header row on the trend tabs
FIRST_DATA_ROW = 4
LAST4 = 4                        # weeks shown on the Unique Headcount shot

# Delivery is SLACK DM from Lucy (Carlos 2026-08-30: "I don't want to email
# that. I want it screenshotted to Slack"). TEST tier = Carlos only until he
# signs off, then flip DEFAULT_TO to PROD_TO (adds his brother Rafael —
# dm_users_with_file tries one group DM first, falls back to individual DMs).
CARLOS_SLACK_ID = "U046G04P5LG"          # same id team_tree DMs
TEST_TO = [CARLOS_SLACK_ID]
PROD_TO = [CARLOS_SLACK_ID, "raffi127@gmail.com"]
DEFAULT_TO = TEST_TO

ORG_TAB = "Org Snapshot (by Campaign)"       # xlsx tab (same one Megan's report reads)
ICD_TAB = "ICD Headcount (by Campaign)"

HILITE = {"red": 1.0, "green": 0.973, "blue": 0.882}
BAND = {"red": 0.902, "green": 0.925, "blue": 0.953}
EPOCH = dt.date(1899, 12, 30)                # Sheets serial-date epoch


class EmailNotLanded(Exception):
    """This week's Archey email hasn't arrived yet — the later launchd firing
    will retry. Exit 0 so the agent doesn't look broken."""


class TrackerNotShared(RuntimeError):
    """The tracker workbook isn't readable by THIS machine's Sheets identity.

    Its own class because it is the one failure a re-run can never fix: someone
    has to add a share in Drive. gspread turns Sheets' 403/404 into a BARE
    `PermissionError` / `SpreadsheetNotFound` whose str() is EMPTY, so the
    first live failure (2026-08-30, Carlos's probe on Lucy 2) reported itself as
    `probe: tracker open ✗ ()` — an empty pair of brackets and nothing else.
    Every Lucy signs in as a DIFFERENT Google account (Lucy 1 raffi127@, Lucy 2
    Carlos's, Lucy 3 alphaletereporting@), so "it opens in my browser" says
    nothing about the machine the agent runs on."""


# --------------------------------------------------------------- email → data
def _expected_week_ending(today: dt.date) -> dt.date:
    """Most recent Saturday strictly before today (Archey labels weeks by
    their ending Saturday) — same rule as residential_rep_count.run."""
    offset = (today.weekday() - 5) % 7 or 7
    return today - dt.timedelta(days=offset)


def fetch_xlsx(tmpdir: str, week_ending: Optional[dt.date]) -> Tuple[Path, dt.date]:
    from automations.residential_rep_count import email_source
    if week_ending:
        xlsx, week, _subj = email_source.fetch_latest(tmpdir, week_ending=week_ending)
        return xlsx, week
    expected = _expected_week_ending(dt.date.today())
    latest = email_source.latest_week_ending()
    if latest is None or latest < expected:
        raise EmailNotLanded(
            f"expected WE {expected} but newest email is WE {latest} — retrying later")
    xlsx, week, _subj = email_source.fetch_latest(tmpdir)
    return xlsx, week


def _headers(ws, max_row=6):
    """(1-based header row, {lower_header: col_idx}) for an openpyxl sheet."""
    import itertools
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row,
                                          values_only=True), 1):
        low = [(str(c).strip().lower() if c else "") for c in row]
        if "org leader" in low and any("unique headcount" in c for c in low):
            return ri, {h: j for j, h in enumerate(low) if h}
    raise ValueError(f"{ws.title!r}: no 'Org Leader' + 'Unique Headcount' header")


def parse_snapshot(xlsx_path: Path) -> Tuple[Dict[str, dict], Dict[str, dict], List[str]]:
    """Return (leaders, totals, notes).

    leaders: {leader name: {'hc': int, 'icds': int|None, 'rpi': float|None}}
    totals: same shape keyed by the three TOTAL_NAMES (whichever exist).
    ICD count comes from the snapshot tab when it has an ICD-count column;
    otherwise it is DERIVED by counting the leader's rows on the ICD tab and a
    note says so.
    """
    import openpyxl
    notes: List[str] = []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if ORG_TAB not in wb.sheetnames:
        raise ValueError(f"tab {ORG_TAB!r} not found; tabs: {wb.sheetnames}")
    ws = wb[ORG_TAB]
    hrow, ci = _headers(ws)
    cl = ci["org leader"]
    chc = next(j for h, j in ci.items() if "unique headcount" in h)
    # An ICD-count column: mentions 'icd' but is not the office/owner/company
    # name column and not a campaign.
    cicd = None
    for h, j in ci.items():
        if "icd" in h and not any(w in h for w in ("office", "owner", "company", "name")):
            cicd = j
            break
    crpi = next((j for h, j in ci.items() if "per icd" in h), None)
    if cicd is None:
        notes.append(f"{ORG_TAB!r} has no ICD-count column "
                     f"(headers: {sorted(ci)}); deriving from {ICD_TAB!r} row counts")

    def _num(row, j, cast=int):
        if j is None or j >= len(row) or row[j] in (None, ""):
            return None
        try:
            return cast(float(row[j]))
        except (TypeError, ValueError):
            return None

    leaders: Dict[str, dict] = {}
    totals: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        nm = row[cl] if cl < len(row) else None
        if not nm or not str(nm).strip():
            continue
        nm = str(nm).strip()
        rec = {"hc": _num(row, chc) or 0, "icds": _num(row, cicd),
               "rpi": _num(row, crpi, float)}
        tn = next((t for t in TOTAL_NAMES if t.lower() in nm.lower()), None)
        if tn:
            totals[tn] = rec
        elif nm.lower().startswith("total") or "grand" in nm.lower():
            notes.append(f"unrecognized summary row skipped: {nm!r}")
        else:
            leaders[nm] = rec

    if cicd is None:
        # Derive ICD counts: rows per Org Leader on the ICD tab.
        from automations.residential_rep_count.parse import parse_headcounts
        counts: Dict[str, int] = {}
        for rec in parse_headcounts(xlsx_path).values():
            org = rec.get("org_leader") or ""
            if org:
                counts[org] = counts.get(org, 0) + 1
        for nm in leaders:
            leaders[nm]["icds"] = counts.get(nm)
        if totals:
            legacy = sum(v["icds"] or 0 for v in leaders.values())
            if "Total (Legacy D2D)" in totals:
                totals["Total (Legacy D2D)"]["icds"] = legacy

    for nm, rec in list(leaders.items()) + list(totals.items()):
        if rec["rpi"] is None and rec["icds"]:
            rec["rpi"] = rec["hc"] / rec["icds"]
    return leaders, totals, notes


# --------------------------------------------------------------- sheet append
def _client_sheet():
    from automations.recruiting_report import fill as rfill
    try:
        return rfill._retry(lambda: rfill.open_by_key(TRACKER_ID))
    except Exception as e:  # noqa: BLE001 — re-raised, just spelled out first
        if type(e).__name__ not in ("PermissionError", "SpreadsheetNotFound"):
            raise
        cause = getattr(e, "__cause__", None)
        resp = getattr(cause, "response", None)
        code = getattr(resp, "status_code", None)
        body = (getattr(resp, "text", "") or "").strip()[:200].replace("\n", " ")
        raise TrackerNotShared(
            f"cannot open the tracker {TRACKER_ID} "
            f"(http={code} {type(e).__name__}: {body or 'no body'}). THIS "
            f"machine's Sheets token (~/.config/recruiting-report/"
            f"oauth-token.json) is not on the file — each Lucy is a different "
            f"Google account. Fix: share {TRACKER_URL} with that account as "
            f"Editor, then re-run. Which account is it: "
            f"`lucy --machine \"Lucy 2\" sheets_whoami`.") from e


def _serial(d: dt.date) -> int:
    return (d - EPOCH).days


def _week_cols(ws) -> Dict[dt.date, int]:
    """{week_ending_date: 1-based col} from the header row (serial dates)."""
    row = ws.get(f"B{HDR_ROW}:ZZ{HDR_ROW}",
                 value_render_option="UNFORMATTED_VALUE")
    out: Dict[dt.date, int] = {}
    for j, v in enumerate(row[0] if row else []):
        if isinstance(v, (int, float)) and v > 40000:
            out[EPOCH + dt.timedelta(days=int(v))] = 2 + j
    return out


def _name_rows(ws) -> Tuple[Dict[str, int], Dict[str, int], int, int]:
    """({leader: row}, {total name: row}, first_totals_row, grand_total_row),
    rows 1-based."""
    col = ws.get(f"A{FIRST_DATA_ROW}:A200")
    names = [(r[0].strip() if r else "") for r in col]
    rows: Dict[str, int] = {}
    tot_rows: Dict[str, int] = {}
    for i, nm in enumerate(names):
        r = FIRST_DATA_ROW + i
        if not nm:
            continue
        if nm in TOTAL_NAMES:
            tot_rows[nm] = r
        else:
            rows[nm] = r
    if "Grand Total" not in tot_rows:
        raise RuntimeError(f"{ws.title!r}: totals rows not found in col A")
    return rows, tot_rows, min(tot_rows.values()), tot_rows["Grand Total"]


def _colletter(c: int) -> str:
    s = ""
    while c:
        c, r = divmod(c - 1, 26)
        s = chr(65 + r) + s
    return s


def append_week(sh, week: dt.date, leaders: Dict[str, dict],
                totals: Dict[str, dict], dry: bool) -> dict:
    """Write the week into all three trend tabs. Returns layout info for the
    screenshot step. Idempotent: an existing column for `week` is overwritten
    in place."""
    metric = {"Unique Headcount": "hc", "ICDs": "icds", "Reps Per ICD": "rpi"}
    info: dict = {}
    for tab in TREND_TABS:
        ws = sh.worksheet(tab)
        cols = _week_cols(ws)
        existed = week in cols
        if existed:
            col = cols[week]
        else:
            if cols and week < max(cols):
                raise RuntimeError(
                    f"{tab}: WE {week} is older than the sheet's newest column "
                    f"({max(cols)}) — backfill by hand, the appender only adds "
                    f"the next week")
            col = (max(cols.values()) + 1) if cols else 2
            if not dry:
                sh.batch_update({"requests": [{"insertDimension": {
                    "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                              "startIndex": col - 1, "endIndex": col},
                    "inheritFromBefore": True}}]})
        rows, tot_rows, tot_first, grand = _name_rows(ws)

        new_leaders = [n for n in leaders if n not in rows]
        if new_leaders and not dry:
            # Insert INSIDE the leader block (before its last row) so banding,
            # filters and the Week View formulas all auto-expand.
            at = tot_first - 1
            sh.batch_update({"requests": [{"insertDimension": {
                "range": {"sheetId": ws.id, "dimension": "ROWS",
                          "startIndex": at - 1, "endIndex": at - 1 + len(new_leaders)},
                "inheritFromBefore": True}}]})
            ws.update(values=[[n] for n in new_leaders],
                      range_name=f"A{at}:A{at + len(new_leaders) - 1}")
            rows, tot_rows, tot_first, grand = _name_rows(ws)

        m = metric[tab]
        updates = [{"range": f"{_colletter(col)}{HDR_ROW}",
                    "values": [[f"{week.month}/{week.day}/{week.year}"]]}]
        cells: Dict[int, object] = {}
        for nm, rec in leaders.items():
            if nm in rows:                      # dry-run can't add the new rows
                v = rec.get(m)
                cells[rows[nm]] = "" if v is None else v
        for r in set(range(FIRST_DATA_ROW, tot_first)) - set(cells):
            cells[r] = ""                       # leaders absent this week
        for tn, tr in tot_rows.items():
            rec = totals.get(tn)
            v = rec.get(m) if rec else None
            cells[tr] = "" if v is None else v
        for r, v in sorted(cells.items()):
            updates.append({"range": f"{_colletter(col)}{r}", "values": [[v]]})
        if not dry:
            ws.batch_update(updates, value_input_option="USER_ENTERED")
            _restyle(sh, ws, col, existed, tot_first, grand)
        info[tab] = {"col": col, "tot_first": tot_first, "grand": grand,
                     "existed": existed, "new_leaders": new_leaders}
        print(f"  {tab}: col {_colletter(col)} "
              f"({'updated' if existed else 'appended'})"
              + (f", new leaders {new_leaders}" if new_leaders else ""))
    return info


def _restyle(sh, ws, col: int, existed: bool, tot_first: int, grand: int) -> None:
    """Move the amber latest-week wash to `col` and re-issue the sort filter."""
    c0 = col - 1                                 # 0-based new column
    reqs = []
    g = lambda sr, er, sc, ec: {"sheetId": ws.id, "startRowIndex": sr,
                                "endRowIndex": er, "startColumnIndex": sc,
                                "endColumnIndex": ec}
    if not existed and col > 2:
        # Prior latest column: leader rows back to banding, totals back to BAND.
        reqs.append({"repeatCell": {"range": g(FIRST_DATA_ROW - 1, tot_first - 1,
                                               c0 - 1, c0),
                     "cell": {}, "fields": "userEnteredFormat.backgroundColor,"
                                           "userEnteredFormat.textFormat.bold"}})
        reqs.append({"repeatCell": {"range": g(tot_first - 1, grand, c0 - 1, c0),
                     "cell": {"userEnteredFormat": {"backgroundColor": BAND,
                              "textFormat": {"bold": True}}},
                     "fields": "userEnteredFormat(backgroundColor,textFormat.bold)"}})
    reqs.append({"repeatCell": {"range": g(FIRST_DATA_ROW - 1, grand, c0, c0 + 1),
                 "cell": {"userEnteredFormat": {"backgroundColor": HILITE,
                          "textFormat": {"bold": True}}},
                 "fields": "userEnteredFormat(backgroundColor,textFormat.bold)"}})
    reqs.append({"setBasicFilter": {"filter": {"range":
                 g(HDR_ROW - 1, tot_first - 1, 0, col)}}})
    sh.batch_update({"requests": reqs})


def point_week_view(sh, week: dt.date, dry: bool) -> None:
    if dry:
        return
    wv = sh.worksheet(WEEK_VIEW_TAB)
    wv.update(values=[[f"{week.month}/{week.day}/{week.year}"]],
              range_name="B2", raw=False)
    uh = sh.worksheet(TREND_TABS[0])
    last = _colletter(max(_week_cols(uh).values()))
    sh.batch_update({"requests": [{"setDataValidation": {
        "range": {"sheetId": wv.id, "startRowIndex": 1, "endRowIndex": 2,
                  "startColumnIndex": 1, "endColumnIndex": 2},
        "rule": {"condition": {"type": "ONE_OF_RANGE", "values": [
            {"userEnteredValue": f"='Unique Headcount'!$B$3:${last}$3"}]},
            "showCustomUi": True, "strict": False}}}]})


# --------------------------------------------------------------- screenshots
def _token() -> str:
    from google.auth.transport.requests import Request as _GARequest
    from google.oauth2.credentials import Credentials
    from automations.recruiting_report.fill import OAUTH_TOKEN_PATH, SCOPES
    creds = Credentials.from_authorized_user_file(str(OAUTH_TOKEN_PATH), SCOPES)
    if not creds.valid:
        creds.refresh(_GARequest())
    return creds.token


def shoot(sh, info: dict, out_dir: Path) -> List[Path]:
    from automations.org_sales_board.screenshot_email import _export_png
    out_dir.mkdir(parents=True, exist_ok=True)
    token = _token()
    shots: List[Path] = []

    wv = sh.worksheet(WEEK_VIEW_TAB)
    n = len([r for r in wv.get("B6:B90") if r and r[0].strip()])
    shots.append(_export_png(wv.id, f"A1:F{6 + n}", out_dir / "week_view.png",
                             token, spreadsheet_id=TRACKER_ID))

    uh = sh.worksheet(TREND_TABS[0])
    i = info[TREND_TABS[0]]
    col, grand = i["col"], i["grand"]
    first_shown = max(2, col - LAST4 + 1)
    hide = {"requests": [{"updateDimensionProperties": {
        "range": {"sheetId": uh.id, "dimension": "COLUMNS",
                  "startIndex": 1, "endIndex": first_shown - 1},
        "properties": {"hiddenByUser": True}, "fields": "hiddenByUser"}}]}
    show = {"requests": [{"updateDimensionProperties": {
        "range": {"sheetId": uh.id, "dimension": "COLUMNS",
                  "startIndex": 1, "endIndex": first_shown - 1},
        "properties": {"hiddenByUser": False}, "fields": "hiddenByUser"}}]}
    if first_shown > 2:
        sh.batch_update(hide)
    try:
        shots.append(_export_png(
            uh.id, f"A{HDR_ROW}:{_colletter(col)}{grand}",
            out_dir / "unique_headcount_last4.png", token,
            spreadsheet_id=TRACKER_ID))
    finally:
        if first_shown > 2:
            sh.batch_update(show)
    return shots


# --------------------------------------------------------------------- slack
def send_slack(week: dt.date, shots: List[Path], to: List[str],
               dry: bool) -> None:
    """DM the screenshots from Lucy. First shot carries the header comment
    with the tracker link; the rest follow bare so the DM reads as one drop."""
    from automations.shared.slack_metrics_post import (
        dm_user_with_file, dm_users_with_file)
    header = (f"*Residential Rep Count Tracker — WE {week.month}/{week.day}*\n"
              f"Week View + Unique Headcount (last {LAST4} weeks). "
              f"Full tracker: {TRACKER_URL}")
    if dry:
        print(f"  DRY-RUN: would Slack-DM {to}: "
              f"{', '.join(x.name for x in shots)}")
        return
    for i, p in enumerate(shots):
        comment = header if i == 0 else ""
        if len(to) == 1:
            r = dm_user_with_file(p, user=to[0], comment=comment)
        else:
            r = dm_users_with_file(p, users=to, comment=comment)
        if not (r.get("ok") or r.get("sent")):
            raise RuntimeError(f"Slack upload failed for {p.name}: {r}")
    print(f"  Slack-DM'd {to}: {len(shots)} screenshots (WE {week})")


# --------------------------------------------------------------------- probe
def probe() -> int:
    """No-write recon: creds, deps, email access, xlsx structure."""
    ok = True
    for mod in ("fitz", "PIL", "openpyxl", "gspread"):
        try:
            __import__(mod)
            print(f"probe: import {mod} ✓")
        except Exception as e:  # noqa: BLE001
            ok = False
            print(f"probe: import {mod} ✗ ({e})")
    try:
        sh = _client_sheet()
        print(f"probe: tracker opens ✓ ({sh.title!r}); tabs "
              f"{[w.title for w in sh.worksheets()]}")
    except Exception as e:  # noqa: BLE001
        ok = False
        # type + message: gspread's 403/404 carry an EMPTY str(), so printing
        # the message alone said literally nothing (2026-08-30).
        print(f"probe: tracker open ✗ ({type(e).__name__}: {e or 'no message'})")
    try:
        from automations.residential_rep_count import email_source
        latest = email_source.latest_week_ending()
        print(f"probe: IMAP ✓ — newest Archey email is WE {latest}")
        with tempfile.TemporaryDirectory() as td:
            xlsx, week = fetch_xlsx(td, latest)
            import openpyxl
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            print(f"probe: xlsx WE {week}; tabs {wb.sheetnames}")
            if ORG_TAB in wb.sheetnames:
                hrow, ci = _headers(wb[ORG_TAB])
                print(f"probe: {ORG_TAB!r} header row {hrow}: {sorted(ci)}")
            leaders, totals, notes = parse_snapshot(xlsx)
            gt = totals.get("Grand Total", {})
            print(f"probe: parsed {len(leaders)} leaders; grand total "
                  f"hc={gt.get('hc')} icds={gt.get('icds')}")
            top = sorted(leaders.items(), key=lambda kv: -kv[1]["hc"])[:3]
            for nm, rec in top:
                print(f"probe:   {nm}: hc={rec['hc']} icds={rec['icds']} "
                      f"rpi={None if rec['rpi'] is None else round(rec['rpi'], 1)}")
            for n in notes:
                print(f"probe: NOTE {n}")
    except Exception as e:  # noqa: BLE001
        ok = False
        print(f"probe: email/xlsx ✗ ({type(e).__name__}: {e})")
    print("=== done ===" if ok else "=== probe FAILED ===")
    return 0 if ok else 1


# ---------------------------------------------------------------------- main
def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--probe", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-send", "--no-email", dest="no_send",
                    action="store_true")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--week-ending", help="YYYY-MM-DD (a Saturday)")
    ap.add_argument("--to", action="append",
                    help="Slack recipient override: user id, email or name "
                         "(repeatable)")
    args = ap.parse_args(argv)
    if args.probe:
        return probe()

    week = (dt.datetime.strptime(args.week_ending, "%Y-%m-%d").date()
            if args.week_ending else None)
    try:
        with tempfile.TemporaryDirectory() as td:
            xlsx, week = fetch_xlsx(td, week)
            flag = FLAG_DIR / f"sent-{week}.flag"
            if flag.exists() and not args.force:
                print(f"WE {week} already appended + emailed ({flag}) — nothing to do")
                print("=== done ===")
                return 0
            print(f"RCNC tracker: WE {week} "
                  f"({'DRY-RUN' if args.dry_run else 'LIVE'})")
            leaders, totals, notes = parse_snapshot(xlsx)
            for n in notes:
                print(f"  NOTE: {n}")
            gt = totals.get("Grand Total", {})
            print(f"  parsed {len(leaders)} leaders, grand total hc={gt.get('hc')}")
            if not leaders or not gt.get("hc"):
                raise RuntimeError("parse produced no leaders / no grand total — "
                                   "refusing to write")
            sh = _client_sheet()
            info = append_week(sh, week, leaders, totals, args.dry_run)
            point_week_view(sh, week, args.dry_run)
            if args.no_send:
                print("  --no-send: skipping screenshots + Slack DM")
            else:
                shots = ([] if args.dry_run else
                         shoot(sh, info, FLAG_DIR / f"shots-{week}"))
                send_slack(week, shots, args.to or DEFAULT_TO, args.dry_run)
            if not args.dry_run and not args.no_send:
                FLAG_DIR.mkdir(parents=True, exist_ok=True)
                flag.write_text(dt.datetime.now().isoformat())
        print("=== done ===")
        return 0
    except EmailNotLanded as e:
        print(f"not ready: {e}")
        print("=== done ===")           # not a failure — the next firing retries
        return 0
    except Exception:
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
