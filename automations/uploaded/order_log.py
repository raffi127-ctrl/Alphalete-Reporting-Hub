"""
Order Log Report — single-file standalone version.

WHAT THIS DOES
  Logs into ownerville → navigates to the ATT Tracker 2.1 - D2D workbook
  in Tableau → opens the ORDER LOG tab → filters Owner Name to a single
  owner and Date Range to "last 1 month → today" → downloads the
  Crosstab CSV → cleans, sorts, color-codes the data → saves a polished
  .xlsx to your Downloads folder ready to drop in Slack.

HOW TO USE
  1. Fill in OWNERVILLE_USERNAME and OWNERVILLE_PASSWORD below.
  2. Install deps (one-time):
         pip install patchright pandas openpyxl python-dateutil
         patchright install chromium
  3. Run:
         python -m automations.order_log_report.run
     (or from inside the folder: `python run.py`)
  4. The .xlsx lands in ~/Downloads/.
"""

# ====================================================================
#  CONFIG — EDIT THIS SECTION
# ====================================================================

# Your ownerville.com login is read from a gitignored local file
# (automations.shared.creds -> ownerville-creds.json at the repo root), NOT
# hardcoded here — the repo was public, so the password must never live in
# source. See _validate_credentials().

# Tableau coordinates.
WORKBOOK_NAME = "ATT Tracker 2.1 - D2D"
TAB_NAME = "ORDER LOG"
OWNER_NAME = "Rafael Hidalgo"

# Saved Tableau Custom View that already has Owner Name, Product Type and
# Captain's Bonus Teams v2 filtered correctly for the default owner. When the
# run is for OWNER_NAME we navigate straight to this view's URL and only adjust
# the date range; the manual filter steps stay as a fallback if the view can't
# be loaded. A non-default --owner skips it, since the view is owner-specific.
CUSTOM_VIEW_NAME = "Order Log - Rafael Hidalgo"
# Direct URL to the Custom View on the ORDER LOG sheet (per-user, lives under
# Rafael's account raffi127@gmail.com — the account the bot logs in as).
# Navigating to it loads the view with its filters already applied, which is
# simpler and more robust than driving the Manage Custom Views dialog.
CUSTOM_VIEW_URL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "ATTTRACKER2_1-D2D/ORDERLOG/c942259a-61b6-47f3-b16f-1ba634c04140/"
    "OrderLog-RafaelHidalgo?:iid=1"
)

# --- Shared crosstab path (storage_state, no Turnstile) ----------------------
# The DEFAULT download path (main(), unattended) pulls the crosstab straight
# from this view via tableau_patchright.download_crosstab_patchright, which
# authenticates by the exported ownerville storage_state — no Cloudflare
# Turnstile, no www/V2 sidebar nav. This is the path proven in
# output/_scratch_orderlog_publish.py.
#
# Re-saved Custom View with a RELATIVE 'Last 1 month' date range (new GUID after
# the re-save). NOTE: this path uses whatever range is baked into the view; it
# does NOT apply main()'s 'last month -> yesterday' override (the shared pull
# can't drive the date textareas the way the legacy filter path does).
CROSSTAB_VIEW_URL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "ATTTRACKER2_1-D2D/ORDERLOG/"
    "96696e4a-8aac-44da-9eff-87f17092435c/OrderLog-RafaelHidalgo?:iid=1"
)
# The worksheet inside the view to export (Tableau Crosstab "sheet" name).
CROSSTAB_SHEET = "A.Order Log"

# ====================================================================
#  Implementation — no changes needed below this line for normal use.
# ====================================================================

import asyncio
import json
import os
import re
import tempfile
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import AsyncIterator, Optional

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from patchright.async_api import (
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)


# === Date range — last 1 month → yesterday (calendar-month aware) ===
# End on yesterday, not today: the current day's orders are still landing /
# syncing in Tableau, so a today-bound pull captures a partial, churning
# last day (Eve 2026-05-31).
END_DATE = date.today() - timedelta(days=1)
START_DATE = END_DATE - relativedelta(months=1)


# === Output location =================================================
# Downloads folder works for any user on macOS or Windows.
OUTPUT_DIR = Path.home() / "Downloads"

# Per-run screenshot folder (optional debugging aid). Set to None to disable.
SCREENSHOT_DIR = Path(__file__).resolve().parent / "screenshots"

# Persistent Chrome profile so Cloudflare / ownerville remember us
# between runs. Lives next to this script.
PROFILE_DIR = Path(__file__).resolve().parent / ".browser_profile"


# ====================================================================
#  Browser launcher (patchright + persistent profile)
# ====================================================================

@asynccontextmanager
async def launch_stealth_browser(headless: bool = False) -> AsyncIterator[Page]:
    """
    Yield an anti-detection Playwright Page backed by patchright.

    Patchright is a Playwright fork with stealth patches baked in —
    Cloudflare and similar bot detectors have a much harder time
    spotting it as automated. Using a persistent profile folder
    makes us look like a returning human, not a fresh bot session.
    """
    PROFILE_DIR.mkdir(exist_ok=True)
    async with async_playwright() as p:
        # Prefer the system Chrome binary - patchright's stealth patches
        # are tuned to it and Cloudflare-style detectors leave it alone
        # more often. If Chrome isn't installed on the machine (common
        # on a clean Windows install or in CI), fall back to patchright's
        # bundled Chromium so the run doesn't die at launch.
        try:
            context = await p.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                channel="chrome",
                headless=headless,
                no_viewport=True,
            )
        except Exception as e:
            print(f"[order_log] system Chrome unavailable ({e!r}) — "
                  "falling back to bundled Chromium", flush=True)
            context = await p.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                headless=headless,
                no_viewport=True,
            )
        page = context.pages[0] if context.pages else await context.new_page()
        try:
            yield page
        finally:
            await context.close()


# ====================================================================
#  Ownerville login
# ====================================================================

LOGIN_URL = "https://ownerville.com"
# Timing knobs — overridable via env for a HEADED manual run where a human
# clears the Cloudflare 'verify you are human' check by hand. Defaults keep the
# automated/unattended behavior byte-identical.
CLOUDFLARE_WAIT_SECONDS = int(os.environ.get("ORDER_LOG_CLOUDFLARE_WAIT_S", "10"))
PRE_SUBMIT_PAUSE_SECONDS = 3
# Post-submit landing wait — ALSO where ownerville's post-login security check
# can surface, so a manual run extends it to give the human time to tilt it.
POST_LOGIN_WAIT_MS = int(os.environ.get("ORDER_LOG_POST_LOGIN_WAIT_S", "45")) * 1000

USERNAME_SELECTOR = (
    'input[type="email"], input[name="username"], input[name="email"], '
    'input[type="text"]'
)
PASSWORD_SELECTOR = 'input[type="password"]'
LOGIN_BUTTON_NAME = re.compile(r"log\s*in|sign\s*in", re.IGNORECASE)
NEXT_BUTTON_NAME = re.compile(r"^\s*next\s*$", re.IGNORECASE)
FINAL_SUBMIT_NAME = re.compile(
    r"sign\s*in|log\s*in|submit|continue|enter", re.IGNORECASE
)


def _validate_credentials() -> tuple[str, str]:
    """Read the ownerville login from the gitignored creds file (not source).
    creds.ownerville_*() raise a clear 'create ownerville-creds.json' error if
    the file/env var is missing."""
    from automations.shared import creds
    return creds.ownerville_username().strip(), creds.ownerville_password().strip()


async def _open_login_form(page: Page) -> None:
    """Click 'Log in' / 'Sign in' on the homepage if visible; otherwise no-op."""
    for role in ("link", "button"):
        candidate = page.get_by_role(role, name=LOGIN_BUTTON_NAME).first
        try:
            if await candidate.is_visible(timeout=2000):
                await candidate.click()
                return
        except PlaywrightTimeoutError:
            continue


async def _attempt_login(page: Page) -> None:
    """One login attempt: homepage -> username -> NEXT -> password -> submit
    -> wait for the post-login landing. Raises PlaywrightTimeoutError if the
    post-login page never renders (e.g. a stale 'Session expired' bounce)."""
    username, password = _validate_credentials()

    print(f"-> Navigating to {LOGIN_URL}")
    await page.goto(LOGIN_URL, wait_until="domcontentloaded")
    await _open_login_form(page)

    print("-> Filling username")
    await page.wait_for_selector(USERNAME_SELECTOR, timeout=15_000)
    await page.fill(USERNAME_SELECTOR, username)

    print("-> Clicking NEXT")
    await page.get_by_role("button", name=NEXT_BUTTON_NAME).first.click()

    await page.wait_for_selector(PASSWORD_SELECTOR, timeout=60_000)

    print(f"-> Letting Cloudflare run for {CLOUDFLARE_WAIT_SECONDS}s "
          "(don't touch the browser)...")
    await asyncio.sleep(CLOUDFLARE_WAIT_SECONDS)

    print("-> Filling password")
    await page.fill(PASSWORD_SELECTOR, password)

    print(f"-> Pausing {PRE_SUBMIT_PAUSE_SECONDS}s before final submit...")
    await asyncio.sleep(PRE_SUBMIT_PAUSE_SECONDS)

    print("-> Clicking final submit button")
    await page.get_by_role("button", name=FINAL_SUBMIT_NAME).first.click()

    print("-> Waiting for post-login page to render...")
    await page.wait_for_load_state("load")
    # Confirm the login landed. The control historically read "Logout", but
    # an exact-match regex (^\s*logout\s*$) silently misses "Log Out" (the
    # space breaks the token) — broaden it to allow an optional space, and
    # race it against the post-login sidebar nav (a.waves-effect) so a
    # cosmetic label change can't block an otherwise-successful login.
    # .first on the COMBINED locator — both match post-login, so racing
    # them without .first trips strict mode.
    logout = page.get_by_text(re.compile(r"^\s*log\s*out\s*$", re.IGNORECASE))
    sidebar = page.locator("a.waves-effect")
    await logout.or_(sidebar).first.wait_for(state="visible",
                                             timeout=POST_LOGIN_WAIT_MS)


async def _inject_ownerville_storage_state(page: Page) -> bool:
    """Seed the context with the manually-exported ownerville session so the
    fast-path 'Log Out' check below finds a live login WITHOUT driving the
    Turnstile form. Reuses tableau_patchright's storage_state file + path
    (imported, not duplicated). A missing/unreadable file is a no-op — the
    caller then fails fast. Returns True iff cookies were injected.

    Lazy import keeps order_log's async stack from pulling the sync patchright
    module at import time (and sidesteps any import cycle via opt_phase)."""
    from automations.shared.tableau_patchright import OWNERVILLE_STORAGE_STATE
    if not OWNERVILLE_STORAGE_STATE.exists():
        print(f"-> no storage_state at {OWNERVILLE_STORAGE_STATE.name}")
        return False
    try:
        state = json.loads(OWNERVILLE_STORAGE_STATE.read_text())
    except Exception as e:
        print(f"-> storage_state unreadable ({e!r}) — ignoring")
        return False
    cookies = state.get("cookies", [])
    if cookies:
        try:
            await page.context.add_cookies(cookies)
        except Exception as e:
            print(f"-> add_cookies failed ({e!r})")
            return False
    print(f"-> storage_state: {len(cookies)} cookie(s) injected")
    return True


async def login(page: Page, allow_form_login: bool = False) -> None:
    """Sign into Ownerville via the exported storage_state session.

    Default path (since 2026-06-17): inject the manually-exported ownerville
    cookies, then reuse them via the fast-path 'Log Out' check below — no login
    form is driven, because ownerville's Cloudflare 'verify you are human' check
    can't be cleared unattended. A missing/expired session FAILS FAST (re-export
    via output/_scratch_ownerville_export_state.py). allow_form_login=True
    re-enables the legacy two-step form-drive (interactive/debug ONLY — it hits
    the Turnstile and stalls unattended)."""
    # Seed the exported session first so the fast-path can reuse it without
    # touching the form.
    await _inject_ownerville_storage_state(page)

    # Fast path: confirm the injected cookies still mint a fresh `rqst` SSO
    # token at v2.ownerville.com — the SAME signal tableau_patchright's
    # _ownerville_session_valid uses (and that Telemapper Knocks/Raf's own
    # order_log path rely on). The old check loaded the public ownerville.com
    # root and waited 6s for the literal "Log Out" text, which the marketing
    # root may not render even when authenticated, and which timed out in
    # unattended runs (Rashad runner 2026-06-28). Same CFID/CFTOKEN cookies,
    # reliable question. On a genuinely dead session this falls through to the
    # same fail-fast RuntimeError below.
    from automations.shared.tableau_patchright import OWNERVILLE_V2_URL
    try:
        await page.goto(OWNERVILLE_V2_URL, wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)
        rqst = bool(re.search(r"rqst=([A-Za-z0-9_]+)", page.url or ""))
        if not rqst:
            href = await page.evaluate(
                "() => { const a=[...document.querySelectorAll('a')]"
                ".find(x=>/rqst=/.test(x.getAttribute('href')||'')); "
                "return a?a.getAttribute('href'):''; }")
            rqst = bool(re.search(r"rqst=([A-Za-z0-9_]+)", href or ""))
        if rqst:
            print("-> Already logged into Ownerville (reusing session) — "
                  "skipping login")
            return
    except PlaywrightTimeoutError:
        pass  # not logged in — fall through

    # Unattended default: never touch the Turnstile form. Fail fast + clear,
    # pointing at the same re-export helper the ownerville fail-fast names.
    if not allow_form_login:
        from automations.shared.tableau_patchright import OWNERVILLE_STORAGE_STATE
        raise RuntimeError(
            "ownerville session expired or missing — run "
            "output/_scratch_ownerville_export_state.py to re-export "
            f"{OWNERVILLE_STORAGE_STATE.name}. (storage_state reuse path; the "
            "login form is disabled because its Cloudflare 'verify you are "
            "human' check can't be cleared unattended.)")

    # Legacy opt-in form-drive — interactive/debug ONLY (hits the Turnstile).
    for attempt in (1, 2):
        try:
            await _attempt_login(page)
            print("-> Login complete")
            return
        except PlaywrightTimeoutError:
            try:
                body = (await page.inner_text("body"))[:300].replace("\n", " | ")
            except Exception:
                body = "(unreadable)"
            stale = "session" in body.lower() and "expired" in body.lower()
            if attempt == 1:
                print(f"   [login] attempt 1 did not land "
                      f"({'stale session detected' if stale else 'post-login timeout'})"
                      " — clearing cookies and retrying from a fresh session...")
                await page.context.clear_cookies()
                continue
            # Second failure — capture diagnostics, then give up.
            try:
                shot = SCREENSHOT_DIR / "order_log_login_stuck.png"
                await page.screenshot(path=str(shot))
                print(f"   [debug] stuck URL: {page.url}")
                print(f"   [debug] title:    {await page.title()}")
                print(f"   [debug] body[:400]: "
                      f"{(await page.inner_text('body'))[:400]}")
                print(f"   [debug] screenshot: {shot}")
            except Exception as _e:
                print(f"   [debug] could not capture page state: {_e}")
            raise RuntimeError(
                "Ownerville login did not complete after 2 attempts (cookies "
                "cleared before the retry). Check the credentials file / "
                "Cloudflare, or the post-login page layout changed."
            )


# ====================================================================
#  Tableau navigation
# ====================================================================

SIDEBAR_TABLEAU_PARENT_SELECTOR = 'a.waves-effect:has(i.fa-t)'
SIDEBAR_TABLEAU_SUB_ITEM_NAME = "Tableau"
LOGIN_TO_TABLEAU_TEXT = "Login to Tableau"
TABLEAU_SEARCH_SELECTOR = "#search-omnibox"
TABLEAU_VIZ_SELECTOR = (
    "div.tab-vizContainer, div.tabCanvas, iframe[src*='tableau']"
)

WORKBOOK_FOLDER_SLUGS = {
    "ATT Tracker 2.1 - D2D": "d2d_tracker",
}


def _slugify(text: str) -> str:
    text = re.sub(r'[^a-zA-Z0-9]+', '_', text.lower())
    return text.strip('_')


def _get_csv_save_path(workbook_name: str, tab_name: str, data_dir: Path) -> Path:
    folder_slug = WORKBOOK_FOLDER_SLUGS.get(
        workbook_name, _slugify(workbook_name)
    )
    folder = data_dir / folder_slug
    folder.mkdir(parents=True, exist_ok=True)
    tab_slug = _slugify(tab_name)
    today = date.today().isoformat()
    path = folder / f"{tab_slug}_{today}.csv"
    if path.exists():
        timestamp = datetime.now().strftime("%H%M%S")
        path = folder / f"{tab_slug}_{today}_{timestamp}.csv"
    return path


async def _shot(page: Page, screenshot_dir: Optional[Path], filename: str) -> None:
    if screenshot_dir is None:
        return
    screenshot_dir.mkdir(exist_ok=True)
    await page.screenshot(path=str(screenshot_dir / filename), full_page=True)


async def _resolve_toolbar_scope(page: Page):
    """
    Return the scope (page or iframe) holding the Tableau viz-viewer
    toolbar. Modern Tableau Cloud puts the toolbar inside an iframe.
    """
    iframes = page.locator("iframe")
    count = await iframes.count()
    for i in range(count):
        src = (await iframes.nth(i).get_attribute("src")) or ""
        if any(m in src.lower() for m in ("tableau", "/views/", "/embed/")):
            frame = page.frame_locator("iframe").nth(i)
            try:
                await frame.locator(
                    '[data-tb-test-id="viz-viewer-toolbar-button-download"], #download'
                ).first.wait_for(state="attached", timeout=2000)
                return frame
            except Exception:
                continue
    return page


async def _dismiss_tableau_announcement_modals(page: Page) -> None:
    """
    Dismiss any Tableau Cloud "Upgrade Coming Soon"-style modal that's
    blocking the page. Ticks "Do not show this notification again" so
    it stays gone across runs. No-op if no modal is on screen.
    """
    close_button = page.locator(
        '[data-tb-test-id="postlogin-footer-close-Button"]'
    ).first
    try:
        await close_button.wait_for(state="visible", timeout=2500)
    except Exception:
        try:
            await page.get_by_text(
                "Tableau Cloud Upgrade", exact=False
            ).first.wait_for(state="visible", timeout=1500)
        except Exception:
            return

    print("-> Dismissing Tableau postlogin announcement modal")

    suppress_strategies = [
        page.locator(
            '[data-tb-test-id*="do-not-show" i], '
            '[data-tb-test-id*="dont-show" i], '
            '[data-tb-test-id*="suppress" i]'
        ),
        page.get_by_label("Do not show this notification again"),
        page.get_by_text(
            "Do not show this notification again", exact=False
        ),
    ]
    for locator in suppress_strategies:
        try:
            target = locator.first
            await target.wait_for(state="visible", timeout=1500)
            try:
                if await target.is_checked():
                    break
            except Exception:
                pass
            await target.click()
            break
        except Exception:
            continue

    try:
        await close_button.click()
    except Exception:
        try:
            await page.get_by_role("button", name="Close").first.click()
        except Exception as exc:
            print(f"   (Close click failed: {exc})")

    await page.wait_for_timeout(600)


async def navigate_to_workbook(
    page: Page,
    workbook_name: str,
    tab_name: Optional[str] = None,
    screenshot_dir: Optional[Path] = None,
) -> Page:
    """
    From the post-login Ownerville page, navigate into Tableau and
    open the specified workbook + dashboard tab. Returns the Page
    where the viz renders (Tableau opens in a new browser tab).
    """
    print('-> Clicking sidebar Tableau parent')
    # login() only confirms SOME sidebar link (a.waves-effect) is visible — not
    # this Tableau-specific parent (:has(i.fa-t)), which can lag behind on a
    # stale/slow session (worked at 6am, went "not visible" by noon → the daily
    # run's failure point). Wait for THIS item; if it never paints, reload the
    # post-login page once to force a clean sidebar render, then click.
    parent = page.locator(SIDEBAR_TABLEAU_PARENT_SELECTOR).first
    try:
        await parent.wait_for(state="visible", timeout=15_000)
    except Exception:
        print('   Tableau sidebar item not visible — reloading to re-render')
        await page.reload(wait_until="load")
        await page.wait_for_timeout(1500)
        await parent.wait_for(state="visible", timeout=20_000)
    await parent.scroll_into_view_if_needed()
    await parent.click()
    await page.wait_for_timeout(800)

    print('-> Clicking sidebar Tableau sub-item')
    sub_links = page.get_by_role(
        "link", name=SIDEBAR_TABLEAU_SUB_ITEM_NAME, exact=True
    )
    count = await sub_links.count()
    clicked = False
    for i in range(count):
        candidate = sub_links.nth(i)
        if await candidate.is_visible():
            await candidate.click()
            clicked = True
            break
    if not clicked:
        raise RuntimeError("No visible 'Tableau' sub-link found.")

    await page.wait_for_load_state("domcontentloaded")
    await _shot(page, screenshot_dir, "01_sidebar_tableau_clicked.png")
    await _shot(page, screenshot_dir, "02_ownerville_tableau_page.png")

    print(f'-> Clicking "{LOGIN_TO_TABLEAU_TEXT}" — expecting a new tab')
    async with page.context.expect_page() as new_page_info:
        await page.get_by_role(
            "link", name=LOGIN_TO_TABLEAU_TEXT, exact=True
        ).first.click()
    tableau_page = await new_page_info.value
    await tableau_page.wait_for_load_state("domcontentloaded")
    print(f"-> New tab opened at {tableau_page.url}")
    await _shot(tableau_page, screenshot_dir, "03_tableau_landed.png")

    await _dismiss_tableau_announcement_modals(tableau_page)

    print(f'-> Searching Tableau for "{workbook_name}"')
    search_box = tableau_page.locator(TABLEAU_SEARCH_SELECTOR)
    await search_box.wait_for(state="visible", timeout=30_000)
    await search_box.fill(workbook_name)
    await search_box.press("Enter")
    await tableau_page.get_by_text(workbook_name, exact=False).first.wait_for(
        state="visible", timeout=30_000
    )
    await _shot(tableau_page, screenshot_dir, "04_search_results.png")

    await _dismiss_tableau_announcement_modals(tableau_page)

    print(f'-> Clicking workbook tile "{workbook_name}"')
    tile_strategies = [
        tableau_page.locator(
            f'a[data-tb-test-id="Thumbnail"][aria-label="{workbook_name}"]'
        ),
        tableau_page.locator(
            f'[data-tb-test-id="WorkbookCardName"]:has-text("{workbook_name}")'
        ),
        tableau_page.get_by_text(workbook_name, exact=False),
    ]
    tile_clicked = False
    for locator in tile_strategies:
        try:
            await locator.first.wait_for(state="visible", timeout=5000)
            await locator.first.click()
            tile_clicked = True
            break
        except Exception:
            continue
    if not tile_clicked:
        raise RuntimeError(
            f"Couldn't click the workbook tile for '{workbook_name}'."
        )

    print("-> Waiting for the workbook page to settle...")
    await tableau_page.wait_for_timeout(2500)
    if not tab_name:
        try:
            await tableau_page.wait_for_selector(
                TABLEAU_VIZ_SELECTOR, timeout=15_000
            )
        except Exception:
            pass

    await _shot(tableau_page, screenshot_dir, "05_workbook_loaded.png")

    if tab_name:
        print(f'-> Clicking tab/view "{tab_name}"')
        view_strategies = [
            tableau_page.locator(f'a[title="{tab_name}"]'),
            tableau_page.locator(
                f'[data-tb-test-id="name-cell"] a[title="{tab_name}"]'
            ),
            tableau_page.get_by_role("link",   name=tab_name, exact=False),
            tableau_page.get_by_role("option", name=tab_name, exact=False),
            tableau_page.locator("a").filter(has_text=tab_name),
            tableau_page.get_by_text(tab_name, exact=False),
        ]
        clicked = False
        for locator in view_strategies:
            try:
                await locator.first.wait_for(state="visible", timeout=10_000)
                await locator.first.click()
                clicked = True
                break
            except Exception:
                continue
        if not clicked:
            raise RuntimeError(
                f"Couldn't find a visible view tile for '{tab_name}'."
            )

        print("-> Waiting for the viz to render after tab click...")
        await tableau_page.wait_for_timeout(2000)
        try:
            await tableau_page.wait_for_selector(
                TABLEAU_VIZ_SELECTOR, timeout=45_000
            )
        except Exception:
            pass
        await _shot(tableau_page, screenshot_dir, "06_tab_loaded.png")

    print("-> Navigation complete")
    return tableau_page


# ====================================================================
#  Custom View (saved per-owner filter state)
# ====================================================================

async def apply_custom_view(page: Page, view_url: str, view_name: str) -> bool:
    """Load the saved Custom View by navigating straight to its Tableau URL.

    'Order Log - Rafael Hidalgo' is a per-user Custom View on the ORDER LOG
    sheet under Rafael's account (raffi127@gmail.com — the account the bot logs
    in as). Going directly to the view's URL loads it with its Owner Name /
    Product Type / Captain's Bonus Teams v2 filters already applied — simpler
    and more robust than driving the Manage Custom Views dialog (which didn't
    even list the view). Returns True if the view loaded, False otherwise, in
    which case the caller falls back to setting the filters manually."""
    print(f"-> Loading Custom View {view_name!r} via URL")

    # The current page (ORDER LOG sheet) and the custom-view URL share the same
    # origin+path and differ only in the hash fragment — and a hash-only goto
    # does NOT reload Tableau's single-page app. So navigate, then reload to
    # force the SPA to actually load the custom-view route.
    try:
        await page.goto(view_url, wait_until="domcontentloaded")
        await page.reload(wait_until="domcontentloaded")
    except Exception as e:
        print(f"   (navigation to the custom-view URL failed: {e} — "
              "falling back)")
        return False

    # Wait for the viz to come up.
    try:
        await page.wait_for_selector(TABLEAU_VIZ_SELECTOR, timeout=45_000)
    except Exception:
        pass
    viz = await _viz_scope(page)
    await _wait_for_viz_ready(viz, page)
    await page.wait_for_timeout(1500)

    # Verify the custom view actually applied: the toolbar's "View:" label
    # names the current view. If it still reads 'Original', the view didn't
    # load — fall back rather than silently pull the wrong (unfiltered) data.
    label = ""
    try:
        scope = await _resolve_toolbar_scope(page)
        label = (await scope.locator(
            '[data-tb-test-id="viz-viewer-toolbar-button-manage-customviews"]'
        ).first.inner_text()).strip()
        print(f"   toolbar view label: {label!r}")
    except Exception:
        pass
    if re.search(r"\boriginal\b", label, re.IGNORECASE):
        print("   (view label still 'Original' — custom view didn't load; "
              "falling back to manual filters)")
        return False

    print(f"   loaded custom view {view_name!r}")
    return True


# ====================================================================
#  Filter automation (Owner Name + Date Range)
# ====================================================================

def _fmt_date(d: date) -> str:
    """Tableau's date textareas accept M/D/YYYY with no leading zeros."""
    return f"{d.month}/{d.day}/{d.year}"


async def _viz_scope(page: Page):
    """
    Return the scope (page or iframe) holding the Tableau viz DOM.
    Filter widgets live wherever the viz lives.
    """
    iframes = page.locator("iframe")
    count = await iframes.count()
    print(f"   (probing {count} iframe(s) for the viz)")
    for i in range(count):
        src = (await iframes.nth(i).get_attribute("src")) or ""
        name = (await iframes.nth(i).get_attribute("name")) or ""
        if any(
            m in (src + name).lower()
            for m in ("tableau", "/views/", "/embed/", "viz")
        ):
            print(f"   (viz lives inside iframe #{i})")
            return page.frame_locator("iframe").nth(i)
    return page


async def _wait_for_viz_ready(viz, page: Page, timeout_ms: int = 60_000) -> None:
    """
    Block until no Tableau loading overlay is visible. `.tab-glass` is
    permanent (transparent click-capture), so we only wait for the
    real loading masks (`.wcGlassPane`, `#loadingGlassPane`).
    """
    glass = ".wcGlassPane:visible, #loadingGlassPane:visible"
    interval_ms, elapsed = 300, 0
    while elapsed < timeout_ms:
        try:
            count = await viz.locator(glass).count()
        except Exception:
            count = 0
        if count == 0:
            return
        await page.wait_for_timeout(interval_ms)
        elapsed += interval_ms
    print(f"   (warning: glass panes still visible after {timeout_ms}ms — proceeding)")


async def _toggle_filter_row(page: Page, row, label: str, want_checked: bool) -> None:
    """
    Toggle a Tableau categorical-filter row to a desired state.

    The outer `.FIItem` div carries `role="checkbox"` and `aria-checked`,
    but Tableau's Dojo widget wires its real click handler to inner
    elements. We try a few targets and verify after each click.
    """
    current = await row.get_attribute("aria-checked")
    wanted = "true" if want_checked else "false"
    if current == wanted:
        print(f'   "{label}" already {wanted} — no toggle needed')
        return

    print(f'   toggling "{label}": aria-checked {current} → {wanted}')
    targets = [
        ("FIText anchor",   "a.FIText"),
        ("fakeCheckBox",    "div.fakeCheckBox"),
        ("FICheckRadio",    "input.FICheckRadio"),
        ("facetOverflow",   "div.facetOverflow"),
        ("row itself",      None),
    ]
    for label_text, sub_selector in targets:
        target = row.locator(sub_selector).first if sub_selector else row
        try:
            await target.click(force=True, timeout=2000)
        except Exception:
            continue
        await page.wait_for_timeout(350)
        new_state = await row.get_attribute("aria-checked")
        if new_state == wanted:
            print(f"     after {label_text}: aria-checked={new_state}")
            return

    raise RuntimeError(
        f'Could not toggle "{label}" to aria-checked={wanted}.'
    )


async def set_owner_name_filter(page: Page, owner_name: str) -> None:
    """Restrict the Owner Name filter to a single owner."""
    print(f'-> Setting Owner Name filter to "{owner_name}"')
    viz = await _viz_scope(page)

    owner_filter = viz.locator(
        '.CategoricalFilter:has(h3.FilterTitle[title="Owner Name"])'
    )
    combobox = owner_filter.locator('[role="combobox"]').first
    await combobox.wait_for(state="visible", timeout=15_000)
    await combobox.click()
    await page.wait_for_timeout(800)

    all_item = viz.locator(".FIItem.all-item").first
    await all_item.wait_for(state="visible", timeout=5000)
    # Normalize to a clean slate before selecting the owner: SELECT (All) so
    # every value is checked, THEN deselect (All) so nothing is. This clears
    # whatever owner the shared dashboard loaded pre-selected — otherwise just
    # "deselect (All) + add Rafael" ADDS Rafael to a pinned subset and leaks
    # other offices' reps (Eve 2026-05-31). Each step no-ops when the row is
    # already in the target state.
    await _toggle_filter_row(page, all_item, "(All)", want_checked=True)
    await _toggle_filter_row(page, all_item, "(All)", want_checked=False)

    owner_row = viz.locator(
        f'.FIItem:has(a.FIText[title="{owner_name}"])'
    ).first
    await owner_row.wait_for(state="visible", timeout=5000)
    await owner_row.scroll_into_view_if_needed()
    await _toggle_filter_row(page, owner_row, owner_name, want_checked=True)

    print("   clicking Apply")
    enabled_apply = viz.locator(
        'button[title="Apply"]:not([disabled])'
    ).first
    await enabled_apply.wait_for(state="visible", timeout=10_000)
    await enabled_apply.click()

    print("   waiting for viz to finish reloading")
    await _wait_for_viz_ready(viz, page)
    await page.wait_for_timeout(1500)


async def reset_team_filter_to_all(page: Page) -> None:
    """Force the 'Captain's Bonus Teams v2' filter to (All).

    The shared ORDER LOG dashboard can load with this filter pinned to a
    specific team (e.g. 'Starr's Team'), which excludes the requested
    owner's orders and leaves the data table empty — the Crosstab dialog
    then only offers the 'Last Refresh' caption sheet, so the export has no
    data (Eve/Megan 2026-05-30). Resetting to (All) makes the pull immune to
    whatever the dashboard's saved filter state happens to be.

    The dashboard carries a legacy 'Captain's Bonus Teams' AND the current
    'Captain's Bonus Teams v2'; the bare 'Captain' prefix is ambiguous and can
    grab the wrong (hidden) one, so the dropdown never opens and the reset
    silently no-ops (Eve 2026-05-31). Pin the match to the 'v2' filter, retry
    the open (it sits far right and can load partially off-screen), and read
    back that (All) actually took so a stuck filter can't pass unnoticed."""
    print("-> Resetting 'Captain's Bonus Teams v2' filter to (All)")
    viz = await _viz_scope(page)

    team_filter = viz.locator(
        '.CategoricalFilter:has(h3.FilterTitle[title^="Captain"][title*="v2"])'
    ).first
    try:
        await team_filter.wait_for(state="visible", timeout=15_000)
    except PlaywrightTimeoutError:
        print("   (filter not found — skipping; dashboard layout may differ)")
        return

    combobox = team_filter.locator('[role="combobox"]').first

    # The filter sits far right and can load partially off-screen, so a single
    # force-click sometimes fails to open the dropdown — retry until the (All)
    # row is actually visible.
    all_item = viz.locator(".FIItem.all-item").first
    opened = False
    for attempt in (1, 2, 3):
        await combobox.scroll_into_view_if_needed()
        await combobox.click(force=True, timeout=10_000)
        await page.wait_for_timeout(800)
        try:
            await all_item.wait_for(state="visible", timeout=5000)
            opened = True
            break
        except PlaywrightTimeoutError:
            print(f"   (dropdown didn't open on attempt {attempt} — retrying)")

    if opened:
        try:
            await _toggle_filter_row(page, all_item, "(All)", want_checked=True)
        except Exception as e:
            print(f"   (could not select (All) on the team filter: {e})")
        # Read back so a stuck filter can't pass silently.
        state = await all_item.get_attribute("aria-checked")
        if state == "true":
            print("   verified: team filter '(All)' is selected")
        else:
            print(f"   WARNING: team filter '(All)' aria-checked={state!r} after "
                  "reset — it may still be pinned to a subset")
    else:
        print("   WARNING: could not open the team filter dropdown after 3 tries "
              "— it may still be pinned to a subset")

    # Commit + CLOSE the dropdown — otherwise the open overlay covers the
    # next filter's combobox. 'Show Apply button' filters expose an Apply
    # (scoped to this filter); 'apply immediately' filters have none and
    # commit on each click, so we just press Escape to close the list.
    applied = False
    apply = team_filter.locator('button[title="Apply"]:not([disabled])').first
    try:
        await apply.wait_for(state="visible", timeout=4000)
        await apply.click()
        applied = True
        print("   clicked Apply on team filter")
    except PlaywrightTimeoutError:
        pass
    if not applied:
        await page.keyboard.press("Escape")
        print("   closed team dropdown (apply-immediate)")

    await page.wait_for_timeout(500)
    await _wait_for_viz_ready(viz, page)
    await page.wait_for_timeout(1000)


async def reset_product_type_filter_to_all(page: Page) -> None:
    """Force the 'Product Type' filter to (All).

    The shared ORDER LOG dashboard loads with this filter pinned to a single
    product (observed: 'NEW INTERNET'), so an export silently contains only
    that product type even though the script applies no product filter
    (Eve 2026-05-31). Resetting to (All) makes the pull immune to whatever the
    dashboard's saved filter state happens to be. Matched by a title prefix so
    'Product Type' vs 'Product Type (Broken Out)' both resolve."""
    print("-> Resetting 'Product Type' filter to (All)")
    viz = await _viz_scope(page)

    product_filter = viz.locator(
        '.CategoricalFilter:has(h3.FilterTitle[title^="Product Type"])'
    ).first
    try:
        await product_filter.wait_for(state="visible", timeout=15_000)
    except PlaywrightTimeoutError:
        print("   (filter not found — skipping; dashboard layout may differ)")
        return

    combobox = product_filter.locator('[role="combobox"]').first
    # Like the team filter, this one can sit partially off-screen → scroll in
    # and force-click past Tableau's transparent click-capture.
    await combobox.scroll_into_view_if_needed()
    await combobox.click(force=True, timeout=10_000)
    await page.wait_for_timeout(800)

    all_item = viz.locator(".FIItem.all-item").first
    try:
        await all_item.wait_for(state="visible", timeout=5000)
        await _toggle_filter_row(page, all_item, "(All)", want_checked=True)
    except Exception as e:
        print(f"   (could not select (All) on the product type filter: {e})")

    # Commit + CLOSE the dropdown so its overlay doesn't cover the next widget.
    applied = False
    apply = product_filter.locator(
        'button[title="Apply"]:not([disabled])'
    ).first
    try:
        await apply.wait_for(state="visible", timeout=4000)
        await apply.click()
        applied = True
        print("   clicked Apply on product type filter")
    except PlaywrightTimeoutError:
        pass
    if not applied:
        await page.keyboard.press("Escape")
        print("   closed product type dropdown (apply-immediate)")

    await page.wait_for_timeout(500)
    await _wait_for_viz_ready(viz, page)
    await page.wait_for_timeout(1000)


async def set_date_range(page: Page, start_date: date, end_date: date) -> None:
    """Set the Start Date and End Date textareas at the top of the view."""
    print(
        f"-> Setting date range: {_fmt_date(start_date)} → {_fmt_date(end_date)}"
    )
    viz = await _viz_scope(page)

    for label, d in [("Start Date", start_date), ("End Date", end_date)]:
        await _wait_for_viz_ready(viz, page)
        box = viz.locator(f'textarea[aria-label="{label}"]').first
        await box.wait_for(state="visible", timeout=10_000)
        # force=True bypasses Tableau's permanent transparent click-capture.
        await box.click(force=True)
        await box.fill(_fmt_date(d))
        await box.press("Enter")

    await _wait_for_viz_ready(viz, page)


# ====================================================================
#  Download (Crosstab → CSV)
# ====================================================================

async def download_dashboard_csv(
    page: Page,
    workbook_name: str,
    tab_name: str,
    data_dir: Path,
) -> Path:
    """Trigger Download → Crosstab (CSV) and save to `data_dir`."""
    csv_path = _get_csv_save_path(workbook_name, tab_name, data_dir)
    scope = await _resolve_toolbar_scope(page)

    print('-> Opening Tableau Download menu')
    download_button = scope.locator(
        '[data-tb-test-id="viz-viewer-toolbar-button-download"], #download'
    ).first
    await download_button.wait_for(state="visible", timeout=15_000)
    await download_button.click()

    print('-> Clicking Crosstab option')
    crosstab_item = scope.locator(
        '[data-tb-test-id="download-flyout-TextMenuItem"]:has-text("Crosstab")'
    ).first
    await crosstab_item.wait_for(state="visible", timeout=10_000)
    await crosstab_item.click()
    await page.wait_for_timeout(1500)

    print('-> Selecting CSV format')
    # Click the label, not the input — label intercepts clicks.
    csv_label = scope.locator(
        '[data-tb-test-id="crosstab-options-dialog-radio-csv-Label"]'
    ).first
    await csv_label.wait_for(state="visible", timeout=10_000)
    await csv_label.click()

    print(f'-> Triggering download → {csv_path}')
    dialog_download = scope.locator(
        '[data-tb-test-id="export-crosstab-export-Button"]'
    ).first
    async with page.expect_download() as download_info:
        await dialog_download.click()

    download = await download_info.value
    await download.save_as(str(csv_path))
    print(f"-> Saved {csv_path}")
    return csv_path


def _set_order_log_dates_sync(start_date: date, end_date: date,
                              verbose: bool = True):
    """Build a sync `pre_export(page, viz)` hook that forces the ORDER LOG
    view's Start/End Date textareas to an EXACT window.

    The shared crosstab path navigates straight to CROSSTAB_VIEW_URL, whose
    baked 'Last 1 month' relative range ends on TODAY — so it drags in today's
    still-churning orders and lands one day past the intended 'last month →
    yesterday' window. Driving the textareas pins the export to the exact
    START_DATE → END_DATE regardless of the view's baked range. This is the
    sync twin of the async set_date_range() the legacy filter path uses; the
    shared pull is sync (sync-patchright), so it can't reuse that coroutine.

    The hook runs after the viz hydrates and on every retry attempt, so a
    re-navigation (which resets to the baked range) re-applies the dates."""
    def _hook(page, viz) -> None:
        if verbose:
            print(f"-> Forcing ORDER LOG date range: {_fmt_date(start_date)} "
                  f"→ {_fmt_date(end_date)}", flush=True)
        for label, d in (("Start Date", start_date), ("End Date", end_date)):
            box = viz.locator(f'textarea[aria-label="{label}"]').first
            box.wait_for(state="visible", timeout=15_000)
            # force=True bypasses Tableau's transparent click-capture overlay.
            box.click(force=True)
            box.fill(_fmt_date(d))
            box.press("Enter")
            page.wait_for_timeout(1200)
        # Let the viz recompute against the new window before the export reads
        # its data; the crosstab dialog's own sheet-poll tolerates the rest.
        page.wait_for_timeout(6000)
    return _hook


def download_order_log_crosstab(
    out_path: Path,
    *,
    allow_form_login: bool = False,
    verbose: bool = True,
) -> Path:
    """Pull the 'A.Order Log' crosstab via the SHARED Tableau storage_state
    session — the same path proven in output/_scratch_orderlog_publish.py, and
    the canonical helper that scratch now imports (so the two can't desync).

    Authenticates by the manually-exported ownerville storage_state: no
    Cloudflare Turnstile, no www/V2 sidebar nav. The date window is whatever
    relative range is baked into CROSSTAB_VIEW_URL ('Last 1 month'); this path
    can't drive the date textareas the legacy filter path used.

    SYNC (patchright sync API). Callers inside an asyncio loop MUST invoke this
    via asyncio.to_thread so sync_playwright doesn't trip the running loop.

    allow_form_login=True re-enables tableau_session's legacy ownerville
    form-drive (interactive/debug ONLY — it hits the Turnstile). Default off: a
    missing/expired session FAILS FAST inside tableau_session, pointing at
    output/_scratch_ownerville_export_state.py to re-export."""
    from automations.shared.tableau_patchright import (
        download_crosstab_patchright,
        tableau_session,
    )
    # Own the session so allow_form_login flows through; download_crosstab_
    # patchright reuses the page (and keeps its own 3x render-flake retry).
    with tableau_session(allow_form_login=allow_form_login,
                         verbose=verbose) as page:
        return download_crosstab_patchright(
            CROSSTAB_VIEW_URL, CROSSTAB_SHEET, out_path,
            verbose=verbose, page=page,
            # Pin the export to the exact 'last month → yesterday' window
            # instead of the view's baked 'Last 1 month' (which ends on today).
            pre_export=_set_order_log_dates_sync(START_DATE, END_DATE,
                                                 verbose=verbose),
        )


# ====================================================================
#  CSV → cleaned + color-coded .xlsx
# ====================================================================
#
# Translated 1:1 from the original Google Apps Script generateCleanedLog().
# Filter columns, drop empty-Rep rows, sort by Rep → Status order → Order
# Date, then write a formatted .xlsx with status-driven row colors.
# ====================================================================

# COLUMNS_TO_KEEP and FRIENDLY_HEADERS are positional twins — index N of one
# maps to index N of the other (df.columns = FRIENDLY_HEADERS in _load_and_clean).
# Keep them aligned when adding/moving a column.
COLUMNS_TO_KEEP = [
    "Rep",
    "sp.Order Date (copy)",
    "Customer Name",
    "sp.Customer Phone",
    "sp.SPM Number",
    "Product Type (Broken Out)",
    "Package",
    "spe.Status",
    "spe.Install Date",
    "Activatoin Date (order log)",  # typo matches source data
    "Tech Install",
]

FRIENDLY_HEADERS = [
    "Rep",
    "Order Date",
    "Customer Name",
    "Customer Phone",
    "SPM Number",
    "Product Type",
    "Package",
    "Status",
    "Install Date",
    "Activation Date",
    "Tech Install",
]

STATUS_ORDER = [
    "Active",
    "Cancelled",
    "Delivered",
    "Disconnected",
    "Open",
    "Pending Shipment",
    "Scheduled",
    "Shipped",
    "",
]

STATUS_COLORS = {
    "Active":           "57BB8A",
    "Cancelled":        "E67C73",
    "Disconnected":     "E67C73",
    "Scheduled":        "FFD666",
    "Open":             "FFD666",
    "Pending Shipment": "FFD666",
    "Shipped":          "FFD666",
    "Delivered":        "F6B26B",
}

STATUS_TEXT_COLORS = {
    "Cancelled":    "CC0000",
    "Disconnected": "CC0000",
    "Delivered":    "B45F06",
}

# Derived from FRIENDLY_HEADERS by label (1-based) so inserting/moving a column
# can't silently shift date formatting onto the wrong column.
DATE_COLUMN_NAMES = ("Order Date", "Install Date", "Activation Date")
DATE_COLUMN_INDEXES = tuple(
    FRIENDLY_HEADERS.index(name) + 1 for name in DATE_COLUMN_NAMES
)


def _status_sort_key(value) -> int:
    cleaned = str(value or "").strip()
    return STATUS_ORDER.index(cleaned) if cleaned in STATUS_ORDER else 999


def _load_and_clean(csv_path: Path) -> pd.DataFrame:
    # Tableau exports UTF-16 LE with tab separators despite the .csv name.
    # 2026-05-30: the ORDER LOG view now prepends a title/caption line
    # ("Last Server Update… / Data Source Sales Date Range…") with no tabs,
    # which a naive read treats as the header. Find the real header row
    # dynamically — the first tab-separated line whose cells include "Rep" —
    # and parse from there, so a leading caption can't shift the columns.
    with open(csv_path, "r", encoding="utf-16") as f:
        lines = f.readlines()
    header_idx = next(
        (i for i, ln in enumerate(lines)
         if "\t" in ln and any(c.strip().strip('"') == "Rep"
                                for c in ln.split("\t"))),
        None,
    )
    # A caption-only export (no "Rep" header row) happens when the owner has
    # zero order-log rows in the window — common for a young office like
    # Rashad's. Tableau's Crosstab then offers only the "Last Refresh" caption
    # sheet, so the saved .csv has no data columns. Return an empty, correctly-
    # typed frame instead of crashing on df["Rep"] (KeyError) downstream.
    if header_idx is None:
        print(f"  (no 'Rep' header in {csv_path.name} — empty/caption-only "
              "export; owner likely has no orders in the window)")
        return pd.DataFrame(columns=FRIENDLY_HEADERS)

    df = pd.read_csv(csv_path, encoding="utf-16", sep="\t", skiprows=header_idx)

    rep_blank = df["Rep"].isna() | (df["Rep"].astype(str).str.strip() == "")
    df = df.loc[~rep_blank].copy()

    # Belt-and-suspenders: if the export has a "Rep" header but is missing some
    # expected columns (another sparse-export shape), don't KeyError on the
    # select — treat it as empty, same as the caption-only case above.
    missing = [c for c in COLUMNS_TO_KEEP if c not in df.columns]
    if missing:
        print(f"  (export missing column(s) {missing} — treating as empty)")
        return pd.DataFrame(columns=FRIENDLY_HEADERS)
    df = df[COLUMNS_TO_KEEP].copy()

    for src_col in ("sp.Order Date (copy)", "spe.Install Date",
                    "Activatoin Date (order log)"):
        df[src_col] = pd.to_datetime(df[src_col], errors="coerce")

    df["_status_order"] = df["spe.Status"].map(_status_sort_key)
    df = df.sort_values(
        by=["Rep", "_status_order", "sp.Order Date (copy)"],
        kind="mergesort",
        na_position="last",
    ).drop(columns="_status_order")

    df.columns = FRIENDLY_HEADERS
    return df


def _thin_border() -> Border:
    side = Side(border_style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _autosize_columns(ws, df: pd.DataFrame) -> None:
    for col_idx, header in enumerate(FRIENDLY_HEADERS, start=1):
        col_series = df[header]
        # Use a defensive str() conversion per-cell. pandas's .astype(str)
        # leaves Arrow-backed nulls as `None`/NaN floats on newer pandas
        # versions, and .map(len) then trips on TypeError: 'float' has no
        # len. A plain generator + str() per value is bulletproof.
        max_len = max(
            (len(header),)
            + tuple(len(str(v)) for v in col_series if v is not None
                    and not (isinstance(v, float) and pd.isna(v)))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4


def csv_to_xlsx(csv_path: Path, output_dir: Path) -> Path:
    """Build the cleaned, color-coded .xlsx ready for Slack."""
    output_dir.mkdir(parents=True, exist_ok=True)
    df = _load_and_clean(csv_path)

    filename = f"Order Log {date.today():%m-%d-%Y}.xlsx"
    out_path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Order Log"

    border = _thin_border()
    header_fill = PatternFill("solid", fgColor="434343")
    header_font = Font(bold=True, color="FFFFFF")
    center_v = Alignment(vertical="center")

    for col_idx, header in enumerate(FRIENDLY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_v
        cell.border = border

    status_col_position = FRIENDLY_HEADERS.index("Status")

    for row_offset, row in enumerate(df.itertuples(index=False), start=2):
        status = str(row[status_col_position] or "").strip()
        bg = STATUS_COLORS.get(status, "FFFFFF")
        fc = STATUS_TEXT_COLORS.get(status, "000000")
        fill = PatternFill("solid", fgColor=bg)
        font = Font(color=fc)

        for col_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            elif col_idx in DATE_COLUMN_INDEXES and hasattr(value, "to_pydatetime"):
                value = value.to_pydatetime()

            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.fill = fill
            cell.font = font
            cell.alignment = center_v
            cell.border = border
            if col_idx in DATE_COLUMN_INDEXES:
                cell.number_format = "m/d/yyyy"

    _autosize_columns(ws, df)
    ws.freeze_panes = "A2"

    # Resilient save: if today's file is still open in Excel, Windows locks it
    # and wb.save() raises PermissionError. Fall back to a timestamped name so
    # the run still produces a deliverable instead of crashing (Eve 2026-06-01).
    try:
        wb.save(out_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%H%M%S")
        out_path = output_dir / f"Order Log {date.today():%m-%d-%Y} {timestamp}.xlsx"
        print(f"  (target .xlsx was locked — likely open in Excel; saving to "
              f"{out_path.name} instead)")
        wb.save(out_path)
    return out_path


# ====================================================================
#  Entry point
# ====================================================================

async def _legacy_filter_download(owner_name: str, tmp_dir: Path,
                                  allow_form_login: bool) -> Path:
    """Legacy download path — the original self-contained flow, preserved as a
    fallback. Drives ownerville login (form-drive only if allow_form_login;
    storage_state reuse otherwise), navigates Tableau, applies the per-owner
    Custom View or manual filters, sets the 'last month -> yesterday' date
    range, and downloads the Crosstab.

    Still the ONLY path that serves a non-default --owner, since the shared
    crosstab view (CROSSTAB_VIEW_URL) is Rafael-specific."""
    async with launch_stealth_browser() as page:
        await login(page, allow_form_login=allow_form_login)

        tableau_page = await navigate_to_workbook(
            page,
            workbook_name=WORKBOOK_NAME,
            tab_name=TAB_NAME,
            screenshot_dir=SCREENSHOT_DIR,
        )

        # Prefer the saved per-owner Custom View: it already has Owner Name,
        # Product Type and Captain's Bonus Teams v2 set correctly, so we
        # only adjust dates. The view is owner-specific, so try it only for
        # the default owner; any other --owner uses the manual filter path.
        # If the view can't be loaded, fall back to driving the filters by
        # hand (the original, self-contained path).
        view_applied = False
        if owner_name == OWNER_NAME:
            view_applied = await apply_custom_view(
                tableau_page, CUSTOM_VIEW_URL, CUSTOM_VIEW_NAME
            )

        if not view_applied:
            # Owner Name first — proves the filter panel has rendered, so
            # the team-filter reset below reliably finds its dropdown.
            await set_owner_name_filter(tableau_page, owner_name)
            await reset_team_filter_to_all(tableau_page)
            await reset_product_type_filter_to_all(tableau_page)

        # Dates always get set last — the Custom View may bake in its own
        # range, so we override regardless of which path ran above.
        await set_date_range(tableau_page, START_DATE, END_DATE)

        return await download_dashboard_csv(
            tableau_page,
            workbook_name=WORKBOOK_NAME,
            tab_name=TAB_NAME,
            data_dir=tmp_dir,
        )


def _post_order_log(xlsx_path: Path, rep_png: "Optional[Path]", *,
                    is_empty: bool, owner_name: str, today,
                    post_to_slack: bool) -> None:
    """Post Order Log + Rep Activations into today's Metrics thread.

    - Empty/no-orders day: post a 'No data available' note in place of the empty
      .xlsx (mirrors Total Knocks) so the metric still appears in the thread.
    - Rep Activations image posts whenever it rendered (even on an empty day).
    - Owner/channel gate: skip ONLY when a non-default owner would post into the
      SHARED #alphalete-sales thread. A scoped channel override (e.g. Rashad ->
      #elevate-sales via METRICS_CHANNEL_ID) is an intentional per-owner report,
      so we DO post there.
    - post_to_slack=False (dry-run): DESCRIBE the planned posts, post nothing.
    """
    import os as _os
    DEFAULT_METRICS_CHANNEL = "C068PH3RFSM"   # #alphalete-sales
    metrics_channel = _os.environ.get("METRICS_CHANNEL_ID", DEFAULT_METRICS_CHANNEL)
    if owner_name != OWNER_NAME and metrics_channel == DEFAULT_METRICS_CHANNEL:
        print(f"  (Skipping Slack post — non-default owner {owner_name!r} into the "
              f"shared thread; that thread is {OWNER_NAME!r}-only.)")
        return
    dry = not post_to_slack
    from automations.shared.slack_metrics_post import (
        post_reply_with_file, post_reply_with_image, post_reply_text_only,
        SlackPostError,
    )
    # Order Log: a 'No data' note on an empty day, else the .xlsx.
    if is_empty:
        note = f"📋 Order Log — {today.strftime('%b')} {today.day} — No data available"
        if dry:
            print(f"  --dry-run — would post Order Log note: {note!r}")
        else:
            try:
                post_reply_text_only(note, react_emoji="clipboard")
                print("  ✓ Slack: posted Order Log 'No data' note")
            except SlackPostError as e:
                print(f"  ⚠ Slack post failed (Order Log note): {e}")
    else:
        slack_filename = f"Order Log {today:%m-%d-%Y}.xlsx"
        if dry:
            print(f"  --dry-run — would post Order Log file: {slack_filename}")
        else:
            try:
                result = post_reply_with_file(
                    xlsx_path, comment="📋 Order Log",
                    react_emoji="clipboard", file_name=slack_filename)
                print(f"  ✓ Slack: posted Order Log (file={result.get('file')})")
            except SlackPostError as e:
                print(f"  ⚠ Slack post failed: {e}")
                print("    .xlsx is in Downloads — drag it into Slack manually.")
    # Rep Activations image (rendered even on empty days).
    if rep_png is not None:
        if dry:
            print(f"  --dry-run — would post Rep Activations image: {rep_png.name}")
        else:
            try:
                post_reply_with_image(
                    rep_png, comment="🆕 Rep Activations — Last & This Week",
                    react_emoji="new",
                    file_name=f"Rep Activations {today:%m-%d-%Y}.png")
                print("  ✓ Slack: posted Rep Activations summary")
            except SlackPostError as e:
                print(f"  ⚠ Slack post failed (Rep Activations): {e}")
    else:
        print("  (Rep Activations image not rendered — nothing to post.)")


async def main(owner_name: str = OWNER_NAME, post_to_slack: bool = True,
               allow_form_login: bool = False) -> None:
    # Intermediate CSV goes to a tempdir that auto-cleans on exit.
    # Final .xlsx is the only artifact left on disk (in Downloads).
    xlsx_path: Optional[Path] = None
    # Companion Rep Activations summary PNG (built from the same crosstab).
    rep_png: Optional[Path] = None
    # Whether the cleaned export is empty (no orders) — drives a 'No data'
    # note in place of an empty .xlsx. Defaults False (post the .xlsx).
    is_empty: bool = False
    with tempfile.TemporaryDirectory(prefix="order_log_") as tmp:
        tmp_dir = Path(tmp)

        # Default (unattended) path: pull the crosstab through the SHARED
        # Tableau storage_state session — no Cloudflare Turnstile, no www/V2
        # sidebar nav. Used for the default owner (Rafael) when not explicitly
        # forcing the legacy form-drive. The legacy filter path is kept as a
        # fallback: it runs when allow_form_login is set OR for a non-default
        # --owner (the shared crosstab view is Rafael-specific).
        if owner_name == OWNER_NAME and not allow_form_login:
            csv_path = tmp_dir / "order_log_crosstab.csv"
            print("-> Downloading 'A.Order Log' crosstab via shared path "
                  "(storage_state auth, no Turnstile)...")
            # sync_playwright can't run inside this asyncio loop — offload the
            # sync shared pull to a worker thread.
            await asyncio.to_thread(
                download_order_log_crosstab, csv_path,
                allow_form_login=False,
            )
        else:
            csv_path = await _legacy_filter_download(
                owner_name, tmp_dir, allow_form_login)

        xlsx_path = csv_to_xlsx(csv_path, OUTPUT_DIR)
        print(f"\n✓ Saved to Downloads: {xlsx_path.name}")

        # Rep Activations summary — two Sun-Sat tables (last week + running
        # week) of Posted / Pending / Total / Canceled per rep, built from the
        # SAME crosstab. Rendered to Downloads (so it survives tempdir cleanup)
        # and posted as a 2nd image below the Order Log .xlsx. Never let a
        # summary hiccup take down the Order Log itself.
        try:
            from automations.rep_activations.aggregate import build_week_tables
            from automations.rep_activations.render import render as render_rep_tables
            _cleaned = _load_and_clean(csv_path)
            is_empty = bool(_cleaned.empty)
            summary = build_week_tables(_cleaned, date.today())
            rep_png = render_rep_tables(
                summary,
                OUTPUT_DIR / f"Rep Activations {date.today():%m-%d-%Y}.png",
            )
            print(f"✓ Saved Rep Activations image: {rep_png.name}")
        except Exception as e:
            rep_png = None
            print(f"  ⚠ Rep Activations summary skipped: {e}")

    # Slack post happens AFTER the browser context is torn down so a failing
    # Slack call can never strand a logged-in patchright session. The
    # owner/channel gate, empty-day 'No data' note, Rep Activations image, and
    # dry-run describe all live in _post_order_log.
    if xlsx_path is None:
        return
    _post_order_log(xlsx_path, rep_png, is_empty=is_empty,
                    owner_name=owner_name, today=date.today(),
                    post_to_slack=post_to_slack)


if __name__ == "__main__":
    # `--owner "Carlos Hidalgo"` lets the same script back a separate
    # 'Carlos Order Log' Hub card without duplicating the codebase.
    # Defaults to OWNER_NAME (Rafael Hidalgo) so the existing Raf card
    # keeps working with no args.
    import argparse
    _ap = argparse.ArgumentParser(description="Download a filtered Order Log")
    _ap.add_argument("--owner", default=OWNER_NAME,
                     help=f"Tableau 'Owner Name' filter value (default: {OWNER_NAME!r})")
    _ap.add_argument("--no-slack", "--no-post", "--dry-run", dest="no_slack",
                     action="store_true",
                     help="Skip the Slack post (just save the .xlsx to "
                          "Downloads). Aliases: --no-post, --dry-run.")
    _ap.add_argument("--allow-form-login", action="store_true",
                     help="Enable the legacy ownerville form-drive (HEADED manual "
                          "run ONLY — you clear the Cloudflare check by hand).")
    _args = _ap.parse_args()
    asyncio.run(main(_args.owner, post_to_slack=not _args.no_slack,
                     allow_form_login=_args.allow_form_login))