"""NDS block for Khalil Mansour + Maxamad-Amin Aden — NDS-SN (RES-ATT-OOF)
Tableau workbook pulls for the org Focus Report campaign zone.

Sources (probed live 2026-08-23; recipes copied from alphalete_org_report/
opt_nds.py and office_metrics/nds_churn.py):

  NDSDailyTracker / 'TT-LineN/P Detail'        current-week owner table:
      Ranking | Rep Count | Sales per Rep | Phone | Air Sold | Air %
  NDSDailyTracker / 'TT-LineN/P Detail (LW)'   the SAME table for last week
      ('Phone (LW)', 'Sales Per Rep (N/P/A)') — stamped under last week's
      Sunday so the zone's WK columns carry a real history.
  NDSDailyTracker / 'Rep Summary'              Total row 'Sales per Rep'
      = National Sales per Rep (also the Sales-per-Rep goal seed).
  LowMetricsLastWeek / 'Low Metrics Owner'     one row per ICD owner with the
      whole quality battery (Metric Flag Count, Trailing Act. Rate, 0-30/90
      churn, ABP %, High/Med Credit Risk %, AT&T Lead Usage Rate). The view is
      last-week/rolling, so its numbers stamp under the CURRENT week's Sunday
      (latest-known convention).
  CHURNRATES / ChurnALlExp / 'Churn Rates (ICD)'  the 'Office/Organization
      Average' Churn Rate row, '0-30 Day Churn' column = National 0-30 Churn
      (same grand row office_metrics/nds_churn.py skips when slicing).
  NDSWeeklyMetricsRep (HTTP .csv)              long format (Measure Names /
      Owner & Office / Rep Name / Measure Values); Rep Name == 'All' is the
      owner total row → 'Next Up %-' and 'Extra & Premium %-' as 0-1
      fractions.
  RepTierBonus-DNQRepCounts (HTTP .csv)        long format (ICD Owner /
      Measure Names / Measure Values) → 'Total DNQ Rep Count' per owner.

Values are emitted as display TEXT per layout.py's convention: counts plain
ints, Sales per Rep one decimal, percents int + '%', churn rates one decimal.
A sub-pull that fails (after the shared downloader's internal retries + one
retry here for the HTTP pulls) is logged and its tuples omitted — partial
data beats none.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from pathlib import Path

from automations.org_campaign_metrics import layout as L
from automations.org_campaign_metrics.run import week_sunday
from automations.shared.tableau_patchright import (
    download_crosstab_patchright,
    requests_session_from_page,
)

CAMPAIGN = "nds"

# Tableau owner spelling (normalized: [suffix] stripped, lowercased) ->
# the org Focus Report picker spelling (layout.MANAGER_CAMPAIGN keys).
MANAGERS = {
    "khalil mansour": "Khalil Mansour",
    "maxamad aden": "Maxamad-Amin Aden",
}

WORKSPACE = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = WORKSPACE / "output" / "org_campaign_metrics"

DAILY_URL = ("https://us-east-1.online.tableau.com/#/site/sci/views/"
             "NDS-SNRES-ATT-OOFWorkbook/NDSDailyTracker?:iid=1")
LOWMETRICS_URL = ("https://us-east-1.online.tableau.com/#/site/sci/views/"
                  "NDS-SNRES-ATT-OOFWorkbook/LowMetricsLastWeek?:iid=1")
# Shared all-owners per-rep churn custom view (see office_metrics/nds_churn.py).
CHURN_URL = ("https://us-east-1.online.tableau.com/#/site/sci/views/"
             "NDS-SNRES-ATT-OOFWorkbook/CHURNRATES/"
             "4fdc2f44-19b3-4941-a9c2-831e88172728/ChurnALlExp?:iid=1")
CHURN_SHEET = "Churn Rates (ICD)"

TABLEAU_WB = "NDS-SNRES-ATT-OOFWorkbook"


# ------------------------------------------------------------ small helpers

def _norm_owner(s):
    """'KHALIL MANSOUR\\n[Dallas, TX]' -> 'khalil mansour' (same rule as
    opt_nds._norm_owner)."""
    s = (s or "").strip()
    s = re.split(r"[\[\n]", s, maxsplit=1)[0].strip()
    return " ".join(s.lower().split())


def _read_tab_csv(path):
    """Tableau UI crosstab reader — UTF-16 tab-delimited, with the XLSX
    fallback for the dialog's 'default' format branch. Trimmed copy of
    opt_nds._read_tab_csv."""
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        return []
    with open(path, "rb") as f:
        magic = f.read(4)
    if magic.startswith(b"PK"):
        from openpyxl import load_workbook
        with open(path, "rb") as f:
            buf = io.BytesIO(f.read())
        wb = load_workbook(buf, read_only=True, data_only=True)
        return [["" if v is None else str(v) for v in row]
                for row in wb.active.iter_rows(values_only=True)]
    for enc in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc) as f:
                rows = list(csv.reader(f, delimiter="\t"))
            if rows and any(len(r) > 1 for r in rows):
                return rows
        except (UnicodeError, csv.Error):
            continue
    return []


def _read_comma_csv(path):
    """Tableau HTTP .csv reader — comma-delimited, ISO-8859-1 (same quirk
    tableau_http.parse_csv handles)."""
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        return []
    text = path.read_text(encoding="iso-8859-1", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def _norm_hdr(h):
    """Normalize a crosstab column header: Tableau's exports carry trailing
    spaces and dangling '-' ('ABP % -', 'Phone (LW)', 'New/Port Trailing
    Act. Rate ')."""
    s = " ".join((h or "").split()).strip().lower()
    return s.rstrip("-").strip()


def _find_col(header, *needles):
    """First column whose normalized header CONTAINS every needle (all
    lowercase). Tolerates the (LW)/(N/P/A) suffix drift between the current
    and last-week owner tables."""
    for i, h in enumerate(header):
        hn = _norm_hdr(h)
        if hn and all(n in hn for n in needles):
            return i
    return None


def _num(s):
    """'74.9%' -> 74.9, '1,062' -> 1062.0, '' -> None."""
    s = str(s or "").strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "").replace("%", "").replace("$", ""))
    except ValueError:
        return None


def _fmt_int(s):
    v = _num(s)
    return None if v is None else str(int(round(v)))


def _fmt_1dp(s):
    v = _num(s)
    return None if v is None else "%.1f" % v


def _fmt_pct_int(s):
    """'74.9%' -> '75%'. Input must already be in percent points."""
    v = _num(s)
    return None if v is None else "%d%%" % int(round(v))


def _fmt_pct_1dp(s):
    """Churn-style: '3.44%' -> '3.4%'."""
    v = _num(s)
    return None if v is None else "%.1f%%" % v


# ------------------------------------------------------------ sub-parsers

def _parse_owner_table(rows):
    """TT-LineN/P Detail (current or LW flavor) -> {norm owner: {field: text}}.
    Column names differ slightly between the two ('Phone' vs 'Phone (LW)',
    'Sales per Rep' vs 'Sales Per Rep (N/P/A)'), hence the contains-matching."""
    if not rows:
        return {}
    header = rows[0]
    cols = {
        "ranking": _find_col(header, "ranking"),
        "rep_count": _find_col(header, "rep count"),
        "spr": _find_col(header, "sales per rep"),
        "phone": _find_col(header, "phone"),
        "air_sold": _find_col(header, "air sold"),
        "air_pct": _find_col(header, "air %"),
    }
    out = {}
    for r in rows[1:]:
        owner = _norm_owner(r[0] if r else "")
        if not owner or owner == "total":
            continue
        rec = {}
        for key, ci in cols.items():
            if ci is not None and ci < len(r) and str(r[ci]).strip() != "":
                rec[key] = str(r[ci]).strip()
        if rec:
            out[owner] = rec
    return out


def _parse_rep_summary_spr(rows):
    """Rep Summary Total row, 'Sales per Rep' column -> national SPR text."""
    if not rows:
        return None
    ci = _find_col(rows[0], "sales per rep")
    if ci is None:
        return None
    for r in reversed(rows):
        if r and (r[0] or "").strip().lower() == "total" and ci < len(r):
            return _fmt_1dp(r[ci])
    return None


def _parse_low_metrics(rows):
    """Low Metrics Owner -> {norm owner: {label: display text}} for the
    quality battery, already formatted per the store convention."""
    if not rows:
        return {}
    header = rows[0]
    spec = [
        # (layout label, formatter, header needles)
        ("Metric Flag Count", _fmt_int, ("metric flag count",)),
        ("Trailing Activation Rate (30–60 Day)", _fmt_pct_int,
         ("trailing act",)),
        ("0–30 Day Churn", _fmt_pct_1dp, ("0-30 day churn",)),
        ("90 Day Churn", _fmt_pct_1dp, ("90 day churn",)),
        ("ABP %", _fmt_pct_int, ("abp %",)),
        ("High/Med Credit Risk %", _fmt_pct_int, ("high/med credit risk",)),
        ("AT&T Lead Usage Rate", _fmt_pct_int, ("lead usage",)),
    ]
    cols = [(lab, fmt, _find_col(header, *needles))
            for lab, fmt, needles in spec]
    out = {}
    for r in rows[1:]:
        owner = _norm_owner(r[0] if r else "")
        if not owner or owner in ("total", "all"):
            continue
        rec = {}
        for lab, fmt, ci in cols:
            if ci is None or ci >= len(r):
                continue
            v = fmt(r[ci])
            if v is not None:
                rec[lab] = v
        if rec:
            out[owner] = rec
    return out


def _parse_national_churn(rows):
    """ChurnALlExp crosstab -> national 0-30 churn text from the
    'Office/Organization Average' / 'Churn Rate' row. Header row 0 names the
    period columns; find '0-30 Day Churn' rather than hardcoding col 4."""
    if not rows:
        return None
    ci = _find_col(rows[0], "0-30 day churn")
    if ci is None:
        ci = 4          # observed layout fallback
    for r in rows[1:]:
        owner = (r[0] if r else "") or ""
        if "office/organization" not in owner.lower():
            continue
        metric = (r[3] if len(r) > 3 else "").strip().lower()
        if not metric.startswith("churn rate"):
            continue
        if ci < len(r):
            return _fmt_pct_1dp(r[ci])
    return None


def _parse_weekly_metrics_all(rows):
    """NDSWeeklyMetricsRep HTTP csv -> {norm owner: {label: text}} from the
    Rep Name == 'All' owner-total rows. Fractions (0-1) -> int percents."""
    if not rows:
        return {}
    header = rows[0]
    m_i = _find_col(header, "measure names")
    o_i = _find_col(header, "owner")
    r_i = _find_col(header, "rep name")
    v_i = _find_col(header, "measure values")
    if None in (m_i, o_i, r_i, v_i):
        return {}
    wanted = {"next up %": "Next Up %",
              "extra & premium %": "Extra / Premium %"}
    out = {}
    for r in rows[1:]:
        if len(r) <= max(m_i, o_i, r_i, v_i):
            continue
        if (r[r_i] or "").strip().lower() != "all":
            continue
        lab = wanted.get(_norm_hdr(r[m_i]))
        if not lab:
            continue
        owner = _norm_owner(r[o_i])
        frac = _num(r[v_i])
        if owner and frac is not None:
            out.setdefault(owner, {})[lab] = "%d%%" % int(round(frac * 100))
    return out


def _parse_dnq(rows):
    """RepTierBonus-DNQRepCounts HTTP csv -> {norm owner: count text} from
    the 'Total DNQ Rep Count' measure rows."""
    if not rows:
        return {}
    header = rows[0]
    o_i = _find_col(header, "icd owner")
    m_i = _find_col(header, "measure names")
    v_i = _find_col(header, "measure values")
    if None in (o_i, m_i, v_i):
        return {}
    out = {}
    for r in rows[1:]:
        if len(r) <= max(o_i, m_i, v_i):
            continue
        if _norm_hdr(r[m_i]) != "total dnq rep count":
            continue
        owner = _norm_owner(r[o_i])
        if not owner or owner == "all":
            continue
        v = _fmt_int(r[v_i])
        if v is not None:
            out[owner] = v
    return out


# ------------------------------------------------------------ collect

def collect(page, today, log=print):
    """-> (values, goals_seed) for the Campaign Log upsert.

    page: live tableau_patchright page from the orchestrator's session.
    values: (manager, week_iso, slot, text); goals_seed: one Sales-per-Rep
    goal per manager, seeded from the national SPR."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    slots = L.slots_by_label(CAMPAIGN)

    def slot(label):
        s = slots.get(label)
        if s is None:
            log("  [nds] layout has no %r row — value dropped" % label)
        return s

    week_cur = week_sunday(today).isoformat()
    week_prev = (week_sunday(today) - dt.timedelta(days=7)).isoformat()

    misses = []

    def _crosstab(url, sheet, fname):
        """One UI crosstab (the shared downloader retries 3x internally).
        Returns parsed rows or [] on failure (logged + reported)."""
        out = OUTPUT_DIR / fname
        try:
            download_crosstab_patchright(url, sheet, out, verbose=False,
                                         page=page)
        except Exception as e:  # noqa: BLE001 — partial data beats none
            log("  [nds] pull failed %r: %s: %s"
                % (sheet, type(e).__name__, str(e).splitlines()[0][:120]))
            misses.append(sheet)
            return []
        rows = _read_tab_csv(out)
        log("  [nds] %s: %d rows" % (fname, len(rows)))
        return rows

    def _http(view, fname):
        """One HTTP .csv view, retried once (its own session per attempt is
        overkill — same cookies both times)."""
        from automations.alphalete_org_report import tableau_http
        out = OUTPUT_DIR / fname
        last = None
        for attempt in (1, 2):
            try:
                tableau_http.download_view_csv(
                    TABLEAU_WB, view, out,
                    session=requests_session_from_page(page))
                rows = _read_comma_csv(out)
                log("  [nds] %s: %d rows" % (fname, len(rows)))
                return rows
            except Exception as e:  # noqa: BLE001
                last = e
                if attempt == 1:
                    log("  [nds] HTTP %s failed (%s) — retrying once"
                        % (view, type(e).__name__))
        log("  [nds] pull failed %r: %s: %s"
            % (view, type(last).__name__, str(last).splitlines()[0][:120]))
        misses.append(view)
        return []

    tt = _parse_owner_table(
        _crosstab(DAILY_URL, "TT-LineN/P Detail", "nds_tt_detail.csv"))
    tt_lw = _parse_owner_table(
        _crosstab(DAILY_URL, "TT-LineN/P Detail (LW)", "nds_tt_detail_lw.csv"))
    nat_spr = _parse_rep_summary_spr(
        _crosstab(DAILY_URL, "Rep Summary", "nds_rep_summary.csv"))
    lmo = _parse_low_metrics(
        _crosstab(LOWMETRICS_URL, "Low Metrics Owner", "nds_low_metrics.csv"))
    nat_churn = _parse_national_churn(
        _crosstab(CHURN_URL, CHURN_SHEET, "nds_churn_all.csv"))
    wm = _parse_weekly_metrics_all(
        _http("NDSWeeklyMetricsRep", "nds_weekly_metrics.csv"))
    dnq = _parse_dnq(
        _http("RepTierBonus-DNQRepCounts", "nds_dnq.csv"))

    values = []
    goals_seed = []

    def put(mgr, week_iso, label, text):
        if text is None:
            return
        s = slot(label)
        if s is not None:
            values.append((mgr, week_iso, s, text))

    for norm, mgr in MANAGERS.items():
        cur = tt.get(norm, {})
        lw = tt_lw.get(norm, {})
        if not cur:
            log("  [nds] %s missing from the current-week owner table" % mgr)

        # TEAM / PRODUCTION — current week.
        put(mgr, week_cur, "Active Selling Reps", _fmt_int(cur.get("rep_count")))
        put(mgr, week_cur, "New/Ports Sold", _fmt_int(cur.get("phone")))
        put(mgr, week_cur, "Sales per Rep", _fmt_1dp(cur.get("spr")))
        put(mgr, week_cur, "Rank on the Tracker", _fmt_int(cur.get("ranking")))
        put(mgr, week_cur, "AIR Sold", _fmt_int(cur.get("air_sold")))
        put(mgr, week_cur, "Air %", _fmt_pct_int(cur.get("air_pct")))

        # Last week's finals from the LW owner table.
        put(mgr, week_prev, "New/Ports Sold", _fmt_int(lw.get("phone")))
        put(mgr, week_prev, "Sales per Rep", _fmt_1dp(lw.get("spr")))

        # QUALITY battery — rolling/last-week views, stamped under the
        # current week's Sunday (latest-known convention).
        for label, text in lmo.get(norm, {}).items():
            put(mgr, week_cur, label, text)
        for label, text in wm.get(norm, {}).items():
            put(mgr, week_cur, label, text)
        put(mgr, week_cur, "DNQ Rep Count", dnq.get(norm))

        # NATIONAL BENCHMARK rows (same value on both managers' zones).
        put(mgr, week_cur, "National Sales per Rep", nat_spr)
        put(mgr, week_cur, "National 0–30 Churn", nat_churn)

        if nat_spr is not None and "Sales per Rep" in slots:
            goals_seed.append((mgr, slots["Sales per Rep"], nat_spr))

    log("  [nds] %d values, %d goal seeds%s"
        % (len(values), len(goals_seed),
           " — MISSED: %s" % ", ".join(misses) if misses else ""))
    return values, goals_seed
