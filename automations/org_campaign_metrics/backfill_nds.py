"""One-shot NDS history backfill for the org Focus Report campaign zone.

Walks Mon-Sun weeks BACKWARDS from the last completed week and stamps, per
manager (Khalil Mansour, Maxamad-Amin Aden) per week, the three production
slots the NDS-SN workbook can actually reconstruct historically:

    New/Ports Sold   WIRELESS product count      "139"
    AIR Sold         AIR product count           "17"
    Air %            AIR/(AIR+WIRELESS) int %    "11%"

Source: ProductSalesSummaryRep / 'Sales By ICD (Weekly View)' in the NDS-SN
(RES-ATT-OOF) workbook — the same crosstab alphalete_org_report/opt_nds.py
pulls, week-pinned by DRIVING the 'Sales Week' quick filter in the UI per
week (the je_pull pre_export pattern). Probed live 2026-08-23: 'Sales Week'
is a multi-select string filter whose members are the week-ending SUNDAY as
M/D/YY ('1/4/26', '12/28/25', ...) for finished weeks, except the two newest
weeks which are the literal labels 'This Week' / 'Last Week' (domain observed
3/2/25..today plus a stray 6/16/24). URL params CANNOT pin it — 'Sales Week'
(ISO and M/D/YY spellings) and the fiber twin's 'Sale Date Week Ending
(mon-sun)' caption are all silently ignored, so the field's real name must
differ from its caption. The pre_export hook drives the member selection and
verifies the combobox before each export; on top of that every downloaded
crosstab must carry EXACTLY the requested week in its week-ending banner row
(a mid-flight render can briefly export a mix — observed live) — a mismatch
retries and then counts as a MISS, a mismatch on the very first week aborts
the run (the drive itself is broken; nothing gets written), and a member
missing from the dropdown counts as an EMPTY week (pre-history).

SPR / rank / rep-count / quality history is NOT backfillable: the NDS daily
tracker and Low Metrics views are current/rolling-window only, so only the
three product-count slots above are emitted.

    python -m automations.org_campaign_metrics.backfill_nds              # dry
    python -m automations.org_campaign_metrics.backfill_nds --write
    python -m automations.org_campaign_metrics.backfill_nds --write --ssid <id>

Dry by default: prints the per-week tuples and touches nothing. --write
upserts into the Campaign Log values section via sheet.upsert_values
(merge-by-key — additive, never deletes). --ssid overrides the target
workbook (sets ORG_CAMPAIGN_SSID before the sheet module loads).

Walk stops at TWO consecutive weeks with no data for either manager, or at
the funnel's first week (WE 2026-01-04; Monday 2025-12-29). A week whose
pull FAILS (after the downloader's internal retries + one retry here) is
logged as MISSED and skipped — it does not count as empty. Python 3.9.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import os
import random
import re
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:  # noqa: BLE001
    pass

CAMPAIGN = "nds"

# Tableau owner spelling (normalized) -> Focus Report picker spelling.
# Copied from pull_nds.MANAGERS (import deferred: see main()'s env dance).
MANAGERS = {
    "khalil mansour": "Khalil Mansour",
    "maxamad aden": "Maxamad-Amin Aden",
}

# The funnel's first week — Monday 2025-12-29, week ending Sunday 2026-01-04.
FLOOR_SUNDAY = dt.date(2026, 1, 4)

# Default ProductSalesSummaryRep view. The 'Sales Week' quick filter is DRIVEN
# in the UI per week (JE je_pull pattern) — URL params don't reach it: the
# caption's param ('Sales Week', tried ISO and the M/D/YY member spelling) and
# the fiber twin's caption are all ignored silently, so the underlying field
# name must differ from the caption. The pre_export hook clicks the member
# tile and verifies the combobox before each export; the banner-date check
# below verifies the export itself.
PSS_URL = ("https://us-east-1.online.tableau.com/#/site/sci/views/"
           "NDS-SNRES-ATT-OOFWorkbook/ProductSalesSummaryRep?:iid=1")
PSS_SHEET = "Sales By ICD (Weekly View)"

# Distinctive marker for "the requested week is not in the filter's domain"
# (pre-history) — mapped to an EMPTY week by the walk, not a MISS.
NOT_IN_DOMAIN = "SALES_WEEK_NOT_IN_DOMAIN"


def week_member(sunday, cur_sunday):
    """The 'Sales Week' member string for a week-ending Sunday: the two
    newest weeks are labeled, older ones are M/D/YY with no zero-padding."""
    if sunday == cur_sunday:
        return "This Week"
    if sunday == cur_sunday - dt.timedelta(days=7):
        return "Last Week"
    return "%d/%d/%d" % (sunday.month, sunday.day, sunday.year % 100)


def _drive_sales_week(member, log):
    """pre_export hook: drive the MULTI-select 'Sales Week' quick filter to
    exactly `member`. Mechanics adapted from org_sales_board/je_pull's
    _drive_week_selection (probed here 2026-08-23: clicking one tile ADDS it
    -> '(Multiple values)', so the JE tick-target-first-then-untick-others
    dance is required, with the hover 'Only' link as the fast path). Raises
    on failure so the downloader's retry ladder re-navigates and re-applies
    the selection on a fresh load."""
    _ROW = 'div.FIItem[role="checkbox"]'

    def _rows(viz):
        """Visible dropdown rows preferred — Tableau keeps hidden copies of
        menu markup in the DOM and a click on one of those is a silent no-op
        (je_pull's hard-won lesson)."""
        try:
            vis = viz.locator(_ROW + ":visible")
            if vis.count():
                return vis
        except Exception:  # noqa: BLE001
            pass
        return viz.locator(_ROW)

    def _item(viz, txt):
        return _rows(viz).filter(
            has_text=re.compile(r"^\s*%s\s*$" % re.escape(txt))).first

    def _checked(it):
        try:
            if (it.get_attribute("aria-checked") or "") == "true":
                return True
            return "FIChecked" in (it.get_attribute("class") or "")
        except Exception:  # noqa: BLE001
            return False

    def _members(viz):
        """[(text, checked)] for every member row except '(All)'."""
        items = _rows(viz)
        out = []
        try:
            n = items.count()
        except Exception:  # noqa: BLE001 — menu closed under us
            return out
        for j in range(n):
            it = items.nth(j)
            try:
                txt = (it.inner_text() or "").strip()
            except Exception:  # noqa: BLE001
                continue
            if not txt or txt == "(All)":
                continue
            out.append((txt, _checked(it)))
        return out

    def _set(viz, page, txt, want):
        it = _item(viz, txt)
        if _checked(it) == want:
            return
        glyph = it.locator(".FICheckRadio").first
        try:
            if glyph.count():
                glyph.scroll_into_view_if_needed()
                glyph.click(timeout=10_000)
            else:
                it.scroll_into_view_if_needed()
                it.click(timeout=10_000)
        except Exception:  # noqa: BLE001 — glyph gone mid-click; row instead
            it.click(timeout=10_000)
        page.wait_for_timeout(900)

    def _only(viz, page, txt):
        """The per-row hover 'Only' link — one click selects just this member.
        Best-effort: False falls through to the tick/untick loop."""
        try:
            it = _item(viz, txt)
            it.scroll_into_view_if_needed()
            it.hover(timeout=5_000)
            page.wait_for_timeout(400)
            link = it.locator("a, button, span").filter(
                has_text=re.compile(r"^\s*Only\s*$", re.I)).first
            if link.count() == 0:
                return False
            link.click(timeout=5_000)
            page.wait_for_timeout(1_500)
            return True
        except Exception:  # noqa: BLE001
            return False

    def _close(page, viz):
        """Escape past the tab-glass outside-click catcher (clicking the
        combobox again is intercepted and burns the full timeout)."""
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(600)
        except Exception:  # noqa: BLE001
            pass
        try:
            glass = viz.locator("div.tab-glass").first
            if glass.count():
                glass.click(timeout=3_000)
                page.wait_for_timeout(400)
        except Exception:  # noqa: BLE001
            pass

    def pre_export(page, viz):
        panel = viz.locator(".QuickFilterPanel", has_text="Sales Week").first
        panel.wait_for(timeout=30_000)
        combo = panel.locator(".tabComboBox").first
        cur = (combo.inner_text() or "").strip()
        if cur == member:
            page.wait_for_timeout(3_000)    # already on the target week —
            return                          # just let the viz settle
        combo.click(timeout=15_000)
        page.wait_for_timeout(1_500)
        members = _members(viz)
        if member not in [t for t, _c in members]:
            _close(page, viz)
            raise RuntimeError("%s: %r not among %d dropdown members"
                               % (NOT_IN_DOMAIN, member, len(members)))

        _only(viz, page, member)            # fast path; loop below verifies
        for _ in range(4):
            members = _members(viz)
            on = any(t == member and c for t, c in members)
            wrong = [t for t, c in members if c and t != member]
            if on and not wrong:
                break
            # Tick the target FIRST: Tableau re-selects everything when a
            # filter would be left with nothing selected.
            if not on:
                _set(viz, page, member, True)
                continue
            for t in wrong:
                _set(viz, page, t, False)
        members = _members(viz)             # last look, menu still open
        ticked = [t for t, c in members if c]
        _close(page, viz)
        final = ""
        for _ in range(12):                 # the box lags on a slow viz
            final = (combo.inner_text() or "").strip()
            if final == member:
                break
            page.wait_for_timeout(1_000)
        if final != member:
            # The box text can lag ('Last Week' / '(Multiple values)' read
            # seconds after the clicks) even when the tick state is already
            # correct — observed live on 8/2 and 6/21. When exactly the
            # target is ticked, proceed: the banner check on the export is
            # the hard gate either way.
            if ticked == [member]:
                log("      box reads %r but exactly %r is ticked — "
                    "proceeding" % (final, member))
            else:
                raise RuntimeError(
                    "Sales Week select failed: box=%r expected %r; ticked: %s "
                    "(of %d member(s))"
                    % (final, member, ", ".join(ticked or ["(none)"]),
                       len(members)))
        # 6s settle: with 3s the dialog occasionally exported the PREVIOUS
        # selection's render (caught by the banner check + retry, but slow).
        page.wait_for_timeout(6_000)
        log("      Sales Week set to %r" % final)
    return pre_export

WORKSPACE = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = WORKSPACE / "output" / "org_campaign_metrics"


# ------------------------------------------------------------ small helpers

def _norm_owner(s):
    """'KHALIL MANSOUR\\n[Dallas, TX]' -> 'khalil mansour' (pull_nds rule)."""
    s = (s or "").strip()
    s = re.split(r"[\[\n]", s, maxsplit=1)[0].strip()
    return " ".join(s.lower().split())


def _read_tab_csv(path):
    """Tableau UI crosstab reader — UTF-16 tab-delimited, XLSX fallback
    (trimmed copy of pull_nds._read_tab_csv)."""
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        return []
    with open(path, "rb") as f:
        magic = f.read(4)
    if magic.startswith(b"PK"):
        from openpyxl import load_workbook
        with open(path, "rb") as f:
            buf = io.BytesIO(f.read())
        wb = load_workbook(buf, read_only=True, data_only=True)
        return [["" if v is None else str(v) for v in row]
                for row in wb.active.iter_rows(values_only=True)]
    for enc in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc) as f:
                rows = list(csv.reader(f, delimiter="\t"))
            if rows and any(len(r) > 1 for r in rows):
                return rows
        except (UnicodeError, csv.Error):
            continue
    return []


_DATE_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")
_ISO_RE = re.compile(r"^\s*(\d{4})-(\d{2})-(\d{2})(?:[ T].*)?$")


def _cell_date(s):
    """'6/7/2026' or '2026-06-07[ 00:00:00]' -> date (XLSX fallback emits ISO)."""
    s = str(s or "")
    m = _DATE_RE.match(s)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = _ISO_RE.match(s)
        if not m:
            return None
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return dt.date(y, mo, d)
    except ValueError:
        return None


def _crosstab_dates(rows):
    """All m/d/yyyy dates appearing in the first two rows (the week-ending
    banner row of the 'Sales By ICD (Weekly View)' crosstab)."""
    out = []
    for r in rows[:2]:
        for c in r:
            d = _cell_date(c)
            if d is not None:
                out.append(d)
    return out


def parse_week_counts(rows):
    """Crosstab rows -> {norm owner: {'WIRELESS': n, 'AIR': n, ...}}.

    Default-view layout (probed 2026-08-23): banner row of week-ending dates,
    then 'Owner & Office | Product Type (Broken Out) | Monday..Sunday | Total
    | Total(Grand)'. Columns are found BY HEADER NAME (the custom views add a
    'Rep Name' column at index 1, so fixed indices mis-parse); the FIRST
    'Total' column is the week total. Skips the Grand Total row, per-owner
    'Total' product rows, and (when a Rep column exists) per-owner rep
    subtotal rows, then sums per owner per product type."""
    if not rows:
        return {}
    hdr_i = owner_i = ptype_i = total_i = rep_i = None
    for i, r in enumerate(rows[:4]):
        low = [str(c or "").strip().lower() for c in r]
        o = next((j for j, c in enumerate(low) if "owner" in c), None)
        p = next((j for j, c in enumerate(low) if "product type" in c), None)
        totals = [j for j, c in enumerate(low) if c == "total"]
        if o is not None and p is not None and totals:
            hdr_i, owner_i, ptype_i, total_i = i, o, p, totals[0]
            rep_i = next((j for j, c in enumerate(low) if "rep name" in c),
                         None)
            break
    if total_i is None:
        return {}
    out = {}
    for r in rows[hdr_i + 1:]:
        if len(r) <= total_i:
            continue
        owner_raw = (r[owner_i] or "").strip()
        ptype = (r[ptype_i] or "").strip().upper()
        if not owner_raw or owner_raw.lower().startswith("grand total"):
            continue
        if ptype == "TOTAL":
            continue                    # per-owner pre-summed product row
        if rep_i is not None and (r[rep_i] or "").strip().lower() == "total":
            continue                    # per-owner rep subtotal row
        owner = _norm_owner(owner_raw)
        if not owner:
            continue
        try:
            val = int(float((r[total_i] or "0").replace(",", "") or "0"))
        except (ValueError, AttributeError):
            continue
        rec = out.setdefault(owner, {})
        rec[ptype] = rec.get(ptype, 0) + val
    return out


def week_tuples(counts, slots, week_iso):
    """One week's {owner: {ptype: n}} -> Campaign Log value tuples."""
    out = []
    for norm, mgr in MANAGERS.items():
        rec = counts.get(norm)
        if rec is None:
            continue                    # owner absent that week — no tuples
        wireless = rec.get("WIRELESS", 0)
        air = rec.get("AIR", 0)
        out.append((mgr, week_iso, slots["New/Ports Sold"], str(wireless)))
        out.append((mgr, week_iso, slots["AIR Sold"], str(air)))
        if air + wireless > 0:
            pct = "%d%%" % int(round(air * 100.0 / (air + wireless)))
            out.append((mgr, week_iso, slots["Air %"], pct))
    return out


# ------------------------------------------------------------ main

def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--write", action="store_true",
                    help="actually write (default: dry preview)")
    ap.add_argument("--ssid", help="override the Campaign Log spreadsheet id "
                                   "(e.g. the test workbook)")
    ap.add_argument("--max-weeks", type=int, default=60,
                    help="safety cap on weeks walked (default 60)")
    ap.add_argument("--reuse-downloads", action="store_true",
                    help="reuse this run-dir's saved backfill_nds_<week>.csv "
                         "files (banner-verified per file) instead of pulling "
                         "Tableau again — so the test-workbook write, the "
                         "spot-check, and the live write all stamp the SAME "
                         "verified pull. Weeks with no valid file still pull.")
    a = ap.parse_args(argv)
    dry = not a.write

    if a.ssid:
        os.environ["ORG_CAMPAIGN_SSID"] = a.ssid
    # Project imports AFTER the env override — sheet.py freezes SSID at import.
    from automations.org_campaign_metrics import layout as L
    from automations.org_campaign_metrics import sheet as CL
    from automations.org_campaign_metrics.run import week_sunday
    from automations.shared.tableau_patchright import (
        download_crosstab_patchright, tableau_session)

    log = print
    slots = L.slots_by_label(CAMPAIGN)
    for lab in ("New/Ports Sold", "AIR Sold", "Air %"):
        if lab not in slots:
            raise SystemExit("nds layout lost its %r row — aborting" % lab)

    today = dt.date.today()
    start = week_sunday(today) - dt.timedelta(days=7)   # last COMPLETED week
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    log("== backfill_nds -> %s%s" % (CL.SSID, " (DRY RUN)" if dry else ""))
    log("   walking Sundays %s -> %s (stop: 2 consecutive empty weeks)"
        % (start.isoformat(), FLOOR_SUNDAY.isoformat()))

    values = []
    per_week = {}                       # iso -> tuple count (for the report)
    missed = []
    empty_streak = 0
    pinned_ok = False                   # any week verified pinned yet?

    # The Tableau session opens LAZILY — a --reuse-downloads run that finds
    # every week's file valid never launches a browser at all.
    from contextlib import ExitStack
    with ExitStack() as stack:
        holder = {}

        def _page():
            if "p" not in holder:
                holder["p"] = stack.enter_context(tableau_session(verbose=True))
            return holder["p"]

        sunday = start
        walked = 0
        while sunday >= FLOOR_SUNDAY and walked < a.max_weeks:
            iso = sunday.isoformat()
            member = week_member(sunday, week_sunday(today))
            out = OUTPUT_DIR / ("backfill_nds_%s.csv" % iso)
            rows, err = [], None

            if a.reuse_downloads and out.exists():
                cached = _read_tab_csv(out)
                if cached and set(_crosstab_dates(cached)) == {sunday}:
                    rows = cached
                    log("   %s: reusing %s" % (iso, out.name))
                else:
                    log("   %s: saved file invalid for this week — repulling"
                        % iso)

            for attempt in () if rows else (1, 2):   # one extra retry atop
                try:                                 # the internal 3
                    download_crosstab_patchright(
                        PSS_URL, PSS_SHEET, out, verbose=False, page=_page(),
                        pre_export=_drive_sales_week(member, log))
                    rows = _read_tab_csv(out)
                    # Banner verification — the export must carry EXACTLY the
                    # requested week. A different week means the selection
                    # didn't take; a MIX of weeks is a mid-flight render (the
                    # viz recomputing while the dialog exported — observed
                    # once live 2026-08-23). Both retry, then count as a MISS.
                    got = set(_crosstab_dates(rows))
                    if got and got != {sunday}:
                        raise RuntimeError(
                            "export carries week(s) %s, wanted %s"
                            % (", ".join(d.isoformat() for d in sorted(got)),
                               iso))
                    err = None
                    break
                except Exception as e:  # noqa: BLE001
                    err = e
                    rows = []
                    if NOT_IN_DOMAIN in str(e):
                        break           # pre-history week — retrying won't help
                    if attempt == 1:
                        log("   %s pull failed (%s: %s) — retrying once"
                            % (iso, type(e).__name__,
                               str(e).splitlines()[0][:100]))
            if err is not None and NOT_IN_DOMAIN in str(err):
                err, rows = None, []    # not in the filter domain == no data
                log("   %s: member %r not in the Sales Week domain"
                    % (iso, member))
            if err is not None:
                if not pinned_ok and "wanted" in str(err):
                    raise SystemExit(
                        "PIN FAILED on the very first week %s (member %r): %s"
                        " — the Sales Week drive is broken; nothing written."
                        % (iso, member, err))
                log("   %s MISSED: %s: %s" % (
                    iso, type(err).__name__, str(err).splitlines()[0][:120]))
                missed.append(iso)
                sunday -= dt.timedelta(days=7)
                walked += 1
                continue
            if rows:
                pinned_ok = True

            counts = parse_week_counts(rows)
            counts = {o: c for o, c in counts.items() if o in MANAGERS}
            tups = week_tuples(counts, slots, iso)
            values.extend(tups)
            per_week[iso] = len(tups)
            if counts:
                empty_streak = 0
                log("   %s: %s" % (iso, "; ".join(
                    "%s W=%d A=%d" % (o, c.get("WIRELESS", 0), c.get("AIR", 0))
                    for o, c in sorted(counts.items()))))
            else:
                empty_streak += 1
                log("   %s: EMPTY (%d consecutive)" % (iso, empty_streak))
                if empty_streak >= 2:
                    log("   two consecutive empty weeks — stopping the walk")
                    break
            sunday -= dt.timedelta(days=7)
            walked += 1

    weeks_with_data = sorted(w for w, n in per_week.items() if n)
    log("-- %d weeks walked, %d with data (%s..%s), %d tuples, %d missed%s"
        % (len(per_week) + len(missed), len(weeks_with_data),
           weeks_with_data[0] if weeks_with_data else "-",
           weeks_with_data[-1] if weeks_with_data else "-",
           len(values), len(missed),
           " (%s)" % ", ".join(missed) if missed else ""))

    if dry:
        for row in values:
            log("   %s" % (row,))
        log("DRY RUN — nothing written. Re-run with --write to stamp %s."
            % CL.SSID)
        return 0

    from automations.funnel_board.auth import session as sheets_session
    S = sheets_session(verbose=False)
    changed = CL.upsert_values(S, values, dry_run=False, log=log)
    log("WROTE %d tuples (%d changed) to spreadsheet %s ('%s' tab)"
        % (len(values), changed, CL.SSID, CL.TAB))

    # Spot-check aid: pick 3 random data weeks to verify by hand.
    if weeks_with_data:
        picks = random.sample(weeks_with_data, min(3, len(weeks_with_data)))
        log("spot-check candidates: %s" % ", ".join(sorted(picks)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
