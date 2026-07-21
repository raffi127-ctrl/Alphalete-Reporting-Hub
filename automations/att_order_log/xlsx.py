"""The daily ATT B2B Order Log workbook: an overall tab + one tab per rep,
grouped by paycheck (week-ending) weeks — the AT&T counterpart of
box_order_log.xlsx (Megan 2026-07-20: "per rep breakdown with the paycheck
weeks like we did for Box").

Same house look as the BOX + Fiber logs (Georgia 12pt, dark header band, blue
week banners, status-coloured rows), so Carlos's two logs read identically.
Built off the un-pivoted AT&T sales lines (att_order_log.clean), the same data
the sheet + the Slack thread use — one source, no second pull.

Grouped by PAYCHECK week — a sale sits in the week of its POSTED date, because
that is the paycheck it's on (Carlos 2026-07-20). Sales with no posted date
haven't hit a paycheck yet and get their own 'Not Yet Posted' section rather
than being dropped (Megan 2026-07-20).

Tabs, in order:
  1. "All Reps"        — every sale, newest paycheck week first, plus a
                         Not-Yet-Posted section.
  2. "Paycheck by Week"— reps down, paycheck weeks across, sales counted in the
                         week they posted (structural twin of BOX's Accepted-by-
                         Supplier payout). Not-yet-posted excluded (counted in
                         the subtitle) since they're on no paycheck.
  3. one tab per rep   — that rep's sales, same paycheck-week grouping.
"""
from __future__ import annotations

import collections
import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import colors
from .sheet import DISPLAY_HEADERS, DISPLAY_LABELS, _parse_date

FONT_NAME = "Georgia"
FONT_SIZE = 12
HEADER_BG = "434343"
WEEK_BG = "2563EB"
# What determines the paycheck week of an AT&T sale: its POSTED DATE (Carlos,
# 2026-07-20 — "What determines what paycheck an at&t sale is on? Posted date").
# So the paycheck matrix keys off this column's week, not the order date.
POSTED_DATE_COL = "spe.dtr Posted Date (copy)"

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Log columns: Rep first on the summary; the per-rep tabs drop it.
_REP_LABEL = DISPLAY_LABELS[0]
_LOG_LABELS = DISPLAY_LABELS            # 17 friendly names
_LABEL_TO_HEADER = dict(zip(DISPLAY_LABELS, DISPLAY_HEADERS))

_BAD_TITLE = re.compile(r"[:\\/?*\[\]]")


def week_ending(d: dt.date) -> dt.date:
    """Saturday of the Sun-Sat week containing d — same convention as
    box_order_log.clean.week_ending, so the two logs' weeks line up."""
    sunday = d - dt.timedelta(days=(d.weekday() + 1) % 7)
    return sunday + dt.timedelta(days=6)


def _font(color: str = "000000", *, bold: bool = False,
          italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, italic=italic,
                color=color)


def _border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_title(name: str, used: set) -> str:
    title = (_BAD_TITLE.sub("-", name).strip() or "Rep")[:31]
    base, n = title, 2
    while title.lower() in used:
        sfx = " ({})".format(n)
        title = base[:31 - len(sfx)] + sfx
        n += 1
    used.add(title.lower())
    return title


def _line_paycheck_week(ln: dict) -> Optional[dt.date]:
    """POSTED-date week — the paycheck week (Carlos 2026-07-20). A sale with no
    posted date hasn't posted yet, so it is not on any paycheck."""
    d = _parse_date(ln.get(POSTED_DATE_COL))
    return week_ending(d) if d else None


def by_week(lines: Sequence[dict]) -> "collections.OrderedDict":
    """Lines grouped by PAYCHECK week (posted date), NEWEST first, with the
    not-yet-posted sales collected under None LAST.

    Grouping by posted date (not order date) is what makes this a paycheck view
    (Carlos: pay is on the posted date). Sales with no posted date haven't hit a
    paycheck yet; they still belong on the rep's tab (Megan 2026-07-20 — "for
    the not posted sales we should still put them on the rep's breakdown tab in
    its own section"), so they get their own 'Not Yet Posted' section rather
    than being dropped or mixed into a week."""
    out: Dict[Optional[dt.date], list] = collections.defaultdict(list)
    for ln in lines:
        out[_line_paycheck_week(ln)].append(ln)
    ordered = collections.OrderedDict()
    for wk in sorted((w for w in out if w), reverse=True):
        ordered[wk] = out[wk]
    if None in out:                     # not-yet-posted, last
        ordered[None] = out[None]
    return ordered


def _cell(ln: dict, label: str):
    raw = str(ln.get(_LABEL_TO_HEADER[label], "") or "").strip()
    if label in ("Order Date", "Status Date", "Install Date"):
        return _parse_date(raw) or raw
    return raw


def _write_header(sh, row: int, labels: Sequence[str]) -> None:
    b = _border()
    for c, label in enumerate(labels, start=1):
        cell = sh.cell(row=row, column=c, value=label)
        cell.font = _font("FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment, cell.border = CENTER, b


def _write_row(sh, row: int, ln: dict, labels: Sequence[str]) -> None:
    b = _border()
    hexfill = colors.fill_for(ln.get("DTR Status (enriched)", ""))
    fill = PatternFill("solid", fgColor=hexfill.lstrip("#")) if hexfill else None
    for c, label in enumerate(labels, start=1):
        cell = sh.cell(row=row, column=c, value=_cell(ln, label))
        cell.font = _font()
        cell.alignment = LEFT if label in ("Rep", "Customer Name",
                                           "Package") else CENTER
        cell.border = b
        if fill is not None:
            cell.fill = fill
        if label in ("Order Date", "Status Date", "Install Date"):
            cell.number_format = "mm/dd/yyyy"


def _banner(sh, row: int, text: str, span: int) -> None:
    sh.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = sh.cell(row=row, column=1, value=text)
    cell.font = _font("FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor=WEEK_BG)
    cell.alignment = LEFT


def _write_log(sh, lines, labels, *, freeze=True) -> None:
    row = 1
    for wk, group in by_week(lines).items():
        n = len(group)
        if wk:
            text = "Paycheck Week Ending {}  •  {} order{}".format(
                wk.strftime("%m/%d/%Y"), n, "" if n == 1 else "s")
        else:
            text = "Not Yet Posted  •  {} order{} (no paycheck yet)".format(
                n, "" if n == 1 else "s")
        _banner(sh, row, text, len(labels))
        row += 1
        _write_header(sh, row, labels)
        row += 1
        for ln in group:
            _write_row(sh, row, ln, labels)
            row += 1
        row += 2
    _autosize(sh, labels)
    if freeze:
        sh.freeze_panes = "A3"


def _write_flat(sh, lines, labels, *, freeze=True) -> None:
    """A flat overall view SORTED BY REP — no week banners (Megan 2026-07-20:
    "the all reps tab shouldn't be broken into payweeks, just an overall view
    sorted by rep"). Within a rep, newest order first."""
    ordered = sorted(
        lines,
        key=lambda ln: (str(ln.get("Rep", "") or "").lower(),
                        -( _parse_date(ln.get("sp.Order Date (copy)"))
                           or dt.date.min).toordinal()))
    _write_header(sh, 1, labels)
    row = 2
    for ln in ordered:
        _write_row(sh, row, ln, labels)
        row += 1
    _autosize(sh, labels)
    if freeze:
        sh.freeze_panes = "A2"


def _autosize(sh, labels: Sequence[str]) -> None:
    """Widen every column to fit its longest value so nothing is clipped
    (Megan 2026-07-20: "expand all cells to fit the text"). Scans ALL rows, not
    a 400-row sample, and the cap is high enough for the longest real values
    (Package strings ~40 chars); banners are merged so they don't force width."""
    banner_rows = set()
    for r in range(1, sh.max_row + 1):
        v = sh.cell(r, 1).value
        if isinstance(v, str) and ("Paycheck Week" in v or "Not Yet Posted" in v):
            banner_rows.add(r)
    for c, label in enumerate(labels, start=1):
        width = len(str(label)) + 2
        for r in range(1, sh.max_row + 1):
            if r in banner_rows:            # merged banner text, not a column
                continue
            v = sh.cell(row=r, column=c).value
            if v is not None:
                width = max(width, len(str(v)) + 2)
        sh.column_dimensions[get_column_letter(c)].width = min(width, 55) * 1.1


def build(lines: Sequence[dict], out_path: Path, *,
          today: Optional[dt.date] = None) -> Path:
    """Write the workbook: All Reps summary, Posted-by-Week, then a tab per rep."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    today = today or dt.date.today()
    wb = Workbook()
    used: set = set()

    # ---- 1. All Reps — flat overall view, sorted by rep -----------------
    sh = wb.active
    sh.title = _safe_title("All Reps", used)
    _write_flat(sh, lines, _LOG_LABELS)

    # ---- 2. Paycheck by Week (posted-date matrix) -----------------------
    _write_paycheck_matrix(wb, lines, used)

    # ---- 3. one tab per rep --------------------------------------------
    by_rep: Dict[str, list] = collections.defaultdict(list)
    for ln in lines:
        rep = str(ln.get("Rep", "") or "").strip()
        if rep:
            by_rep[rep].append(ln)
    # Per-rep tab drops the Rep column (redundant).
    rep_labels = [l for l in _LOG_LABELS if l != _REP_LABEL]
    for rep in sorted(by_rep):
        rsh = wb.create_sheet(_safe_title(rep, used))
        _write_log(rsh, by_rep[rep], rep_labels)

    wb.save(out_path)
    return out_path


def _write_paycheck_matrix(wb, lines, used) -> None:
    """Reps down, week-endings across, orders by their POSTED-date week — the
    paycheck matrix (Carlos 2026-07-20: pay is determined by the posted date).
    AT&T's twin of BOX's Accepted-by-Supplier payout. A sale with no posted date
    isn't on any paycheck yet, so it's excluded here (but still in the log)."""
    weeks = collections.Counter()
    posted = collections.defaultdict(int)     # (rep, posted-week) -> count
    reps = set()
    unposted = 0
    for ln in lines:
        rep = str(ln.get("Rep", "") or "").strip()
        if not rep:
            continue
        reps.add(rep)
        wk = _line_paycheck_week(ln)
        if not wk:
            unposted += 1
            continue
        posted[(rep, wk)] += 1
        weeks[wk] += 1
    if not reps:
        return
    weeks_desc = sorted(weeks, reverse=True)
    psh = wb.create_sheet(_safe_title("Paycheck by Week", used))
    psh.cell(row=1, column=1,
             value="Orders by paycheck week (posted date)").font = _font(bold=True)
    psh.cell(row=2, column=1,
             value=("Each column is the week a sale POSTED — that's the "
                    "paycheck it's on. Sales not yet posted ({} of them) aren't "
                    "on a paycheck and are excluded here, though they still "
                    "show in the log.").format(unposted)
             ).font = _font(italic=True)
    headers = ["Rep"] + [w.strftime("%m/%d") for w in weeks_desc] + ["Total"]
    _write_header(psh, 4, headers)
    r = 5
    for rep in sorted(reps,
                      key=lambda x: -sum(posted.get((x, w), 0)
                                         for w in weeks_desc)):
        vals = [rep]
        tot = 0
        for w in weeks_desc:
            n = posted.get((rep, w), 0)
            tot += n
            vals.append(n)
        vals.append(tot)
        for c, v in enumerate(vals, start=1):
            cell = psh.cell(row=r, column=c, value=v)
            cell.font = _font(bold=(c == len(headers)))
            cell.alignment = LEFT if c == 1 else CENTER
            cell.border = _border()
        r += 1
    psh.freeze_panes = "B5"
    _autosize(psh, headers)
