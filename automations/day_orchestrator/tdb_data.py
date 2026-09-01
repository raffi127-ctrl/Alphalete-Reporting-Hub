
"""Texas de Brazil competition — DATA LAYER (git-tracked).

Config + all data reads (sales board, recruiting, leadership, terminated reps)
and build_board. Extracted from the Report-Library report module (which is a
git-ignored Sheet cell capped at 50K chars) so the heavy logic lives in git and
the cell stays lean. The report does `from automations.day_orchestrator.tdb_data
import *` and keeps only the HTML/PDF rendering + Slack/iMessage delivery."""

import os, re, sys, html, glob, shutil, subprocess, tempfile, unicodedata, datetime, importlib, json, calendar, argparse, time
from collections import defaultdict

def _ensure(pkg):
    """Import a package, pip-installing it on first run if missing."""
    try:
        return importlib.import_module(pkg)
    except ImportError:
        print(f"Installing {pkg} (one-time)...")
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", pkg], check=True)
        return importlib.import_module(pkg)

_ensure("openpyxl")
_ensure("pypdf")
import openpyxl
from pypdf import PdfWriter

SALES_SHEET_ID   = "1MC9pfKryQrRtcMthUBL2hOciDCaa83U059pz0N2CmHc"
RECRUIT_SHEET_ID = "1Ez-mbROADd5aCWbLak6kQkNapb-BEk9W81n2ln6DVB4"
SALES_GLOB   = os.path.expanduser("~/Downloads/Alphalete SALES BOARD 2025*.xlsx")
RECRUIT_GLOB = os.path.expanduser("~/Downloads/All in One Local Office - Raf*.xlsx")
RECRUIT_TAB  = "2nd rds %s"
ESTIMATED_MINUTES = 2

_anchor    = datetime.date.today() - datetime.timedelta(days=1)
COMP_YEAR  = int(os.environ.get("TDB_COMP_YEAR")  or _anchor.year)
COMP_MONTH = int(os.environ.get("TDB_COMP_MONTH") or _anchor.month)
MONTH_NAME = datetime.date(COMP_YEAR, COMP_MONTH, 1).strftime("%B")
MONTH_UP   = MONTH_NAME.upper()
MONTH_LAST = calendar.monthrange(COMP_YEAR, COMP_MONTH)[1]
WIN        = 10
PERIOD     = f"{COMP_YEAR}-{COMP_MONTH:02d}"
# Best Car Ride Leader retired from AUGUST on (Raf, 2026-07-31). Gated on the
# competition PERIOD, not today, so July still renders with it.
CAR_ON     = PERIOD < "2026-08"

MANUAL_INPUTS = os.path.expanduser("~/recruiting-report/output/texas_de_brazil_manual.json")

DINNER_DAY_DEFAULT  = "TO BE DETERMINED"
DINNER_TIME_DEFAULT = ""

LEADERS_STATE = os.path.expanduser("~/recruiting-report/output/texas_de_brazil_leaders_state.json")

# Not competing reps (owners/mgmt). A name here scores nothing and never shows
# on the board — not even through Break-a-Leader (their promoted rep still gets
# paid; only the excluded side is skipped).
EXCLUDE      = {"Rafael Hidalgo", "Joshua Mascorro",
                "Basil Elhassan"}   # Bas doesn't compete (Eve, 2026-09-01)
ALIAS        = {"Andrew Sanborn Roadtrip": "Andrew Sanborn", "Randy Amoo": "Randy Amoa",
                "Sebastian Guerrero": "SABASTIN GUERRERO",
                "Chole Johnson": "Chloe Johnson",

                "Drew": "Andrew Sanborn", "D": "Deavion Allen", "Zoey": "Zoria Johnson",
                "Al": "Algemar Kennel", "Bas": "Basil Elhassan",
                "Keiah": "Lakeaih Gregory"}

PROMOTIONS_BY_MONTH = {
    "2026-07": [
        ("Willie Henderson", "Jessie Gomez"),
        ("Willie Henderson", "Jordan Ruiz"),
        ("Safiya Mahmoud", "Abel Mireles"),
    ],
}
SOLO_LEADERS_BY_MONTH = {
}

# Leaders who run a crew and no longer sell. They have NO row on the sales board,
# and the board's column B IS the roster — so they were invisible to every point
# that isn't a sale: Break-a-Leader, 2nd rounds, new starts, adjustments all land
# through resolve_roster() against that roster. Listed here, they join with a
# zeroed board line and everything else pays normally.
# (Basil Elhassan was here for August 2026; he's in EXCLUDE now — doesn't compete.)
NON_SELLING_LEADERS = [
]
CAR_RIDE_LEADERS_BY_MONTH = {
    "2026-07": [
        "Jordan Ruiz",
        "Kaleb Muvunyi",
    ],
}

# Manual point corrections, scoped BY MONTH so a one-off fix never leaks into the
# next month. Key = 'YYYY-MM'. Empty months need no entry.
ADJUSTMENTS_BY_MONTH = {
    "2026-07": {"Algemar Kennel": 15},
}

EXCLUDE_NEW_LEADERS = {"Giselle Loredo"}   # dropped even if auto-detected

# Promoter corrections, keyed by NEW LEADER, scoped BY MONTH.
# Auto-detect credits the Break-a-Leader to whoever the sales board's 'Trainer'
# column names at the moment the rep flips to Level 1 — so a wrong/stale Trainer
# pays the wrong person and the flyer prints the wrong arrow. The pair is already
# frozen in the local json AND the shared store by then (union-only, first
# sighting wins), and a manual pair would ADD a second row rather than replace
# it, so the fix has to be a rewrite at read time. Applied to BOTH readers
# (update_leaders_state for points, load_leaders_state for the flyer), never
# written back to the state file.
PROMOTER_FIXES_BY_MONTH = {
    # Board Trainer said Anthony Coca; Algemar: it's Anthony Marchetti (2026-08-28).
    "2026-08": {"Giovanna Santos": "Anthony Marchetti"},
}

POS_HERE = {"Here", "H+DC", "RT", "H+LM"}
LATE_PEN = {"Late"}
OFF_PEN  = {"Off", "STF", "O-NA"}
REMOVE   = {"T"}

HERE_PTS      = 3
SHOW_PTS      = 10
ENERGY_PTS    = 1
ENERGY5_BONUS = 10
CARRIDE_PTS   = 10

INT_OFF    = 6
DTV_OFF    = 4
NL_OFF     = 3
ENERGY_OFF = 2
DAY_OFF    = 7

# Header text -> the day-block metric it feeds. Each 'Roll Call' column closes a
# 7-cell day block (Apps / Int / Int Up / DTV / NL / EN / Cx); we locate the four
# we score BY HEADER, because the offsets are not stable: on the 'Sales Board
# WE 9.6' tab the 6th cell was renamed EN (energy sales) -> TK (total knocks),
# and reading it positionally scored ~200 knocks/day as ~200 energy points
# (2026-09-01 — it re-ranked the whole final August board). 'TK' is deliberately
# absent from this map: an unmapped header scores NOTHING instead of guessing.
DAY_COL_HEADERS = {"int": "int", "dtv": "dtv", "nl": "nl", "en": "energy"}

def norm(name):
    n = re.sub(r"\([^)]*\)", " ", str(name))
    n = unicodedata.normalize("NFKD", n)
    n = "".join(c for c in n if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", n).strip()

def akey(raw):
    # norm + ALIAS: one ALIAS row fixes a name in every reader
    s = str(raw).strip()
    if s in ALIAS:
        return norm(ALIAS[s])
    n = norm(s)
    return norm(ALIAS.get(n, n))

def resolve_roster(name, rized):
    """Name/nickname -> roster key: exact, else unique first-name prefix."""
    if not name:
        return None
    n = akey(name)
    if not n:
        return None
    if n in rized:
        return n
    toks = n.lower().split()
    first = toks[0]
    cands = []
    for k in rized:
        kt = k.lower().split()
        if not kt:
            continue
        if kt[0].startswith(first) or first.startswith(kt[0]):
            cands.append(k)
    if len(toks) >= 2 and len(cands) > 1:
        nar = [k for k in cands if len(k.lower().split()) >= 2
               and k.lower().split()[1].startswith(toks[1])]
        if nar:
            cands = nar
    cands = list(dict.fromkeys(cands))
    return cands[0] if len(cands) == 1 else None

def numv(x):
    return float(x) if isinstance(x, (int, float)) else 0.0

def fv(x):
    try: return float(x)
    except (TypeError, ValueError): return 0.0

def roster_end(raw):
    """True at the TOTALS row closing a roster (blank rows are GAPS, not the end)."""
    return isinstance(raw, str) and raw.strip().upper().startswith("TOTAL")

def newest(pattern):
    fs = glob.glob(pattern)
    return max(fs, key=os.path.getmtime) if fs else None

# Drive fetch: how hard to try before giving up. One attempt used to be enough
# "most days" — until 2026-08-13, when a single ReadTimeout on the recruiting
# workbook killed the whole 8am post (the ~/Downloads fallback below is a
# LAPTOP path; the mini has no such file, so a failed fetch there is fatal).
FETCH_TRIES    = int(os.environ.get("TDB_FETCH_TRIES") or 3)
FETCH_TIMEOUT  = int(os.environ.get("TDB_FETCH_TIMEOUT") or 120)
FETCH_BACKOFF  = (5, 15)     # seconds to wait before try 2, try 3, ...

def fetch_from_drive(sheet_id, label, workdir):
    """Sheet -> temp .xlsx via the Hub login. None if unavailable, so the caller
    falls back to ~/Downloads.

    Retries transient failures (FETCH_TRIES attempts, FETCH_BACKOFF between) —
    Google's export endpoint stalls for a minute now and then, and each workbook
    is independent, so a slow recruiting pull no longer discards a good sales
    pull. Lives HERE (git-tracked) rather than in the Report-Library cell so the
    retry has a diff and a history; the cell must NOT redefine it, or the cell's
    copy shadows this one."""
    try:
        from automations.recruiting_report import fill as _fill
    except Exception:
        return None                      # no creds layer at all — retrying won't help
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    for attempt in range(1, FETCH_TRIES + 1):
        try:
            sh = _fill.open_by_key(sheet_id)
            sess = sh.client.session
            r = sess.get(url, timeout=FETCH_TIMEOUT); r.raise_for_status()
            path = os.path.join(workdir, f"{label}.xlsx")
            with open(path, "wb") as fh:
                fh.write(r.content)
            note = "" if attempt == 1 else f" (attempt {attempt}/{FETCH_TRIES})"
            print(f"{label:11}: live from Google Drive{note}")
            return path
        except Exception as e:
            why = f"{type(e).__name__}: {e}"
            if attempt < FETCH_TRIES:
                wait = FETCH_BACKOFF[min(attempt - 1, len(FETCH_BACKOFF) - 1)]
                print(f"({label} fetch attempt {attempt}/{FETCH_TRIES} failed — "
                      f"{why}; retrying in {wait}s)")
                time.sleep(wait)
            else:
                print(f"({label} live fetch unavailable after {FETCH_TRIES} "
                      f"attempts: {why}; trying Downloads)")
    return None

def find_chrome():
    """Locate Chrome/Chromium (mac/win/linux)."""
    for name in ("google-chrome", "google-chrome-stable", "chromium", "chromium-browser", "chrome"):
        p = shutil.which(name)
        if p:
            return p
    cands = []
    if sys.platform == "darwin":
        cands += [os.path.join(os.sep, "Applications", "Google Chrome.app", "Contents", "MacOS", "Google Chrome"),
                  os.path.join(os.sep, "Applications", "Chromium.app", "Contents", "MacOS", "Chromium")]
    elif sys.platform.startswith("win"):
        for base in (os.environ.get("PROGRAMFILES", r"C:\Program Files"),
                     os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"),
                     os.environ.get("LOCALAPPDATA", "")):
            if base:
                cands.append(os.path.join(base, "Google", "Chrome", "Application", "chrome.exe"))
    else:
        cands += ["/usr/bin/google-chrome", "/usr/bin/chromium-browser", "/usr/bin/chromium",
                  "/snap/bin/chromium"]
    for c in cands:
        if c and os.path.exists(c):
            return c
    sys.exit("ERROR: Google Chrome not found. Install Chrome, then re-run.")

def day_block_cols(hdr, rc):
    """The scored columns of the day block closing at Roll Call column `rc`,
    located by HEADER TEXT (see DAY_COL_HEADERS) -> {'int','dtv','nl','energy'}.

    A metric whose header isn't in the block is simply ABSENT from the result
    (the caller scores it 0) — that's how a renamed column stops paying points
    for the wrong number. Only when the block has NO recognizable header at all
    (an old tab, a blank header row) do we fall back to the fixed offsets."""
    out = {}
    for j in range(max(0, rc - DAY_OFF), rc):
        c = hdr[j] if j < len(hdr) else None
        if isinstance(c, str):
            key = DAY_COL_HEADERS.get(c.strip().lower())
            if key and key not in out:
                out[key] = j
    if not out:
        return {"int": rc - INT_OFF, "dtv": rc - DTV_OFF,
                "nl": rc - NL_OFF, "energy": rc - ENERGY_OFF}
    return out

def sales_week_tabs(wb):
    tabs = [n for n in wb.sheetnames if re.match(r"Sales Board WE \d", n)]
    def k(n):
        m = re.search(r"WE\s+(\d{1,2})\.(\d{1,2})", n)
        return (int(m.group(1)), int(m.group(2))) if m else (99, 99)
    return sorted(tabs, key=k)

def tab_week_dates(tabname):
    """The 7 dates of the week ENDING on the WE date."""
    m = re.search(r"WE\s+(\d{1,2})\.(\d{1,2})", tabname)
    if not m:
        raise ValueError("no WE date in tab name")
    mo, dy = int(m.group(1)), int(m.group(2))
    end = datetime.date(COMP_YEAR, mo, dy)
    return [end - datetime.timedelta(days=6 - i) for i in range(7)]

def read_sales(sales_file):
    wb = openpyxl.load_workbook(sales_file, read_only=True, data_only=True)
    sales = defaultdict(lambda: {"int": 0.0, "dtv": 0.0, "nl": 0.0, "energy": 0.0,
                                 "here": 0, "late": 0, "off": 0, "int3": 0, "energy5": 0})
    daily = defaultdict(dict)   # name -> {iso_date: {per-day raw + attendance}} for the drill-down
    removed = {}
    through = None
    today = datetime.date.today()
    for tab in sales_week_tabs(wb):
        try:
            dates = tab_week_dates(tab)
        except ValueError:
            continue
        if not any(d.month == COMP_MONTH and d.year == COMP_YEAR for d in dates):
            continue
        ws = wb[tab]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        try:
            shr = next(i for i, r in enumerate(rows)
                       if any(isinstance(c, str) and c.strip() == "Roll Call" for c in r))
        except StopIteration:
            continue
        rc_cols = [j for j, c in enumerate(rows[shr]) if isinstance(c, str) and c.strip() == "Roll Call"]
        daterow = rows[shr - 1]
        label2date = {d.day: d for d in dates}

        rc_date = {}
        rc_cmap = {}
        for i, rc in enumerate(rc_cols):
            rc_cmap[rc] = day_block_cols(rows[shr], rc)
            lbl = daterow[rc - DAY_OFF] if rc - DAY_OFF >= 0 else None
            dt = None
            if lbl is not None:
                try:
                    dt = label2date.get(int(lbl))
                except (TypeError, ValueError):
                    dt = None
            if dt is None and i < len(dates):
                dt = dates[i]
            rc_date[rc] = dt
        for r in rows[shr + 1:]:
            raw = r[2] if len(r) > 2 else None
            # Roster ends at TOTALS; a blank row is a GAP. Breaking on one cut
            # this to 16 of 75 reps after a re-sort left holes (2026-07-31).
            if roster_end(raw):
                break
            if not (isinstance(raw, str) and raw.strip()):
                continue
            name = akey(raw); rec = sales[name]
            for rc in rc_cols:
                dt = rc_date.get(rc)
                if dt is None or dt.month != COMP_MONTH or dt.year != COMP_YEAR:
                    continue
                if dt >= today:
                    continue
                is_sun = dt.weekday() == 6
                cm = rc_cmap[rc]
                def _cell(key, _r=r, _cm=cm):
                    j = _cm.get(key)
                    return numv(_r[j]) if (j is not None and j < len(_r)) else 0.0
                dayint = _cell("int"); dayeng = _cell("energy")
                daydtv = _cell("dtv"); daynl = _cell("nl")
                rec["int"] += dayint; rec["dtv"] += daydtv; rec["nl"] += daynl
                rec["energy"] += dayeng
                if dayint >= 3:
                    rec["int3"] += 1
                if dayeng >= 3:
                    rec["energy5"] += 1
                got = dayint > 0 or dayeng > 0
                v = r[rc]
                att = None   # (label, pts) for this day's attendance code, if any
                if isinstance(v, str) and v.strip():
                    got = True
                    v = v.strip()
                    if v in REMOVE: removed[name] = v
                    elif v in POS_HERE:
                        rec["here"] += 1; att = ("Here / On-Time / Dress", HERE_PTS)
                    elif v in LATE_PEN:
                        if not is_sun:
                            rec["late"] += 1; att = ("Late", -5)
                    elif v in OFF_PEN:
                        if not is_sun:
                            rec["off"] += 1; att = ("Off / STF / No-Answer", -10)
                if dayint or daydtv or daynl or dayeng or att:
                    daily[name][dt.isoformat()] = {
                        "int": dayint, "dtv": daydtv, "nl": daynl, "eng": dayeng,
                        "int3": dayint >= 3, "eng3": dayeng >= 3, "att": att,
                    }
                if got and (through is None or dt > through):
                    through = dt
    wb.close()
    return sales, removed, through, daily


def _rep_days(daymap):
    """Per-rep day-by-day point items for the drill-down. Each item is
    [label, points, count, unit] — count/unit are None for flat bonuses and
    attendance. Each day's items sum to that day's swing; every point here is
    date-attributable (sales board)."""
    out = []
    for iso in sorted(daymap):
        d = daymap[iso]
        items = []
        if d["int"]:  items.append(["Internet", d["int"] * 2, int(d["int"]), 2])
        if d["int3"]: items.append(["3+ Internet", 10, None, None])
        if d["eng"]:  items.append(["Energy", d["eng"] * ENERGY_PTS, int(d["eng"]), ENERGY_PTS])
        if d["eng3"]: items.append(["3+ Energy", ENERGY5_BONUS, None, None])
        if d["dtv"]:  items.append(["DTV", d["dtv"], int(d["dtv"]), 1])
        if d["nl"]:   items.append(["New Line", d["nl"], int(d["nl"]), 1])
        if d["att"]:  items.append([d["att"][0], d["att"][1], None, None])
        if items:
            out.append({"d": iso, "items": items, "tot": sum(p for _, p, _, _ in items)})
    return out

def terminated_reps(sales_file):
    """akey names with a 'T' (terminate) roll-call code anywhere on the board.

    Month-independent on purpose: a rep terminated in ANY week is dropped from the
    flyer going forward, so we never celebrate a promotion / car-ride for someone
    who's no longer with the company. Board points are handled separately (read_sales
    already excludes T'd reps), so this only cleans up the promotions display."""
    out = set()
    try:
        wb = openpyxl.load_workbook(sales_file, read_only=True, data_only=True)
    except Exception:
        return out
    for tab in sales_week_tabs(wb):
        ws = wb[tab]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        try:
            shr = next(i for i, r in enumerate(rows)
                       if any(isinstance(c, str) and c.strip() == "Roll Call" for c in r))
        except StopIteration:
            continue
        rc_cols = [j for j, c in enumerate(rows[shr]) if isinstance(c, str) and c.strip() == "Roll Call"]
        for r in rows[shr + 1:]:
            raw = r[2] if len(r) > 2 else None
            if roster_end(raw):
                break
            if not (isinstance(raw, str) and raw.strip()):
                continue
            for c in rc_cols:
                v = r[c] if c < len(r) else None
                if isinstance(v, str) and v.strip() in REMOVE:
                    out.add(akey(raw)); break
    wb.close()
    return out

def read_recruiting(recruit_file):
    """Per-month 2nd-round accepts + new-starts-showed for the COMPETITION month.

    The tab lays out one block PER MONTH side by side, each headed
    '<Month> BOB Call / Stats' (Accepted col) with a 'New starts showed' col and
    its own 'Leader Name' column. We find the block whose header names MONTH_NAME
    and read only that one — so July counts July, August counts August, and a
    month with no block yet correctly scores ZERO (no stale carry-over). Columns
    are located by header, never fixed index."""
    wb = openpyxl.load_workbook(recruit_file, read_only=True, data_only=True)
    ws = wb[RECRUIT_TAB]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    out = {}
    if len(rows) < 2:
        wb.close(); return out
    h0, h1 = rows[0], rows[1]
    target = MONTH_NAME.lower()
    acc_col = name_col = show_col = None
    for j, c in enumerate(h0):
        if isinstance(c, str) and "bob call" in c.lower() and target in c.lower():
            acc_col = j
            for k in range(j, -1, -1):                       # block's Leader Name (to the left)
                if isinstance(h0[k], str) and h0[k].strip().lower() == "leader name":
                    name_col = k; break
            for k in range(j, min(j + 8, len(h0))):          # 'New starts showed' (to the right)
                if isinstance(h0[k], str) and "new starts showed" in h0[k].strip().lower():
                    show_col = k; break
            break
    if acc_col is None or name_col is None:
        print(f"Recruiting: no '{MONTH_NAME}' block on {RECRUIT_TAB!r} yet — 0 recruiting points this month")
        wb.close(); return out
    for r in rows[2:]:
        nm = r[name_col] if name_col < len(r) else None
        if roster_end(nm):
            break
        if not (isinstance(nm, str) and nm.strip()):
            continue
        acc = fv(r[acc_col]) if acc_col < len(r) else 0.0
        show = fv(r[show_col]) if (show_col is not None and show_col < len(r)) else 0.0
        out[nm.strip()] = (acc, show)
    wb.close()
    return out

def read_leadership(sales_file):
    """Latest weekly tab -> {name: {status, trainer, best}}. Columns found by HEADER
    TEXT, not fixed letters."""
    wb = openpyxl.load_workbook(sales_file, read_only=True, data_only=True)
    out = {}
    for tab in reversed(sales_week_tabs(wb)):
        ws = wb[tab]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        try:
            shr = next(i for i, r in enumerate(rows)
                       if any(isinstance(c, str) and c.strip() == "Roll Call" for c in r))
        except StopIteration:
            continue

        def find_col(label):
            for r in rows[:shr + 1]:
                for j, c in enumerate(r):
                    if isinstance(c, str) and c.strip().lower() == label:
                        return j
            return None
        c_status = find_col("leadership status")
        c_train = find_col("trainer")
        c_best = find_col("best car ride leader")
        if c_status is None:
            continue
        for r in rows[shr + 1:]:
            raw = r[2] if len(r) > 2 else None
            if roster_end(raw):
                break
            if not (isinstance(raw, str) and raw.strip()):
                continue
            name = akey(raw)
            status = r[c_status] if c_status < len(r) else None
            trainer = r[c_train] if (c_train is not None and c_train < len(r)) else None
            best = r[c_best] if (c_best is not None and c_best < len(r)) else None
            out[name] = {
                "status": status.strip() if isinstance(status, str) else "",
                "trainer": trainer.strip() if isinstance(trainer, str) else "",
                "best": isinstance(best, str) and best.strip().upper() == "BEST",
            }
        break
    wb.close()
    return out

def _fix_promoters(pairs):
    """Rewrite the promoter of any pair listed in PROMOTER_FIXES_BY_MONTH."""
    fixes = {norm(k): v for k, v in
             PROMOTER_FIXES_BY_MONTH.get(_current_period(), {}).items()}
    if not fixes:
        return [list(p) for p in pairs]
    return [[fixes.get(norm(p[1]), p[0]), p[1]] for p in pairs]

def update_leaders_state(leadership):
    """Accumulate promotions to 'Level 1' + car-ride 'BEST'; first sighting seeds."""
    try:
        state = json.loads(open(LEADERS_STATE).read())
    except Exception:
        state = {}

    if (state.get("period") or _current_period()) != _current_period():
        state = {}
    baseline = state.get("baseline") or {}
    promos = state.get("new_leaders") or []
    cars = state.get("car_ride") or []
    first_run = not baseline
    detected = {p[1] for p in promos}
    car_set = set(cars)
    for nm, d in leadership.items():
        if nm not in baseline:
            baseline[nm] = d["status"]
        elif d["status"] == "Level 1" and baseline[nm] != "Level 1" and nm not in detected:
            promos.append([d["trainer"], nm]); detected.add(nm)
        if d.get("best") and nm not in car_set:
            cars.append(nm); car_set.add(nm)
    try:
        os.makedirs(os.path.dirname(LEADERS_STATE), exist_ok=True)
        with open(LEADERS_STATE, "w") as fh:
            json.dump({"period": _current_period(), "baseline": baseline,
                       "new_leaders": promos, "car_ride": cars}, fh, indent=2)
    except Exception as e:
        print(f"(couldn't save leaders state: {e})")
    if first_run:
        print(f"Leaders baseline set for {len(baseline)} reps")
    return _fix_promoters(promos), cars

def load_leaders_state():
    """Read-only view of auto-detected leaders, for the flyer."""
    try:
        state = json.loads(open(LEADERS_STATE).read())
    except Exception:
        return [], []
    if (state.get("period") or _current_period()) != _current_period():
        return [], []
    return _fix_promoters(state.get("new_leaders") or []), state.get("car_ride") or []

def load_manual_inputs():
    """Hub-typed additions -> (promotions, solo, car_ride). Never raises."""
    prom, solo, car = [], [], []
    try:
        data = json.loads(open(MANUAL_INPUTS).read())
    except Exception:
        return prom, solo, car
    for line in str(data.get("new_leaders_text", "") or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if ">" in line:
            a, b = line.split(">", 1)
            a, b = a.strip(), b.strip()
            if a and b:
                prom.append((a, b))
            elif b:
                solo.append(b)
        else:
            solo.append(line)
    for line in str(data.get("car_ride_text", "") or "").splitlines():
        line = line.strip()
        if line:
            car.append(line)
    return prom, solo, car

def load_dinner():
    """This month's dinner date from dinner_schedule["YYYY-MM"], else the legacy
    dinner_day/time, else "TO BE DETERMINED"."""
    day, time = DINNER_DAY_DEFAULT, DINNER_TIME_DEFAULT
    try:
        data = json.loads(open(MANUAL_INPUTS).read())
    except Exception:
        return day, time
    entry = (data.get("dinner_schedule") or {}).get(_current_period()) or {}
    d = str(entry.get("day", "") or "").strip() or str(data.get("dinner_day", "") or "").strip()
    t = str(entry.get("time", "") or "").strip() or str(data.get("dinner_time", "") or "").strip()
    return (d or day), (t or time)

def _current_period():
    """The competition period key ('YYYY-MM') used to auto-reset month state."""
    return f"{COMP_YEAR}-{COMP_MONTH:02d}"

def build_board(sales_file, recruit_file):
    sales, removed, through, daily = read_sales(sales_file)
    recruit = read_recruiting(recruit_file)
    m_prom, m_solo, m_car = load_manual_inputs()
    a_prom, a_car = update_leaders_state(read_leadership(sales_file))
    _mp = _current_period()
    _excl = {norm(x) for x in EXCLUDE_NEW_LEADERS}
    promotions   = [(p, q) for (p, q) in dict.fromkeys(PROMOTIONS_BY_MONTH.get(_mp, []) + m_prom + [tuple(p) for p in a_prom]) if norm(q) not in _excl]
    solo_leaders = [q for q in dict.fromkeys(SOLO_LEADERS_BY_MONTH.get(_mp, []) + m_solo) if norm(q) not in _excl]
    car_leaders  = list(dict.fromkeys(CAR_RIDE_LEADERS_BY_MONTH.get(_mp, []) + m_car + a_car))
    if a_prom or a_car:
        print(f"Auto-detected: {len(a_prom)} new leaders, {len(a_car)} car-ride")
    if m_prom or m_solo or m_car:
        print(f"Manual overrides: {len(m_prom)} promotions, {len(m_solo)} solo, {len(m_car)} car-ride")

    rized = {}
    for name, s in sales.items():
        if name in EXCLUDE or name in removed:
            continue
        rized[name] = {
            "name": name, "int_p": s["int"] * 2, "dtv_p": s["dtv"], "nl_p": s["nl"],
            "eng_p": s["energy"] * ENERGY_PTS,
            "att_p": s["here"] * HERE_PTS - s["late"] * 5 - s["off"] * 10,
            # att_p split out so the scoring log can show losses on their own
            # (tdb_scoring_log). here_p + late_p + off_p == att_p.
            "here_p": s["here"] * HERE_PTS, "late_p": -s["late"] * 5, "off_p": -s["off"] * 10,
            "i3_p": s["int3"] * 10, "e5_p": s["energy5"] * ENERGY5_BONUS,
            "acc": 0.0, "show": 0.0, "acc_p": 0.0, "show_p": 0.0, "brk_p": 0.0, "car_p": 0.0,
            "adj_p": 0.0,
        }
    # Crew leaders with no board row join with a zeroed sales line (see
    # NON_SELLING_LEADERS) so the non-sales points can find them. A leader who is
    # excluded, terminated, or already on the board is left alone.
    for _nm in NON_SELLING_LEADERS:
        _k = akey(_nm)
        if _k in EXCLUDE or _k in removed or _k in rized:
            continue
        rized[_k] = {
            "name": _k, "int_p": 0.0, "dtv_p": 0.0, "nl_p": 0.0, "eng_p": 0.0,
            "att_p": 0.0, "here_p": 0.0, "late_p": 0.0, "off_p": 0.0,
            "i3_p": 0.0, "e5_p": 0.0,
            "acc": 0.0, "show": 0.0, "acc_p": 0.0, "show_p": 0.0, "brk_p": 0.0,
            "car_p": 0.0, "adj_p": 0.0,
        }
    ci = {k.lower(): k for k in rized}
    for rn, (acc, show) in recruit.items():
        key = akey(rn)
        # case-only spelling differences between the two sheets must still match
        key = key if key in rized else ci.get(key.lower(), key)
        if key in EXCLUDE or key not in rized:
            continue
        rized[key]["acc"] += acc; rized[key]["show"] += show
        rized[key]["acc_p"] += acc * 5; rized[key]["show_p"] += show * SHOW_PTS
    unmatched = []
    for promoter, newleader in promotions:
        # A pair whose two sides resolve to the SAME rep is a solo promotion, not
        # a Break-a-Leader: the board's Trainer cell named the rep themself (or a
        # nickname of them). Paying both sides handed them +30 for one promotion
        # — Edgar Camunez and Juan Pablo Deleon in August 2026.
        paid = set()
        for nm in (promoter, newleader):
            if akey(nm) in EXCLUDE:
                continue   # excluded side pays nothing and isn't an unmatched name
            key = resolve_roster(nm, rized)
            if key:
                if key in paid:
                    continue
                paid.add(key)
                rized[key]["brk_p"] += 15
            elif str(nm).strip():
                unmatched.append(nm)
    for nm in solo_leaders:
        key = resolve_roster(nm, rized)
        if key:
            rized[key]["brk_p"] += 15
    for nm in (car_leaders if CAR_ON else []):
        key = resolve_roster(nm, rized)
        if key:
            rized[key]["car_p"] += CARRIDE_PTS
        elif str(nm).strip():
            unmatched.append(nm)
    if unmatched:
        print("Unmatched names (no points; check spelling / give me the nickname): "
              + ", ".join(sorted(set(str(u).strip() for u in unmatched))))
    for nm, pts in ADJUSTMENTS_BY_MONTH.get(_mp, {}).items():
        key = resolve_roster(nm, rized)
        if key:
            rized[key]["adj_p"] += pts

    board = list(rized.values())
    for r in board:
        r["total"] = (r["int_p"] + r["dtv_p"] + r["nl_p"] + r["att_p"] + r["eng_p"]
                      + r["acc_p"] + r["show_p"] + r["brk_p"] + r["i3_p"]
                      + r["e5_p"] + r["car_p"] + r["adj_p"])
        # Drill-down: date-attributable day rows + the month-level bonuses that
        # have no single date. days' swings + month_extra == total.
        r["days"] = _rep_days(daily.get(r["name"], {}))
        # [label, points, count, unit] — count/unit None where there's no natural
        # multiplier (Adjustment). Lets the drill-down show "2nd Round 9 × +5".
        month_extra = [
            ["2nd Round", r["acc_p"], int(r["acc"]), 5],
            ["New Start Showed", r["show_p"], int(r["show"]), 10],
            ["Break-a-Leader", r["brk_p"], int(round(r["brk_p"] / 15)), 15],
            ["Adjustment", r["adj_p"], None, None],
        ]
        if CAR_ON:
            month_extra.insert(2, ["Car Ride", r["car_p"], int(round(r["car_p"] / CARRIDE_PTS)), CARRIDE_PTS])
        r["month_extra"] = [it for it in month_extra if it[1]]
    board.sort(key=lambda x: (-x["total"], -x["acc"], x["name"]))
    return board, through

