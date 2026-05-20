"""Parse the emailed B2B.D2D First/Last Sale .xlsx file.

The file has 3 channel tabs (B2B / RES IF / RES OOF). Each row =
(Channel, Captains?, Owner Name, day-label, first-sale-time,
last-sale-time, order-count). RES OOF skips the 'Captains' column —
columns are matched by HEADER NAME, not position.

The week the data covers is encoded in the FILENAME:
  'B2B.D2D First Last Sale WE 5.10.2026.xlsx' -> date(2026, 5, 10)
We read the filename, not the file's modified date — the modified date is
unreliable (download time, not data week).

Public API:
  parse_filename_week(path) -> datetime.date
  parse_all(path) -> {owner_norm: {channel, raw_name, days{day: (first,last,orders)}, _sheets}}
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Dict, Set, Tuple, Any

import openpyxl


def _norm(s: str) -> str:
    """Strip all non-alphanum + lowercase. Matches the rep-name normalization
    used elsewhere; tolerant of hyphen vs space vs extra whitespace."""
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def parse_filename_week(path: Path) -> dt.date:
    """Filename pattern: 'B2B.D2D First Last Sale WE M.D.YYYY.xlsx'."""
    m = re.search(r"WE\s*(\d{1,2})\.(\d{1,2})\.(\d{4})", Path(path).name, re.I)
    if not m:
        raise ValueError(f"can't parse WE date from filename: {Path(path).name}")
    mo, d, y = (int(x) for x in m.groups())
    return dt.date(y, mo, d)


def parse_all(path: Path) -> Dict[str, Dict[str, Any]]:
    """Open the .xlsx and return per-owner data.

    Match columns by HEADER NAME (so the RES OOF tab without 'Captains'
    works). An owner appearing in 2+ sheets is flagged via `_sheets` so the
    caller can warn — Megan: each ICD should be in only 1 channel."""
    wb = openpyxl.load_workbook(Path(path), data_only=True)
    out: Dict[str, Dict[str, Any]] = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [ws.cell(row=1, column=c).value
                   for c in range(1, ws.max_column + 1)]
        try:
            i_chan = headers.index("Channel")
            i_owner = headers.index("Owner Name")
            i_label = headers.index("Week Avg")
            i_first = headers.index("First Sale Avg Office")
            i_last = headers.index("Last Sale Avg Office")
            i_orders = headers.index("Order Count")
        except ValueError:
            continue
        i_cap = headers.index("Captains") if "Captains" in headers else None
        for r in range(2, ws.max_row + 1):
            owner = ws.cell(row=r, column=i_owner + 1).value
            if not owner:
                continue
            key = _norm(owner)
            ent = out.setdefault(key, {
                "channel": ws.cell(row=r, column=i_chan + 1).value,
                "captain": (ws.cell(row=r, column=i_cap + 1).value
                            if i_cap is not None else None),
                "raw_name": str(owner).strip(),
                "days": {},
                "_sheets": set(),  # type: Set[str]
            })
            ent["_sheets"].add(sname)
            day = (ws.cell(row=r, column=i_label + 1).value or "").strip()
            if day:
                ent["days"][day] = (
                    ws.cell(row=r, column=i_first + 1).value,
                    ws.cell(row=r, column=i_last + 1).value,
                    ws.cell(row=r, column=i_orders + 1).value,
                )
    return out
