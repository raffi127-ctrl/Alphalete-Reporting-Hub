"""Vantura Tableau pulls — Order Log per owner + Churn Rates reconciliation.

Views (ATTTRACKER-B2B workbook, same sci site the Hub already uses):
  * ORDERLOG — the raw order crosstab. Filtered per owner via URL params
    (Start Date / End Date / Owner & Office), same technique as
    country_metrics — no fighting the date-box UI.
  * CHURNRATES — the dashboard the humans reconcile against. One crosstab
    carries every owner (rows per Owner & Office, sub-rows 'Activated
    SPE/SP' / disconnect count / 'Churn Rate', bucket columns '0-30 Day'…),
    so a single download covers both Carlos and Atef.

The 60-day order window is deliberate (runbook Part 3.2): the Order Log
filters by ORDER date but churn buckets by POSTED date, so a 30-day pull
silently drops customers ordered >30 days ago who activated recently.
"""
from __future__ import annotations

import datetime as dt
import re
import warnings
from pathlib import Path
from urllib.parse import quote

from automations.shared.tableau_patchright import download_crosstab_patchright

ORDERLOG_URL_TMPL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "ATTTRACKER-B2B/ORDERLOG"
    "?:iid=1&Start%20Date={start}&End%20Date={end}")
ORDERLOG_SHEET = "Order Log"

CHURNRATES_URL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "ATTTRACKER-B2B/CHURNRATES/429cb06d-a32e-4d0e-bf06-9acb77587afd/"
    "ALLTEAMCHURN?:iid=1")
CHURNRATES_SHEET = "ICD Churn"

# The product breakout added to CHURNRATES on 2026-08-30. Probed live off
# ALLTEAMCHURN: 5 values — Total, AIR/AWB, BYOD WIRELESS, NON BYOD WIRELESS,
# NEW INTERNET. 'Total' is a roll-up and is never summed.
PRODUCT_COL = "Product Type (Broken Out) (BYOD/Non BYOD)"
PRODUCT_TOTAL = "TOTAL"

# The crosstab's unnamed measure column. 'Disconnect count (SPE/SP)' is the name
# the export actually uses — this module looked for "Disconnects", which has
# never matched anything; both are accepted so a Tableau rename can't silently
# take the derived rate away again.
_DISCONNECT_MEASURES = ("Disconnect count (SPE/SP)", "Disconnects")
_COUNT_MEASURES = ("Activated SPE/SP",) + _DISCONNECT_MEASURES
_MEASURES = (("Activated SPE/SP", "Calculation1 (1)", "Churn Rate")
             + _DISCONNECT_MEASURES)


def _to_num(v):
    """Crosstab cell -> float, or None when it isn't a plain number."""
    t = str(v if v is not None else "").replace(",", "").replace("%", "").strip()
    return float(t) if re.fullmatch(r"-?\d+(\.\d+)?", t) else None

# Exact Tableau filter values (runbook Part 10). The Owner & Office member
# renders with an embedded newline in crosstabs; the FILTER value is the
# flat one-line form shown in the quick-filter dropdown.
OWNERS = {
    "carlos": "CARLOS HIDALGO [alphalete specialized marketing inc(tx]",
    "atef": "ATEF CHOUDHURY [domin8 acquisitions, inc.]",
}


def orderlog_url(owner_key: str, today: dt.date) -> str:
    # All-reps download (no owner URL filter — it displays but doesn't filter
    # the rows, and compute.load_orderlog filters by owner name anyway). One
    # download serves both owners.
    start = today - dt.timedelta(days=60)
    return ORDERLOG_URL_TMPL.format(
        start=start.isoformat(), end=today.isoformat())


def download_orderlog(owner_key: str, today: dt.date, out_path: Path,
                      page=None, verbose: bool = True) -> Path:
    return download_crosstab_patchright(
        orderlog_url(owner_key, today), ORDERLOG_SHEET, out_path,
        page=page, verbose=verbose)


def download_churnrates(out_path: Path, page=None, verbose: bool = True) -> Path:
    return download_crosstab_patchright(
        CHURNRATES_URL, CHURNRATES_SHEET, out_path,
        page=page, verbose=verbose)


def _load_grid(path: Path) -> list[list]:
    """Crosstab → grid, back-filling merged/blanked row-header cells. Handles
    the manual .xlsx (merged) AND the automated CSV download (UTF-16 tab), so
    the same parser works for both. Detected by the zip magic bytes."""
    with open(path, "rb") as f:
        magic = f.read(2)
    if magic == b"PK":
        import openpyxl
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path)
        ws = wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        for mr in ws.merged_cells.ranges:
            v = grid[mr.min_row - 1][mr.min_col - 1]
            for rr in range(mr.min_row - 1, min(mr.max_row, len(grid))):
                for cc in range(mr.min_col - 1, mr.max_col):
                    grid[rr][cc] = v
        wb.close()
        return grid
    import csv as _csv
    for enc in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc, newline="") as fh:
                rows = list(_csv.reader(fh, delimiter="\t"))
            if rows and len(rows[0]) > 1:
                # forward-fill blanked left (owner) column so measure sub-rows
                # inherit their owner name
                prev = ""
                for r in rows[1:]:
                    if r and str(r[0]).strip():
                        prev = r[0]
                    elif r:
                        r[0] = prev
                return rows
        except Exception:
            continue
    raise RuntimeError(f"Could not parse CHURNRATES crosstab at {path}")


def parse_churnrates(path: Path, owner_prefix: str) -> dict:
    """Extract one owner's 0-30 Day reconciliation numbers.

    Returns {'base': int|None, 'rate': float|None (fraction), 'raw': {...}}
    with every measure sub-row found, so run.py can log exactly what the
    dashboard said.
    """
    grid = _load_grid(path)
    hdr_row = None
    col_030 = None
    for ri, row in enumerate(grid[:10]):
        for ci, v in enumerate(row):
            if str(v or "").strip() == "0-30 Day":
                hdr_row, col_030 = ri, ci
                break
        if col_030 is not None:
            break
    if col_030 is None:
        raise RuntimeError("CHURNRATES crosstab: no '0-30 Day' column found")

    # PRODUCT BREAKOUT (2026-08-30). CHURNRATES gained a "Product Type (Broken
    # Out)" row dimension, which split every owner's block into one row PER
    # PRODUCT. This loop used to see exactly one 0-30 reading per measure per
    # owner and assign it; with the breakout it saw four and kept the LAST, so
    # `base` became one product bucket instead of the owner's total — CARLOS
    # read 22 against a real 401, and the reconcile gate correctly refused to
    # write all day. (The old comment claimed "first"; the code always assigned,
    # so it was really the last. Identical while there was only ever one.)
    #
    # Summing the REAL product rows is right whether or not each owner also
    # carries its own 'Total' roll-up: the roll-ups are excluded either way, so
    # this can never double-count, and it needs no assumption about a per-owner
    # Total row that the probe could not confirm exists.
    col_prod = None
    for ci, v in enumerate(grid[hdr_row]):
        if str(v or "").strip() == PRODUCT_COL:
            col_prod = ci
            break

    raw: dict = {}          # measure -> last non-empty reading (rates, labels)
    sums: dict = {}         # measure -> summed reading (counts)
    for row in grid[hdr_row + 1:]:
        owner = ""
        measure = ""
        for v in row[:col_030]:
            s = str(v or "").split("\n")[0].strip()
            if s.upper().startswith(owner_prefix.upper()):
                owner = s
            if str(v or "").strip() in _MEASURES:
                measure = str(v).strip()
        # Skip the roll-up rows — adding a total to its own parts doubles it.
        if col_prod is not None and col_prod < len(row):
            if str(row[col_prod] or "").strip().upper() == PRODUCT_TOTAL:
                continue
        # An owner spans multiple color-band blocks (Red/Yellow/Green); each
        # holds only some day-buckets, so the 0-30 value is present in exactly
        # one block and BLANK in the others. Never overwrite a real value with
        # a blank.
        if owner and measure and str(row[col_030] or "").strip():
            raw[measure] = row[col_030]
            if measure in _COUNT_MEASURES:
                n = _to_num(row[col_030])
                if n is not None:
                    sums[measure] = sums.get(measure, 0.0) + n
    if not raw:
        raise RuntimeError(
            f"CHURNRATES crosstab: no rows for owner '{owner_prefix}'")

    def _num(v):
        s = str(v if v is not None else "").replace(",", "").replace("%", "").strip()
        m = re.fullmatch(r"-?\d+(\.\d+)?", s)
        return float(s) if m else None

    # Counts come from the SUM across products; the rate is DERIVED, never
    # summed — adding four percentages together is meaningless, and the
    # dashboard's own per-product rates cannot be averaged into an owner rate
    # without their weights, which is exactly what base and disc are.
    base = sums.get("Activated SPE/SP")
    if base is None:
        base = _num(raw.get("Activated SPE/SP"))
    disc = next((sums[m] for m in _DISCONNECT_MEASURES if m in sums), None)

    rate = None
    if disc is not None and base:
        rate = disc / base
    if rate is None:
        rate = _num(raw.get("Churn Rate"))
        if rate is not None and rate > 1:
            rate = rate / 100.0
    return {"base": int(base) if base is not None else None,
            "rate": rate, "raw": raw}
