"""Vantura churn & activations math — pure logic over an Order Log crosstab.

Implements the runbook exactly (Vantura Master Sales Board, 2026-07-13):
  * Churn counts by POSTED date (last 30 days); Activations by ORDER date.
  * Product mapping: WIRELESS→Wireless, AIR/AWB→Air, NEW INTERNET→Internet;
    every other product type is ignored for churn.
  * Helper block (hidden R:AE) = one row per disconnected account
    (product+customer+posted-date), sorted soonest-to-fall-off first
    within each product, products in Wireless/Air/Internet order.
  * AE = lines still on the 0-30 list AFTER this row's disconnect date
    rolls off (drives the "Churn % after rolloff" formula in U).

Validated against the hand-reconciled 7/13/2026 numbers: Carlos
321/58/67 bases + 8/5/2 disconnects (15/446 = 3.4%), Atef 549/111/1 +
45/7/0 (52/661 = 7.9%) — exact match on both owners.
"""
from __future__ import annotations

import datetime as dt
import re
import warnings
from collections import defaultdict
from pathlib import Path
from typing import Optional

PRODUCT_MAP = {"WIRELESS": "Wireless", "AIR/AWB": "Air", "NEW INTERNET": "Internet"}
PRODUCT_ORDER = ["Wireless", "Air", "Internet"]

# Crosstab column headers we consume (ATTTRACKER-B2B/ORDERLOG, "Order Log" sheet).
COLS = {
    "owner": "Owner & Office",
    "rep": "Rep",
    "customer": "Customer Name",
    "order_date": "sp.Order Date (copy)",
    "spm": "sp.SPM Number",
    "ban": "spe.Account BAN",
    "product": "Product Type (Broken Out)",
    "cru_iru": "CRU/IRU",
    "status": "DTR Status (enriched)",
    "status_date": "DTR Status Date",
    "posted": "spe.dtr Posted Date (copy)",
    "spe": "spe.Name",
    "tn": "spe.TN",
    "tn_type": "spe.TN Type",
    "phone": "spe.Phone",
    "wireless_ip": "Wireless Installment Plan",
    "abp": "Auto Bill Pay",
}


def _parse_date(v) -> Optional[dt.date]:
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return dt.date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    return None


def _fmt_date(d: Optional[dt.date]) -> str:
    return f"{d.month}/{d.day}/{d.year}" if d else ""


# Row-header columns Tableau merges (Excel) / blanks-on-continuation (CSV):
# the account-level fields that span an account's product lines. Forward-fill
# these so every line row carries them; per-line fields stay as-is.
_GROUP_COLS = ("owner", "rep", "customer", "order_date", "spm", "ban", "spe")


def _load_grid(path: Path) -> list[list]:
    """Read a Tableau crosstab export into a cell grid, back-filling the merged
    / blanked row-header cells. Handles BOTH formats: the manual download is a
    real .xlsx (merged cells); the automated crosstab download is UTF-16
    tab-delimited CSV (row-headers blanked on continuation rows). Detected by
    the zip magic bytes, since both may carry an .xlsx name."""
    with open(path, "rb") as f:
        head = f.read(4)
    if head[:2] == b"PK":  # xlsx (zip)
        import openpyxl
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path)  # need merged_cells
        ws = wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        for mr in ws.merged_cells.ranges:
            v = grid[mr.min_row - 1][mr.min_col - 1]
            for rr in range(mr.min_row - 1, min(mr.max_row, len(grid))):
                for cc in range(mr.min_col - 1, mr.max_col):
                    grid[rr][cc] = v
        wb.close()
        return grid
    # CSV (tab-delimited, usually UTF-16)
    import csv as _csv
    rows = None
    for enc in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc, newline="") as fh:
                rows = list(_csv.reader(fh, delimiter="\t"))
            if rows and len(rows[0]) > 1:
                break
        except Exception:
            continue
    if not rows:
        raise RuntimeError(f"Could not parse Order Log crosstab at {path}")
    hdr = [str(h or "").strip() for h in rows[0]]
    gidx = [hdr.index(COLS[k]) for k in _GROUP_COLS if COLS[k] in hdr]
    prev = {}
    for r in rows[1:]:
        for ci in gidx:
            if ci < len(r) and str(r[ci]).strip():
                prev[ci] = r[ci]
            elif ci in prev:
                if ci >= len(r):
                    r.extend([""] * (ci - len(r) + 1))
                r[ci] = prev[ci]
    return rows


def load_orderlog(path: Path, owner_prefix: str) -> list[dict]:
    """Parse the Order Log crosstab (xlsx or CSV) into line-level dicts for
    one owner. Grand Total / blank-owner rows are dropped; the Owner & Office
    cell carries an embedded newline before the office suffix, so match on the
    person-name prefix only."""
    grid = _load_grid(path)
    rows_iter = iter(grid)
    hdr = [str(h or "").strip() for h in next(rows_iter)]
    missing = [c for c in COLS.values() if c not in hdr]
    if missing:
        raise RuntimeError(f"Order Log crosstab missing columns: {missing}")
    ix = {k: hdr.index(c) for k, c in COLS.items()}

    ncol = len(hdr)
    out = []
    for r in rows_iter:
        if len(r) < ncol:  # CSV rows can be ragged
            r = list(r) + [""] * (ncol - len(r))
        owner = str(r[ix["owner"]] or "").split("\n")[0].strip().upper()
        if not owner.startswith(owner_prefix.upper()):
            continue
        rec = {k: ("" if r[i] is None else str(r[i]).strip())
               for k, i in ix.items()}
        rec["order_date"] = _parse_date(r[ix["order_date"]])
        rec["posted"] = _parse_date(r[ix["posted"]])
        out.append(rec)
    if not out:
        raise RuntimeError(f"Order Log has 0 rows for owner '{owner_prefix}' — "
                           "wrong file or the Tableau owner filter didn't apply.")
    return out


def _churn_units(lines: list[dict], today: dt.date) -> list[dict]:
    """The churn-countable units: DISTINCT SPEs (per product) posted in the
    last 30 days. The dashboard's 0-30 base is literally 'Activated SPE/SP'
    — an account can have two crosstab leaf rows for ONE SPE (verified
    7/13/2026: YAMAYNA C, New Internet), and counting raw rows overshoots
    the dashboard by one. Dedupe on (product, spe.Name)."""
    cutoff = today - dt.timedelta(days=30)
    seen, units = set(), []
    for ln in lines:
        p = PRODUCT_MAP.get(ln["product"].upper())
        if not p or not ln["posted"] or ln["posted"] < cutoff:
            continue
        key = (p, ln["spe"])
        if key in seen:
            continue
        seen.add(key)
        units.append(dict(ln, prod=p))
    return units


def churn_summary(lines: list[dict], today: dt.date) -> dict:
    """Per-product activation bases + disconnect counts (0-30 by posted date)."""
    base = {p: 0 for p in PRODUCT_ORDER}
    disc = {p: 0 for p in PRODUCT_ORDER}
    for u in _churn_units(lines, today):
        base[u["prod"]] += 1
        if u["status"] == "Disconnected":
            disc[u["prod"]] += 1
    return {"base": base, "disc": disc,
            "base_total": sum(base.values()), "disc_total": sum(disc.values())}


def _join_uniq(vals, sep=" / "):
    seen, out = set(), []
    for v in vals:
        v = (v or "").strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return sep.join(out)


def helper_block(lines: list[dict], today: dt.date) -> list[list]:
    """The hidden R:AE rows (formulas U/AC left as None — fill.py injects
    them with per-row references). One row per product+customer+posted-date."""
    # One row per product+ACCOUNT (matches the hand-built tab): an account
    # whose lines posted on different days still gets ONE row, keyed to its
    # EARLIEST in-window posted date (verified 7/13: ANTHONY A, lines posted
    # 7/2 + 7/4, shown as one 2-line row rolling off 8/1).
    # BAN is part of the key: two different accounts can share a customer
    # name (verified 7/13: two 'Brenda R' Air accounts on Atef's tab).
    groups: dict = defaultdict(list)
    for u in _churn_units(lines, today):
        if u["status"] != "Disconnected":
            continue
        groups[(u["prod"], u["customer"], u["ban"])].append(u)

    def _device(g):
        # Sheet's "Phone / BYOD" column: BYOD when the line has no
        # installment device, else the device name.
        return "BYOD" if g["wireless_ip"] == "BYOD" else g["phone"]

    rows = []
    for p in PRODUCT_ORDER:
        prod = [(k, v) for k, v in groups.items() if k[0] == p]
        prod.sort(key=lambda kv: (min(u["posted"] for u in kv[1]), kv[0][1]))
        for (_, customer, _ban), grp in prod:
            posted = min(u["posted"] for u in grp)
            disc_date = posted + dt.timedelta(days=30)
            remaining = sum(len(v) for _, v in prod
                            if min(u["posted"] for u in v) > posted)
            rows.append([
                f"{(disc_date - today).days}d",          # R Days Left
                _fmt_date(disc_date),                    # S Disconnect Date
                len(grp),                                # T Lines
                None,                                    # U churn-after formula
                customer,                                # V Customer
                _join_uniq(g["rep"] for g in grp),       # W Sales Rep
                _fmt_date(min((g["order_date"] for g in grp
                               if g["order_date"]), default=None)),  # X Order Date
                _fmt_date(posted),                       # Y Activation Date
                _join_uniq((g["cru_iru"] for g in grp), sep="/"),  # Z CRU/IRU
                _join_uniq(_device(g) for g in grp),     # AA Phone/BYOD
                p,                                       # AB Product Type
                None,                                    # AC notes formula
                p,                                       # AD Product key
                remaining,                               # AE remaining after
            ])
    return rows


# Activations tab: one row per SPM+BAN+status, orders (ORDER date) in last
# 30 days, all products & statuses. 16 columns A..P (P = preserved notes,
# filled by fill.py).
def activations_rows(lines: list[dict], today: dt.date) -> list[list]:
    cutoff = today - dt.timedelta(days=30)
    groups: dict = defaultdict(list)
    for ln in lines:
        if not ln["order_date"] or ln["order_date"] < cutoff:
            continue
        groups[(ln["spm"], ln["ban"], ln["status"])].append(ln)

    rows = []
    for (spm, ban, status), grp in groups.items():
        # Scalar columns mirror the hand-built tab: the FIRST leaf row's
        # value in crosstab order, even when later lines differ (and even
        # when it's blank — ABP/posted included). Only TN and TN Type are
        # joins (unique, ascending).
        g0 = grp[0]
        order = min((g["order_date"] for g in grp if g["order_date"]), default=None)
        tns = sorted({g["tn"] for g in grp if g["tn"]})
        tn_types = sorted({g["tn_type"] for g in grp if g["tn_type"]})
        rows.append([
            g0["rep"],                                   # A Rep
            g0["customer"],                              # B Customer Name
            len(grp),                                    # C Total Apps
            status,                                      # D DTR Status
            _fmt_date(order),                            # E Order Date
            _fmt_date(g0["posted"]),                     # F Posted Date
            g0["status_date"],                           # G DTR Status Date
            ban,                                         # H BAN
            g0["cru_iru"],                               # I CRU/IRU
            g0["product"],                               # J Product Type
            " / ".join(tns),                             # K TN (ascending)
            " / ".join(tn_types),                        # L TN Type (sorted)
            g0["wireless_ip"],                           # M Wireless IP
            g0["abp"],                                   # N ABP
            spm,                                         # O SPM
            "",                                          # P Notes (preserved)
        ])
    rows.sort(key=lambda r: (str(r[0]).upper(), str(r[3]).upper()))
    return rows
