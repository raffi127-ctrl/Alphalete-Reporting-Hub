"""NDS OPT phase for the Alphalete Org 1on1s - Focus Reports sheet.

Downloads the seven NDS Tableau crosstabs, parses them, and writes per-rep
OPT metrics to each NDS tab on the Alphalete Org sheet. Mirrors the shape
of automations/recruiting_report/opt_phase.py (Raf's pipeline) but points
at a different Tableau workbook because the NDS reps aren't in Raf's
'AUTOMATION PULL ICD' view.

Source map (resources/opt-section/alphalete-org-campaign-sources.md):
  - NDSDailyTracker / TT-LineN/P Detail → Active Selling Heads + Scorecard
    Ranking (per-rep). Reads "Rep Count" + "Ranking" columns.
  - NDSDailyTracker / Rep Summary → National AVG for sales (org total
    of New/Port per Rep; SAME value on every NDS tab).
  - SARAPLUSSALESSUMMARY iid=1 / 'Sara Plus Sales Summary' → org totals
    (ATV / Internet / New/Port Lines / Next Up / Premium/Elite / Extra).
    Used for the office-level shared metrics (Next Up %, Extra/Premium %).
  - ProductSalesSummaryRep / ALLPRODUCTS-EXPANDEDREPS iid=1 → per-rep
    breakdown across all sale types (WIRELESS / NEW INTERNET / VIDEO / AIR).
    Feeds Personal Production. Megan's custom view sets the Product Type
    filter to (All) so all four types appear in the crosstab.
  - NDSWeeklyMetricsRep / THISWEEK → 0-30 Day Cancel Rate 4wk avg
    ('Cancel Fraud Review %', office total).
  - CHURNRATES / THISWEEK → 0-30 / 60 / 90 Day Churn (per-rep, look for
    rows where col-2 = 'Green' for the headline rate).
  - ACTIVATIONRATES → Activation % by Week (per-rep, '60+ Days' column).
  - LeadPenetrationOverview / THISWEEK → Total Leads (sum of 'Lead Count'
    rows assigned to each ICD).
  - DirectDeposit / PROGRAMSUMMARY / DOWNLINEVIEW → Direct Deposit
    ('Grand Total to ICD', per-ICD; shared with every other Alphalete
    Org program).

Computed metrics (no separate download):
  - AVG Apps Per Active Headcount = New Lines / Active Selling Heads
  - Next Up % = (office Next Up count) / (office New/Port Lines count)
  - Extra/Premium % = (office Premium/Elite + Extra) / (office New/Port Lines)

Rep Breakdown chart at the bottom of each NDS tab — separate module
(production_breakdown style), wireless-only, comes from
ProductSalesSummaryRep / REPEXPANDED / 'Sales By ICD (Weekly View)'.

Week scoping: Megan confirmed (2026-05-20) that Tableau's THISWEEK custom
view on Monday morning resolves to the just-completed week's totals.
Today's data (mid-week) reflects 5/18-5/24 partial; Monday-morning runs
will reflect the full 5/18-5/24 week.
"""
from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import gspread

from automations.shared import sheet_flags as _sheet_flags

from automations.recruiting_report import fill as rfill
from automations.recruiting_report.opt_phase import download_crosstab as _download_crosstab_inline
from automations.alphalete_org_report import tableau_http
from automations.shared.tableau_patchright import (
    tableau_session,
    download_crosstab_patchright,
    requests_session_from_page,
)


def _download_crosstab_subprocess(url: str, sheet: str, out_path: Path,
                                  verbose: bool = False) -> Path:
    """Download via a fresh Python subprocess. Workaround for the
    Python 3.9 + Playwright sync API asyncio race that surfaces when
    multiple sync_playwright() contexts run in the same process.

    Each subprocess does ONE download then exits, so each playwright
    runtime gets a clean lifecycle. Slower than inline but reliable."""
    import subprocess
    import sys
    # Cross-platform: use the interpreter that's already running this report
    # (the Hub launches reports with the venv's python on macOS AND Windows)
    # and the repo root derived from this file's location — never a hardcoded
    # absolute path. The previous '/Users/megan/...' literal worked only on
    # Megan's Mac and would crash on Eve's Windows box.
    workspace = Path(__file__).resolve().parent.parent.parent
    script = (
        "from automations.recruiting_report.opt_phase import download_crosstab; "
        "from pathlib import Path; "
        f"download_crosstab({url!r}, {sheet!r}, Path({str(out_path)!r}), verbose={verbose!r})"
    )
    result = subprocess.run(
        [sys.executable, "-c", script],
        cwd=str(workspace),
        capture_output=True, text=True, timeout=300,
    )
    if result.returncode != 0:
        # Surface the inner traceback's last few lines for diagnostics
        err = (result.stderr or result.stdout or "").strip().splitlines()
        tail = " | ".join(err[-3:]) if err else "(no stderr)"
        raise RuntimeError(f"subprocess download failed: {tail}")
    return out_path


# Default to the subprocess wrapper — it's the reliable path. Set
# OPT_NDS_INLINE_DOWNLOAD=1 to use the in-process version for debugging.
import os as _os
if _os.environ.get("OPT_NDS_INLINE_DOWNLOAD") == "1":
    download_crosstab = _download_crosstab_inline
else:
    download_crosstab = _download_crosstab_subprocess

WORKSPACE = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = WORKSPACE / "output"

ALPHALETE_ORG_SHEET_ID = "1C6BLttOSZhs_dREySac19XkxnMl-Ab_sYacNSl2l6AQ"

# Tableau Crosstab views — all pulled via patchright in one session.
# (url, crosstab_worksheet_name, output_filename)
# Org-wide Direct Deposit — DD BY OWNER (ORG) / DOWNLINEVIEW. THE canonical DD
# source for EVERY campaign: the grid holds all ICD owners across every campaign
# (RES-ATT, NDS Wireless, B2B, BOX, Retail, …), so each campaign fills its reps'
# Direct Deposit from this one view via parse_direct_deposit(). Standardized
# 2026-05-25 (Megan) so BOX/Retail/JE/B2B all pull DD from Tableau, not per-
# campaign DD views (which missed BOX + Retail reps entirely).
ORG_DD_URL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "DirectDepositICDVIEWVersion2_0/DDBYOWNERORG/"
    "796feca0-272f-459f-a665-63ac9aec3af8/DOWNLINEVIEW?:iid=1"
)
ORG_DD_SHEET = "Sheet 7 (5)"

NDS_VIEWS: List[Tuple[str, str, str]] = [
    (
        "https://us-east-1.online.tableau.com/#/site/sci/views/NDS-SNRES-ATT-OOFWorkbook/NDSDailyTracker?:iid=1",
        "TT-LineN/P Detail",
        "opt_nds_tt_detail.csv",
    ),
    (
        "https://us-east-1.online.tableau.com/#/site/sci/views/NDS-SNRES-ATT-OOFWorkbook/NDSDailyTracker?:iid=1",
        "Rep Summary",
        "opt_nds_rep_summary.csv",
    ),
    (
        "https://us-east-1.online.tableau.com/#/site/sci/views/NDS-SNRES-ATT-OOFWorkbook/CHURNRATES/c289786d-e0d4-4de7-825a-264c21e133c1/THISWEEK?:iid=1",
        "Churn Rates (ICD)",
        "opt_nds_churn.csv",
    ),
    (
        "https://us-east-1.online.tableau.com/#/site/sci/views/DropshipV_2/SARAPLUSSALESSUMMARY?:iid=1",
        "Sara Plus Sales Summary",
        "opt_nds_sara_plus.csv",
    ),
    (
        # Personal Production — per-rep breakdown across all sale types.
        # Custom view 'ALLPRODUCTS-EXPANDEDREPS' filters Product Type to
        # (All) and expands the Rep dimension.
        "https://us-east-1.online.tableau.com/#/site/sci/views/NDS-SNRES-ATT-OOFWorkbook/ProductSalesSummaryRep/c6d0a461-f8ac-49ed-bb38-27a807328a70/ALLPRODUCTS-EXPANDEDREPS?:iid=1",
        "Sales By ICD (Weekly View)",
        "opt_nds_personal_production.csv",
    ),
    (
        # DD BY OWNER (ORG) / DOWNLINEVIEW — per-ICD direct deposit grid.
        "https://us-east-1.online.tableau.com/#/site/sci/views/DirectDepositICDVIEWVersion2_0/DDBYOWNERORG/796feca0-272f-459f-a665-63ac9aec3af8/DOWNLINEVIEW?:iid=1",
        "Sheet 7 (5)",
        "opt_nds_direct_deposit.csv",
    ),
    (
        # ABP % — same source Retail uses (DropshipV_2 / ABPCONVERSIONS,
        # "ABP National Average (2)"). Honors Min/Max Date URL params, so the
        # download loop date-pins it to the target week like SARA. Parsed by
        # opt_retail.parse_abp_conversions (Megan 2026-05-25).
        "https://us-east-1.online.tableau.com/#/site/sci/views/DropshipV_2/ABPCONVERSIONS?:iid=1",
        "ABP National Average (2)",
        "opt_nds_abp.csv",
    ),
]
NDS_ABP_FILENAME = "opt_nds_abp.csv"

# The Rep Breakdown chart at the bottom of each tab — its OWN download
# (separate from the metric crosstabs above).
NDS_REP_BREAKDOWN_URL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "NDS-SNRES-ATT-OOFWorkbook/ProductSalesSummaryRep/"
    "b86d7862-bfc7-4966-a0a4-7803432a6444/REPEXPANDED?:iid=2"
)
NDS_REP_BREAKDOWN_SHEET = "Sales By ICD (Weekly View)"
NDS_REP_BREAKDOWN_PATH = OUTPUT_DIR / "opt_nds_product_sales_rep.csv"


# HTTP-direct sources — these use Tableau's /views/.../.csv endpoint
# (tableau_http.download_view_csv). Much faster than the UI Crosstab flow
# (~1s per view vs 60-90s) and not subject to the UI's intermittent
# Download-button-stays-disabled bug. We use HTTP for everything that
# DOESN'T require sub-worksheet selection from a multi-worksheet
# dashboard. Direct Deposit + Rep Breakdown still need the UI path
# because their .csv URL only returns a header-label stub
# ([[reference-tableau-worksheet-names]]).
#
# Tuples: (workbook_slug, view_slug, output_filename)
NDS_HTTP_VIEWS: List[Tuple[str, str, str]] = [
    ("DropshipV_2",                  "ACTIVATIONRATES",
     "opt_nds_activation_http.csv"),
    ("NDS-SNRES-ATT-OOFWorkbook",    "NDSWeeklyMetricsRep",
     "opt_nds_weekly_metrics_http.csv"),
    ("NDS-SNRES-ATT-OOFWorkbook",    "LeadPenetrationOverview",
     "opt_nds_lead_penetration_http.csv"),
    ("DropshipV_2",                  "SARAPLUSSALESSUMMARYBYDAY",
     "opt_nds_sara_plus_byday.csv"),
]


# Map our sheet row labels (col B) → metric keys used in fill_nds_tab.
# Format follows Raf's opt_phase convention: lookup by label string, never
# by hardcoded row index ([[feedback_no_hardcoded_columns]]).
NDS_METRIC_LABELS = {
    "Active Selling Heads":              "active_selling_heads",
    "New Lines":                          "new_lines",
    "AVG Apps Per Active Headcount":     "avg_apps_per_hc",
    "National AVG for sales":            "national_avg_sales",
    "Scorecard Ranking":                 "scorecard_ranking",
    "Personal Production":               "personal_production",
    "0-30 Day Cancel Rate  4wk avg":     "cancel_rate_30",
    "Activation % by Week":              "activation_pct",
    "0-30 Day Churn":                    "churn_30",
    "60 Day Churn":                      "churn_60",
    "90 Day Churn":                      "churn_90",
    "Total Leads":                       "total_leads",
    "Next Up %":                          "next_up_pct",
    "Extra/Premium %":                   "extra_premium_pct",
    "Direct Deposit":                    "direct_deposit",
}


# Some NDS tabs label a metric's row differently than its `values` key. Most
# notably the headcount row is "Active Headcount on Tableau" on tabs that
# follow the B2B/Retail/Frontier layout, not "Active Selling Heads" (verified
# 2026-06-01: Colten Wright / Drew Tepper - NDS both carry the former). The
# writer tries each spelling in order (mirrors opt_box._row_any). Add an entry
# here only when a tab's row LABEL differs — NOT when the row is simply absent
# (e.g. ABP %, which most NDS tabs have no row for at all).
NDS_ROW_LABEL_ALTERNATES = {
    "Active Selling Heads": ["Active Selling Heads", "Active Headcount on Tableau"],
}


# ------------------------------------------------------------ parsers


def _read_tab_csv(path: Path) -> List[List[str]]:
    """Read a Tableau crosstab into rows. Handles UTF-16 tab-delimited
    CSV (Tableau's CSV format) AND XLSX (returned when the Crosstab
    dialog's 'default' format branch fires — happens when the CSV radio
    click silently no-ops). Detection is by magic bytes, not extension,
    since patchright's 'default' path saves xlsx content under a .csv
    name. Returns [] if the file is missing or empty."""
    if not path.exists() or path.stat().st_size == 0:
        return []
    with open(path, "rb") as f:
        magic = f.read(4)
    if magic.startswith(b"PK"):
        # Wrap in BytesIO so openpyxl skips its filename-extension check
        # (it refuses .csv even when the bytes are valid XLSX).
        import io as _io
        from openpyxl import load_workbook
        with open(path, "rb") as f:
            buf = _io.BytesIO(f.read())
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        out: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            out.append(["" if v is None else str(v) for v in row])
        return out
    for enc in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc) as f:
                rows = list(csv.reader(f, delimiter="\t"))
            if rows and any(len(r) > 1 for r in rows):
                return rows
        except UnicodeDecodeError:
            continue
    return []


def _norm_owner(s: str) -> str:
    """Normalize a Tableau owner name to its canonical form for matching.
    Tableau formats names as 'ALEX GUZMAN BENITEZ\n[Tampa, FL]' — we want
    just 'alex guzman benitez'."""
    s = (s or "").strip()
    # Strip trailing "[city, state]" if present (may be after a \n)
    s = re.split(r"[\[\n]", s, maxsplit=1)[0].strip()
    return " ".join(s.lower().split())


def parse_tt_detail(path: Path) -> Dict[str, Dict[str, str]]:
    """Parse TT-LineN/P Detail: {normalized owner: {Rep Count: ..., Ranking: ...}}."""
    rows = _read_tab_csv(path)
    if not rows:
        return {}
    header = rows[0]
    # Find column indices by header label (no hardcoded positions).
    def col(label: str) -> Optional[int]:
        for i, h in enumerate(header):
            if (h or "").strip().lower() == label.lower():
                return i
        return None
    rank_i = col("Ranking")
    rc_i = col("Rep Count")
    out: Dict[str, Dict[str, str]] = {}
    for r in rows[2:]:  # skip header + Total row
        owner = _norm_owner(r[0] if r else "")
        if not owner:
            continue
        rec = {}
        if rank_i is not None and rank_i < len(r):
            rec["ranking"] = r[rank_i]
        if rc_i is not None and rc_i < len(r):
            rec["rep_count"] = r[rc_i]
        if rec:
            out[owner] = rec
    return out


def parse_rep_summary_total(path: Path) -> Dict[str, str]:
    """Parse the Total row from Rep Summary — gives the org-wide averages
    that go in 'National AVG for sales'. Returns {column: value}."""
    rows = _read_tab_csv(path)
    if not rows:
        return {}
    header = rows[0]
    # Total row is the LAST row with first col == 'Total'
    total_row = next((r for r in reversed(rows)
                      if r and (r[0] or "").strip().lower() == "total"), None)
    if not total_row:
        return {}
    out: Dict[str, str] = {}
    for i, h in enumerate(header):
        if i < len(total_row) and h:
            out[h.strip()] = total_row[i]
    return out


def parse_sara_plus_total(path: Path) -> Dict[str, str]:
    """Parse the (single-row) totals from Sara Plus Sales Summary iid=1.
    Returns {column: value} — includes Next Up, New/Port Lines, etc."""
    rows = _read_tab_csv(path)
    if not rows or len(rows) < 2:
        return {}
    header = rows[0]
    data = rows[1]
    return {(h or "").strip(): (data[i] if i < len(data) else "")
            for i, h in enumerate(header)}


# Sheet display abbreviations for Personal Production text. Order = output
# order, so "NI" comes before "NL" before "DTV" before "AIR" in the sheet
# cell — mirrors Megan's manual entries like "1 NI, 15 NL".
PP_PRODUCT_ABBREV = [
    ("NEW INTERNET", "NI"),
    ("WIRELESS",     "NL"),
    ("VIDEO",        "DTV"),
    ("AIR",          "AIR"),
]


def parse_personal_production(path: Path) -> Dict[str, str]:
    """Parse the ALLPRODUCTS-EXPANDEDREPS crosstab and return
    {normalized ICD owner: 'X NI, Y NL, Z DTV, W AIR'} — but ONLY counting
    the ICD's *own* sales (rows where Rep Name == Owner name), not the
    sum of their downline reps. Per Megan 2026-05-22: Personal Production
    is the ICD's personal selling activity, not their team total.

    Row layout (confirmed 2026-05-22):
      col 0 = 'Owner & Office'  ('ISAIAH REVELLE\\n[legacy specialized...]')
      col 1 = 'Rep Name'        ('Brandon Childs', 'Isaiah Revelle', or 'Total')
      col 2 = 'Product Type (Broken Out)'  (WIRELESS / NEW INTERNET / VIDEO / AIR)
      cols 3-7 = per-weekday counts (Mon-Fri)
      col 8 = 'Total' (week sum for this rep/type)

    Owners with no matching self-row (no personal sales this week) come
    back as empty strings."""
    rows = _read_tab_csv(path)
    if not rows or len(rows) < 3:
        return {}
    # Shared ICD alias list so rep-name spelling variants resolve to the ICD
    # (Megan 2026-05-24: 'Maxamad-Amin Aden' IS Max). Identity fallback if the
    # alias Sheet is unreachable.
    try:
        from automations.focus_office_att.aliases import (
            load_aliases as _load_aliases, alias_to_canonical)
        _aliases = _load_aliases()
    except Exception:
        _aliases = {}
        def alias_to_canonical(name, raw):
            return name
    OWNER_I, REP_I, TYPE_I = 0, 1, 2
    # The view's day-column count varies with the week's progress (only Mon-Wed
    # early in the week, full Mon-Sun later), so the Total is NOT at a fixed
    # index. Find it from the header — the rightmost 'Total'/'Grand Total'
    # column. (Hardcoding 8 silently skipped EVERY row when the current week had
    # < 5 days → 0 PP for all NDS ICDs. Found 2026-06-03.)
    _hdr = rows[1] if len(rows) > 1 else []
    _totals = [i for i, c in enumerate(_hdr) if "total" in (c or "").lower()]
    TOTAL_I = _totals[-1] if _totals else (len(_hdr) - 1 if _hdr else 8)
    bucket: Dict[str, Dict[str, int]] = {}
    seen_owners: set = set()
    for r in rows[2:]:
        if len(r) <= TOTAL_I:
            continue
        owner_raw = (r[OWNER_I] or "").strip()
        rep_raw = (r[REP_I] or "").strip()
        ptype = (r[TYPE_I] or "").strip().upper()
        if not owner_raw or owner_raw.lower().startswith("grand total"):
            continue
        # Skip per-ICD subtotal rows
        if rep_raw.lower() == "total" or ptype == "TOTAL":
            continue
        # Bucket key = the CANONICAL ICD name (alias-resolved, office-suffix
        # stripped) so the caller's lookup by sheet-tab name matches even when
        # the view spells it differently ('MAXAMAD ADEN' vs tab 'Maxamed Aden').
        owner = _norm_owner(
            alias_to_canonical(owner_raw.split("\n")[0].strip(), _aliases))
        seen_owners.add(owner)
        # Count only the ICD's OWN rows (personal production). The owner cell
        # carries an office suffix ('MAXAMAD ADEN\n[maximal..]') the rep cell
        # never has, so compare just the NAME — and resolve spelling variants
        # through the shared ICD alias list (Megan 5/24: 'Maxamad-Amin Aden'
        # IS Max). The old compare matched owner+suffix vs the bare rep name,
        # so NO self-row ever matched and every ICD's PP came back 0.
        owner_name = alias_to_canonical(
            owner_raw.split("\n")[0].strip(), _aliases).strip().lower()
        rep_name = alias_to_canonical(rep_raw, _aliases).strip().lower()
        if rep_name != owner_name:
            continue
        try:
            val = int(float((r[TOTAL_I] or "0").replace(",", "")))
        except (ValueError, AttributeError):
            continue
        if val == 0:
            continue
        rec = bucket.setdefault(owner, {})
        rec[ptype] = rec.get(ptype, 0) + val

    out: Dict[str, str] = {}
    # Include all owners we saw in the CSV (even those with no personal
    # sales) so the caller can distinguish "no PP data" from "0 sales".
    # When the owner has zero personal sales of every type, write "0"
    # rather than blank so the column shows a real value for the week.
    for owner in seen_owners:
        totals = bucket.get(owner, {})
        parts = []
        for tableau_type, abbrev in PP_PRODUCT_ABBREV:
            n = totals.get(tableau_type, 0)
            if n > 0:
                parts.append(f"{n} {abbrev}")
        out[owner] = ", ".join(parts) if parts else "0"
    return out


DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday",
             "Friday", "Saturday", "Sunday"]


def parse_rep_breakdown_per_owner(path: Path,
                                  ptype_filter: str = "WIRELESS"
                                  ) -> Dict[str, List[Dict]]:
    """Parse the same ALLPRODUCTS-EXPANDEDREPS crosstab as
    parse_personal_production, but return per-rep daily breakdowns
    filtered to a single product type (wireless by default — the format
    Megan's Rep Breakdown chart expects).

    Returns {normalized owner: [{rep, days{day_name: int}, total}, ...]}
    where each owner's reps are sorted by total DESC (then rep name asc).

    Skips zero-total reps so the chart only lists reps who actually sold."""
    rows = _read_tab_csv(path)
    if not rows or len(rows) < 3:
        return {}
    header = rows[1]
    OWNER_I, REP_I, TYPE_I = 0, 1, 2
    # Map day name → column index
    day_cols: Dict[str, int] = {}
    for j, h in enumerate(header):
        h_clean = (h or "").strip()
        if h_clean in DAY_ORDER:
            day_cols[h_clean] = j
    # First "Total" column AFTER the day columns is the per-rep weekly total
    max_day_col = max(day_cols.values()) if day_cols else TYPE_I
    total_i: Optional[int] = None
    for j in range(max_day_col + 1, len(header)):
        if (header[j] or "").strip().lower() == "total":
            total_i = j
            break
    if total_i is None:
        return {}

    target = ptype_filter.upper()
    bucket: Dict[str, List[Dict]] = {}
    for r in rows[2:]:
        if len(r) <= total_i:
            continue
        owner_raw = (r[OWNER_I] or "").strip()
        rep_raw = (r[REP_I] or "").strip()
        ptype = (r[TYPE_I] or "").strip().upper()
        if not owner_raw or owner_raw.lower().startswith("grand total"):
            continue
        if rep_raw.lower() == "total" or ptype == "TOTAL":
            continue
        if ptype != target:
            continue
        try:
            total = int(float((r[total_i] or "0").replace(",", "")))
        except (ValueError, AttributeError):
            continue
        if total == 0:
            continue
        days: Dict[str, int] = {}
        for day, j in day_cols.items():
            if j < len(r):
                try:
                    n = int(float((r[j] or "0").replace(",", "")))
                    if n != 0:
                        days[day] = n
                except (ValueError, AttributeError):
                    pass
        owner = _norm_owner(owner_raw)
        bucket.setdefault(owner, []).append({
            "rep": rep_raw,
            "days": days,
            "total": total,
        })

    # Sort each owner's reps by total desc, then name asc
    for owner in bucket:
        bucket[owner].sort(key=lambda x: (-x["total"], x["rep"]))
    return bucket


def parse_direct_deposit(path: Path) -> Dict[str, float]:
    """Parse DD BY OWNER (ORG) Sheet 7 (5): {normalized ICD owner:
    sum of Total $ to ICD across all per-line rows for that owner}.

    The crosstab has a PIVOTED layout that the column headers misrepresent:
      - col[1] = 'cl.ICD Owner Name'
      - col[2] = 'Sales Rep'
      - col[4] = 'Commission Description' ('Total' on subtotal rows)
      - col[6] = measure-name pivot column (no header). Values are
        'Distinct count of cl.ID' or 'Total $ to ICD'.
      - col[8] = the dollar/count VALUE (the header reads 'RES-ATT' but
        Tableau actually puts the rolled-up Grand Total here)

    Each (account, ICD, rep, commission) generates TWO rows: one for
    count, one for dollars. We filter to dollar rows (col[6]) and skip
    'Total' subtotals (col[4]) to avoid double-counting, then sum col[8]
    per ICD owner.

    Confirmed against 8039-row sample 2026-05-21. Maxamed Aden has 0
    rows — either he doesn't appear as an ICD owner in DD or his name
    differs ([[feedback-alias-list]] candidate)."""
    rows = _read_tab_csv(path)
    if not rows or len(rows) < 2:
        return {}
    # Fixed column positions per the observed pivot layout.
    OWNER_I, REP_I, DESC_I, MEASURE_I = 1, 2, 4, 6
    # Value cells live in cols [7]-[20], each col representing a campaign
    # (Grand Total, RES-ATT, NDS Wireless, B2B-ATT-SBS, ...). EACH ROW
    # puts its dollar value in EXACTLY ONE of these columns — the
    # campaign that commission belongs to. Cody Cannon's ICD downline
    # gets dollars in col[8] (RES-ATT); the NDS reps (Isaiah, Colten,
    # etc.) get dollars in col[9] (NDS Wireless). Summing across all
    # value cols handles both cases. Confirmed 2026-05-21.
    VALUE_COLS = range(7, 21)
    DOLLAR_MEASURE = "total $ to icd"
    out: Dict[str, float] = {}
    for r in rows[1:]:
        if not r or len(r) <= MEASURE_I:
            continue
        if (r[MEASURE_I] or "").strip().lower() != DOLLAR_MEASURE:
            continue
        # Skip per-rep subtotal rows (Commission Description == 'Total')
        # to avoid summing per-line + subtotal together
        if DESC_I < len(r) and (r[DESC_I] or "").strip().lower() == "total":
            continue
        owner = _norm_owner(r[OWNER_I])
        if not owner:
            continue
        row_total = 0.0
        for ci in VALUE_COLS:
            if ci >= len(r):
                break
            raw = (r[ci] or "").strip().lstrip("$").replace(",", "")
            if not raw:
                continue
            try:
                row_total += float(raw)
            except ValueError:
                continue
        if row_total:
            out[owner] = out.get(owner, 0.0) + row_total
    return out


def match_dd_owner(dd: Dict[str, float], name: str,
                   aliases_map: Optional[Dict[str, List[str]]] = None
                   ) -> Optional[float]:
    """Look up an owner's Direct Deposit dollar in the org-wide DD grid.
    Match order:
      1. exact _norm_owner
      2. ICD alias group (when `aliases_map` given) — the canonical fix for
         name spelling/nickname variants (tab 'Joe Delgado' -> grid 'joseph
         delgado'; tab 'Selena' -> 'selena powers'). Checked BEFORE the last-
         name fallback so it stays precise even when siblings share a surname
         (Joe vs George Delgado — match the alias, never the other brother).
      3. first+last name (handles middle names: 'Roshan Amin Ahmad' -> 'roshan ahmad')
      4. a UNIQUE last-name match (skipped when the surname isn't unique).
    Returns None if no confident match (caller leaves DD untouched)."""
    key = _norm_owner(name)
    if key in dd:
        return dd[key]
    if aliases_map:
        for canon, al in aliases_map.items():
            group = [canon] + list(al)
            if key in {_norm_owner(g) for g in group}:
                for g in group:
                    v = dd.get(_norm_owner(g))
                    if v is not None:
                        return v
    parts = key.split()
    if len(parts) >= 2:
        fl = f"{parts[0]} {parts[-1]}"
        if fl in dd:
            return dd[fl]
        last = parts[-1]
        cands = [k for k in dd if k.split() and k.split()[-1] == last]
        if len(cands) == 1:
            return dd[cands[0]]
    return None


def parse_churn_icd(path: Path) -> Dict[str, Dict[str, str]]:
    """Parse Churn Rates (ICD): {normalized owner: {churn_30/60/90: pct}}.
    Headline rates come from the 'Green' status rows (per-rep, NOT
    Office/Organization Average). Each rep has FOUR rows of metric type —
    we want the 'Churn Rate' one only."""
    rows = _read_tab_csv(path)
    if not rows:
        return {}
    # Header: cols 4-7 are 0-30 / 30 / 60 / 90 Day Churn
    header = rows[0]
    c_30 = c_60 = c_90 = None
    for i, h in enumerate(header):
        h_clean = (h or "").strip().lower()
        if h_clean == "0-30 day churn":
            c_30 = i
        elif h_clean == "60 day churn":
            c_60 = i
        elif h_clean == "90 day churn":
            c_90 = i
    out: Dict[str, Dict[str, str]] = {}
    for r in rows[1:]:
        if len(r) < 3:
            continue
        owner = _norm_owner(r[0])
        status = (r[1] or "").strip()
        metric = (r[2] or "").strip().lower()
        if not owner or "office/organization" in owner:
            continue
        if metric != "churn rate":
            continue
        # Take Green status row (the headline rate); fall back to first
        # row seen for this owner if no Green.
        rec = out.setdefault(owner, {})
        is_green = status.lower() == "green"
        for key, col in (("churn_30", c_30), ("churn_60", c_60), ("churn_90", c_90)):
            if col is None or col >= len(r):
                continue
            val = (r[col] or "").strip()
            if not val:
                continue
            if key not in rec or is_green:
                rec[key] = val
    return out


# ------------------------------------------------------------ fill helpers


def _find_row_by_label(grid: List[List[str]], label: str,
                       label_col: int = 1) -> Optional[int]:
    """Find a row whose col-B (default) matches `label` (case-insensitive,
    whitespace-normalized). Returns 0-indexed row or None.

    Matching is intentionally loose — exact, OR all words from the label
    appear in the cell (in any order). Handles legacy tab variants like
    'Churn 0 -30 Day' vs 'Churn 0-30 Day' vs '0-30 Day Churn' without
    needing a per-tab patch."""
    target_norm = " ".join(label.lower().split())
    target_words = set(label.lower().replace("-", " ").split())
    target_compact = label.lower().replace(" ", "").replace("-", "")

    # Pass 1: exact match wins across the whole tab. Otherwise a later
    # canonical row can be shadowed by an earlier fuzzy-equivalent typo
    # (e.g. Colten's 'Churn 0 -30 Day' on row 41 shadowing the real
    # '0-30 Day Churn' on row 42).
    for ri, r in enumerate(grid):
        cell = r[label_col] if len(r) > label_col else ""
        if " ".join((cell or "").lower().split()) == target_norm:
            return ri

    # Pass 2: fuzzy fallbacks. Compact-string equality wins over word-set
    # match so trivial punctuation/spacing differences ('Extra / Premium %'
    # vs 'Extra/Premium %') beat looser matches.
    for ri, r in enumerate(grid):
        cell = r[label_col] if len(r) > label_col else ""
        cell_norm = " ".join((cell or "").lower().split())
        cell_compact = cell_norm.replace(" ", "").replace("-", "")
        if target_compact and target_compact == cell_compact:
            return ri

    for ri, r in enumerate(grid):
        cell = r[label_col] if len(r) > label_col else ""
        cell_norm = " ".join((cell or "").lower().split())
        cell_words = set(cell_norm.replace("-", " ").split())
        if not cell_words:
            continue
        # All target words present in cell (handles word-order swap + cell
        # having extra words — e.g. 'Churn 0-30 Day' vs '0-30 Day Churn').
        if target_words and target_words.issubset(cell_words):
            return ri
        # Cell words are a subset of target (cell has fewer words —
        # e.g. '0-30 Day Cancel Rate' vs '0-30 Day Cancel Rate 4wk avg').
        # Require ≥3 cell words AND ≥half the target so single-word cells
        # like 'Rate' don't match every percent metric.
        if (cell_words.issubset(target_words)
                and len(cell_words) >= 3
                and len(cell_words) >= len(target_words) // 2):
            return ri
    return None


def _find_week_col(grid: List[List[str]], week_label: str) -> Optional[int]:
    """Find the column whose header (row 0) equals `week_label` (e.g.
    '5/24/26'). Returns 0-indexed col or None."""
    if not grid:
        return None
    for ci, h in enumerate(grid[0]):
        if (h or "").strip() == week_label:
            return ci
    return None


def _current_target_week_end(today: Optional[dt.date] = None) -> dt.date:
    """Sheet WE-Sunday of the week we're filling = the most recent Sunday
    ON OR BEFORE today. This is the SAME column the recruiting funnel writes
    (dashboard._last_completed_we_sunday uses the identical formula), so OPT
    and the funnel always agree.

    The week flips to the new column the day it ends (Sunday), then stays
    put through the following Saturday — so a Sunday-evening or Monday-
    morning run both fill the just-ended week, and we never overwrite the
    prior week's finalized column.

    Examples:
      - Sun 5/24 → 5/24 (the week ending today — fill it now, don't touch 5/17)
      - Mon 5/25 → 5/24 (last Sunday)
      - Tue 5/26 → 5/24
      - Sat 5/23 → 5/17 (week ending 5/24 not reached yet → still 5/17)
      - Fri 5/22 → 5/17
    """
    today = today or dt.date.today()
    # Days back to the most recent Sunday, counting today if it IS Sunday.
    days_back = (today.weekday() + 1) % 7
    return today - dt.timedelta(days=days_back)


def _current_target_week_col_label(today: Optional[dt.date] = None) -> str:
    """Sheet column label for the most-recently-completed WE Sunday —
    formatted to match the header text on the Alphalete Org sheet
    (e.g. '5/17/26'). See _current_target_week_end for the date math."""
    target = _current_target_week_end(today)
    return f"{target.month}/{target.day}/{target.year % 100}"


def _target_week_date_range(today: Optional[dt.date] = None) -> Dict[str, str]:
    """Min/Max Date URL params (ISO) for the target week — Mon..Sun ending
    on the target WE Sunday. Used to pin the SARA views to the right week
    instead of relying on the dashboard's default 'current week' (which is
    only correct on a Monday run). Unlike NDSDailyTracker, the SARA
    dashboard honors these params. (2026-05-24.)"""
    end = _current_target_week_end(today)        # Sunday
    start = end - dt.timedelta(days=6)           # Monday
    return {"Min Date": start.isoformat(), "Max Date": end.isoformat()}




def _find_rep_breakdown_anchor(grid: List[List[str]]) -> Optional[Tuple[int, int]]:
    """Locate the Rep Breakdown chart header on a worksheet grid. Returns
    (row, col) of the 'Rep' header cell (0-indexed), or None if absent.

    The anchor is the 'Rep' cell with 'Product Type (Broken Out)' to its
    immediate right — matches Isaiah's tab layout (DZ57 on his sheet).
    Tabs without the chart skeleton return None and the filler skips them."""
    for ri, row in enumerate(grid):
        for ci, cell in enumerate(row):
            if (cell or "").strip() != "Rep":
                continue
            next_cell = row[ci + 1] if ci + 1 < len(row) else ""
            if (next_cell or "").strip() == "Product Type (Broken Out)":
                return (ri, ci)
    return None


REP_CHART_LIGHT_GRAY = {"red": 243/255, "green": 243/255, "blue": 243/255}
REP_CHART_WHITE = {"red": 1.0, "green": 1.0, "blue": 1.0}


def fill_rep_breakdown_chart(ws: gspread.Worksheet, owner_norm: str,
                              breakdown: Dict[str, List[Dict]],
                              week_end: dt.date,
                              dry_run: bool = False,
                              logfn=print,
                              max_rep_rows: int = 30,
                              ) -> List[str]:
    """Write the per-rep wireless daily breakdown into the chart at the
    bottom of an NDS tab. Layout (anchored on 'Rep' header):

      anchor-1 row, anchor col:        WE M.D week label
      anchor row:    Rep | Product Type | Mon | Tue | ... | Sun | Product Total
      anchor+1 row:  Total | Total | sum_mon | ... | sum_sun | grand_total
      anchor+2 row+: per-rep wireless rows sorted by total desc

    Tabs without the chart skeleton are skipped silently — Megan adds the
    skeleton to each NDS tab over time."""
    log: List[str] = []
    grid = rfill._retry(ws.get_all_values)
    if not grid:
        return []
    anchor = _find_rep_breakdown_anchor(grid)
    if anchor is None:
        return []   # no chart on this tab yet

    anchor_row, anchor_col = anchor    # 0-indexed
    header_row = grid[anchor_row]
    # Map day name → 0-indexed column position
    day_cols: Dict[str, int] = {}
    for j, cell in enumerate(header_row):
        if j <= anchor_col:
            continue
        if (cell or "").strip() in DAY_ORDER:
            day_cols[(cell or "").strip()] = j
    # Product Total column
    total_col: Optional[int] = None
    for j in range(anchor_col + 2, len(header_row)):
        if (header_row[j] or "").strip() == "Product Total":
            total_col = j
            break
    if not day_cols or total_col is None:
        return [f"[skip-chart] {ws.title}: chart header malformed"]

    reps = breakdown.get(owner_norm, [])

    # Build aggregate Total row
    day_totals: Dict[str, int] = {}
    grand_total = 0
    for r in reps:
        for day, n in r["days"].items():
            day_totals[day] = day_totals.get(day, 0) + n
        grand_total += r["total"]

    updates: List[Dict] = []

    # Week label one row above the anchor in the Rep column
    we_label = f"WE {week_end.month}.{week_end.day}"
    if anchor_row - 1 >= 0:
        updates.append({
            "range": gspread.utils.rowcol_to_a1(anchor_row, anchor_col + 1),
            "values": [[we_label]],
        })

    # Total row (anchor + 1, 0-indexed → +2 for 1-indexed A1)
    total_row_1based = anchor_row + 2
    row_width = total_col - anchor_col + 1
    total_row_values = ["Total", "Total"] + [""] * (row_width - 2)
    for day, j in day_cols.items():
        n = day_totals.get(day, 0)
        total_row_values[j - anchor_col] = str(n) if n else ""
    total_row_values[total_col - anchor_col] = str(grand_total) if grand_total else ""
    updates.append({
        "range": (f"{gspread.utils.rowcol_to_a1(total_row_1based, anchor_col + 1)}:"
                  f"{gspread.utils.rowcol_to_a1(total_row_1based, total_col + 1)}"),
        "values": [total_row_values],
    })

    # Per-rep rows (sorted by total desc — already sorted upstream)
    for idx, r in enumerate(reps):
        row_1based = anchor_row + 3 + idx   # anchor+2 + idx (0-indexed) → +3 for 1-indexed
        row_values = [r["rep"], "WIRELESS"] + [""] * (row_width - 2)
        for day, j in day_cols.items():
            n = r["days"].get(day, 0)
            row_values[j - anchor_col] = str(n) if n else ""
        row_values[total_col - anchor_col] = str(r["total"])
        updates.append({
            "range": (f"{gspread.utils.rowcol_to_a1(row_1based, anchor_col + 1)}:"
                      f"{gspread.utils.rowcol_to_a1(row_1based, total_col + 1)}"),
            "values": [row_values],
        })

    # Clear any leftover rows from prior larger weeks. Blank out
    # max_rep_rows rows below the last written rep row, capped at the
    # current grid height.
    leftover_start = anchor_row + 3 + len(reps)   # 1-indexed: anchor_row + 2 + len(reps) + 1
    leftover_end = anchor_row + 2 + max_rep_rows
    if leftover_start <= leftover_end:
        blank_row = [""] * row_width
        for r1 in range(leftover_start, leftover_end + 1):
            if r1 - 1 >= len(grid):
                break
            # Only blank rows where SOMETHING is in the chart columns;
            # otherwise skip to save API calls.
            row = grid[r1 - 1]
            has_content = any(
                (row[c] or "").strip()
                for c in range(anchor_col, total_col + 1)
                if c < len(row)
            )
            if has_content:
                updates.append({
                    "range": (f"{gspread.utils.rowcol_to_a1(r1, anchor_col + 1)}:"
                              f"{gspread.utils.rowcol_to_a1(r1, total_col + 1)}"),
                    "values": [blank_row],
                })

    # Formatting strategy — chart must look uniform regardless of how many
    # rep slots were pre-formatted in Megan's template:
    #   1. Copy the format of the FIRST rep row (template's canonical style)
    #      to every other rep row we wrote. Ensures borders, font, alignment
    #      etc. match for all reps even when len(reps) > template slot count.
    #   2. Apply zebra striping (overrides backgroundColor on alternate rows).
    #   3. Set THICK bottom border on the actual last rep row so the chart's
    #      bottom edge tracks the current rep count.
    #   4. Clear all formatting on rows below the last rep row in the chart's
    #      column range (eliminates leftover bordered/empty rows from prior
    #      larger weeks).
    fmt_requests: List[Dict] = []
    first_rep_row0 = anchor_row + 2   # 0-indexed first rep row
    ws_id = ws.id

    # 1. Replicate template's first-rep-row formatting across all rep rows.
    if len(reps) > 1:
        fmt_requests.append({"copyPaste": {
            "source": {"sheetId": ws_id,
                       "startRowIndex": first_rep_row0,
                       "endRowIndex": first_rep_row0 + 1,
                       "startColumnIndex": anchor_col,
                       "endColumnIndex": total_col + 1},
            "destination": {"sheetId": ws_id,
                            "startRowIndex": first_rep_row0 + 1,
                            "endRowIndex": first_rep_row0 + len(reps),
                            "startColumnIndex": anchor_col,
                            "endColumnIndex": total_col + 1},
            "pasteType": "PASTE_FORMAT",
            "pasteOrientation": "NORMAL"}})

    # 2. Zebra background — applies AFTER copyPaste so it survives.
    for idx in range(len(reps)):
        row0 = first_rep_row0 + idx
        color = REP_CHART_LIGHT_GRAY if idx % 2 == 0 else REP_CHART_WHITE
        fmt_requests.append({"repeatCell": {
            "range": {"sheetId": ws_id,
                      "startRowIndex": row0, "endRowIndex": row0 + 1,
                      "startColumnIndex": anchor_col,
                      "endColumnIndex": total_col + 1},
            "cell": {"userEnteredFormat": {"backgroundColor": color}},
            "fields": "userEnteredFormat.backgroundColor"}})

    # 3. Move the chart's THICK bottom border to the last rep row. Without
    # this, copyPaste in step 1 gives the last rep a regular (thin) bottom
    # border like the middle reps, so the chart visually "bleeds" into the
    # rows below.
    if len(reps) > 0:
        last_rep_row0 = first_rep_row0 + len(reps) - 1
        fmt_requests.append({"updateBorders": {
            "range": {"sheetId": ws_id,
                      "startRowIndex": last_rep_row0,
                      "endRowIndex": last_rep_row0 + 1,
                      "startColumnIndex": anchor_col,
                      "endColumnIndex": total_col + 1},
            "bottom": {"style": "SOLID_THICK",
                       "color": {"red": 0, "green": 0, "blue": 0}}}})

    # 4. Fully clear formatting on every leftover row inside the chart's
    # column range so it's not visible as part of the chart. Uses
    # updateCells with empty userEnteredFormat to wipe borders + bg + font
    # at once. Limited to max_rep_rows past the last written rep row.
    leftover_first0 = first_rep_row0 + len(reps)
    leftover_last0 = first_rep_row0 + max_rep_rows - 1
    if leftover_first0 <= leftover_last0:
        end_row_excl = min(leftover_last0 + 1, len(grid))
        if leftover_first0 < end_row_excl:
            fmt_requests.append({"updateCells": {
                "range": {"sheetId": ws_id,
                          "startRowIndex": leftover_first0,
                          "endRowIndex": end_row_excl,
                          "startColumnIndex": anchor_col,
                          "endColumnIndex": total_col + 1},
                "fields": "userEnteredFormat"}})

    log.append(f"  chart: {len(reps)} rep(s), total={grand_total}, header={we_label!r}")
    if dry_run:
        return [f"[DRY-RUN chart] {ws.title}: would write {len(updates)} range(s) + {len(fmt_requests)} format(s)"] + log
    if updates:
        rfill._retry(ws.batch_update, updates, value_input_option="USER_ENTERED")
    if fmt_requests:
        rfill._retry(ws.spreadsheet.batch_update, {"requests": fmt_requests})
    if updates or fmt_requests:
        return [f"[OK chart] {ws.title}: wrote {len(updates)} range(s) + {len(fmt_requests)} format(s)"] + log
    return [f"[skip-chart] {ws.title}: nothing to write"]


def fill_nds_tab(ws: gspread.Worksheet, owner_norm: str,
                 tt: Dict[str, Dict[str, str]],
                 rep_summary_total: Dict[str, str],
                 sara_totals: Dict[str, str],
                 churn: Dict[str, Dict[str, str]],
                 week_col_label: str,
                 dry_run: bool = False,
                 logfn=print,
                 # HTTP-sourced metrics (optional — when not provided, the
                 # corresponding rows just stay blank rather than erroring)
                 activation: Optional[Dict[str, str]] = None,
                 cancel: Optional[Dict[str, str]] = None,
                 leads: Optional[Dict[str, int]] = None,
                 sara_byday: Optional[Dict[str, Dict[str, int]]] = None,
                 personal_production: Optional[Dict[str, str]] = None,
                 direct_deposit: Optional[Dict[str, float]] = None,
                 abp: Optional[Dict[str, str]] = None,
                 aliases_map: Optional[Dict[str, List[str]]] = None,
                 backfill: bool = False,
                 ) -> List[str]:
    """Write all available metrics for this rep into the target week column.
    Skips metrics whose source data isn't available.

    `backfill=True` (past-week catch-up): don't write the metrics whose
    dashboards have NO date control and so can't be pinned to a past week
    — NDSDailyTracker (Active Selling Heads, Scorecard Ranking, National
    AVG, AVG Apps), Direct Deposit, and Total Leads. Those cells are left
    blank rather than filled with wrong-week data. (Megan 2026-05-24.)"""
    log: List[str] = []
    grid = rfill._retry(ws.get_all_values)
    if not grid:
        return [f"[skip] {ws.title}: empty tab"]

    week_col = _find_week_col(grid, week_col_label)
    if week_col is None:
        return [f"[skip] {ws.title}: no column for week {week_col_label}"]

    # Gather values
    rep = tt.get(owner_norm, {})
    crow = churn.get(owner_norm, {})

    values: Dict[str, str] = {}
    # NDSDailyTracker-sourced metrics. That dashboard has no URL date
    # control, so on a past-week backfill we leave these blank rather
    # than write the wrong (current) week.
    if not backfill:
        if "rep_count" in rep:
            values["Active Selling Heads"] = rep["rep_count"]
        if "ranking" in rep:
            values["Scorecard Ranking"] = rep["ranking"]
        # 'National AVG for sales' = Rep Summary Total row, 'New/Port per
        # Rep' column (Megan confirmed via screenshot 2026-05-24).
        if rep_summary_total.get("New/Port per Rep"):
            values["National AVG for sales"] = rep_summary_total["New/Port per Rep"]
    if "churn_30" in crow:
        values["0-30 Day Churn"] = crow["churn_30"]
    if "churn_60" in crow:
        values["60 Day Churn"] = crow["churn_60"]
    # 90 Day Churn: Tableau leaves this empty for newer ICDs whose
    # customers haven't aged 90 days yet (e.g. Maxamed). Write 'N/A' so
    # the cell reads as a known gap, not a missed pull.
    values["90 Day Churn"] = crow.get("churn_90") or "N/A"

    # HTTP-sourced per-rep metrics
    if activation and activation.get(owner_norm):
        values["Activation % by Week"] = activation[owner_norm]
    if cancel and cancel.get(owner_norm) is not None:
        # Tableau exports the cancel rate as a decimal fraction
        # (e.g. 0.004975 = 0.50%). Convert to a percent string for display.
        raw = (cancel[owner_norm] or "").strip()
        try:
            values["0-30 Day Cancel Rate  4wk avg"] = f"{float(raw):.2%}"
        except ValueError:
            values["0-30 Day Cancel Rate  4wk avg"] = raw
    # Total Leads: Jairo Ruiz isn't in the Lead Penetration view at all
    # (his ICD may be rolled under a parent or filtered out). Write 'N/A'
    # for any owner missing from the leads dict so the gap is explicit.
    # Skipped on backfill — LeadPenetrationOverview isn't date-pinnable.
    if leads is not None and not backfill:
        v = leads.get(owner_norm)
        values["Total Leads"] = str(v) if v is not None else "N/A"

    # Sara Plus By Day → per-rep weekly totals for the 5 wireless metrics.
    # 'New Lines' = Wireless Lines (Megan's mapping). Per-rep new lines
    # feeds AVG Apps Per Active Headcount = new_lines / active_selling_heads.
    rep_byday = (sara_byday or {}).get(owner_norm, {})
    new_lines_rep = rep_byday.get("Wireless Lines")
    if new_lines_rep is not None:
        values["New Lines"] = str(new_lines_rep)
        # AVG Apps Per Active Headcount = New Lines / Active Selling Heads.
        # Active Selling Heads is NDSDailyTracker-sourced, so this is
        # skipped in backfill mode (no headcount for a past week).
        if not backfill:
            try:
                heads = float(str(rep.get("rep_count", "")).strip())
                if heads > 0:
                    values["AVG Apps Per Active Headcount"] = f"{new_lines_rep / heads:.2f}"
            except (ValueError, TypeError):
                pass

    # Personal Production — the ICD's OWN per-rep production from
    # ProductSalesSummaryRep (0 for managers like Jairo/Isaiah, whose
    # teams sell but who don't personally). NOT the office total — that
    # was the SARA By Day bug Megan caught 2026-05-24. ProductSalesSummaryRep
    # has no date control, so it's skipped on a past-week backfill.
    if (personal_production and personal_production.get(owner_norm)
            and not backfill):
        values["Personal Production"] = personal_production[owner_norm]

    # Direct Deposit — per-ICD-owner dollar total. Skipped on backfill —
    # the DD dashboard isn't date-pinnable (Megan 2026-05-24).
    if direct_deposit and not backfill:
        # match_dd_owner resolves nickname/spelling variants via the alias list
        # (tab 'Joe Delgado' -> grid 'joseph delgado'; 'Selena' -> 'selena
        # powers') without conflating same-surname siblings (Joe vs George).
        _dd_v = match_dd_owner(direct_deposit, owner_norm, aliases_map)
        if _dd_v is not None:
            # Format as dollar amount with comma separators (matches sheet style)
            values["Direct Deposit"] = f"${_dd_v:,.2f}"

    # ABP % — from ABPCONVERSIONS (same source Retail uses). Date-pinned, so
    # it's safe on a backfill too (Megan 2026-05-25).
    if abp and abp.get(owner_norm):
        values["ABP %"] = abp[owner_norm]

    # Office-level shared metrics computed from Sara Plus org totals.
    # Tableau crosstabs comma-separate large numbers ("1,062") — strip
    # them before float-conversion so the division doesn't silently fail.
    def _num(s):
        try:
            return float(str(s).replace(",", "").strip())
        except (ValueError, AttributeError):
            return None
    next_up = _num(sara_totals.get("Next Up"))
    new_port = _num(sara_totals.get("New/Port Lines"))
    premium = _num(sara_totals.get("Premium/Elite"))
    extra = _num(sara_totals.get("Extra"))
    # Next Up % = Next Up / New/Port Lines (office total, shared)
    if next_up is not None and new_port and new_port > 0:
        values["Next Up %"] = f"{next_up / new_port:.2%}"
    # Extra/Premium % = (Premium/Elite + Extra) / New/Port Lines (shared)
    if premium is not None and extra is not None and new_port and new_port > 0:
        values["Extra/Premium %"] = f"{(premium + extra) / new_port:.2%}"

    if not values:
        return [f"[skip] {ws.title}: no metrics available for {owner_norm}"]

    # Build batch update — only writes cells where we found a row. Some tabs
    # label a row differently than its `values` key (see NDS_ROW_LABEL_
    # ALTERNATES); try each known spelling before giving up.
    updates = []
    for label, val in values.items():
        row = None
        for cand in NDS_ROW_LABEL_ALTERNATES.get(label, [label]):
            row = _find_row_by_label(grid, cand)
            if row is not None:
                break
        if row is None:
            log.append(f"  [miss] {ws.title}: no row for {label!r}")
            continue
        a1 = gspread.utils.rowcol_to_a1(row + 1, week_col + 1)
        updates.append({"range": a1, "values": [[val]]})
        log.append(f"  {a1} ({label}) ← {val!r}")

    if dry_run:
        return [f"[DRY-RUN] {ws.title}: would write {len(updates)} cells"] + log
    if updates:
        rfill._retry(ws.batch_update, updates, value_input_option="USER_ENTERED")
        _red = _sheet_flags.weird_ranges(updates)   # fill-but-flag weird %s
        if _red:
            _sheet_flags.apply_red_font(ws, _red, retry=rfill._retry)
        return [f"[OK] {ws.title}: wrote {len(updates)} cells"
                + (f" ({len(_red)} flagged)" if _red else "")] + log
    return [f"[skip] {ws.title}: nothing to write"]


# ------------------------------------------------------------ entry point


def run_nds_opt(dry_run: bool = False, only_rep: Optional[str] = None,
                skip_download: bool = False, backfill: bool = False,
                logfn=print) -> dict:
    """Download all NDS Tableau views, parse, and fill each NDS tab on
    Alphalete Org. Returns {filled: [...], skipped: [...], errors: [...]}.

    `backfill=True`: a past-week run (e.g. catching up WE 5/17 after the
    dashboards rolled to the current week). Pins the SARA views to the
    target week via Min/Max Date params, and SKIPS the NDSDailyTracker
    metrics + Rep Breakdown chart (their dashboards have no date control,
    so a past week can't be reconstructed — those cells are left blank)."""
    download_errors: List[str] = []
    date_params = _target_week_date_range()
    # Views whose dashboards honor Min/Max Date URL params (pin to the
    # target week). NDSDailyTracker is NOT here — it ignores the params.
    SARA_BYDAY_FNAME = "opt_nds_sara_plus_byday.csv"

    # All downloads run inside ONE patchright stealth session — no CDP /
    # Report Chrome dependency, so NDS is fully unattended/schedulable.
    #   - HTTP-direct CSV pulls (fast) reuse the patchright cookies.
    #   - UI Crosstab pulls drive the Download dialog. The
    #     drive_crosstab_dialog click logic stops at the first strategy
    #     that enables Export (instead of firing every strategy and
    #     accidentally deselecting) — that fixed the dialogs that used
    #     to silent-no-op under both CDP and the old patchright path.
    # (Megan 2026-05-24.)
    def _fallback_existing(filename: str) -> bool:
        target = OUTPUT_DIR / filename
        if target.exists() and target.stat().st_size > 500:
            logfn(f"OPT NDS: using existing {filename} "
                  f"({target.stat().st_size:,} bytes) as fallback")
            return True
        return False

    if not skip_download:
        try:
            with tableau_session(verbose=False) as page:
                # 1a: HTTP-direct CSV pulls via patchright cookies. SARA By
                # Day is pinned to the target week (it honors Min/Max Date);
                # the others are date-robust (rolling / 4wk-avg) so left as-is.
                http_session = requests_session_from_page(page)
                for wb, view, fname in NDS_HTTP_VIEWS:
                    out = OUTPUT_DIR / fname
                    params = date_params if fname == SARA_BYDAY_FNAME else None
                    try:
                        logfn(f"OPT NDS: HTTP downloading {fname}…")
                        tableau_http.download_view_csv(wb, view, out,
                                                       session=http_session,
                                                       params=params)
                    except Exception as e:
                        msg = f"{fname}: {type(e).__name__}: {str(e)[:120]}"
                        logfn(f"OPT NDS: ✗ HTTP {msg}")
                        if not _fallback_existing(fname):
                            download_errors.append(msg)

                # 1b: UI Crosstab pulls
                for url, sheet, fname in NDS_VIEWS:
                    # Backfill skips the views with no date control — their
                    # past-week data can't be reconstructed.
                    if backfill and ("NDSDailyTracker" in url
                                     or "ProductSalesSummaryRep" in url):
                        logfn(f"OPT NDS: [backfill skip] {fname} (no date control)")
                        continue
                    # Pin date-aware views (SARA total, ABP) to the target week.
                    dl_url = url
                    if "SARAPLUSSALESSUMMARY" in url or "ABPCONVERSIONS" in url:
                        sep = "&" if "?" in dl_url else "?"
                        dl_url = (f"{dl_url}{sep}Min%20Date={date_params['Min Date']}"
                                  f"&Max%20Date={date_params['Max Date']}")
                    out = OUTPUT_DIR / fname
                    last_err = None
                    for attempt in (1, 2):
                        tag = f"attempt {attempt}/2 "
                        logfn(f"OPT NDS: patchright Crosstab {tag}→ "
                              f"{sheet!r} → {fname}…")
                        try:
                            download_crosstab_patchright(
                                dl_url, sheet, out, verbose=False, page=page)
                            last_err = None
                            break
                        except Exception as e:
                            last_err = e
                            if attempt == 1:
                                logfn(f"OPT NDS:   {type(e).__name__}, "
                                      "resetting page + retrying…")
                                try:
                                    page.goto("about:blank",
                                              wait_until="domcontentloaded",
                                              timeout=10_000)
                                    page.wait_for_timeout(3_000)
                                except Exception:
                                    pass
                    if last_err is not None:
                        msg = f"{fname}: {type(last_err).__name__}: {str(last_err)[:120]}"
                        logfn(f"OPT NDS: ✗ patchright {msg}")
                        if not _fallback_existing(fname):
                            download_errors.append(msg)
        except Exception as e:
            logfn(f"OPT NDS: patchright session failed: "
                  f"{type(e).__name__}: {str(e)[:160]}")
            download_errors.append(
                f"patchright session: {type(e).__name__}: {str(e)[:120]}")
    else:
        logfn("OPT NDS: --skip-download — reusing cached crosstabs")

    # Step 2: parse what we have
    tt = parse_tt_detail(OUTPUT_DIR / "opt_nds_tt_detail.csv")
    rep_summary = parse_rep_summary_total(OUTPUT_DIR / "opt_nds_rep_summary.csv")
    sara_totals = parse_sara_plus_total(OUTPUT_DIR / "opt_nds_sara_plus.csv")
    personal_production = parse_personal_production(
        OUTPUT_DIR / "opt_nds_personal_production.csv")
    rep_breakdown = parse_rep_breakdown_per_owner(
        OUTPUT_DIR / "opt_nds_personal_production.csv")
    churn = parse_churn_icd(OUTPUT_DIR / "opt_nds_churn.csv")
    direct_deposit = parse_direct_deposit(OUTPUT_DIR / "opt_nds_direct_deposit.csv")
    # ABP % — reuse Retail's parser (same ABPCONVERSIONS source). Lazy import:
    # opt_retail imports opt_nds at module load, so importing it at the top here
    # would be circular; by run-time both modules are loaded.
    try:
        from automations.alphalete_org_report.opt_retail import parse_abp_conversions
        abp = parse_abp_conversions(OUTPUT_DIR / NDS_ABP_FILENAME)
    except Exception as e:
        logfn(f"OPT NDS: ABP parse skipped ({type(e).__name__})")
        abp = {}
    # HTTP-sourced parses
    activation = tableau_http.parse_activation(
        OUTPUT_DIR / "opt_nds_activation_http.csv")
    cancel = tableau_http.parse_weekly_metrics_cancel(
        OUTPUT_DIR / "opt_nds_weekly_metrics_http.csv")
    leads = tableau_http.parse_lead_penetration(
        OUTPUT_DIR / "opt_nds_lead_penetration_http.csv")
    sara_byday = tableau_http.parse_sara_plus_byday(
        OUTPUT_DIR / "opt_nds_sara_plus_byday.csv")
    logfn(f"OPT NDS: parsed {len(tt)} TT-Detail, "
          f"{len(churn)} Churn, "
          f"{len(activation)} Activation, "
          f"{len(cancel)} Cancel, "
          f"{len(leads)} Leads, "
          f"{len(sara_byday)} Sara-ByDay, "
          f"{len(personal_production)} PP, "
          f"{len(direct_deposit)} DD, "
          f"{len(abp)} ABP")

    # Step 3: open sheet + walk NDS-suffixed tabs
    client = rfill._client()
    sh = rfill.open_by_key(ALPHALETE_ORG_SHEET_ID, client)
    # Skip hidden tabs (same convention as recruiting + financial)
    resp = sh.client.request(
        "get",
        f"https://sheets.googleapis.com/v4/spreadsheets/{sh.id}",
        params={"fields": "sheets(properties(title,hidden))"},
    )
    hidden = {s["properties"]["title"] for s in resp.json().get("sheets", [])
              if s["properties"].get("hidden")}

    week_col_label = _current_target_week_col_label()
    logfn(f"OPT NDS: target week column = {week_col_label!r}")

    # Shared ICD alias list — lets DD resolve nickname/spelling variants
    # (Joe Delgado -> Joseph Delgado, Selena -> Selena Powers). Loaded once.
    try:
        from automations.focus_office_att.aliases import load_aliases as _la
        nds_aliases = _la()
    except Exception:
        nds_aliases = {}

    filled: List[str] = []
    skipped: List[str] = []
    for ws in rfill._retry(sh.worksheets):
        title = ws.title
        if not title.endswith(" - NDS") or title in hidden or title.startswith("x"):
            continue
        rep_name = title[: -len(" - NDS")].strip()
        owner_norm = _norm_owner(rep_name)
        if only_rep and only_rep.lower() not in rep_name.lower():
            continue
        # Owner aliases: try the tab name + a couple fallbacks. Tableau
        # spellings often differ slightly from sheet tab names (e.g. tab
        # 'Maxamed Aden' vs Tableau 'MAXAMAD ADEN'). We try:
        #   1. Direct match against the normalized tab name
        #   2. Same last-name + same-first-letter-of-first-name match
        #      (catches 'Maxamed' ↔ 'Maxamad', 'Selena' ↔ 'Selena Powers')
        #   3. Same last-name only (last-resort)
        candidates = [owner_norm]
        tw = owner_norm.split()
        for k in tt:
            kw = k.split()
            if not kw or not tw:
                continue
            if kw[-1] == tw[-1] and kw[0][:1] == tw[0][:1]:
                if k not in candidates:
                    candidates.append(k)
        # Final fallback: same last name only
        for k in tt:
            kw = k.split()
            if kw and tw and kw[-1] == tw[-1] and k not in candidates:
                candidates.append(k)
        match = next((c for c in candidates if c in tt), None) or owner_norm

        lines = fill_nds_tab(ws, match, tt, rep_summary, sara_totals,
                             churn, week_col_label, dry_run, logfn,
                             activation=activation, cancel=cancel,
                             leads=leads, sara_byday=sara_byday,
                             personal_production=personal_production,
                             direct_deposit=direct_deposit,
                             abp=abp,
                             aliases_map=nds_aliases,
                             backfill=backfill)
        for ln in lines:
            logfn(f"OPT NDS: {ln}")
        if lines and lines[0].startswith(("[OK]", "[DRY-RUN]")):
            filled.append(title)
        else:
            skipped.append(title)

        # Rep Breakdown chart — sourced from ProductSalesSummaryRep, which
        # has no date control, so it's skipped on a past-week backfill.
        if not backfill:
            chart_lines = fill_rep_breakdown_chart(
                ws, match, rep_breakdown, _current_target_week_end(),
                dry_run=dry_run, logfn=logfn,
            )
            for ln in chart_lines:
                logfn(f"OPT NDS: {ln}")

    return {"filled": filled, "skipped": skipped, "errors": download_errors}


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--only", help="Only this rep (substring match).")
    ap.add_argument("--skip-download", action="store_true",
                    help="Reuse cached crosstabs instead of re-pulling.")
    ap.add_argument("--backfill", action="store_true",
                    help="Past-week catch-up: pin SARA to the target week, "
                         "skip the no-date-control metrics (NDSDailyTracker "
                         "+ Rep Breakdown chart).")
    args = ap.parse_args()
    result = run_nds_opt(dry_run=args.dry_run, only_rep=args.only,
                         skip_download=args.skip_download,
                         backfill=args.backfill)
    print(f"\nFilled: {len(result['filled'])} tab(s); "
          f"Skipped: {len(result['skipped'])}; "
          f"Download errors: {len(result['errors'])}")
    for err in result["errors"]:
        print(f"  ✗ {err}")
