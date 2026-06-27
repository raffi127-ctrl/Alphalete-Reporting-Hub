"""Parse Archey's Residential Rep Count .xlsx attachment.

The data we use lives on the `ICD Headcount (by Campaign)` tab. Header is on
row 2:
  Ownerville # | Account ID | ICD Owner Name | ICD Company Name | NC | RC |
  Org Leader | ICD Company Name | PO | Unique Headcount | <17 campaign cols> ...

Metric = `Unique Headcount`, matched by `ICD Owner Name`. (Validated 2026-06-27
cell-for-cell 30/30 against the sheet's existing WE 6/13/26 column. The sibling
`ICD Owner Snapshot` tab's `Headcount` column is a 5-week churn-cohort number —
do NOT use it.)
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict

import openpyxl

HEADCOUNT_TAB = "ICD Headcount (by Campaign)"


def norm_name(s: str) -> str:
    """Lowercase, drop everything but letters — so 'Aya Al-Khafaji',
    'Aya Al Khafaji', 'aya  al-khafaji' all collapse to one key."""
    return re.sub(r"[^a-z]", "", (s or "").lower())


def _header_row(ws):
    """Find the (1-based row, header list) carrying 'ICD Owner Name' +
    'Unique Headcount'. The tab has a title row above the header."""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        low = [(str(c).lower() if c else "") for c in row]
        if any("icd owner name" in c for c in low) and any(
            "unique headcount" in c for c in low
        ):
            return ri, list(row)
    raise ValueError(
        f"'{HEADCOUNT_TAB}': couldn't find a header row with "
        "'ICD Owner Name' + 'Unique Headcount'")


def parse_headcounts(xlsx_path: str | Path) -> Dict[str, dict]:
    """Return {norm_name: {'name', 'headcount', 'org_leader'}} for every ICD
    owner row on the `ICD Headcount (by Campaign)` tab."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if HEADCOUNT_TAB not in wb.sheetnames:
        raise ValueError(
            f"tab {HEADCOUNT_TAB!r} not found. Tabs: {wb.sheetnames}")
    ws = wb[HEADCOUNT_TAB]
    hrow, header = _header_row(ws)
    ci = {(str(h).strip().lower() if h else ""): j for j, h in enumerate(header)}
    cn = ci.get("icd owner name")
    chc = ci.get("unique headcount")
    corg = ci.get("org leader")
    if cn is None or chc is None:
        raise ValueError(
            f"'{HEADCOUNT_TAB}': missing required columns "
            f"(owner={cn}, unique headcount={chc})")

    out: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        nm = row[cn] if cn < len(row) else None
        if not nm or not str(nm).strip():
            continue
        raw = row[chc] if chc < len(row) else None
        try:
            hc = int(float(raw)) if raw not in (None, "") else 0
        except (TypeError, ValueError):
            hc = 0
        org = ""
        if corg is not None and corg < len(row) and row[corg]:
            org = str(row[corg]).strip()
        out[norm_name(str(nm))] = {
            "name": str(nm).strip(),
            "headcount": hc,
            "org_leader": org,
        }
    return out
