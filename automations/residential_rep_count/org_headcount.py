"""SCI RC/NC org-level headcount → the 'SCI RC/NC Headcount' tab.

Source: the xlsx attachment's 'Org Snapshot (by Campaign)' tab — one row per
Org Leader with that week's 'Unique Headcount' (the SAME number shown in the
email body's 'RC/NC Org Snapshot' table; NOT the 5-week cohort 'Headcount' on
the RC/NC Org Snapshot tabs). Megan 2026-06-27.

Tab layout (Megan owns it; read by LABEL, never fixed indices):
  A 'Org Leader' | B 'Group' | then one 'WE m/d/yy' column per week
  active leaders … | 'No longer active' divider | inactive leaders | 'TOTALS'

Weekly upkeep mirrors Rep Count 24-26: append the week column (inherit prior
week's format), auto-add new org leaders, move 0-for-2-weeks → inactive and
inactive-with-data → active, recompute TOTALS, sort both sections + collapse
the inactive group.
"""
from __future__ import annotations

import datetime as dt
from typing import Dict, List, Optional, Tuple

import openpyxl
from gspread.utils import rowcol_to_a1

from automations.recruiting_report import fill as rfill
from automations.residential_rep_count.parse import norm_name
from automations.residential_rep_count.fill import (
    SPREADSHEET_ID, we_label, we_label_to_date)

ORG_TAB = "Org Snapshot (by Campaign)"        # tab inside the xlsx attachment
SCI_TAB = "SCI RC/NC Headcount"               # destination tab
SCI_SANDBOX_TAB = "SCI RC/NC Headcount SANDBOX"
ZERO_STREAK_ORG = 2                           # 0 for N weeks running -> inactive


def parse_org_snapshot(xlsx_path) -> Dict[str, dict]:
    """{norm_leader: {name, headcount, office, group}} from the
    'Org Snapshot (by Campaign)' tab (one row per Org Leader)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if ORG_TAB not in wb.sheetnames:
        raise ValueError(f"tab {ORG_TAB!r} not found. Tabs: {wb.sheetnames}")
    ws = wb[ORG_TAB]
    hrow = hdr = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        low = [(str(c).lower() if c else "") for c in row]
        if any("org leader" in c for c in low) and any(
                "unique headcount" in c for c in low):
            hrow, hdr = ri, list(row)
            break
    if hdr is None:
        raise ValueError(f"{ORG_TAB!r}: no 'Org Leader' + 'Unique Headcount' header")
    ci = {(str(h).strip().lower() if h else ""): j for j, h in enumerate(hdr)}
    cl, chc = ci["org leader"], ci["unique headcount"]
    co, cg = ci.get("icd office"), ci.get("group")
    out: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        nm = row[cl] if cl < len(row) else None
        if not nm or not str(nm).strip():
            continue
        # Skip the snapshot's summary rows (Total (Legacy D2D) / Non Legacy
        # Org Leaders / Grand Total) — not real org leaders.
        nl = str(nm).strip().lower()
        if "total" in nl or "non legacy" in nl:
            continue
        try:
            hc = int(float(row[chc])) if (chc < len(row) and row[chc] not in (None, "")) else 0
        except (TypeError, ValueError):
            hc = 0
        out[norm_name(str(nm))] = {
            "name": str(nm).strip(),
            "headcount": hc,
            "office": (str(row[co]).strip() if co is not None and co < len(row) and row[co] else ""),
            "group": (str(row[cg]).strip() if cg is not None and cg < len(row) and row[cg] else ""),
        }
    return out


def build_rows(weeks: List[Tuple[dt.date, Dict[str, dict]]]) -> List[list]:
    """weeks oldest→newest → full A1 block: A Org Leader | B Group | week cols.
    One row per leader (union; newest office/group), sorted by name."""
    latest: Dict[str, dict] = {}
    for _d, parsed in weeks:
        latest.update(parsed)
    order = sorted(latest.values(), key=lambda r: norm_name(r["name"]))
    rows = [["Org Leader", "Group"] + [we_label(d) for d, _ in weeks]]
    for rec in order:
        k = norm_name(rec["name"])
        line = [rec["name"], rec["group"]]
        for _d, parsed in weeks:
            r = parsed.get(k)
            line.append(r["headcount"] if r else "")
        rows.append(line)
    return rows


def open_tab(sandbox: bool = True):
    sh = rfill.open_by_key(SPREADSHEET_ID)
    title = SCI_SANDBOX_TAB if sandbox else SCI_TAB
    try:
        ws = sh.worksheet(title)
    except Exception:
        ws = sh.add_worksheet(title=title, rows=200, cols=80)
    return ws, sh


def write_backfill(ws, weeks, dry_run: bool = True) -> dict:
    """One-time build of an EMPTY tab (Org Leader × weekly grid)."""
    rows = build_rows(weeks)
    if not dry_run:
        rfill._retry(ws.update, values=rows, range_name="A1",
                     value_input_option="USER_ENTERED")
        rfill._retry(ws.spreadsheet.batch_update, {"requests": [
            {"repeatCell": {
                "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True},
                                               "horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat(textFormat,horizontalAlignment)"}},
            {"updateSheetProperties": {
                "properties": {"sheetId": ws.id, "gridProperties": {
                    "frozenRowCount": 1, "frozenColumnCount": 2}},
                "fields": "gridProperties(frozenRowCount,frozenColumnCount)"}}]})
    return {"leaders": len(rows) - 1, "weeks": len(weeks),
            "first_week": weeks[0][0], "last_week": weeks[-1][0]}


# ---------- weekly upkeep (active/inactive + TOTALS) ----------

def resolve_org_layout(grid) -> dict:
    A = lambda r: (r[0] if r else "").strip()
    header = next((i for i, r in enumerate(grid) if A(r).lower() == "org leader"), 0)
    nla = next((i for i in range(header + 1, len(grid))
                if "no longer" in A(grid[i]).lower()), None)
    totals = next((i for i in range(header + 1, len(grid))
                   if A(grid[i]).lower().strip().rstrip("s") == "total"), None)
    hd = grid[header]
    we_cols = {h: j for j, h in enumerate(hd) if (h or "").startswith("WE ")}
    group_col = next((j for j, h in enumerate(hd)
                      if (h or "").strip().lower() == "group"), 1)
    act_end = (nla - 1) if nla is not None else (
        (totals - 1) if totals is not None else len(grid) - 1)
    inactive = ((nla + 1, (totals - 1) if totals is not None else len(grid) - 1)
                if nla is not None else None)
    return {"header": header, "nla": nla, "totals": totals, "we_cols": we_cols,
            "group_col": group_col, "active": (header + 1, act_end),
            "inactive": inactive}


def _row_by_name(grid, name) -> Optional[int]:
    n = norm_name(name)
    for i, r in enumerate(grid):
        if r and norm_name(r[0]) == n:
            return i
    return None


def _org_ops(grid, layout, parsed, target_label) -> List[dict]:
    ops: List[dict] = []
    a0, a1 = layout["active"]
    dated = sorted(((d, c) for l, c in layout["we_cols"].items()
                    if (d := we_label_to_date(l)) is not None), key=lambda x: x[0])
    tgt = we_label_to_date(target_label)
    prior = [c for d, c in dated if tgt is None or d < tgt][-(ZERO_STREAK_ORG - 1):]
    existing = {norm_name(grid[r][0]) for r in range(len(grid))
                if grid[r] and (grid[r][0] or "").strip()}
    # DOWN: active, 0/blank this week + 0/blank the prior week(s).
    for r in range(a0, a1 + 1):
        nm = (grid[r][0] if grid[r] else "").strip()
        if not nm:
            continue
        rec = parsed.get(norm_name(nm))
        if (rec["headcount"] if rec else 0) == 0:
            pv = [str(grid[r][c]).strip() if c < len(grid[r]) else "" for c in prior]
            if prior and all(p in ("", "0") for p in pv):
                ops.append({"type": "down", "name": nm})
    # UP: inactive with data this week.
    if layout["inactive"]:
        i0, i1 = layout["inactive"]
        for r in range(i0, i1 + 1):
            nm = (grid[r][0] if grid[r] else "").strip()
            if not nm:
                continue
            rec = parsed.get(norm_name(nm))
            if rec and rec["headcount"] >= 1:
                ops.append({"type": "up", "name": nm, "group": rec["group"]})
    # ADD: new org leader in the email with data, no row yet.
    for k, rec in parsed.items():
        if rec["headcount"] >= 1 and k not in existing:
            ops.append({"type": "add", "name": rec["name"], "group": rec["group"]})
    return ops


def _org_apply(ws, sid, grid, layout, op) -> str:
    a0, a1 = layout["active"]
    gc = layout["group_col"]
    if op["type"] == "add":
        ins = a0 + 1                       # after 1st leader → inherits data-row fmt
        rfill._retry(ws.spreadsheet.batch_update, {"requests": [{
            "insertDimension": {"range": {"sheetId": sid, "dimension": "ROWS",
                "startIndex": ins, "endIndex": ins + 1}, "inheritFromBefore": True}}]})
        rfill._retry(ws.update, values=[[op["name"]]], range_name=f"A{ins + 1}",
                     value_input_option="USER_ENTERED")
        if gc is not None:
            rfill._retry(ws.update, values=[[op.get("group", "")]],
                         range_name=rowcol_to_a1(ins + 1, gc + 1),
                         value_input_option="USER_ENTERED")
        return f"+ added {op['name']!r}"
    src = _row_by_name(grid, op["name"])
    if src is None:
        return f"  (skip {op['name']!r})"
    if op["type"] == "up":
        rfill._retry(ws.spreadsheet.batch_update, {"requests": [{
            "moveDimension": {"source": {"sheetId": sid, "dimension": "ROWS",
                "startIndex": src, "endIndex": src + 1}, "destinationIndex": a0}}]})
        return f"↑ moved {op['name']!r} to active"
    dest = layout["nla"] + 1
    rfill._retry(ws.spreadsheet.batch_update, {"requests": [{
        "moveDimension": {"source": {"sheetId": sid, "dimension": "ROWS",
            "startIndex": src, "endIndex": src + 1}, "destinationIndex": dest}}]})
    return f"↓ moved {op['name']!r} to 'No longer active'"


def _org_fill_and_totals(ws, week_date, parsed):
    """Write the week's Unique Headcount for each active leader, then recompute
    TOTALS = column sum across all leader rows."""
    grid = rfill._retry(ws.get_all_values)
    layout = resolve_org_layout(grid)
    wk = layout["we_cols"].get(we_label(week_date))
    a0, a1 = layout["active"]
    updates = []
    if wk is not None:
        for r in range(a0, a1 + 1):
            nm = (grid[r][0] if grid[r] else "").strip()
            if not nm:
                continue
            rec = parsed.get(norm_name(nm))
            updates.append({"range": rowcol_to_a1(r + 1, wk + 1),
                            "values": [[rec["headcount"] if rec else ""]]})
    if updates:
        rfill._retry(ws.batch_update, updates, value_input_option="USER_ENTERED")
    if layout["totals"] is None:
        return
    grid = rfill._retry(ws.get_all_values)
    layout = resolve_org_layout(grid)
    tot = layout["totals"]
    tot_updates = []
    for _lbl, c in layout["we_cols"].items():
        s = 0
        for r in range(layout["header"] + 1, tot):
            if r == layout["nla"]:
                continue
            v = grid[r][c] if c < len(grid[r]) else ""
            try:
                s += int(float(v))
            except (TypeError, ValueError):
                pass
        tot_updates.append({"range": rowcol_to_a1(tot + 1, c + 1), "values": [[s]]})
    if tot_updates:
        rfill._retry(ws.batch_update, tot_updates, value_input_option="USER_ENTERED")


def _org_sort_collapse(ws, sid):
    grid = rfill._retry(ws.get_all_values)
    layout = resolve_org_layout(grid)
    ncols = max(len(r) for r in grid)
    a0, a1 = layout["active"]
    inact = layout["inactive"]
    # Active = greatest→least by the latest week's headcount (name as tiebreak).
    latest_col = max(layout["we_cols"].values()) if layout["we_cols"] else 0
    reqs = [{"sortRange": {"range": {"sheetId": sid, "startRowIndex": a0,
             "endRowIndex": a1 + 1, "startColumnIndex": 0, "endColumnIndex": ncols},
             "sortSpecs": [{"dimensionIndex": latest_col, "sortOrder": "DESCENDING"},
                           {"dimensionIndex": 0, "sortOrder": "ASCENDING"}]}}]
    grp = None
    if inact and inact[1] >= inact[0]:
        i0, i1 = inact
        grp = {"sheetId": sid, "dimension": "ROWS", "startIndex": i0, "endIndex": i1 + 1}
        for req in ([{"deleteDimensionGroup": {"range": grp}}],
                    [{"updateDimensionProperties": {"range": grp,
                      "properties": {"hiddenByUser": False}, "fields": "hiddenByUser"}}]):
            try:
                rfill._retry(ws.spreadsheet.batch_update, {"requests": req})
            except Exception:
                pass
        reqs.append({"sortRange": {"range": {"sheetId": sid, "startRowIndex": i0,
             "endRowIndex": i1 + 1, "startColumnIndex": 0, "endColumnIndex": ncols},
             "sortSpecs": [{"dimensionIndex": 0, "sortOrder": "ASCENDING"}]}})
    rfill._retry(ws.spreadsheet.batch_update, {"requests": reqs})
    if grp:
        try:
            rfill._retry(ws.spreadsheet.batch_update, {"requests": [
                {"addDimensionGroup": {"range": grp}},
                {"updateDimensionGroup": {"dimensionGroup": {"range": grp, "depth": 1,
                  "collapsed": True}, "fields": "collapsed"}}]})
        except Exception as e:
            print(f"  (inactive collapse skipped: {e})")


def update_week(ws, week_date, parsed: Dict[str, dict],
                dry_run: bool = True) -> dict:
    """Weekly upkeep: append the week column (inherit prior format), apply the
    active/inactive moves + new-leader adds, fill headcounts, recompute TOTALS,
    sort + collapse. Falls back to a fresh build if the tab is empty."""
    grid = rfill._retry(ws.get_all_values)
    if not grid or not (grid[0] and (grid[0][0] or "").strip()):
        return write_backfill(ws, [(week_date, parsed)], dry_run=dry_run)

    layout = resolve_org_layout(grid)
    label = we_label(week_date)
    sid = ws.id

    # 1) week column (chronological insert + format from the prior week).
    if label not in layout["we_cols"]:
        we = layout["we_cols"]
        dated = sorted(((d, c) for l, c in we.items()
                        if (d := we_label_to_date(l)) is not None), key=lambda x: x[0])
        wk = next((c for d, c in dated if d > week_date), None)
        if wk is None:
            wk = (max(we.values()) + 1) if we else 2
        if not dry_run:
            reqs = [{"insertDimension": {"range": {"sheetId": sid, "dimension": "COLUMNS",
                "startIndex": wk, "endIndex": wk + 1}, "inheritFromBefore": True}}]
            left = grid[layout["header"]][wk - 1] if wk - 1 < len(grid[layout["header"]]) else ""
            if wk >= 1 and (left or "").startswith("WE "):
                reqs.append({"copyPaste": {
                    "source": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": len(grid),
                               "startColumnIndex": wk - 1, "endColumnIndex": wk},
                    "destination": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": len(grid),
                                    "startColumnIndex": wk, "endColumnIndex": wk + 1},
                    "pasteType": "PASTE_FORMAT"}})
            rfill._retry(ws.spreadsheet.batch_update, {"requests": reqs})
            rfill._retry(ws.update, values=[[label]],
                         range_name=rowcol_to_a1(layout["header"] + 1, wk + 1),
                         value_input_option="USER_ENTERED")

    if dry_run:
        ops = _org_ops(grid, layout, parsed, label)
        return {"label": label, "ops": [f"{o['type']} {o['name']}" for o in ops]}

    # 2) structural moves/adds (fixpoint, re-read between each).
    done: List[str] = []
    for _ in range(120):
        g = rfill._retry(ws.get_all_values)
        lay = resolve_org_layout(g)
        ops = _org_ops(g, lay, parsed, label)
        if not ops:
            break
        done.append(_org_apply(ws, sid, g, lay, ops[0]))
    # 3) fill week + TOTALS, then 4) sort + collapse.
    _org_fill_and_totals(ws, week_date, parsed)
    _org_sort_collapse(ws, sid)
    return {"label": label, "ops": done}
