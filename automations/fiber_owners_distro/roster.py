"""Fetch + parse Kelly Pavone's weekly "AT&T ICD Owner Roster".

Kelly (kpavone@thesmartcircle.com) emails an xlsx roster every Friday. It lands in
raffi127@gmail.com. The workbook has four tabs — Fiber / Wireless / Hybrid / B2B —
each with an owner-per-row grid:

    ICD Company Name | ICD OWNER NAME | Email | Phone | HEADCOUNT | Address 1 | ...

We read it over IMAP (unattended, app-password auth) — same pattern as
automations.bg_check_sync.email_source, pointed at the same raffi127 mailbox.

Columns are found by HEADER LABEL, never by index — SEI reshuffles the sheet.
"""
from __future__ import annotations

import datetime as dt
import email
import imaplib
import os
import re
from email.header import decode_header
from pathlib import Path
from typing import Dict, List, Optional, Tuple

IMAP_HOST = "imap.gmail.com"
ACCOUNT = os.environ.get("FIBER_DISTRO_IMAP_ACCOUNT", "raffi127@gmail.com")
APP_PW_PATH = Path(os.environ.get(
    "FIBER_DISTRO_IMAP_APPPW_PATH",
    str(Path.home() / ".config" / "recruiting-report" / "gmail-app-password-raffi127")))

ROSTER_SENDER = "kpavone@thesmartcircle.com"
ROSTER_SUBJECT = "Owner Roster"          # IMAP SUBJECT substring match

# Which tabs feed which campaign. Raf 2026-07-26: Hybrid rolls up under Fiber (it
# is not its own campaign). Wireless (= NDS) and B2B are intentionally ignored.
FIBER_TABS = ("Fiber", "Hybrid")

EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")
# A single Email cell can hold several addresses ("a@x.com; b@y.com") and stray
# trailing punctuation ("c@z.com;"). Split on ; , and whitespace, clean each token.
_EMAIL_TOKEN = re.compile(r"^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$")


def canonical(email: str) -> str:
    """Identity key for comparing addresses. Gmail ignores dots and +tags in the
    local part (aron.corral13@gmail.com == aroncorral13@gmail.com), so canonicalize
    gmail/googlemail; leave every other domain's local part untouched."""
    e = (email or "").strip().lower()
    if "@" not in e:
        return e
    local, dom = e.rsplit("@", 1)
    if dom in ("gmail.com", "googlemail.com"):
        local = local.split("+", 1)[0].replace(".", "")
        dom = "gmail.com"
    return f"{local}@{dom}"


def extract_emails(cell) -> List[str]:
    """All valid, lowercased addresses in one Email cell (handles multi-email cells)."""
    out: List[str] = []
    if cell is None:
        return out
    for tok in re.split(r"[;,\s]+", str(cell)):
        tok = tok.strip().strip(".;,").lower()
        if tok and _EMAIL_TOKEN.match(tok):
            out.append(tok)
    return out


class Owner:
    __slots__ = ("company", "name", "email", "phone", "headcount", "tab")

    def __init__(self, company, name, email_, phone, headcount, tab):
        self.company = company
        self.name = name
        self.email = email_
        self.phone = phone
        self.headcount = headcount
        self.tab = tab

    def __repr__(self):
        return f"Owner({self.name!r}, {self.email!r}, tab={self.tab!r})"


def _app_password() -> str:
    if not APP_PW_PATH.exists():
        raise RuntimeError(
            f"Gmail app password not found at {APP_PW_PATH}. This is the app "
            f"password for {ACCOUNT} (Google Account -> Security -> App passwords).")
    return APP_PW_PATH.read_text(encoding="utf-8-sig").strip().replace(" ", "")


def _decode(s: str) -> str:
    return "".join(
        t.decode(enc or "utf-8", "replace") if isinstance(t, bytes) else t
        for t, enc in decode_header(s or ""))


def fetch_latest_roster(save_dir: Optional[Path] = None,
                        verbose: bool = True) -> Tuple[Path, dt.datetime]:
    """Download the most recent roster attachment. Returns (path, email_date)."""
    M = imaplib.IMAP4_SSL(IMAP_HOST)
    M.login(ACCOUNT, _app_password())
    try:
        M.select('"[Gmail]/All Mail"', readonly=True)
        typ, data = M.search(
            None, f'(FROM "{ROSTER_SENDER}" SUBJECT "{ROSTER_SUBJECT}")')
        ids = (data[0] or b"").split()
        if not ids:
            raise RuntimeError(
                f"No roster emails found from {ROSTER_SENDER} with subject "
                f"containing {ROSTER_SUBJECT!r} in {ACCOUNT}.")
        mid = ids[-1]                       # IMAP returns ascending order → last = newest
        typ, md = M.fetch(mid, "(RFC822)")
        msg = email.message_from_bytes(md[0][1])
        subject = _decode(msg.get("Subject", ""))
        edate = email.utils.parsedate_to_datetime(msg.get("Date"))
        if save_dir is None:
            save_dir = Path(__file__).resolve().parents[2] / "output" / "fiber_owners_distro"
        save_dir.mkdir(parents=True, exist_ok=True)
        out = None
        for part in msg.walk():
            if part.get_content_disposition() == "attachment":
                fn = _decode(part.get_filename() or "roster.xlsx")
                if not fn.lower().endswith((".xlsx", ".xlsm")):
                    continue
                out = save_dir / fn
                out.write_bytes(part.get_payload(decode=True))
                break
        if out is None:
            raise RuntimeError(f"Roster email {subject!r} had no .xlsx attachment.")
        if verbose:
            print(f"✓ roster: {subject!r} ({edate:%Y-%m-%d}) → {out}", flush=True)
        return out, edate
    finally:
        M.logout()


def _header_index(header_row, *labels) -> Optional[int]:
    """First column whose header matches any of `labels` (case-insensitive)."""
    want = {l.strip().lower() for l in labels}
    for i, cell in enumerate(header_row):
        if cell is not None and str(cell).strip().lower() in want:
            return i
    return None


def parse_tab(xlsx_path: Path, tab: str) -> List[Owner]:
    """Parse one tab into Owners, finding columns by header label."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        raise RuntimeError(f"Tab {tab!r} not in roster (tabs: {wb.sheetnames}).")
    ws = wb[tab]
    rows = [r for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    hdr = rows[0]
    ci_email = _header_index(hdr, "Email", "E-mail", "Email Address")
    ci_name = _header_index(hdr, "ICD OWNER NAME", "Owner Name", "Owner")
    ci_co = _header_index(hdr, "ICD Company Name", "Company", "Company Name")
    ci_phone = _header_index(hdr, "Phone", "Phone Number")
    ci_hc = _header_index(hdr, "HEADCOUNT", "Headcount")
    if ci_email is None:
        raise RuntimeError(
            f"Tab {tab!r}: no 'Email' column found. Headers: "
            f"{[str(c) for c in hdr]}")

    def g(row, idx):
        return "" if idx is None or idx >= len(row) or row[idx] is None else str(row[idx]).strip()

    owners: List[Owner] = []
    for row in rows[1:]:
        emails = extract_emails(g(row, ci_email))
        if not emails:
            continue                        # skip blank / malformed rows
        for em in emails:                   # one Owner per address (multi-email cells)
            owners.append(Owner(
                company=g(row, ci_co),
                name=g(row, ci_name),
                email_=em,
                phone=g(row, ci_phone),
                headcount=g(row, ci_hc),
                tab=tab,
            ))
    return owners


def fiber_owners(xlsx_path: Path) -> List[Owner]:
    """All Fiber + Hybrid owners, deduped by email (first tab wins for metadata)."""
    seen: Dict[str, Owner] = {}
    for tab in FIBER_TABS:
        for o in parse_tab(xlsx_path, tab):
            seen.setdefault(o.email, o)
    return list(seen.values())


def all_roster_emails(xlsx_path: Path) -> List[str]:
    """Every address on EVERY tab (Fiber/Wireless/Hybrid/B2B). Used to protect
    still-active owners from removal — we only ADD Fiber+Hybrid, but we never drop
    someone who's a current owner on any tab (e.g. a Wireless owner already on the
    list). Raf/Megan 2026-07-26."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    out: List[str] = []
    for tab in wb.sheetnames:
        out.extend(o.email for o in parse_tab(xlsx_path, tab))
    return out
