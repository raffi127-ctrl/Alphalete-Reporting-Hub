"""Fill per-day sale-type metrics (New INT, Upgrades, DTV, New Lines) from
the Tableau 'Sales By ICD (Weekly View).xlsx' export.

Tableau dashboards are painful to scrape directly (canvas/SVG render). Tableau
Online provides a built-in 'Download as Excel' option that gives us a clean
per-rep per-day per-product crosstab. This script parses that file and writes
to the matching cells on every owner tab.

Excel format (verified 2026-05-14):
  Row 1: ['Owner Name', 'Rep', 'Product Type (Broken Out)', 'Monday', 'Tuesday', ...]
  Row 2: 'Sales Total' summary row (skip)
  Per-owner blocks:
    First row: <Owner Name> | 'Total' | None | <day totals>      (skip — totals)
    Body rows: None | <Rep Name> | <PRODUCT TYPE> | <day counts>
      - Rep cell can be None on continuation rows (same rep, different product)

Product type → Sheet canonical metric (Raf-confirmed):
  NEW INTERNET     → New INT
  UPGRADE INTERNET → Upgrades
  VIDEO            → DTV
  WIRELESS         → New Lines

Run:
    .venv/bin/python -m automations.focus_office_att.step6_fill_tableau \\
        --file ~/Downloads/'Sales By ICD (Weekly View).xlsx'
    # Cody-only test:
    .venv/bin/python -m automations.focus_office_att.step6_fill_tableau \\
        --file <path> --only "Cody Cannon"
    # Dry-run (no writes):
    .venv/bin/python -m automations.focus_office_att.step6_fill_tableau \\
        --file <path> --dry-run
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

from automations.recruiting_report import fill as _fill
from automations.focus_office_att.aliases import (
    load_aliases, alias_to_canonical, get_search_candidates)
from automations.focus_office_att.columns import resolve_layout, _normalize
from automations.focus_office_att.step5_fill_one_owner import (
    _col_letter, design_cosmetic_ops, strip_rep_mark,
    TT_FIELD_TO_CANONICAL, DISP_FIELD_TO_CANONICAL,
)

DEST_SPREADSHEET_ID = "1xgVE_e8bZimACgPdqcdNCr1qo4sedWect_zzEcUgEJY"

PRODUCT_TO_METRIC = {
    "NEW INTERNET":     "New INT",
    "UPGRADE INTERNET": "Upgrades",
    "VIDEO":            "DTV",
    "WIRELESS":         "New Lines",
}

# Day name in Tableau → weekday index used by columns.py.
DAY_TO_WEEKDAY = {
    "Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
    "Friday": 4, "Saturday": 5, "Sunday": 6,
}


def _rows_from_xlsx(path: Path) -> list[list]:
    """Load an xlsx workbook's first sheet into a list of row lists."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[list] = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows


def _detect_csv_encoding_and_delim(path: Path) -> tuple[str, str]:
    """Sniff Tableau-exported CSVs. Tableau's default 'Crosstab CSV' is
    actually UTF-16 LE with a BOM and tab-delimited — not standard CSV.
    Falls back to utf-8 + comma if the BOM is absent (e.g., a hand-edited
    file)."""
    with open(path, "rb") as f:
        head = f.read(4)
    if head[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16", "\t"
    return "utf-8-sig", ","


def _rows_from_csv(path: Path) -> list[list]:
    """Load a CSV into a list of row lists. Auto-detects Tableau's
    UTF-16/tab format vs standard UTF-8/comma. Strips comma thousands
    separators from numeric strings so values like '1,705' parse as int."""
    encoding, delim = _detect_csv_encoding_and_delim(path)
    rows: list[list] = []
    with open(path, newline="", encoding=encoding) as f:
        reader = csv.reader(f, delimiter=delim)
        for raw in reader:
            row: list = []
            for cell in raw:
                s = (cell or "").strip()
                if s == "":
                    row.append(None)
                    continue
                # Strip thousands-separator commas before numeric attempt.
                # Tableau formats numbers like '1,705' / '12,345.67'.
                num_candidate = s.replace(",", "") if "," in s and not s.endswith(",") else s
                try:
                    row.append(int(num_candidate))
                    continue
                except ValueError:
                    pass
                try:
                    row.append(float(num_candidate))
                    continue
                except ValueError:
                    pass
                row.append(s)
            rows.append(row)
    return rows


def _parse_tableau_rows(rows: list[list]) -> dict:
    """Parser shared between xlsx + csv readers. Operates on a list of row
    lists (1-based row R = rows[R-1])."""
    if not rows:
        return {}
    headers = rows[0]
    day_cols: dict[int, int] = {}  # weekday_idx → 0-based col
    for col_idx, h in enumerate(headers):
        if isinstance(h, str) and h in DAY_TO_WEEKDAY:
            day_cols[DAY_TO_WEEKDAY[h]] = col_idx

    out: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))

    current_owner: str | None = None
    current_rep: str | None = None

    for row in rows[1:]:
        owner_cell = row[0] if len(row) > 0 else None
        rep_cell   = row[1] if len(row) > 1 else None
        prod_cell  = row[2] if len(row) > 2 else None

        # Global summary row — skip
        if owner_cell == "Sales Total":
            continue

        # XLSX has merged cells: owner_cell is None on body/continuation rows,
        # only set on the per-owner header (Total) row.
        # CSV has no merged cells: owner_cell + rep_cell are REPEATED on every
        # body row.
        # Unified handling: update current_owner when a NEW owner name appears;
        # treat repeated names as no-ops, not as new totals to skip.
        if owner_cell and owner_cell != current_owner:
            current_owner = owner_cell
            current_rep = None  # reset rep tracking for new owner

        # Per-owner Total row (XLSX: prod=None; CSV: prod="Total"). Detected
        # via rep_cell="Total" — that's stable across both formats.
        if rep_cell == "Total":
            continue

        # XLSX continuation rows: rep_cell is None → keep current_rep.
        # CSV: rep_cell is set on every body row → updates to same value (no-op).
        if rep_cell:
            current_rep = rep_cell

        if not current_owner or not current_rep or not prod_cell:
            continue

        metric = PRODUCT_TO_METRIC.get(str(prod_cell).strip().upper())
        if not metric:
            continue

        for wd, col in day_cols.items():
            val = row[col] if col < len(row) else None
            if isinstance(val, (int, float)) and val:
                out[current_owner][current_rep][wd][metric] += int(val)

    return {
        owner: {
            rep: {wd: dict(metrics) for wd, metrics in days.items()}
            for rep, days in reps.items()
        }
        for owner, reps in out.items()
    }


def parse_tableau_xlsx(path: Path) -> dict:
    """Returns {owner_name: {rep_name: {weekday_idx: {metric: count}}}}.

    Dispatches by file extension — supports both .xlsx (Crosstab Excel
    download) and .csv (Crosstab CSV download). The Tableau Crosstab
    download has the same row/column shape regardless of format.

    Skips Sales Total row + per-owner Total rows. Handles continuation
    rows (rep cell is None on second/third product line for the same rep).
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _rows_from_csv(path)
    elif suffix in (".xlsx", ".xlsm"):
        rows = _rows_from_xlsx(path)
    else:
        raise ValueError(f"Unsupported file extension {suffix!r} (want .csv or .xlsx)")
    return _parse_tableau_rows(rows)


def fill_tableau_for_owner(ws, owner_data: dict, layout, dry_run: bool = False) -> dict:
    """owner_data is {rep_name: {weekday_idx: {metric: count}}}.

    Reps that appear in Tableau but not in this owner's Sheet tab get
    APPENDED as new rep rows. Their production cells get filled with the
    Tableau data; their activity cells stay empty and get 'x' filled in
    by the downstream apply_empty_cell_defaults step.

    Returns stats with: written, unmatched_reps (empty if all matched),
    new_reps (list of names appended).
    """
    rep_col_vals = ws.col_values(layout.rep_name_col)
    # Build {lowercase_rep_name: row}. Skip the OFFICE TOTALS row.
    # Strip the Tableau-only marker emoji when keying so a previously-
    # marked rep still matches their Tableau name on the next run.
    #
    # SCOPE TO THE CURRENT (top) ZONE: stop at the frozen 'LAST WEEK'
    # label. Without this, a rep present in both weeks maps to their
    # frozen-block row (last occurrence wins) and this week's production
    # gets written INTO last week's frozen block, corrupting it.
    _lw_row = next(
        (i for i, v in enumerate(rep_col_vals, start=1)
         if isinstance(v, str) and v.strip().upper() == "LAST WEEK"),
        None,
    )
    sheet_reps: dict[str, int] = {}
    for i, name in enumerate(rep_col_vals, start=1):
        if i < 3 or not name or not name.strip():
            continue
        if _lw_row is not None and i >= _lw_row:
            break  # reached the frozen LAST WEEK block — never match into it
        if name.strip().upper() == "OFFICE TOTALS":
            continue
        sheet_reps[strip_rep_mark(name).lower().strip()] = i

    # AUTHORITATIVE fill so the office totals ALWAYS match Tableau (Megan
    # 2026-06-19). Clear every current-zone sale cell for the EXISTING reps
    # first, then overlay the Tableau values — so a rep who sold last week but
    # 0 this week (or any stale leftover) goes to 0 instead of keeping the old
    # number. The old fill only overwrote NON-zero, which silently inflated
    # totals. SAFE BY DESIGN: only owners WITH a Tableau block reach this
    # function (step6 loops the export), so an owner missing from a partial /
    # failed pull is never cleared — a bad pull can't wipe a whole tab to zero.
    SALE_METRICS = ("New INT", "Upgrades", "DTV", "New Lines")
    cells: dict[str, object] = {}
    for row in list(sheet_reps.values()):
        for wd_cols in layout.day_cols.values():
            for m in SALE_METRICS:
                c = wd_cols.get(m)
                if c:
                    cells[f"{_col_letter(c)}{row}"] = ""   # clear stale, overlay below

    written = 0
    new_reps: list[str] = []
    # Track next available row so multiple new reps in one call don't
    # collide. Starts at the row after the last existing rep.
    next_new_row = (max(sheet_reps.values()) + 1) if sheet_reps else 3

    for tableau_rep, days in owner_data.items():
        key = tableau_rep.lower().strip()
        row = sheet_reps.get(key)
        if row is None:
            # Unmatched — append as new rep row. Write the rep name into
            # the rep_name_col; production cells get filled in below; the
            # downstream apply_empty_cell_defaults will write 'x' into any
            # leftover empty activity cells.
            row = next_new_row
            next_new_row += 1
            sheet_reps[key] = row
            new_reps.append(tableau_rep)
            cells[f"{_col_letter(layout.rep_name_col)}{row}"] = tableau_rep

        for wd, metric_counts in days.items():
            if wd not in layout.day_cols:
                continue
            wd_cols = layout.day_cols[wd]
            for metric, count in metric_counts.items():
                col = wd_cols.get(metric)
                if col is None:
                    continue
                cells[f"{_col_letter(col)}{row}"] = int(count)   # overlay wins over the clear
                written += 1

    if cells and not dry_run:
        data = [{"range": f"'{ws.title}'!{a1}", "values": [[v]]} for a1, v in cells.items()]
        ws.spreadsheet.values_batch_update({"valueInputOption": "USER_ENTERED", "data": data})

    return {"written": written, "unmatched_reps": [], "new_reps": new_reps}


def main() -> int:
    # Pace every Sheets call under Google's quota so Phase 3 never
    # 429-storms (and the 65s retry never has to fire).
    from automations.focus_office_att._ratelimit import install as _install_pacing
    _install_pacing()
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Path to the Tableau xlsx download.")
    ap.add_argument("--only", default="",
                    help="Comma-separated owner tab names to fill (rest skipped).")
    ap.add_argument("--dry-run", action="store_true", help="Parse + report; no Sheet writes.")
    args = ap.parse_args()

    path = Path(args.file).expanduser()
    if not path.exists():
        print(f"❌ File not found: {path}")
        return 1
    print(f"Parsing {path.name}…")
    tableau = parse_tableau_xlsx(path)
    print(f"  ✓ {len(tableau)} owner(s) in file.")

    sh = _fill._client().open_by_key(DEST_SPREADSHEET_ID)
    all_tabs = {t.title: t for t in sh.worksheets()}
    only = {s.strip() for s in args.only.split(",") if s.strip()}

    # Tableau-name → Sheet-tab-name alias (e.g. "Tony Chavez" → "Jose Antonio Chavez").
    # Load owner aliases from the shared 'ICD Aliases' Sheet tab. Returns
    # {canonical: [aliases]} — we use alias_to_canonical() for the reverse
    # lookup (Tableau name → Sheet tab name).
    aliases_raw = load_aliases()
    if aliases_raw:
        total = sum(len(v) for v in aliases_raw.values())
        print(f"  Loaded {total} alias(es) for {len(aliases_raw)} ICD(s) from shared Sheet")

    def _resolve_tab(owner: str) -> str:
        """Tableau owner name → actual Sheet tab name. The shared ICD Aliases
        sheet's canonical may be EITHER form: the DD report names its tabs with
        the canonical (Tableau) name, but this Daily report names them with the
        alias spelling (tab 'Lamar Mitchell' for Tableau 'Tre Mitchell', etc.).
        So resolve to the canonical, then pick whichever candidate name
        (canonical OR any alias) is a real tab here. Falls back to the canonical
        when none match, so the 'no Sheet tab' report still reads sensibly."""
        canonical = alias_to_canonical(owner, aliases_raw)
        for cand in get_search_candidates(canonical, aliases_raw):
            if cand in all_tabs:
                return cand
        return canonical

    summary: dict[str, dict] = {}
    metrics_for_layout = (
        list(TT_FIELD_TO_CANONICAL.values())
        + list(DISP_FIELD_TO_CANONICAL.values())
        + ["Total Apps", "New INT", "Upgrades", "DTV", "New Lines"]
    )

    def _is_quota_error(e: Exception) -> bool:
        return "429" in str(e) and "Quota exceeded" in str(e)

    def _process_owner_with_retry(owner, owner_data, sheet_tab_name, ws):
        """Process one owner; on 429 quota, sleep 65s and retry once."""
        for attempt in (1, 2):
            try:
                layout = resolve_layout(ws, metrics=metrics_for_layout, interactive=False)
                stats = fill_tableau_for_owner(ws, owner_data, layout, dry_run=args.dry_run)
                if stats.get("new_reps"):
                    print(f"    + added {len(stats['new_reps'])} new rep row(s): {stats['new_reps']}")
                if not args.dry_run and stats["written"] > 0:
                    # SPEED (Megan 2026-06-07): Phase 2 already applied each
                    # tab's full design (cosmetic + structure), formatting
                    # persists day-to-day, the live '=' formulas + persistent
                    # conditional-format rules auto-recompute when we write
                    # production, and Phase 2 skips design on no-change owners
                    # anyway. So in Phase 3 the ONLY thing that must re-run for
                    # an unchanged owner is the STATIC aggregate rows that are
                    # computed from the production we just wrote (office totals
                    # + summary, written RAW). Run the FULL pass only when a new
                    # rep row was appended (then sort/format/formulas must be
                    # re-laid). This cuts ~16 ops/owner → 2 for most owners,
                    # which is what was blowing the Phase-3 time cap.
                    # Escape hatch: FOCUS_PHASE3_FULL_DESIGN=1 forces the old
                    # full pass on every owner.
                    ops = design_cosmetic_ops(ws, layout)
                    if (not stats.get("new_reps")
                            and not os.environ.get("FOCUS_PHASE3_FULL_DESIGN")):
                        _keep = {"write_office_totals_row",
                                 "write_office_summary_block"}
                        ops = [(l, f) for l, f in ops if l in _keep]
                    for label, fn in ops:
                        try:
                            fn()
                        except Exception as e:
                            # On a Sheets read-quota 429, wait out the
                            # per-minute window and retry once. Without this
                            # the design op is silently skipped and the tab
                            # ends up half-styled — expanded columns, missing
                            # borders (happened to Jose/Sam/Trang on the
                            # 2026-05-15 run).
                            if _is_quota_error(e):
                                print(f"    ⏳ quota hit on {label} — sleeping 65s + retrying once")
                                time.sleep(65)
                                try:
                                    fn()
                                    continue
                                except Exception as e2:
                                    e = e2
                            print(f"    ⚠ {label} failed (ignoring): {type(e).__name__}: {e}")
                return stats
            except Exception as e:
                if _is_quota_error(e) and attempt == 1:
                    print(f"    ⏳ Sheets read quota hit on {sheet_tab_name} — sleeping 65s + retrying once")
                    time.sleep(65)
                    continue
                raise

    for owner, owner_data in tableau.items():
        # Tableau name → actual Sheet tab name (handles tabs named with either
        # the canonical or an alias spelling — see _resolve_tab).
        sheet_tab_name = _resolve_tab(owner)
        if only and sheet_tab_name not in only:
            continue
        if sheet_tab_name not in all_tabs:
            summary[owner] = {"status": "no Sheet tab — skipped"}
            continue
        ws = all_tabs[sheet_tab_name]
        label = f"{owner}" + (f" → tab {sheet_tab_name!r}" if sheet_tab_name != owner else "")
        print(f"  → {label}…")
        try:
            stats = _process_owner_with_retry(owner, owner_data, sheet_tab_name, ws)
        except Exception as e:
            print(f"    ✗ {type(e).__name__}: {e}")
            summary[owner] = {"status": f"error: {type(e).__name__}", "written": 0,
                               "unmatched_reps": [], "new_reps": []}
            continue
        summary[owner] = {
            "status": "ok",
            "written": stats["written"],
            "unmatched_reps": stats.get("unmatched_reps", []),
            "new_reps": stats.get("new_reps", []),
        }
        verb = "would write" if args.dry_run else "wrote"
        print(f"    ✓ {verb} {stats['written']} cell(s)" + (
            f"; {len(stats['unmatched_reps'])} unmatched rep(s)" if stats["unmatched_reps"] else ""))
        # No per-tab sleep — the process-wide Sheets rate limiter
        # (_ratelimit, installed in main) paces every call under quota.

    print()
    print("=== SUMMARY ===")
    for owner, s in summary.items():
        if s["status"] != "ok":
            print(f"  • {owner}: {s['status']}")
            continue
        print(f"  ✓ {owner}: {s['written']} cell(s) written" + (
            f"; unmatched reps: {s['unmatched_reps']}" if s["unmatched_reps"] else ""))

    # Record which Sheet owner tabs are MISSING from the Tableau export —
    # an owner with a tab but no Tableau data means we can't pull their
    # sale-type numbers. The Hub surfaces these (alongside the ownerville
    # access gaps) in the report's review email so access can be re-pinged.
    _NON_OWNER_TABS = {"Template", "Raf play"}
    sheet_owner_tabs = {t for t in all_tabs if t not in _NON_OWNER_TABS}
    in_tableau = {_resolve_tab(o) for o in tableau}
    missing_from_tableau = sorted(sheet_owner_tabs - in_tableau)
    try:
        results_path = (Path(__file__).resolve().parents[2]
                        / "output" / "focus_office_tableau_results.json")
        results_path.parent.mkdir(parents=True, exist_ok=True)
        results_path.write_text(json.dumps({
            "run_at": dt.datetime.now().isoformat(timespec="seconds"),
            "owners_in_tableau": sorted(in_tableau),
            "missing_from_tableau": missing_from_tableau,
        }, indent=2))
        print(f"  Tableau coverage: {len(missing_from_tableau)} owner tab(s) "
              f"missing from the export.")
    except Exception as e:
        print(f"  ⚠ couldn't write Tableau results file: {e}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
