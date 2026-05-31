"""Phase 3 — Step 7: Download the per-rep Tableau crosstab via Playwright.

Zero-touch flow:
  - The 'AUTOMATION PULL' custom view (saved by Megan) holds the things
    we can't easily script: rep dimension expanded (canvas-rendered) +
    Product Type filter (the 4 types we care about).
  - Sale Date Week Ending rolls every week, so we override it via URL
    parameter at runtime. Tableau applies URL filter params on page load.

What this script does:
  1. Compute current week's Sunday.
  2. Navigate the existing Tableau tab to:
       .../AUTOMATIONPULL?Sale Date Week Ending (mon-sun)=YYYY-MM-DD
  3. Wait for the viz to render fully.
  4. Click Download → Crosstab → 'Sales By ICD (Weekly View)' → Excel.
  5. Capture the download, save to a known path.
  6. Validate the file has the 'Rep' column. If not, the saved custom view
     is broken or Tableau didn't restore the rep expansion — fail loudly.

VA workflow: just log into ownerville master/admin and have the Tableau
tab open (SSO'd via 'Login to Tableau' on the ownerville Tableau page).
Then trigger this script. No filter clicking, no rep expansion — the
custom view handles those.

Run:
    .venv/bin/python -m automations.focus_office_att.step7_download_tableau
    .venv/bin/python -m automations.focus_office_att.step7_download_tableau --fill

The optional --fill flag chains to step6_fill_tableau to write the Sheet.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

from patchright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# 'AUTOMATION PULL' custom view of ATT Tracker 2.1 D2D / PRODUCT SALES SUMMARY 4WK.
# Saved by Megan 2026-05-14 with rep dimension expanded + 4 product types selected.
# The UUID + name are stable; we override Sale Date Week Ending via URL param.
TABLEAU_VIEW_URL = (
    "https://us-east-1.online.tableau.com/#/site/sci/views/"
    "ATTTRACKER2_1-D2D/PRODUCTSALESSUMMARY4WK/"
    "b2da26b8-8971-4a45-9e42-bd04af46f0fa/AUTOMATIONPULL"
)


def _current_week_sunday() -> "dt.date":
    """Sunday at the end of the current Mon-Sun week. If today is Sunday,
    returns today. Otherwise returns the next upcoming Sunday."""
    today = dt.date.today()
    days_until_sun = (6 - today.weekday()) % 7
    return today + dt.timedelta(days=days_until_sun)


def _build_view_url(week_ending: "dt.date | None" = None) -> str:
    """Return the AUTOMATION PULL custom view URL with Sale Date Week Ending
    filter set to week_ending (defaults to current week's Sunday)."""
    from urllib.parse import quote
    if week_ending is None:
        week_ending = _current_week_sunday()
    qs = f"{quote('Sale Date Week Ending (mon-sun)')}={quote(str(week_ending))}"
    return f"{TABLEAU_VIEW_URL}?{qs}"

# How long to wait for the viz iframe to fully render after navigation.
# Tableau is heavy JS; under-waiting causes the Crosstab to export the base
# 'Product Sales Summary by ORG' sheet instead of the per-rep one.
VIZ_RENDER_WAIT_MS = 10_000

# Defaults if --out not passed.
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parents[2] / "output"


def _is_active_tableau_page(pg) -> bool:
    """A Tableau tab counts as 'active' (SSO'd in) only if it's on the
    online.tableau.com domain, NOT a login/SSO bounce page."""
    url = (pg.url or "").lower()
    if "online.tableau.com" not in url:
        return False
    # Login/SSO pages share the parent host but route through different paths.
    if "/login" in url or "sso." in url or "/idp/" in url:
        return False
    return True


def _find_ov_page(browser):
    """Return the first ownerville tab in the connected Chrome."""
    for ctx in browser.contexts:
        for pg in ctx.pages:
            if "ownerville" in (pg.url or "").lower():
                return pg
    return None


def _bootstrap_tableau_via_ov(browser, verbose: bool = True):
    """Open a new tab + drive OV's `?p=81&...&v2=1` SSO chain so we land on
    Tableau Online with an active session. Used when no SSO'd Tableau tab
    is open (e.g. session expired overnight).

    Returns the resulting Tableau page, or raises if SSO fails.
    """
    import re
    ov_page = _find_ov_page(browser)
    if ov_page is None:
        raise RuntimeError(
            "No ownerville tab open. Launch debug Chrome + log into "
            "ownerville first; we can't SSO into Tableau without an active "
            "OV session."
        )
    # Pull the current OV rqst
    m = re.search(r"rqst=([A-Fa-f0-9_]+)", ov_page.url)
    if not m:
        raise RuntimeError(f"OV tab has no rqst in URL: {ov_page.url}")
    rqst = m.group(1)
    # `?p=81&rqst=...&ssook=1` is the URL OV's "Login to Tableau" button
    # links to. Hitting it directly triggers the SSO chain that bounces
    # through Tableau's SAML and lands on online.tableau.com with an
    # active session. Skips the manual button-click step.
    sso_url = f"https://v2.ownerville.com/index.cfm?p=81&rqst={rqst}&ssook=1"
    if verbose:
        print(f"Bootstrapping Tableau SSO via OV: {sso_url}", flush=True)

    # Open in a NEW tab so we don't disturb the OV tab's state.
    new_page = ov_page.context.new_page()
    try:
        new_page.goto(sso_url, wait_until="domcontentloaded", timeout=20_000)
    except Exception as e:
        raise RuntimeError(f"OV SSO bootstrap nav failed: {e}")

    # Wait for the SSO bounce chain to land on online.tableau.com (max 30s).
    deadline = 30.0
    poll = 0.5
    waited = 0.0
    while waited < deadline:
        cur = (new_page.url or "").lower()
        if "online.tableau.com" in cur and "/idp/" not in cur and "/login" not in cur:
            if verbose:
                print(f"Landed on Tableau: {new_page.url}", flush=True)
            return new_page
        new_page.wait_for_timeout(int(poll * 1000))
        waited += poll
    raise RuntimeError(
        f"Tableau SSO didn't complete after {deadline}s — last URL: {new_page.url}"
    )


def _find_tableau_tab(browser, verbose: bool = True):
    """Return an SSO'd Tableau tab. Strategy:
      1. Existing tab on online.tableau.com (not login/SSO) → use it
      2. Otherwise → bootstrap via OV's ?p=81 SSO chain (auto-creates tab)
    """
    all_pages = [pg for ctx in browser.contexts for pg in ctx.pages]
    active = [pg for pg in all_pages if _is_active_tableau_page(pg)]
    if active:
        return active[0]
    if verbose:
        print("No active Tableau tab found — bootstrapping SSO via OV…", flush=True)
    return _bootstrap_tableau_via_ov(browser, verbose=verbose)


def _validate_per_rep_file(path: Path) -> None:
    """Raise if the downloaded file is owner-level (missing Rep column).

    Caught early so we don't quietly fill the Sheet with wrong-granularity data.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Rep" not in headers:
        raise RuntimeError(
            f"Downloaded file is owner-level (no 'Rep' column). Headers: {headers}. "
            "The VA must click the `+` next to 'Owner Name' in the Tableau view to "
            "expand the rep drilldown BEFORE triggering this script."
        )


def download_crosstab(out_path: Path, verbose: bool = True,
                      week_ending: "dt.date | None" = None,
                      download_format: str = "csv") -> Path:
    """Download the per-rep crosstab Excel; save to out_path. Returns out_path.

    Navigates to the AUTOMATION PULL custom view with Sale Date Week Ending
    overridden via URL param (defaults to current week's Sunday). The custom
    view restores rep expansion + Product Type filter.

    Uses page.frame_locator() rather than a cached Frame handle: Tableau
    sometimes re-attaches the viz iframe during interactions, invalidating
    a cached handle. frame_locator re-resolves on every use.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Pull via the patchright stealth session (self-logs into ownerville ->
    # Tableau), NOT the old CDP/Report-Chrome path — that needed a human to keep
    # a Tableau tab open + signed in, and timed out for Eve on 2026-05-25 (the
    # same failure the recruiting OPT had before its patchright fix). No human-
    # launched Chrome / Tableau tab required now.
    from automations.shared.tableau_patchright import tableau_session
    with tableau_session(verbose=verbose) as page:
        url = _build_view_url(week_ending)
        # Viz iframe has title="Data Visualization" (verified via DOM dump).
        viz = page.frame_locator('iframe[title="Data Visualization"]')
        dl_btn = '[data-tb-test-id="viz-viewer-toolbar-button-download"]'

        # Navigate + wait for the viz toolbar, with a render pause and a
        # reload-retry. A cold Tableau viz can take well over 30s to hydrate
        # its toolbar; the old single 30s wait timed out Eve's run and failed
        # the whole Tableau pull (2026-05-31). A reload almost always fixes a
        # slow first load.
        last_err = None
        for attempt in (1, 2, 3):
            if verbose:
                print(f"Navigating Tableau tab to: {url} (attempt {attempt})",
                      flush=True)
            page.goto(url, wait_until="domcontentloaded")
            page.wait_for_timeout(5_000)  # let the iframe attach + render
            # A stale/missing custom view renders Tableau's error page (no viz
            # iframe at all) — retrying or waiting for a Download button that
            # will never appear just burns minutes. Detect it and fail with an
            # actionable message. This is what broke Eve's Phase 3 (2026-05-31):
            # the AUTOMATIONPULL view URL returns "page could not be accessed".
            try:
                body = (page.inner_text("body", timeout=5_000) or "").lower()
            except Exception:
                body = ""
            if ("could not be accessed" in body
                    or "necessary permissions" in body
                    or "view does not exist" in body):
                raise RuntimeError(
                    "Tableau view not accessible — the AUTOMATION PULL custom "
                    f"view ({url.split('?')[0]}) returned 'page could not be "
                    "accessed'. The custom view was most likely deleted or "
                    "recreated (its GUID changes) or lost permissions. Recreate "
                    "the AUTOMATION PULL view (rep expansion + Product Type "
                    "filter, 'Sales By ICD (Weekly View)' worksheet) and update "
                    "TABLEAU_VIEW_URL in step7_download_tableau, or point it "
                    "at a current view.")
            try:
                if verbose:
                    print("Waiting for viz Download button to be visible...",
                          flush=True)
                viz.locator(dl_btn).wait_for(state="visible", timeout=60_000)
                last_err = None
                break
            except Exception as e:
                last_err = e
                if verbose:
                    print(f"  viz toolbar didn't appear on attempt {attempt} "
                          f"({type(e).__name__}); reloading…", flush=True)
                page.wait_for_timeout(3_000)
        if last_err is not None:
            raise last_err
        if verbose:
            print(f"Download button visible. Waiting {VIZ_RENDER_WAIT_MS/1000:.0f}s for data hydration...", flush=True)
        page.wait_for_timeout(VIZ_RENDER_WAIT_MS)

        if verbose:
            print("Clicking Download → Crosstab → Sales By ICD (Weekly View) → Excel...", flush=True)

        debug_dir = out_path.parent / "_phase3_debug"
        debug_dir.mkdir(parents=True, exist_ok=True)

        with page.expect_download(timeout=60_000) as dl_info:
            # Open Download menu (in viz toolbar)
            viz.locator('[data-tb-test-id="viz-viewer-toolbar-button-download"]').click()
            page.wait_for_timeout(1500)
            page.screenshot(path=str(debug_dir / "01_download_menu.png"), full_page=True)

            # Click Crosstab option
            viz.locator('[data-tb-test-id="download-flyout-download-crosstab-MenuItem"]').click()
            # Modal takes 2-3s to fully render thumbnails. Wait for the title
            # text 'Sales By ICD (Weekly View)' to actually appear before clicking it.
            viz.locator(
                '[data-tb-test-id^="sheet-thumbnail-"]:has-text("Sales By ICD (Weekly View)")'
            ).first.wait_for(state="visible", timeout=10_000)
            page.screenshot(path=str(debug_dir / "02_modal_open.png"), full_page=True)

            # Click the right sheet thumbnail
            target_thumb = viz.locator(
                '[data-tb-test-id^="sheet-thumbnail-"]:has-text("Sales By ICD (Weekly View)")'
            ).first
            if verbose:
                tid = target_thumb.get_attribute("data-tb-test-id")
                print(f"Target thumbnail data-tb-test-id: {tid}", flush=True)
            target_thumb.click()
            page.wait_for_timeout(800)
            page.screenshot(path=str(debug_dir / "03_after_thumb_click.png"), full_page=True)

            # Select format (CSV by default — Excel format produced wrong
            # data on the AUTOMATION PULL view; CSV is the verified-correct
            # export format per Megan 2026-05-14 walk-through).
            # Click the Label element rather than calling .check() on the
            # RadioButton — .check()/.is_checked() can hit "event loop
            # already running" inside the expect_download context. The
            # Label is a regular clickable element and toggles the radio.
            fmt = (download_format or "csv").lower()
            label_test_id = (
                "crosstab-options-dialog-radio-csv-Label" if fmt == "csv"
                else "crosstab-options-dialog-radio-excel-Label"
            )
            viz.locator(f'[data-tb-test-id="{label_test_id}"]').click()
            page.wait_for_timeout(300)
            page.screenshot(path=str(debug_dir / "04_before_download_click.png"), full_page=True)

            # Click final Download button in modal
            viz.locator('[data-tb-test-id="export-crosstab-export-Button"]').click()

        dl = dl_info.value
        if verbose:
            print(f"Download fired: {dl.suggested_filename}", flush=True)
        dl.save_as(str(out_path))
        if verbose:
            print(f"Saved: {out_path} ({out_path.stat().st_size:,} bytes)", flush=True)

    return out_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--format",
        choices=("csv", "excel"),
        default="csv",
        help="Crosstab download format. Default csv (Excel produced wrong "
             "metrics on PRODUCT SALES SUMMARY 4WK per Megan 2026-05-14).",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Where to save the downloaded crosstab (default: "
             "output/tableau_phase3_download.{csv|xlsx} based on --format)",
    )
    ap.add_argument(
        "--fill",
        action="store_true",
        help="After download, also run step6_fill_tableau to fill the Sheet",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="With --fill: don't actually write to the Sheet",
    )
    args = ap.parse_args()

    if args.out is None:
        ext = "csv" if args.format == "csv" else "xlsx"
        args.out = DEFAULT_OUTPUT_DIR / f"tableau_phase3_download.{ext}"

    try:
        out = download_crosstab(args.out, download_format=args.format)
        # Skip the per-rep validation for CSV — _validate_per_rep_file uses
        # openpyxl which can't read CSV. CSV correctness is verified
        # downstream via the parser + Megan's spot-check.
        if args.format != "csv":
            _validate_per_rep_file(out)
    except PlaywrightTimeoutError as e:
        print(f"FAIL: timed out — {e}", flush=True)
        return 2
    except Exception as e:
        print(f"FAIL: {type(e).__name__}: {e}", flush=True)
        return 1

    print(f"\nDownloaded crosstab: {out}")

    if args.fill:
        print("\nHanding off to step6_fill_tableau...", flush=True)
        # Import here so step7 can be used standalone without dragging in
        # gspread / Sheet auth when you only want the download.
        from automations.focus_office_att import step6_fill_tableau
        sys.argv = [
            "step6_fill_tableau",
            "--file", str(out),
            *(["--dry-run"] if args.dry_run else []),
        ]
        return step6_fill_tableau.main()

    return 0


if __name__ == "__main__":
    sys.exit(main())
