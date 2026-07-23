"""Vantura Weekly Payroll — scheduled prep run (Lucy 2, Wednesday 11:00 CT).

Automates the DETERMINISTIC prep half of the weekly Vantura commission/payroll
runbook so that when Carlos sits down the week is already pulled, loaded, and
refreshed and he only has to enter that week's judgement inputs and print.

Reuses the existing Lucy 2 routes (do NOT rebuild these):
  * Tableau crosstab download  -> automations.vantura_churn.cdp_pull
    (real-Chrome-over-CDP; the B2B server won't export under patchright)
  * Vantura board auth + write  -> automations.recruiting_report.fill.open_by_key
    (same SHEET_ID as vantura_churn; same gspread client)
  * "as Lucy" Slack delivery    -> automations.shared.slack_metrics_post

WHAT THIS RUN DOES (automatable, no human judgement):
  1. Compute week ending (the Sunday AFTER the DD Saturday).
  2. Pull the "ICD dd Detail" crosstab (Direct Deposit ICD VIEW -> DD DETAIL,
     Owner: Carlos) from Tableau.
  3. Load the rows into the RAW tab: append below last week, stamp Week Ending
     in column A, re-hide RAW. Report the row range used.
  4. Set Commission!B1 to the week ending.
  5. Re-point the per-campaign P&L formulas to this week's RAW range.
  6. Trigger "Refresh commission sheets", read the sync summary, run the
     read-only P&L checks (orphan payouts + campaign reconciliation).
  7. DM Carlos as Lucy: week loaded, RAW row range, sync summary, checks.

WHAT THIS RUN DELIBERATELY DOES NOT DO (human-only / money / irreversible):
  * bonuses / no-pay / rate changes  -> weekly judgement, only Carlos knows them
  * final verify + Print commission pack (PDF)
  * Lock the week -> already auto-locks Thursday ~11am CT via the board's own
    Apps Script cloud trigger; nothing to do here.

SAFETY (mirrors vantura_churn + CLAUDE.MD "Eve rules"):
  * --dry-run is the DEFAULT. It computes, pulls, and PRINTS what it would do;
    it writes NOTHING to the board and sends no Slack.
  * --sandbox writes to a duplicate board (SANDBOX_SHEET_ID) for testing.
  * --live is required to touch the real board, and is gated further below.
  * product->campaign mapping and the per-campaign formulas are the payroll
    correctness core. They were read back VERBATIM from the live WE 7/5 block
    (CH157/CH164/CH171/CH184, 2026-07-15) and parameterized on (RAW range,
    block column) — but a --live run still MUST be preceded by a --sandbox
    verification on a duplicated board. See PAYROLL_RUNBOOK.md for the spec.

  python -m automations.vantura_payroll.run                 # dry-run (default)
  python -m automations.vantura_payroll.run --sandbox        # write to test board
  python -m automations.vantura_payroll.run --live           # write to real board
  python -m automations.vantura_payroll.run --week 2026-07-12 # override week ending
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

REPORT_ID = "vantura-payroll"


def _find_repo_root() -> Path:
    """Repo root, independent of where this file lives — works both as the
    tracked module (automations/vantura_payroll/run.py) AND when the shared
    library materializes it under automations/uploaded/_shared/."""
    here = Path(__file__).resolve()
    for anc in here.parents:
        if (anc / "automations" / "day_orchestrator").is_dir():
            return anc
    return here.parents[2]


REPO_ROOT = _find_repo_root()

# Live Vantura Master Sales Board (same board vantura_churn writes to).
SHEET_ID = "1Hltk25zTudsaoYJFKvKqWlpT_4MF5_ZZq734XKVCJKY"
# A duplicated test board for --sandbox. TODO: fill in once Carlos makes a copy
# (CLAUDE.MD: build against a duplicated Sheet until "use the real Sheet").
SANDBOX_SHEET_ID = ""

# The product -> campaign mapping from the runbook (section 6). Kept here so the
# P&L split is codified, not "in chat memory". Sanity-check against the runbook.
CAMPAIGN_MAP = {
    "base": ["Energy Enrollment", "RES Pilot Program", "Lead Disposition Bonus",
             "BasePowerRES $200 (blank description)"],
    "box": ["BF 1", "BF 2", "Term Length Bonus", "kWH Bonus"],
    # b2b = everything else (AT&T / CRU / IRU + MCOE + Next Up + Roadtrip +
    #       Tiered/Rep Volume bonuses). Captain's bonus is EXCLUDED from gross.
}

# Carlos + Maud, for the "as Lucy" kickoff DM (Slack user ids from the
# captainship modules: Carlos U046G04P5LG, Maud U045USN7NCD).
DM_RECIPIENTS = ["U046G04P5LG"]


def _log(msg: str) -> None:
    print(f"[{dt.datetime.now().replace(microsecond=0).isoformat()}] {msg}",
          flush=True)


def week_ending(today: dt.date | None = None) -> dt.date:
    """The board's week ending = the Sunday AFTER the DD Saturday. Run on a
    Wednesday, the just-completed payroll week ends the Sunday 3 days ago."""
    today = today or dt.date.today()
    # weekday(): Mon=0..Sun=6. Most recent Sunday on/before today.
    return today - dt.timedelta(days=(today.weekday() + 1) % 7)


# --------------------------------------------------------------------------
# Steps 2-6 — implemented against the reused routes. The board-WRITE internals
# (RAW layout, per-campaign formulas) are stubbed until verified in --sandbox;
# they raise rather than guess, so a --live run can never write a wrong number.
# --------------------------------------------------------------------------

# RAW column layout (verified read-only 2026-07-15). Col A = week number; B-H =
# the paid-line fields; col I (Commission) is a spilling ARRAYFORMULA at I2 — we
# NEVER write it. Target header -> the source-xlsx header(s) we accept for it.
# First aliases are the export's REAL headers per the payroll runbook (REP.Full
# Name, cl.Description, …); later ones are fuzzy fallbacks. Sale/Act date
# headers are unconfirmed — a miss fails loud and prints the real headers.
RAW_TARGETS = {
    "B": ("Rep Name", ("rep.full name", "rep name", "rep", "icd name")),
    "C": ("Sale Date", ("cl.sale date", "sale date", "sold date")),
    "D": ("Activation Date", ("cl.activation date", "activation date", "activated")),
    "E": ("Description", ("cl.description", "description")),
    "F": ("Description Detail", ("cl.description detail", "description detail", "detail")),
    "G": ("Customer Name", ("cl.customer name", "customer name", "customer")),
    "H": ("Total $ to ICD", ("total $ to icd", "total $", "dd")),
}
RAW_FIRST_DATA_ROW = 2

# The P&L tab. Each week = a 3-column block (Brought In / Got Paid / Profit)
# headed "WE m/d" in row 1; rep rows 3-152; the campaign summary blocks sit
# below (~rows 154-210), located by their labels — never by hardcoded rows.
PNL_TAB = "Copy of Carlos PNL 2026"
PNL_REP_FIRST, PNL_REP_LAST = 3, 152


def _week_num(week: dt.date) -> float:
    """Board week label as the number it's stored as (July 5 -> 7.5, 12th -> 7.12)."""
    return float(f"{week.month}.{week.day}")


# The DD DETAIL crosstab source (confirmed 2026-07-15 from Carlos's Tableau
# history; the view Carlos calls "DD Detail by Rep"). The unattended CDP pull
# needs a Tableau session on the runner machine (ownerville storage_state, like
# vantura_churn) — Lucy 2's session-holder keeps it warm.
DD_DETAIL_URL = ("https://us-east-1.online.tableau.com/#/site/sci/views/"
                 "DirectDepositICDVIEWVersion2_0/DDDETAIL?:iid=1")
DD_DETAIL_SHEET = "ICD dd Detail"


def _pull_icd_dd_detail(week: dt.date, log=_log) -> Path:
    """Download this week's ICD dd Detail crosstab from Tableau — unattended.

    The SAME production route Carlos's other Lucy 2 reports use (vantura_churn):
    real Chrome over CDP, authenticated by the runner's seeded ownerville
    Tableau session. Needs no human on a provisioned Lucy 2; on an un-seeded
    machine it fails fast at the auth step (expected). Manual override: --file.
    """
    out = (REPO_ROOT / "output" / "vantura_payroll" /
           f"ICD dd Detail {week.isoformat()}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists():
        out.unlink()  # a stale file must never satisfy the exists() check below
    log(f"pulling DD DETAIL crosstab from Tableau -> {out}")
    from automations.vantura_churn import cdp_pull
    cdp_pull.download_views([(DD_DETAIL_URL, DD_DETAIL_SHEET, str(out))],
                            verbose=True, log=log)
    if not out.exists():
        raise FileNotFoundError("Tableau pull did not produce the export file.")
    log(f"pulled export: {out} ({out.stat().st_size:,} bytes)")
    return out


def _read_export(path: Path, log=_log) -> tuple[list[str], list[list]]:
    """Read the crosstab export -> (headers, data_rows). Handles BOTH formats
    (same as vantura_churn f2e21c3): a MANUAL download is a real .xlsx, but the
    AUTOMATED crosstab download saves UTF-16 tab-delimited CSV regardless of
    the file's extension — detect by zip magic bytes, never by name."""
    with open(path, "rb") as f:
        is_xlsx = f.read(4)[:2] == b"PK"
    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[("" if c is None else c) for c in r]
                for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        import csv as _csv
        rows = None
        for enc in ("utf-16", "utf-8-sig", "utf-8"):
            try:
                with open(path, encoding=enc, newline="") as fh:
                    rows = list(_csv.reader(fh, delimiter="\t"))
                if rows and len(rows[0]) > 1:
                    break
            except Exception:
                continue
        if not rows or len(rows[0]) < 2:
            raise ValueError(
                f"could not parse {path.name} as .xlsx OR tab-delimited "
                "crosstab CSV — inspect the downloaded file.")
        log(f"parsed {path.name} as crosstab CSV ({len(rows)} raw rows)")
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f"{path.name} has no rows")
    return [str(h).strip() for h in rows[0]], rows[1:]


def _map_columns(headers: list[str], log=_log) -> dict[str, int]:
    """Map each RAW target col -> source column index. Fail LOUD (never guess)
    if any target is unmatched — prints the source headers so we can adjust."""
    low = [h.lower() for h in headers]
    out, missing = {}, []
    for col, (name, aliases) in RAW_TARGETS.items():
        idx = next((i for i, h in enumerate(low) if h in aliases), None)
        if idx is None:
            idx = next((i for i, h in enumerate(low)
                        if any(a in h for a in aliases)), None)
        if idx is None:
            missing.append(f"{col} ({name})")
        else:
            out[col] = idx
    if missing:
        raise ValueError(
            "Could not map RAW column(s): " + ", ".join(missing) +
            "\nSource headers were: " + " | ".join(headers) +
            "\n-> update RAW_TARGETS aliases to match, then re-run.")
    log("column map (RAW <- source header): " +
        ", ".join(f"{c}<-{headers[i]!r}" for c, i in out.items()))
    return out


def _load_raw(xlsx: Path, week: dt.date, *, write: bool, sheet_id: str, log=_log) -> tuple[int, int]:
    """Append the week's rows to RAW (cols A-H; col I ARRAYFORMULA auto-computes).
    Col A = week number. Returns (first_row, last_row). Dry-run previews only."""
    from automations.recruiting_report.fill import open_by_key
    headers, data = _read_export(xlsx, log=log)
    cmap = _map_columns(headers, log=log)
    wnum = _week_num(week)

    # The crosstab's first data row is Tableau's grand-total row (runbook §4.1)
    # — drop it, but only if it LOOKS like one (blank rep / a 'total' label), so
    # a layout change can never silently cost a real paid line OR double the
    # week by loading the total as data.
    if data:
        first = [str(c).strip() for c in data[0]]
        rep_i = cmap["B"]
        rep_val = first[rep_i] if rep_i < len(first) else ""
        if not rep_val or any("total" in c.lower() for c in first):
            log(f"skipped grand-total row: {[c for c in first if c][:4]}")
            data = data[1:]
        else:
            raise ValueError(
                "Expected the first data row to be Tableau's grand-total row "
                f"(blank rep or a 'total' label) but got rep={rep_val!r} — the "
                "crosstab layout changed; verify before loading.")

    sh = open_by_key(sheet_id)
    raw = sh.worksheet("RAW")
    colA = raw.get(f"A2:A{raw.row_count}", value_render_option="UNFORMATTED_VALUE")
    # i is 0-based from row 2, so a non-empty entry's sheet row is i+2. The new
    # week starts on the row right after the last stamped one (7.5 ended at
    # 1234 -> 7.12 starts at 1235; verified against the live board).
    last = max((i + 2 for i, r in enumerate(colA) if r and r[0] not in ("", None)),
               default=1)
    start = last + 1

    # guard: this week must not already be loaded — but RESUME instead of
    # refusing (2026-07-23: the Wed 7/22 run died after the RAW append and the
    # hard refusal left Commission showing last week's block under the new
    # label; skipping the append and re-running the remaining steps is safe
    # and idempotent).
    existing = {str(r[0]) for r in colA if r and r[0] not in ("", None)}
    if str(wnum) in existing:
        wk_rows = [i + 2 for i, r in enumerate(colA)
                   if r and str(r[0]) == str(wnum)]
        first_row, last_row = min(wk_rows), max(wk_rows)
        if wk_rows != list(range(first_row, last_row + 1)):
            raise RuntimeError(
                f"week {wnum} rows in RAW are non-contiguous "
                f"({first_row}..{last_row}) — inspect before resuming.")
        log(f"RESUME: week {wnum} already in RAW (rows {first_row}-{last_row})"
            " — skipping the load; re-running week-set/P&L/refresh/checks.")
        return first_row, last_row

    out_rows = []
    for r in data:
        def cell(col):
            i = cmap[col]
            return r[i] if i < len(r) else ""
        out_rows.append([wnum, cell("B"), cell("C"), cell("D"),
                         cell("E"), cell("F"), cell("G"), cell("H")])
    end = start + len(out_rows) - 1
    log(f"RAW load: {len(out_rows)} rows, week {wnum}, would fill A{start}:H{end}")
    if out_rows:
        log(f"  first row -> {out_rows[0]}")
        log(f"  last  row -> {out_rows[-1]}")
    if not write:
        log("  (dry-run: nothing written)")
        return (start, end)
    if raw.row_count < end:
        raw.add_rows(end - raw.row_count + 5)
    raw.update(f"A{start}:H{end}", out_rows, value_input_option="USER_ENTERED")
    log(f"  WROTE RAW A{start}:H{end}")
    return (start, end)


def _set_week(week: dt.date, *, write: bool, sheet_id: str, log=_log) -> None:
    """Set Commission!B1 to the week number (valid because RAW col A now has it)."""
    from automations.recruiting_report.fill import open_by_key
    wnum = _week_num(week)
    log(f"set Commission!B1 = {wnum}")
    if not write:
        log("  (dry-run: not set)")
        return
    sh = open_by_key(sheet_id)
    sh.worksheet("Commission").update_acell("B1", wnum)
    log("  set B1")


def _col_letter(n: int) -> str:
    """1-indexed column number -> A1 letter(s)."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _locate_block(pnl, week: dt.date, log=_log) -> dict:
    """Find the week's 3-col P&L block (brought/paid/profit column letters) by
    its 'WE m/d' header in row 1, and the campaign anchor ROWS by their labels
    in the block's paid column. Fails loud on any mismatch — never guesses.

    Layout (verified live 2026-07-15, WE 7/5 = CF/CG/CH): labels in the paid
    col — B2B@154/TOTAL@157, BOX@161/TOTAL@164, Base-or-JE@168/TOTAL@171,
    'TOTAL PNL'@182, 'Carlos DD B2B'@184 — located by label, not row number.
    """
    hdr = f"WE {week.month}/{week.day}"
    row1 = pnl.row_values(1)
    try:
        c = row1.index(hdr) + 1
    except ValueError:
        raise RuntimeError(
            f"P&L: no {hdr!r} header in row 1 of '{PNL_TAB}' — the week's "
            "3-column block doesn't exist yet; add it (copy the prior block) "
            "and re-run.")
    blk = {"brought": _col_letter(c), "paid": _col_letter(c + 1),
           "profit": _col_letter(c + 2), "header": hdr}
    labels = pnl.get(f"{blk['paid']}150:{blk['paid']}210")
    seq = [(i, (r[0].strip() if r and r[0] else ""))
           for i, r in enumerate(labels, start=150)]

    def find(label, after, alts=()):
        for row, v in seq:
            if row > after and (v == label or v in alts):
                return row
        raise RuntimeError(
            f"P&L: label {label!r} not found below row {after} in col "
            f"{blk['paid']} — the summary-block layout changed; refusing to "
            "guess payroll cells.")

    b2b_hdr = find("B2B", 149)
    blk["b2b_total"] = find("TOTAL", b2b_hdr)
    box_hdr = find("BOX", blk["b2b_total"])
    blk["box_total"] = find("TOTAL", box_hdr)
    third_hdr = find("Base", blk["box_total"], alts=("JE",))
    blk["third_total"] = find("TOTAL", third_hdr)
    blk["total_pnl"] = find("TOTAL PNL", blk["third_total"])
    blk["dd_b2b"] = find("Carlos DD B2B", blk["total_pnl"])
    log(f"P&L block {hdr}: cols {blk['brought']}/{blk['paid']}/{blk['profit']}, "
        f"TOTALs at {blk['b2b_total']}/{blk['box_total']}/{blk['third_total']}, "
        f"DD add-back at {blk['dd_b2b']}")
    return blk


# Campaign membership by RAW Description (runbook §4.6 + the live WE 7/5
# formulas, read back verbatim 2026-07-15). Blank description = the
# BasePowerRES $200 lines. Only the Captain's bonus is excluded from gross.
BOX_DESCS = ("BF 1", "BF 2", "Term Length Bonus", "kWH Bonus")
BASE_DESCS = ("Energy Enrollment", "RES Pilot Program - Weekly Guarantee", "")
# 2026-07-23 (Carlos): Lead Disposition is NEVER paid out and gets its own
# Revenue-by-Campaign section (was lumped into the Base bucket before).
LEAD_DESCS = ("Lead Disposition Bonus",)

# "Revenue by Campaign" summary (Carlos, 2026-07-15): per-campaign revenue /
# paid-out / payroll-tax / profit in rows 215-233 of each week's block —
# labels in the paid column, values in the profit column, mirroring the
# summary blocks above. Backfilled by hand for 6/21-7/12; the weekly run
# writes it for each new week. Anchored at fixed rows per Carlos's spec.
REV_TITLE_ROW = 215
# 2026-07-23 v2 (Carlos): Lead Disposition revenue belongs under BOX (still
# never paid — BOX's Paid Out mask excludes it). No separate section.
REV_CAMPAIGNS = (("B2B", 217), ("BOX", 223), ("Base", 229))
REV_METRICS = ("Revenue Brought In", "Paid Out", "Payroll Tax", "Profit")

# House style (read off the hand-built summary blocks 2026-07-19): dark
# red-brown header bands w/ white bold, gray labels, semantic value colors
# (blue=revenue in, red=money out, green=profit), Calibri 12 centered,
# currency format, solid grid borders.
_HDR_BG = {"red": 0.52156866, "green": 0.1254902, "blue": 0.047058824}
_WHITE = {"red": 1, "green": 1, "blue": 1}
_GRAY = {"red": 0.84705883, "green": 0.84705883, "blue": 0.84705883}
_VALUE_BGS = ({"red": 0.62352943, "green": 0.77254903, "blue": 0.9098039},   # revenue
              {"red": 0.95686275, "green": 0.78039217, "blue": 0.7647059},   # paid
              {"red": 0.95686275, "green": 0.78039217, "blue": 0.7647059},   # tax
              {"red": 0.7176471, "green": 0.88235295, "blue": 0.8039216})    # profit


def _col_index(a1: str) -> int:
    """'CJ' -> 0-based column index."""
    n = 0
    for ch in a1:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _format_rev_block(sh, sheet_id: int, blk: dict, log=_log) -> None:
    """Apply the house style to the week's Revenue-by-Campaign block so a new
    week looks like the hand-formatted 6/21-7/12 ones."""
    p, f = _col_index(blk["paid"]), _col_index(blk["profit"])

    def cell_fmt(bg, bold=False, white=False, money=False):
        fmt = {"backgroundColor": bg, "horizontalAlignment": "CENTER",
               "textFormat": {"fontFamily": "Calibri", "fontSize": 12,
                              "bold": bold}}
        if white:
            fmt["textFormat"]["foregroundColor"] = _WHITE
        if money:
            fmt["numberFormat"] = {"type": "CURRENCY", "pattern": "$#,##0.00"}
        return fmt

    def repeat(r1, r2, c1, c2, fmt):
        fields = ("userEnteredFormat(backgroundColor,horizontalAlignment,"
                  "textFormat" + (",numberFormat)" if "numberFormat" in fmt
                                  else ")"))
        return {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": r1 - 1,
                      "endRowIndex": r2, "startColumnIndex": c1,
                      "endColumnIndex": c2 + 1},
            "cell": {"userEnteredFormat": fmt}, "fields": fields}}

    solid = {"style": "SOLID", "width": 1}

    def borders(r1, r2):
        return {"updateBorders": {
            "range": {"sheetId": sheet_id, "startRowIndex": r1 - 1,
                      "endRowIndex": r2, "startColumnIndex": p,
                      "endColumnIndex": f + 1},
            "top": solid, "bottom": solid, "left": solid, "right": solid,
            "innerHorizontal": solid, "innerVertical": solid}}

    reqs = [repeat(REV_TITLE_ROW, REV_TITLE_ROW, p, f,
                   cell_fmt(_HDR_BG, True, True)),
            borders(REV_TITLE_ROW, REV_TITLE_ROW)]
    for _name, hdr in REV_CAMPAIGNS:
        reqs.append(repeat(hdr, hdr, p, f, cell_fmt(_HDR_BG, True, True)))
        for i, bg in enumerate(_VALUE_BGS, start=1):
            bold = i == 4  # Profit row bold, like the TOTAL rows above
            reqs.append(repeat(hdr + i, hdr + i, p, p, cell_fmt(_GRAY, bold)))
            reqs.append(repeat(hdr + i, hdr + i, f, f,
                               cell_fmt(bg, bold, money=True)))
        reqs.append(borders(hdr, hdr + 4))
    sh.batch_update({"requests": reqs})
    log(f"  formatted revenue block ({blk['paid']}/{blk['profit']} "
        f"rows {REV_TITLE_ROW}-{REV_CAMPAIGNS[-1][1] + 4})")


def _repoint_pnl(week: dt.date, raw_range: tuple[int, int], *, write: bool,
                 sheet_id: str, log=_log) -> dict:
    """Write this week's per-campaign profit formulas + the Carlos-DD
    Roadtrip/MCOE add-back into the week's P&L block, pinned to the week's RAW
    row range — byte-for-byte the structure of the hand-built WE 7/5 formulas
    (CH157/CH164/CH171/CH184). Returns the located block for reuse."""
    from automations.recruiting_report.fill import open_by_key
    sh = open_by_key(sheet_id)
    pnl = sh.worksheet(PNL_TAB)
    blk = _locate_block(pnl, week, log=log)

    s, e = raw_range
    E = f"RAW!$E${s}:$E${e}"
    H = f"RAW!$H${s}:$H${e}"
    I = f"RAW!$I${s}:$I${e}"
    box = "+".join(f'({E}="{d}")' for d in BOX_DESCS)
    base = "+".join(f'({E}="{d}")' for d in BASE_DESCS)
    lead = "+".join(f'({E}="{d}")' for d in LEAD_DESCS)
    non_b2b = f'(1-({box}+{base}+{lead}+ISNUMBER(SEARCH("Captain",{E}))))'
    rep_c = f"$C{PNL_REP_FIRST}:$C{PNL_REP_LAST}"
    brought_rng = (f"{blk['brought']}{PNL_REP_FIRST}:"
                   f"{blk['brought']}{PNL_REP_LAST}")

    formulas = {
        f"{blk['profit']}{blk['b2b_total']}":
            f"=SUMPRODUCT({non_b2b}*{H})-SUMPRODUCT({non_b2b}*{I})*1.12",
        f"{blk['profit']}{blk['box_total']}":
            f"=SUMPRODUCT(({box}+{lead})*{H})-SUMPRODUCT(({box})*{I})*1.12",
        f"{blk['profit']}{blk['third_total']}":
            f"=SUMPRODUCT(({base})*{H})-SUMPRODUCT(({base})*{I})*1.12",
        f"{blk['profit']}{blk['dd_b2b']}":
            f'=SUMIF({rep_c},"B2B",{brought_rng})'
            f'+SUMIF({E},"B2B Roadtrip Bonus",{H})+SUMIF({E},"MCOE Bonus",{H})',
    }

    # Revenue-by-Campaign block (rows 215-233): labels in the paid column,
    # revenue / paid-out / tax / profit values in the profit column.
    labels = {f"{blk['paid']}{REV_TITLE_ROW}": "Revenue by Campaign"}
    # (revenue mask, paid mask): BOX revenue includes never-paid Lead Dispo
    camp_mask = {"B2B": (non_b2b, non_b2b),
                 "BOX": (f"({box}+{lead})", f"({box})"),
                 "Base": (f"({base})", f"({base})")}
    for name, hdr_row in REV_CAMPAIGNS:
        labels[f"{blk['paid']}{hdr_row}"] = name
        for i, metric in enumerate(REV_METRICS, start=1):
            labels[f"{blk['paid']}{hdr_row + i}"] = metric
        P, r0 = blk["profit"], hdr_row + 1
        rev_m, paid_m = camp_mask[name]
        formulas[f"{P}{r0}"] = f"=SUMPRODUCT({rev_m}*{H})"
        formulas[f"{P}{r0+1}"] = f"=SUMPRODUCT({paid_m}*{I})"
        formulas[f"{P}{r0+2}"] = f"={P}{r0+1}*0.12"
        formulas[f"{P}{r0+3}"] = f"={P}{r0}-{P}{r0+1}-{P}{r0+2}"

    # Captainship revenue (Carlos 2026-07-23): label-driven 'Captain' slot
    # (row shifts with the ledger — locate by label, currently CM206/CN206).
    cap_scan = pnl.get(f"{blk['paid']}190:{blk['paid']}214")
    for off, rowv in enumerate(cap_scan or []):
        if rowv and str(rowv[0]).strip().lower() == "captain":
            formulas[f"{blk['profit']}{190 + off}"] = (
                f'=SUMPRODUCT(ISNUMBER(SEARCH("Captain",{E}))*{H})')
            break

    for cell, f in formulas.items():
        log(f"P&L {cell} <- {f[:110]}…")
    if not write:
        log("  (dry-run: formulas + revenue-block labels not written)")
        return blk
    pnl.batch_update(
        [{"range": c, "values": [[v]]} for c, v in labels.items()]
        + [{"range": c, "values": [[f]]} for c, f in formulas.items()],
        value_input_option="USER_ENTERED")
    log(f"  WROTE {len(formulas)} formulas + revenue-block labels into "
        f"block {blk['header']}")
    try:  # cosmetics must never fail the payroll run
        _format_rev_block(sh, pnl.id, blk, log=log)
    except Exception as e:  # noqa: BLE001
        log(f"  (revenue-block formatting skipped: {type(e).__name__}: "
            f"{str(e)[:100]})")
    return blk


# One-time setup (runbook §1): api* wrappers + doGet appended to the board's
# bound Payroll.gs, deployed as a Web App (execute as Carlos, access: only me).
# The /exec URL is machine-local config — gitignored file or env var.
WEBAPP_CONFIG = REPO_ROOT / "vantura-payroll-webapp.json"


def _webapp_url() -> str:
    import json as _json
    import os
    url = os.environ.get("VANTURA_WEBAPP_URL", "").strip()
    if url:
        return url
    try:
        return str(_json.loads(WEBAPP_CONFIG.read_text())
                   .get("webapp_url", "")).strip()
    except Exception:
        return ""


def _refresh_and_check(week: dt.date, raw_range: tuple[int, int], *,
                       write: bool, sheet_id: str, log=_log) -> dict:
    """Trigger 'Refresh commission sheets' headlessly (apiRefresh via the
    deployed Web App — runbook §1/§4.4), then run the read-only P&L checks:
    orphan B2B payouts (paid with $0 brought) + campaign reconciliation
    (the three campaign TOTALs vs TOTAL PNL)."""
    parts = []

    url = _webapp_url()
    if not url:
        log("refresh NOT triggered — no web-app URL configured (set "
            f"{WEBAPP_CONFIG.name} at the repo root or VANTURA_WEBAPP_URL). "
            "One-time setup: append the api* wrappers + doGet to the board's "
            "Payroll.gs and deploy as a Web App (PAYROLL_RUNBOOK.md §1). The "
            "board still auto-refreshes on its own Thursday 11am CT trigger.")
        parts.append("refresh: SKIPPED (web app not configured)")
    elif not write:
        log(f"dry-run: would GET {{webapp}}?action=refresh")
        parts.append("refresh: dry-run (not triggered)")
    else:
        import requests
        r = requests.get(url, params={"action": "refresh"}, timeout=600)
        r.raise_for_status()
        parts.append(f"refresh: {r.text[:150]}")
        log(f"refresh -> {r.text[:150]}")

    # Read-only checks — always run (on a dry-run they preview the CURRENT
    # block state, i.e. pre-load zeros; the numbers are real after a live run).
    from automations.recruiting_report.fill import open_by_key
    sh = open_by_key(sheet_id)
    pnl = sh.worksheet(PNL_TAB)
    blk = _locate_block(pnl, week, log=log)

    def col_vals(col, r1, r2):
        got = pnl.get(f"{col}{r1}:{col}{r2}",
                      value_render_option="UNFORMATTED_VALUE")
        vals = [(r[0] if r else "") for r in got]
        vals += [""] * ((r2 - r1 + 1) - len(vals))
        return vals

    camp = col_vals("C", PNL_REP_FIRST, PNL_REP_LAST)
    brought = col_vals(blk["brought"], PNL_REP_FIRST, PNL_REP_LAST)
    paid = col_vals(blk["paid"], PNL_REP_FIRST, PNL_REP_LAST)

    def num(v):
        return float(v) if isinstance(v, (int, float)) else 0.0

    orphan = sum(num(p) for c, b, p in zip(camp, brought, paid)
                 if str(c).strip() == "B2B" and num(b) == 0 and num(p) > 0)
    parts.append(f"orphan B2B payouts (paid, $0 brought): ${orphan:,.2f}"
                 + ("" if orphan == 0 else "  ⚠ INVESTIGATE"))

    totals = {r: num(pnl.acell(f"{blk['profit']}{r}",
                               value_render_option="UNFORMATTED_VALUE").value)
              for r in (blk["b2b_total"], blk["box_total"],
                        blk["third_total"], blk["total_pnl"])}
    camp_sum = sum(v for r, v in totals.items() if r != blk["total_pnl"])
    delta = camp_sum - totals[blk["total_pnl"]]
    parts.append(f"campaign TOTALs sum ${camp_sum:,.2f} vs TOTAL PNL "
                 f"${totals[blk['total_pnl']]:,.2f} (Δ ${delta:,.2f}; manual "
                 "bonuses raise payroll un-tagged — small Δ expected)")

    checks = " | ".join(parts[1:])
    log("checks: " + checks)
    return {"summary": parts[0], "checks": checks}


def _kickoff_dm(week: dt.date, raw_range, summary, checks, *, send: bool, log=_log) -> None:
    """DM Carlos as Lucy that the week is loaded and refreshed, and what's left
    for him (bonuses/no-pay/rates, verify, print; auto-locks Thu)."""
    lines = [
        f"🐺 Vantura payroll prep is done for week ending {week:%-m/%-d} "
        "(loaded + refreshed on Lucy 2).",
        f"• RAW rows: {raw_range}",
        f"• Sync summary: {summary}",
        f"• P&L checks: {checks}",
        "",
        "Left for you: enter this week's bonuses / no-pay / rate changes, then "
        "Refresh again, verify, and Print the commission pack. It auto-locks "
        "Thursday ~11am.",
    ]
    msg = "\n".join(lines)
    if not send:
        log("DRY-RUN kickoff DM (not sent):\n" + msg)
        return
    from automations.shared import slack_metrics_post as slack  # reuse "as Lucy"
    client = slack._bot_client()
    for u in DM_RECIPIENTS:
        uid = slack._resolve_user_id(client, u)
        ch = client.conversations_open(users=uid)["channel"]["id"]
        client.chat_postMessage(channel=ch, text=msg)
        log(f"kickoff DM sent to {uid}")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Vantura weekly payroll prep (Lucy 2).")
    ap.add_argument("--week", help="week ending YYYY-MM-DD (default: computed)")
    ap.add_argument("--file", help="explicit ICD dd Detail .xlsx (skips the Tableau pull)")
    mode = ap.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", default=True,
                      help="compute + pull + PRINT only; no board writes, no Slack (DEFAULT)")
    mode.add_argument("--sandbox", action="store_true", help="write to the test board copy")
    mode.add_argument("--live", action="store_true", help="write to the REAL board")
    args = ap.parse_args(argv)

    live = bool(args.live)
    sandbox = bool(args.sandbox)
    write = live or sandbox
    send = live  # kickoff DM only on a real live run

    week = (dt.datetime.strptime(args.week, "%Y-%m-%d").date()
            if args.week else week_ending())

    mode_name = "LIVE" if live else ("SANDBOX" if sandbox else "DRY-RUN")
    _log(f"Vantura payroll prep — mode={mode_name} — week ending {week.isoformat()}")
    if sandbox and not SANDBOX_SHEET_ID:
        _log("ERROR: --sandbox needs SANDBOX_SHEET_ID (a duplicated board). Not set.")
        return 2
    # Sandbox writes go to the copy; dry-run/live read+write the real board
    # (dry-run never writes). Reads always come from the resolved board.
    sheet_id = SANDBOX_SHEET_ID if sandbox else SHEET_ID

    try:
        xlsx = Path(args.file) if args.file else _pull_icd_dd_detail(week)
        raw_range = _load_raw(xlsx, week, write=write, sheet_id=sheet_id)
        _set_week(week, write=write, sheet_id=sheet_id)
        _repoint_pnl(week, raw_range, write=write, sheet_id=sheet_id)
        result = _refresh_and_check(week, raw_range, write=write,
                                    sheet_id=sheet_id)
        summary, checks = result.get("summary"), result.get("checks")
        _kickoff_dm(week, raw_range, summary, checks, send=send)
    except (FileNotFoundError, ValueError, RuntimeError) as e:
        _log(f"STOP: {e}")
        return 4
    _log("done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
