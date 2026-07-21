"""VZ+FTR Dual-Campaign Wireless Performance — an EMAIL-sourced country tracker.

Every other tracker in this module is a live Tableau view (see pages.py + the
Download→Image capture in capture.py). This one is different: it arrives as a
daily .xlsx attachment from Credico (vzreports@credicousa.com) into the reporting
inbox (alphaletereporting@gmail.com). We fetch the newest workbook, render its
"Combined - Current Week" tab to a clean PNG, and hand it to the SAME slack_post
pipeline as every other tracker — so it lands in the identical "Tableau Country
Trackers" thread, in every org channel, with its own header line.

No Tableau, no PDF, no LibreOffice: openpyxl reads the sheet, we build a styled
HTML table, and a short-lived headless chromium (patchright, the same browser the
Tableau reports use) screenshots it. Rendering runs in its OWN isolated browser so
it's independent of the Tableau session (a cold Tableau login can't block it) and
testable standalone.

Wired into run.py by the spec's `source: "email"` marker (pages.py); capture() has
the same signature as capture.capture_page so run.py dispatches to it transparently.
"""
from __future__ import annotations

import datetime as dt
import html as _html
import re
import tempfile
from pathlib import Path

from automations.shared import email_ingest
from automations.tableau_screenshots import capture as cap

# --- where the report comes from ---
EMAIL_FROM = "vzreports@credicousa.com"
SUBJECT_CONTAINS = "Dual-Campaign Wireless Performance"
FILENAME_GLOB = "Dual-Campaign Wireless Performance Report*VZ+FTR*.xlsx"
SHEET = "Combined - Current Week"

# Columns whose stored value is a 0..1 fraction to render as a percentage, and the
# float column to show with decimals. Everything else numeric renders as an int.
_PERCENT_HEADERS = {"Self Setup %", "Insurance %", "Wireless Attach"}
_AVG_HEADERS = {"Scoring Avg"}
# Left-aligned text columns (the rest right-align as numbers).
_TEXT_HEADERS = {"Primary Campaign", "Owner Name", "Office Name", "State", "City"}


def fetch_xlsx(dest: str | Path, *, since_days: int = 10, verbose: bool = True) -> Path | None:
    """Newest VZ+FTR workbook in the last `since_days`, or None if none has landed
    (partial-safe — the caller decides whether a miss is fatal)."""
    found = email_ingest.fetch_by_globs(
        EMAIL_FROM, [FILENAME_GLOB], dest,
        subject=SUBJECT_CONTAINS, since_days=since_days, verbose=verbose)
    return next(iter(found.values()), None)


def _report_date(name: str) -> "dt.date | None":
    """Pull the report date from the filename ('… - SCI - 20260717.xlsx')."""
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", name)
    if not m:
        return None
    try:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _fmt(header: str, v) -> str:
    """Render one cell to display text, formatting numbers by column."""
    if v is None:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if header in _PERCENT_HEADERS:
            return f"{v * 100:.1f}%"
        if header in _AVG_HEADERS:
            return f"{v:.2f}"
        if isinstance(v, float):
            return str(int(v)) if float(v).is_integer() else f"{v:.2f}"
        return str(v)
    return str(v)


def _campaign_class(v: str) -> str:
    s = (v or "").lower()
    if "frontier" in s and "verizon" in s:
        return "camp-dual"
    if "frontier" in s:
        return "camp-ftr"
    if "verizon" in s:
        return "camp-vz"
    return ""


def _build_html(rows: list, report_date: "dt.date | None") -> str:
    """rows = list of row tuples straight from the sheet (row 1..N). Row 4 (index 3)
    is the header; data starts at row 5 (index 4). Title is row 1."""
    title = ""
    for r in rows[:3]:
        if r and r[0]:
            title = str(r[0])
            break
    header = [("" if c is None else str(c)) for c in rows[3]]
    ncol = len(header)
    body = rows[4:]

    thead = "".join(
        f'<th class="{"num" if h not in _TEXT_HEADERS else "txt"}">{_html.escape(h)}</th>'
        for h in header)

    trs = []
    for i, r in enumerate(body):
        if not any(c is not None and str(c).strip() for c in r):
            continue
        cells = []
        for ci in range(ncol):
            h = header[ci]
            v = r[ci] if ci < len(r) else None
            txt = _fmt(h, v)
            cls = "num" if h not in _TEXT_HEADERS else "txt"
            if ci == 0:  # Primary Campaign chip
                chip = _campaign_class(str(v or ""))
                inner = f'<span class="chip {chip}">{_html.escape(txt)}</span>' if txt else ""
                cells.append(f'<td class="txt camp">{inner}</td>')
            else:
                cells.append(f'<td class="{cls}">{_html.escape(txt)}</td>')
        trs.append(f'<tr class="{"odd" if i % 2 else "even"}">' + "".join(cells) + "</tr>")

    # Portable date (no %-m/%-d — those are POSIX-only and the report must render
    # on Windows too).
    date_str = (f"{report_date.strftime('%a')} {report_date.month}/{report_date.day}/"
                f"{report_date.year}") if report_date else ""

    return f"""<!doctype html><html><head><meta charset="utf-8"><style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 24px; background: #ffffff;
         font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif; }}
  #card {{ display: inline-block; border: 1px solid #d0d7de; border-radius: 10px;
          overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  #head {{ background: #0b2545; color: #fff; padding: 14px 18px; }}
  #head .t {{ font-size: 20px; font-weight: 700; letter-spacing: .2px; }}
  #head .s {{ font-size: 13px; opacity: .85; margin-top: 3px; }}
  table {{ border-collapse: collapse; font-size: 13px; }}
  th, td {{ padding: 6px 10px; white-space: nowrap; border-bottom: 1px solid #eef1f4; }}
  thead th {{ background: #13315c; color: #fff; font-weight: 600; text-align: right;
             position: sticky; top: 0; }}
  thead th.txt {{ text-align: left; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; color: #1b2733; }}
  td.txt {{ text-align: left; color: #33404d; }}
  tr.odd td {{ background: #f7f9fb; }}
  tr.even td {{ background: #ffffff; }}
  td.camp .chip {{ display: inline-block; padding: 2px 8px; border-radius: 10px;
                  font-size: 11px; font-weight: 600; color: #fff; }}
  .camp-vz {{ background: #d52b1e; }}
  .camp-ftr {{ background: #d81e5b; }}
  .camp-dual {{ background: linear-gradient(90deg,#d52b1e 50%,#d81e5b 50%); }}
</style></head><body>
  <div id="card">
    <div id="head">
      <div class="t">{_html.escape(title or 'Dual-Campaign Wireless Performance — VZ+FTR')}</div>
      <div class="s">Combined · Current Week{(' · ' + _html.escape(date_str)) if date_str else ''}</div>
    </div>
    <table><thead><tr>{thead}</tr></thead><tbody>{''.join(trs)}</tbody></table>
  </div>
</body></html>"""


def _screenshot_html(html_str: str, out_path: Path, *, context=None) -> Path:
    """Render `html_str` and screenshot the #card element to `out_path`. Reuses a
    passed browser `context` (a tab in the Tableau session) if given, else launches
    its own isolated headless chromium at 2x for a crisp standalone/test render.

    With context=None this MUST NOT be called while another Playwright sync session
    is open — the sync API refuses to start a second one ("Playwright Sync API
    inside the asyncio loop") and the render dies. run.py therefore captures every
    email-sourced tracker BEFORE it opens the Tableau session."""
    def _shot(pg):
        pg.set_content(html_str, wait_until="load")
        pg.wait_for_timeout(400)
        el = pg.query_selector("#card")
        (el or pg).screenshot(path=str(out_path))

    if context is not None:
        pg = context.new_page()
        try:
            _shot(pg)
        finally:
            pg.close()
        return out_path

    from patchright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            ctx = browser.new_context(viewport={"width": 1600, "height": 1000},
                                      device_scale_factor=2)
            pg = ctx.new_page()
            _shot(pg)
        finally:
            browser.close()
    return out_path


def render_png(xlsx_path: str | Path, out_path: str | Path, *,
               context=None, verbose: bool = True) -> Path:
    """Load the 'Combined - Current Week' tab and render it to `out_path` PNG."""
    import openpyxl
    out_path = Path(out_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"{SHEET!r} tab not in {Path(xlsx_path).name} "
                           f"(has: {wb.sheetnames})")
    ws = wb[SHEET]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    if len(rows) < 5:
        raise RuntimeError(f"{SHEET!r} has only {len(rows)} row(s) — nothing to render")
    html_str = _build_html(rows, _report_date(Path(xlsx_path).name))
    _screenshot_html(html_str, out_path, context=context)
    if verbose:
        print(f"  ✓ rendered {SHEET} -> {out_path.name}", flush=True)
    return out_path


def capture(page, spec: dict, out_dir, force_crop=None, verbose: bool = True) -> Path:
    """Same signature as capture.capture_page so run.py can dispatch to it for an
    email-sourced tracker. Fetches the newest VZ+FTR xlsx and renders it; raises if
    no email has landed (run.py catches → flags this one tracker, others still post).
    `force_crop` is accepted (for a uniform call site) and ignored — there's no viz
    to crop, we render the sheet directly."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    tmp = Path(tempfile.gettempdir()) / "vzftr_tracker"
    xlsx = fetch_xlsx(tmp, verbose=verbose)
    if not xlsx:
        raise RuntimeError(
            f"no VZ+FTR email from {EMAIL_FROM} in the last 10 days — nothing to render")
    out_png = out_dir / f"{cap._sanitize(spec['title'])}.png"
    # Always render in email_tracker's OWN isolated 2x chromium (context=None),
    # never the passed Tableau session's context: device_scale_factor is a
    # context-level setting, so reusing the Tableau context would render at its
    # native scale and produce a lower-res image than the 2x we verified. A short-
    # lived headless chromium is cheap and keeps this fully decoupled from Tableau.
    render_png(xlsx, out_png, context=None, verbose=verbose)
    return out_png
